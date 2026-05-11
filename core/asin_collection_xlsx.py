"""ASIN 评论采集清单 → 2-sheet Excel 输出。"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.asin_collection_planner import OUT_COLUMNS, SELECTION_RULES

C_BLUE = 'FF1F3864'
C_YELLOW = 'FFFFF1B8'
C_GREEN = 'FFD4EDDA'
C_WHITE = 'FFFFFFFF'

THIN = Side(style='thin', color='FFB0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, row, col, text, bg=C_BLUE):
    c = ws.cell(row, col, text)
    c.font = Font(bold=True, color=C_WHITE, size=11)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER


def _val(ws, row, col, text, bold=False, bg=None, fg='FF333333', wrap=False, align='left'):
    c = ws.cell(row, col, text)
    c.font = Font(bold=bold, color=fg, size=10)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = BORDER


def write_asin_collection_xlsx(df: pd.DataFrame, output_path: str, category_name: str) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = '采集清单'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUT_COLUMNS))
    title_c = ws.cell(1, 1, f'{category_name} — 重点 ASIN 评论采集清单')
    title_c.font = Font(bold=True, size=14, color=C_BLUE)
    title_c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(OUT_COLUMNS))
    note_c = ws.cell(2, 1,
        '运营请按本清单去卖家精灵下载对应 ASIN 的评论 .xlsx 文件，'
        '上传后即可生成完整的 10-Sheet 选品评估报告。'
        '点击右侧 URL 可直接跳到对应评论页。')
    note_c.font = Font(italic=True, size=10, color='FF666666')
    note_c.fill = PatternFill('solid', fgColor=C_YELLOW)
    note_c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 36

    for ci, col in enumerate(OUT_COLUMNS, 1):
        _hdr(ws, 3, ci, col)
    ws.row_dimensions[3].height = 28

    for ri, (_, row) in enumerate(df.iterrows(), 4):
        for ci, col in enumerate(OUT_COLUMNS, 1):
            v = row[col]
            wrap = col in ('标题', '推荐理由')
            align = 'center' if col in ('优先级', 'ASIN', '评分', '评论数', '月销量', '月收入($)', '在售天数', 'BuyBox国家') else 'left'
            _val(ws, ri, ci, v, wrap=wrap, align=align)
        ws.row_dimensions[ri].height = 56

    widths = {
        '优先级': 6, 'ASIN': 13, '标题': 36, '品牌': 14, '月销量': 9, '月收入($)': 11,
        '评分': 7, '评论数': 8, 'BuyBox国家': 10, '在售天数': 9,
        '入选标签': 28, '推荐理由': 50, '建议下载评论数': 18,
        'Amazon评论页URL': 42, '卖家精灵评论页URL': 50,
    }
    for ci, col in enumerate(OUT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)

    ws.freeze_panes = 'B4'

    ws2 = wb.create_sheet('筛选规则说明')
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws2.cell(1, 1, '筛选规则（多标签累加打分，按总分取 Top 12-18）')
    t.font = Font(bold=True, size=13, color=C_BLUE)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26

    headers = ['标签', '命中规则', '上限', '分数', '为何选它']
    for ci, h in enumerate(headers, 1):
        _hdr(ws2, 2, ci, h)
    ws2.row_dimensions[2].height = 24

    for ri, rule in enumerate(SELECTION_RULES, 3):
        for ci, v in enumerate(rule, 1):
            wrap = ci in (2, 5)
            align = 'center' if ci in (3, 4) else 'left'
            _val(ws2, ri, ci, v, wrap=wrap, align=align)
        ws2.row_dimensions[ri].height = 36

    rule_widths = [16, 38, 8, 8, 38]
    for ci, w in enumerate(rule_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    extra_row = len(SELECTION_RULES) + 4
    ws2.merge_cells(start_row=extra_row, start_column=1, end_row=extra_row, end_column=5)
    ws2.row_dimensions[extra_row].height = 60
    extra = ws2.cell(extra_row, 1,
        '建议下载评论数：\n'
        '  - 命中"差评异常" → 优先全部 1-2★ 评论 + 最新 100 条好评（用于挖痛点）\n'
        '  - 命中"头部销量"/"评论高活跃" → 最新 200 条（保证样本厚度）\n'
        '  - 仅命中"新品代表"/"价格带代表" → 全部（评论本身不多）'
    )
    extra.font = Font(size=10, color='FF333333')
    extra.fill = PatternFill('solid', fgColor=C_GREEN)
    extra.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb.save(output_path)
