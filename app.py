"""
选品报告生成系统 - Flask Web服务 v2.0
完整版 - 移植原版 generate_led_report.py 所有分析逻辑
"""

import os
import uuid
import shutil
import logging
import time
from datetime import datetime
from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify
from werkzeug.utils import secure_filename

import pandas as pd
import sys
import io

# 顶层 logging 配置：让全部 INFO 级日志（包括 [BSR vision] / [image] 等）落到 server_err.log
# force=True 覆盖 Flask debug 模式可能预设的 handler
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    force=True,
)
log = logging.getLogger(__name__)

# 性能优化：pandas 默认用 calamine 引擎读 xlsx（Rust 实现，比默认 openpyxl 快 5-10 倍）
# 全局 monkey-patch 一次，后续 pd.read_excel / pd.ExcelFile 自动走 calamine，除非显式指定 engine
try:
    import python_calamine  # noqa: F401
    _orig_read_excel = pd.read_excel
    _orig_ExcelFile = pd.ExcelFile

    def _fast_read_excel(*args, **kwargs):
        kwargs.setdefault("engine", "calamine")
        return _orig_read_excel(*args, **kwargs)

    class _FastExcelFile(_orig_ExcelFile):
        def __init__(self, *args, **kwargs):
            kwargs.setdefault("engine", "calamine")
            super().__init__(*args, **kwargs)

    pd.read_excel = _fast_read_excel
    pd.ExcelFile = _FastExcelFile
except ImportError:
    pass  # 没装 calamine 就退回默认 openpyxl

# openpyxl 样式类
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ========== 主图下载工具（BSR TOP100 / 竞品分析嵌入图片用） ==========
# 单例线程池 + 内存缓存，避免重复下载相同 URL（同一 ASIN 可能在多个 sheet 出现）
# 注意：Windows + Python 默认 SSL 与 amazon CDN 协商可能 UNEXPECTED_EOF（SChannel/OpenSSL 问题），
# 必须用 requests + urllib3 自定义 SSLContext 强制重建 TLS 配置，否则全部下载失败。
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
import threading

_IMG_CACHE: dict[str, BytesIO | None] = {}
_IMG_FAIL_COUNT: dict[str, int] = {}     # URL 失败计数：>= 3 才彻底放弃，避免单次抖动永久放弃
_IMG_MAX_FAIL = 3
_HTTP_SESSION = None

# 跨请求共享的图片下载信号量：单请求 ThreadPool 上限 8，但若两个浏览器标签并发，
# 两个 pool 之和会到 16，触发 Amazon CDN 限流。该信号量把"真正在网"的下载钳到 8。
# 缓存命中不占信号量（命中早退在 acquire 之前）。
_IMG_FETCH_SEM = threading.Semaphore(8)


def _get_http_session():
    """复用 requests.Session 避免重建 TLS。修复 amazon-images CDN 在 Windows 上的 SSL EOF 问题。"""
    global _HTTP_SESSION
    if _HTTP_SESSION is not None:
        return _HTTP_SESSION
    import ssl
    import requests
    from urllib3.util.ssl_ import create_urllib3_context

    class _AmazonCdnAdapter(requests.adapters.HTTPAdapter):
        def init_poolmanager(self, *a, **kw):
            ctx = create_urllib3_context()
            ctx.minimum_version = ssl.TLSVersion.TLSv1_2
            kw['ssl_context'] = ctx
            return super().init_poolmanager(*a, **kw)

    s = requests.Session()
    s.mount('https://', _AmazonCdnAdapter())
    s.headers.update({'User-Agent': 'Mozilla/5.0'})
    _HTTP_SESSION = s
    return s


def _fetch_one_image(url: str, max_size_px: int = 90, timeout: float = 8.0) -> BytesIO | None:
    """下载单张图，缩略后返回 BytesIO（PNG）。失败返回 None，不抛异常。
    - 内部一次重试（首次失败后 sleep 1s 再试）
    - 失败计数 _IMG_FAIL_COUNT[url] >= _IMG_MAX_FAIL 才彻底放弃（写 None），避免单次抖动永久缓存失败
    - 所有失败都 log.warning，留 URL + 异常类型供排查
    """
    if not url or not isinstance(url, str) or not url.startswith(('http://', 'https://')):
        return None
    cached = _IMG_CACHE.get(url)
    if cached is not None:
        cached.seek(0)
        return cached
    # 已经写过 None（彻底放弃）→ 直接返回
    if url in _IMG_CACHE and _IMG_CACHE[url] is None:
        return None

    sess = _get_http_session()
    last_exc: Exception | None = None
    # 真正出网时才取信号量，缓存命中已在上方早退，不占用名额
    with _IMG_FETCH_SEM:
        # 3 次尝试 + 指数退避（0.5s / 1.5s）：应对 CDN SSL 握手突发拒绝
        for attempt in (1, 2, 3):
            try:
                resp = sess.get(url, timeout=timeout)
                if resp.status_code != 200 or not resp.content:
                    last_exc = RuntimeError(f"HTTP {resp.status_code}, {len(resp.content)} bytes")
                else:
                    from PIL import Image as PILImage
                    img = PILImage.open(BytesIO(resp.content))
                    img.thumbnail((max_size_px, max_size_px), PILImage.LANCZOS)
                    if img.mode not in ('RGB', 'RGBA'):
                        img = img.convert('RGB')
                    out = BytesIO()
                    img.save(out, format='PNG')
                    out.seek(0)
                    _IMG_CACHE[url] = out
                    return out
            except Exception as e:
                last_exc = e
            if attempt == 1:
                time.sleep(0.5)
            elif attempt == 2:
                time.sleep(1.5)

    # 两次都失败 → 计数 + 警告
    _IMG_FAIL_COUNT[url] = _IMG_FAIL_COUNT.get(url, 0) + 1
    log.warning("[image] 下载失败 (%d/%d) %s: %s", _IMG_FAIL_COUNT[url], _IMG_MAX_FAIL, url[:120], last_exc)
    if _IMG_FAIL_COUNT[url] >= _IMG_MAX_FAIL:
        _IMG_CACHE[url] = None  # 彻底放弃
    return None


def _prefetch_images(urls: list[str], max_workers: int = 8) -> None:
    """并发预下载一批图片到 _IMG_CACHE，加速后续单元格插入。
    并发数 8（曾用 32，但对 Amazon CDN 太激进，触发 SSL 握手 reset → 部分图永久丢失）。
    """
    todo = [u for u in urls if u and isinstance(u, str) and u.startswith(('http://', 'https://')) and u not in _IMG_CACHE]
    if not todo:
        return
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for fut in as_completed([ex.submit(_fetch_one_image, u) for u in todo]):
            try:
                fut.result()
            except Exception:
                pass


def _insert_image_to_cell(ws, row: int, col: int, url: str, cell_height_pt: float = 70.0) -> bool:
    """在 (row, col) 单元格嵌入图片。下载失败时返回 False（调用方可写 '-' 占位）。
    每次都从缓存的 bytes 复制出新的 BytesIO，避免同一图被多个 XLImage 共享时
    openpyxl save 阶段读到已关闭的流（"I/O operation on closed file"）。"""
    cached = _fetch_one_image(url)
    if cached is None:
        return False
    try:
        # 关键：复制 bytes 出来，每个 XLImage 用独立的 BytesIO
        cached.seek(0)
        new_buf = BytesIO(cached.read())
        img = XLImage(new_buf)
        img.width = 80
        img.height = 80
        img.anchor = f'{get_column_letter(col)}{row}'
        ws.add_image(img)
        return True
    except Exception:
        return False

# Packs 运行时（LLM 深度分析 4 份源数据 → Insight Pack）
from core.packs_runtime import (
    Packs, prepare_packs,
    classify_with_packs, neg_keywords_dict, pos_keywords_dict,
    price_band_descriptions, display_name_for_title,
    sheet10_dimension_reasons, sheet10_headline,
    sheet6_priority_matrix, upgrade_directions,
    spec_dimensions, extract_specs_for_title,
    compliance_certs, compliance_return_info,
    sheet1_sales_tier_narratives, sheet2_risk_bullets, sheet2_brand_strategies,
    sheet1_market_conclusions,
)

# 配置
UPLOAD_FOLDER = 'uploads'
REPORT_FOLDER = 'reports'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_CONTENT_LENGTH = 100 * 1024 * 1024

app = Flask(__name__)
app.secret_key = 'product-analysis-v2-secret-key'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['REPORT_FOLDER'] = REPORT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _build_promotion_pressure_fallback_text(packs, kw_summary_txt: str, kw_lowest_spr_for_conclusion: str) -> str:
    """推广压力维度兜底句（Python 直接用 packs.synthesis_stats 里的真实指标拼）。

    卖家精灵 ReverseASIN 没有单列"广告占比"字段；过去"广告占比 0% / 广告高占比词 0"
    是列匹配失败造成的假数字。现在改为引用真实复合指标：
      - promo_ppc_bid_avg: Top20 核心词平均 PPC 竞价（$）
      - promo_ads_competitor_avg: 平均广告竞品数（个）
      - promo_click_share_avg_top: 前 3 ASIN 点击总占比（头部垄断程度，%）
      - promo_conversion_share_avg_top: 前 3 ASIN 转化总占比（%）
      - promo_spr_median: 8 天上首页销量门槛中位数
      - bsr_sp_ads_pct: BSR TOP100 开 SP 广告占比（%）
    任何字段缺失时跳过该片段，不拼"0"或"无数据"；全空时退回到关键词摘要旧逻辑。
    """
    stats = getattr(packs, 'synthesis_stats', {}) if packs is not None else {}
    bits: list[str] = []
    ppc = stats.get('promo_ppc_bid_avg')
    if ppc:
        bits.append(f'Top20 核心词平均 PPC ${ppc:.2f}')
    ads_comp = stats.get('promo_ads_competitor_avg')
    if ads_comp:
        bits.append(f'平均广告竞品 {int(ads_comp)} 个')
    click_share = stats.get('promo_click_share_avg_top')
    if click_share:
        bits.append(f'头部前 3 ASIN 点击占 {click_share:.0f}%')
    conv_share = stats.get('promo_conversion_share_avg_top')
    if conv_share:
        bits.append(f'转化占 {conv_share:.0f}%')
    sp_ads = stats.get('bsr_sp_ads_pct')
    if sp_ads:
        bits.append(f'BSR TOP100 开 SP 广告 {sp_ads:.0f}%')
    spr_med = stats.get('promo_spr_median')
    if spr_med:
        bits.append(f'上首页销量中位 {int(spr_med)}')

    if bits:
        return '；'.join(bits)
    # 最后兜底：所有推广压力指标都拿不到 → 退回老逻辑（关键词摘要）
    return f'核心词：{kw_summary_txt}；最低门槛词：{kw_lowest_spr_for_conclusion or "（需关键词文件）"}'


def _build_supply_chain_fallback_text(packs, difficulty_level: str) -> str:
    """供应链维度 LLM 不可用时的兜底句：
    - 若 Compliance Pack 给出了必备认证，优先引用具体认证名 + 推广周期
    - 否则只引用推广周期 + 通用建议
    绝不出现 "LLM 不可用" 字样（那属于实现细节，不应写入 Excel）
    """
    cycle = {'低': '3-6 月', '中': '6-9 月', '高': '9-12 月'}.get(difficulty_level, '6-9 月')
    certs: list[str] = []
    if packs is not None and packs.is_compliance_real():
        certs = [c.name for c in packs.compliance.required_certifications if c.mandatory][:3]
    if certs:
        return f'新品进入期 {cycle}；核心认证要求：{", ".join(certs)}（需提前排查工厂资质）'
    return f'新品进入期 {cycle}；建议核实工厂资质 + 确认 MOQ 与交期再建仓'


# ============================================================
# 产品分类兜底函数
# ============================================================
# 说明：全品类报告不再写死某个品类的分类规则。
# - 正常流程由 LLM（BSRAnalyzer）识别本次品类的 product_segments，
#   通过 classify_with_packs(title, packs, classify) 归属 ASIN
# - 只有在 LLM 完全不可用时才 fallback 到本函数 —— 此时返回 "未分类"，
#   而不是 LED 工作灯专属的"充电磁吸/三脚架/泛光"等分类，避免跨品类误标
def classify(title):
    return '未分类'


# ============================================================
# 统一的产品类型综合评分（供 Sheet 6 上新方向 和 Sheet 10 首推入场品类 共用）
# 维度：需求(月销)、竞争度(SKU数)、单品收益、新品占比、质量评分
# ============================================================
def rank_product_types(df, type_agg):
    """
    对每个产品类型计算综合得分，返回按得分降序排列的类型列表 + 得分细节 dict。
    评分维度（每项 0-1 归一化后加权求和）：
      - 需求(月销量)：0.30
      - 单品收益($/SKU)：0.25
      - 新品占比(<1年)：0.15
      - 质量评分(平均Rating)：0.15
      - 竞争度(SKU数倒数，少为佳)：0.15
    """
    if type_agg is None or len(type_agg) == 0:
        return [], {}

    rows = []
    for _, tr in type_agg.iterrows():
        pt = tr['product_type']
        if pt in ('其他', '未分类'):
            continue
        type_df_tmp = df[df['product_type'] == pt]
        sku_count = max(len(type_df_tmp), 1)
        total_sales = float(tr.get('total_sales', 0) or 0)
        total_rev = float(tr.get('total_revenue', 0) or 0)
        avg_rating = float(tr.get('avg_rating', 0) or 0)
        avg_rev = total_rev / sku_count if sku_count > 0 else 0
        new_ratio = (len(type_df_tmp[type_df_tmp['Available days'] < 365]) / sku_count
                     if 'Available days' in df.columns else 0.5)
        rows.append({
            'product_type': pt, 'total_sales': total_sales, 'avg_rev': avg_rev,
            'new_ratio': new_ratio, 'avg_rating': avg_rating, 'sku_count': sku_count,
        })
    if not rows:
        return [], {}

    max_sales = max(r['total_sales'] for r in rows) or 1
    max_rev = max(r['avg_rev'] for r in rows) or 1
    max_sku = max(r['sku_count'] for r in rows) or 1

    scores = {}
    for r in rows:
        n_sales = r['total_sales'] / max_sales
        n_rev = r['avg_rev'] / max_rev
        n_new = min(r['new_ratio'], 1.0)
        n_rating = r['avg_rating'] / 5.0 if r['avg_rating'] else 0.6
        n_comp = 1 - (r['sku_count'] / max_sku) * 0.7
        score = n_sales * 0.30 + n_rev * 0.25 + n_new * 0.15 + n_rating * 0.15 + n_comp * 0.15
        scores[r['product_type']] = {
            'score': score, 'total_sales': r['total_sales'], 'avg_rev': r['avg_rev'],
            'new_ratio': r['new_ratio'], 'avg_rating': r['avg_rating'], 'sku_count': r['sku_count'],
        }

    ranked = sorted(scores.keys(), key=lambda k: scores[k]['score'], reverse=True)
    return ranked, scores


# ============================================================
# 基于实际数据动态计算推荐入场价
# ============================================================
def calculate_pricing_recommendations(df, price_col, type_agg, ranked_types=None, top_segment_name=None, packs=None):
    """
    根据实际BSR数据计算各产品类型的推荐入场价
    返回动态生成的价格建议列表
    top_segment_name: 已统一决定的首推品类名（优先 LLM 矩阵 P1）；传了就用它标 ★，
                     避免 Sheet 4 ★ 和 Sheet 6/10 矩阵 P1 矛盾
    ranked_types: rank_product_types() 返回的排序列表，仅当 top_segment_name 未提供时 fallback 用
    packs: Packs 实例，传了就在"理由"列末尾拼 LLM 的 pricing_segment_insights 叙述
    """
    # LLM 按本品类给的每个细分的一句定性叙述（Python 数字 + LLM 叙述混合拼接）
    _seg_narr_map: dict[str, str] = {}
    if packs is not None:
        try:
            from core.packs_runtime import pricing_segment_narratives
            _seg_narr_map = pricing_segment_narratives(packs)
        except Exception:
            _seg_narr_map = {}

    pricing_recommendations = []
    top_type = top_segment_name or (ranked_types[0] if ranked_types else None)
    
    for _, type_row in type_agg.iterrows():
        ptype = type_row['product_type']
        if ptype in ('其他', '未分类'):
            continue
        type_df = df[df['product_type'] == ptype]

        if len(type_df) < 2:
            continue
            
        prices = type_df[price_col].dropna()
        if len(prices) < 2:
            continue
            
        min_price = prices.min()
        p25 = prices.quantile(0.25)
        median_price = prices.median()
        p75 = prices.quantile(0.75)
        max_price = prices.max()
        
        # 计算单品收益
        total_rev = type_row['total_revenue'] if pd.notna(type_row['total_revenue']) else 0
        sku_count = type_row['count']
        avg_rev_per_sku = total_rev / sku_count if sku_count > 0 else 0
        
        # 推荐入场价：选择P40-P60区间，既不太低也不太高
        rec_min = round(median_price * 0.9, 2)
        rec_max = round(median_price * 1.1, 2)
        
        # 理由列：纯数据事实拼接，不用"价格带宽/差异化空间大"这类不可证伪的形容词
        # 每行 4 段：价格四分位 / 竞品 SKU 数 / 单品月收益 / 毛利中位数（若有）
        reasons = [
            f'P25=${p25:.1f} · 中位 ${median_price:.1f} · P75=${p75:.1f}',
            f'{len(type_df)} 个 SKU',
        ]
        if avg_rev_per_sku >= 10000:
            reasons.append(f'单品月均收益 ${avg_rev_per_sku/10000:.1f} 万')
        else:
            reasons.append(f'单品月均收益 ${avg_rev_per_sku:,.0f}')
        # 毛利中位数（如源表有 Gross Margin 列）
        try:
            if 'Gross Margin' in type_df.columns:
                _gm = pd.to_numeric(type_df['Gross Margin'], errors='coerce').dropna()
                if len(_gm) > 0:
                    _gm_val = float(_gm.median())
                    # 源表 Gross Margin 可能是 0-1 小数也可能是 0-100 百分数
                    if _gm_val <= 1.0:
                        reasons.append(f'毛利中位 {_gm_val:.0%}（卖家精灵数据，仅扣佣金+FBA）')
                    else:
                        reasons.append(f'毛利中位 {_gm_val:.0f}%（卖家精灵数据，仅扣佣金+FBA）')
        except Exception:
            pass
        
        # 判断是否是首推品类（用统一综合评分）
        is_top_type = (ptype == top_type) if top_type else (
            ptype == type_agg.iloc[0]['product_type'] if len(type_agg) > 0 else False
        )
        note = '★ 首推入场品类' if is_top_type else ''
        
        # Python 数字事实 + LLM 定性叙述拼接
        _reason_facts = '；'.join(reasons)
        _seg_narr = _seg_narr_map.get(ptype, '')
        _final_reason = f'{_reason_facts}。{_seg_narr}' if _seg_narr else _reason_facts

        pricing_recommendations.append({
            'product_type': ptype,
            'min_price': min_price,
            'p25': p25,
            'median': median_price,
            'p75': p75,
            'max_price': max_price,
            'rec_min': rec_min,
            'rec_max': rec_max,
            'reason': _final_reason,
            'note': note,
            'sku_count': len(type_df),
            'avg_rev': avg_rev_per_sku,
            'is_top': is_top_type
        })
    
    # 按综合评分排序（与 Sheet 6 上新方向保持一致）
    if ranked_types:
        rank_index = {t: i for i, t in enumerate(ranked_types)}
        pricing_recommendations.sort(key=lambda x: rank_index.get(x['product_type'], 999))
    else:
        pricing_recommendations.sort(key=lambda x: x['avg_rev'], reverse=True)
    return pricing_recommendations


# ============================================================
# 基于实际数据动态生成产品上新方向
# ============================================================
def generate_product_directions(df, rev_df, neg_counts, type_agg, price_col, packs=None):
    """
    生成产品上新方向（Sheet 6 核心改进点 / 数据支撑依据 两列的内容）。

    核心原则：优先使用 LLM Synthesizer 产出的 `sheet6_priority_matrix` —— 它的 improvements
    字段是按每个细分的痛点/规格差距定制的，各不相同；同时 action_plan 是品类特异的策略叙述。
    LLM 不可用时才回退到本函数的 Python 启发式模板（仍按不同类型匹配不同痛点，避免 5 行雷同）。

    Sheet 6 表头 9 列：优先级 / 产品方向 / 核心改进点 / 数据支撑依据 / 目标售价 / 预计月销 /
                      竞争难度 / 启动方式 / 预期亮点
    """
    # 盈亏平衡采购上限计算（跨品类通用，跟 Sheet 4 入场利润测算用同一组费率参数）
    _COMMISSION_RATE = 0.15
    _FBA_FEE_USD = 5.15
    _AD_RATE = 0.12
    _OTHER_FEE_USD = 1.50

    def _purchase_upper(median_price: float) -> str:
        """按该 segment 中位价算盈亏平衡采购上限（保本/10%/15%/20% 多档）"""
        if not median_price or median_price <= 0:
            return '—'
        breakeven = median_price * (1 - _COMMISSION_RATE - _AD_RATE) - _FBA_FEE_USD - _OTHER_FEE_USD
        def _fmt(upper):
            return f'≤${upper:.2f}' if upper >= 0 else '<$0'
        return (
            f'保本{_fmt(breakeven)}\n'
            f'保10%{_fmt(breakeven - median_price * 0.10)}\n'
            f'保15%{_fmt(breakeven - median_price * 0.15)}\n'
            f'保20%{_fmt(breakeven - median_price * 0.20)}'
        )

    # 预备：按综合评分排序的产品类型（用于 LLM 不在时的 P1-P4 分配，也用于数据支撑依据的销量数据）
    ranked_types, _type_score_detail = rank_product_types(df, type_agg)
    type_priority = {}
    for i, pt in enumerate(ranked_types):
        if i == 0:
            type_priority[pt] = 'P1'
        elif i == 1:
            type_priority[pt] = 'P2'
        elif i <= 3:
            type_priority[pt] = 'P3'
        else:
            type_priority[pt] = 'P4'

    # 每个 segment 能拿到的 type_agg 行（按产品类型名匹配，用于目标售价/月销数据）
    type_agg_map = {row['product_type']: row for _, row in type_agg.iterrows()}

    def _data_support_for(segment_name: str, sku_count: int, sales: int, pain_names: list) -> str:
        """数据支撑依据列：[月销 + SKU 数] + [差评痛点 TOP3]（后者由 VOC 给）"""
        parts = [f'该类型月总销量{int(sales):,}件，{sku_count}个SKU参与竞争']
        if pain_names:
            parts.append('差评痛点：' + '、'.join(pain_names[:3]))
        return '\n'.join(parts)

    def _seg_targets_from_type_agg(ptype: str):
        """按 product_type 名从 type_agg 拿 median_price / sku_count / 月销均值"""
        type_df = df[df['product_type'] == ptype]
        prices = type_df[price_col].dropna() if price_col else pd.Series(dtype=float)
        median_price = prices.median() if len(prices) > 0 else 30
        sku_count = len(type_df)
        sales = 0
        if ptype in type_agg_map:
            _row = type_agg_map[ptype]
            sales = int(_row['total_sales']) if pd.notna(_row.get('total_sales')) else 0
        avg_monthly = sales / sku_count if sku_count > 0 else 200
        return median_price, sku_count, sales, avg_monthly

    # --- 路径 A：LLM Synthesizer.sheet6_priority_matrix 可用时，直接用 LLM 的 improvements ---
    # 走 helper 拿 sheet6_priority_matrix——helper 内部已对 improvements / action_plan 过 _scrub_schema_paths
    _matrix_scrubbed = sheet6_priority_matrix(packs)
    if _matrix_scrubbed:
        directions = []
        # VOC pain 按细分匹配：对每个 segment 取 pain_clusters 里 name 含该 segment 关键词的，
        # 否则退化到全局 top 3 pain（避免所有行都是同一条）
        all_pain_names = []
        if packs is not None and packs.is_voc_real() and packs.voc.pain_clusters:
            all_pain_names = [c.name for c in packs.voc.pain_clusters[:10]]

        for item in _matrix_scrubbed:
            seg = item.get('segment') or ''
            if not seg:
                continue
            improvements = [f'• {imp}' for imp in (item.get('improvements') or []) if imp.strip()]
            # action_plan 单独也算有效内容——LLM 偶尔只给 action_plan 不给 improvements，
            # 之前 `improvements` 必须非空才追加 action_plan 的逻辑会让该条被 511 行跳过、
            # 导致 Sheet 6 比 Sheet 10「上新优先级矩阵」少展示一些细分（如 P3 木质字母）。
            ap = (item.get('action_plan') or '').strip()
            if ap:
                improvements.append(f'• 策略：{ap}')
            if not improvements:
                # 真的 improvements 和 action_plan 都为空才跳过
                continue

            # 匹配 segment 特异的痛点：
            # 1) VOC.pain_clusters.affected_asins 与本 segment.member_asins 有交集的 pain（最强信号）
            # 2) pain.name 里含 segment 关键词的
            # 3) 本 segment ASIN 的 rev_df 评论关键词出现次数 TOP3（纯 Python 从源数据聚合）
            # 4) 都空时留空"—"，不再 fallback 到全局 top3（避免三个细分抄同一段）
            seg_pains = []
            if packs.is_voc_real() and packs.voc.pain_clusters:
                _seg_asins = set()
                if packs.is_market_real():
                    for _ms in packs.market.product_segments:
                        if _ms.name == seg:
                            _seg_asins = set(_ms.member_asins or [])
                            break
                if _seg_asins:
                    for _pc in packs.voc.pain_clusters[:8]:
                        if _seg_asins & set(_pc.affected_asins or []):
                            seg_pains.append(_pc.name)
            if not seg_pains and all_pain_names:
                lowered_seg = seg.lower()
                for p in all_pain_names:
                    pl = p.lower()
                    if any(tok in pl for tok in lowered_seg.split() if len(tok) >= 2):
                        seg_pains.append(p)

            median_price, sku_count, sales, avg_monthly = _seg_targets_from_type_agg(seg)
            target_low = round(median_price * 0.95, 2)
            target_high = round(median_price * 1.15, 2)
            est_low = int(avg_monthly * 0.6)
            est_high = int(avg_monthly * 1.2)
            if sku_count <= 5:
                difficulty, note = '低', '竞品少，蓝海机会'
            elif sku_count <= 15:
                difficulty, note = '中', '竞争适中'
            else:
                difficulty, note = '中高', '竞争激烈，需差异化'

            priority = item.get('priority') or type_priority.get(seg, 'P3')
            prio_label = f'{priority}\n{"首推" if priority == "P1" else "推荐" if priority == "P2" else "参考" if priority == "P3" else "备选"}'
            directions.append((
                prio_label, seg,
                '\n'.join(improvements),
                _data_support_for(seg, sku_count, sales, seg_pains),
                f'${target_low:.2f}-${target_high:.2f}',
                f'{est_low}-{est_high}件/月',
                difficulty,
                'FBA直发，首批备货300-500pcs',
                note,
            ))

        if directions:
            priority_order = {'P1': 0, 'P2': 1, 'P3': 2, 'P4': 3}
            directions.sort(key=lambda x: priority_order.get(x[0].split('\n')[0], 4))
            return directions[:6]

    # --- 路径 B：LLM 不可用时，Python 启发式模板（按每个 segment 匹配不同痛点避免全部雷同） ---
    directions = []
    neg_sorted = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True)
    top_pains = neg_sorted[:8] if len(neg_sorted) >= 5 else neg_sorted

    for _, type_row in type_agg.iterrows():
        ptype = type_row['product_type']
        if ptype not in type_priority:
            continue
        type_df = df[df['product_type'] == ptype]
        if len(type_df) < 2:
            continue
        prices = type_df[price_col].dropna()
        avg_price = prices.mean() if len(prices) > 0 else 30
        median_price = prices.median() if len(prices) > 0 else 30

        # 按细分索引取不同的 pain（避免所有行都引第一条）
        seg_idx = list(type_priority.keys()).index(ptype)
        seg_pains = top_pains[seg_idx:seg_idx + 3] if top_pains else []
        if not seg_pains and top_pains:
            seg_pains = top_pains[:3]
        pain_summary = '、'.join([f'{p[0]}({p[1]}条)' for p in seg_pains[:3]])

        improvements = []
        for pain_name, _cnt in seg_pains[:2]:
            # 用痛点名本身做改进点（避免模板化"5000mAh/35N/铝合金"）
            improvements.append(f'• 针对「{pain_name}」重点改良：见 Sheet 4 竞品差评 TOP 表')
        if not improvements:
            improvements = ['• 结合 VOC 差评做差异化改良（见 Sheet 4）']

        target_low = round(median_price * 0.95, 2)
        target_high = round(median_price * 1.15, 2)
        type_sales = type_row['total_sales'] if pd.notna(type_row['total_sales']) else 0
        type_sku = type_row['count']
        avg_monthly_sales = type_sales / type_sku if type_sku > 0 else 200
        est_monthly_low = int(avg_monthly_sales * 0.6)
        est_monthly_high = int(avg_monthly_sales * 1.2)
        priority = type_priority.get(ptype, 'P3')
        prio_label = f'{priority}\n{"首推" if priority == "P1" else "推荐" if priority == "P2" else "参考" if priority == "P3" else "备选"}'
        if type_sku <= 5:
            difficulty, note = '低', '竞品少，蓝海机会'
        elif type_sku <= 15:
            difficulty, note = '中', '竞争适中'
        else:
            difficulty, note = '中高', '竞争激烈，需差异化'
        data_support = f'该类型月总销量{int(type_sales):,}件，{type_sku}个SKU参与竞争'
        if pain_summary:
            data_support += f'\n差评痛点：{pain_summary}'

        directions.append((
            prio_label, ptype,
            '\n'.join(improvements),
            data_support,
            f'${target_low:.2f}-${target_high:.2f}',
            f'{est_monthly_low}-{est_monthly_high}件/月',
            difficulty,
            'FBA直发，首批备货300-500pcs',
            note,
        ))

    priority_order = {'P1': 0, 'P2': 1, 'P3': 2, 'P4': 3}
    directions.sort(key=lambda x: priority_order.get(x[0].split('\n')[0], 4))
    return directions[:6]


# ============================================================
# Excel样式函数
# ============================================================
C_BLUE_DARK = 'FF1F3864'
C_BLUE_MID = 'FF2E74B5'
C_BLUE_LIGHT = 'FFD6E4F0'
C_YELLOW = 'FFFFF2CC'
C_GREEN_LIGHT = 'FFE2EFDA'
C_ORANGE = 'FFFCE4D6'
C_WHITE = 'FFFFFFFF'
C_GREY_LIGHT = 'FFF2F2F2'
C_RED_LIGHT = 'FFFDECEA'
C_SECTION_BG = 'FF1F3864'


def hdr(ws, row, col, text, bg=C_BLUE_MID, fg=C_WHITE, bold=True, size=11, wrap=False, h_align='center'):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=wrap)
    return cell


def val(ws, row, col, text, bg=C_WHITE, bold=False, size=10, wrap=True,
        h_align='left', fg='FF000000', italic=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size, italic=italic)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=wrap)
    return cell


def thin_border():
    side = Side(style='thin', color='FFBFBFBF')
    return Border(left=side, right=side, top=side, bottom=side)


def apply_border(ws, min_row, max_row, min_col, max_col):
    border = thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def section_title(ws, row, col, text, span=8, bg=C_SECTION_BG):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Arial', bold=True, color=C_WHITE, size=12)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 24


# ============================================================
# Market 分析文件读取（可选）
# ============================================================
def load_market_data(market_path):
    """
    读取卖家精灵 US-Market 文件。文件缺失或 sheet 缺失时返回 None，不抛异常。
    返回 dict：{sheet_key: DataFrame or None, ...}
    """
    keys = [
        ('market_summary', 'Market Analysis'),
        ('demand_trends', 'Industry Demand and Trends'),
        ('sell_trends', 'Industry Sell Trends'),
        ('listing_concentration', 'Listing Concentration'),
        ('brand_concentration', 'Brand Concentration'),
        ('seller_concentration', 'Seller Concentration'),
        ('fulfillment', 'Fulfillment'),
        ('aplus_video', 'A+ Content and Video'),
        ('origin_of_seller', 'Origin of Seller'),
        ('publication_time', 'Publication Time'),
        ('publication_time_trends', 'Publication Time Trends'),
        ('ratings_dist', 'Ratings'),
        ('rating_dist', 'Rating'),
        ('price_dist', 'Price'),
    ]
    result = {k: None for k, _ in keys}
    result['_available'] = False
    if not market_path or not os.path.exists(market_path):
        return result
    try:
        from llm.analyzers.bsr_analyzer import CN_EN_MARKET_SHEETS, normalize_market_columns
        xl = pd.ExcelFile(market_path)
        sheets = set(xl.sheet_names)
        # 构造"英文标准名 → 原始 sheet 名"的反向映射，使中文 sheet 文件也能按英文 key 读到
        en_to_original = {}
        for orig_name in xl.sheet_names:
            en_name = CN_EN_MARKET_SHEETS.get(orig_name, orig_name)
            en_to_original[en_name] = orig_name
        for key, sheet_name in keys:
            original_sheet = sheet_name if sheet_name in sheets else en_to_original.get(sheet_name)
            if original_sheet and original_sheet in sheets:
                try:
                    _df = pd.read_excel(market_path, sheet_name=original_sheet)
                    # 中文列名 → 英文标准名（仅覆盖下游硬编码的列）
                    result[key] = normalize_market_columns(_df)
                except Exception as e:
                    print(f"读取 Market sheet '{original_sheet}' 出错: {e}")
        result['_available'] = any(v is not None for k, v in result.items() if k != '_available')
    except Exception as e:
        print(f"读取 Market 文件出错: {e}")
    return result


# ============================================================
# 关键词文件读取（可选，默认使用 ExpandKeywords；保留旧 ReverseASIN 兼容）
# ============================================================
def load_keyword_data(keyword_path):
    result = {'keywords': None, 'unique_words': None, '_available': False}
    if not keyword_path or not os.path.exists(keyword_path):
        return result
    try:
        from llm.analyzers.bsr_analyzer import normalize_keyword_columns
        xl = pd.ExcelFile(keyword_path)
        for sheet in xl.sheet_names:
            if 'US-' in sheet or sheet.startswith('US-'):
                result['keywords'] = pd.read_excel(keyword_path, sheet_name=sheet)
                break
        # ExpandKeywords 文件的数据 sheet 名形如 "US-B07799RL2B(20)__"，已被上面匹配到
        # 中文列名版本的数据也会在此处统一规范化（关键词→Keyword 等）
        if result['keywords'] is not None:
            result['keywords'] = normalize_keyword_columns(result['keywords'])
        if 'Unique Words' in xl.sheet_names:
            uw = pd.read_excel(keyword_path, sheet_name='Unique Words')
            result['unique_words'] = normalize_keyword_columns(uw)
        result['_available'] = result['keywords'] is not None
    except Exception as e:
        print(f"读取关键词文件出错: {e}")
    return result


# ============================================================
# Title 特征提取
# ============================================================
import re as _re

def extract_specs_from_title(title):
    """从产品标题中正则提取常见规格，识别不到的字段返回空字符串"""
    t = str(title) if title is not None else ''
    tl = t.lower()
    specs = {
        '功率W': '',
        '光通量lm': '',
        '电池类型': '',
        '续航': '',
        '防水等级': '',
        '光源': '',
        '供电方式': '',
    }
    m = _re.search(r'(\d{2,6})\s*(?:w|watt)s?\b', tl)
    if m:
        specs['功率W'] = f"{m.group(1)}W"
    m = _re.search(r'(\d{2,6})\s*(?:lm|lumens?)\b', tl)
    if m:
        specs['光通量lm'] = f"{m.group(1)}lm"
    if 'lithium' in tl or 'li-ion' in tl or 'li ion' in tl:
        specs['电池类型'] = '锂电池'
    elif 'lifepo4' in tl:
        specs['电池类型'] = '磷酸铁锂'
    elif 'aa ' in tl or 'aaa ' in tl:
        specs['电池类型'] = '干电池'
    m = _re.search(r'(\d{1,2})\s*(?:v|volt)s?\b', tl)
    if m and not specs['电池类型']:
        specs['电池类型'] = f"{m.group(1)}V 电池"
    m = _re.search(r'(\d{1,3})\s*(?:hours?|hrs?|h)\s*(?:run|runtime|battery)?', tl)
    if m:
        specs['续航'] = f"{m.group(1)}h"
    m = _re.search(r'ip\s?(\d{2})\b', tl)
    if m:
        specs['防水等级'] = f"IP{m.group(1)}"
    elif 'waterproof' in tl:
        specs['防水等级'] = '防水'
    elif 'water resistant' in tl or 'water-resistant' in tl:
        specs['防水等级'] = '防泼溅'
    if 'cob' in tl:
        specs['光源'] = 'COB LED'
    elif 'led' in tl:
        specs['光源'] = 'LED'
    if 'rechargeable' in tl or 'cordless' in tl:
        specs['供电方式'] = '充电式'
    elif 'solar' in tl:
        specs['供电方式'] = '太阳能'
    elif 'plug' in tl or 'corded' in tl:
        specs['供电方式'] = '插电'
    elif 'battery' in tl:
        specs['供电方式'] = '电池供电'
    return specs


def extract_specs_from_bullets(bullet_text):
    """从 Bullet Points 列文本中提取详细产品规格（补充标题正则无法覆盖的参数）"""
    t = str(bullet_text) if bullet_text is not None else ''
    tl = t.lower()
    specs = {
        '功率W': '',
        '光通量lm': '',
        '电池类型': '',
        '续航': '',
        '防水等级': '',
        '光源': '',
        '供电方式': '',
        '电池容量mAh': '',
        '充电方式': '',
        '光照模式数': '',
        '磁力/固定方式': '',
        '旋转角度': '',
        '材质': '',
        '重量': '',
    }
    if not tl.strip():
        return specs

    # 功率
    m = _re.search(r'(\d{2,6})\s*(?:w|watt)s?\b', tl)
    if m:
        specs['功率W'] = f"{m.group(1)}W"
    # 光通量
    m = _re.search(r'(\d{2,6})\s*(?:lm|lumens?)\b', tl)
    if m:
        specs['光通量lm'] = f"{m.group(1)}lm"
    # 电池类型
    if 'lithium' in tl or 'li-ion' in tl or 'li ion' in tl:
        specs['电池类型'] = '锂电池'
    elif 'lifepo4' in tl:
        specs['电池类型'] = '磷酸铁锂'
    # 续航
    m = _re.search(r'(?:up\s+to\s+)?(\d{1,3}(?:\.\d)?)\s*(?:hours?|hrs?)\s*(?:of\s+)?(?:run|runtime|battery|light|use)?', tl)
    if m:
        specs['续航'] = f"{m.group(1)}h"
    # 防水等级
    m = _re.search(r'ip\s?[x]?(\d{2})\b', tl)
    if m:
        specs['防水等级'] = f"IP{m.group(1)}"
    elif 'waterproof' in tl:
        specs['防水等级'] = '防水'
    elif 'water resistant' in tl or 'water-resistant' in tl:
        specs['防水等级'] = '防泼溅'
    # 光源
    if 'cob' in tl:
        specs['光源'] = 'COB LED'
    elif 'smd' in tl:
        specs['光源'] = 'SMD LED'
    elif 'led' in tl:
        specs['光源'] = 'LED'
    # 供电方式
    if 'rechargeable' in tl or 'cordless' in tl:
        specs['供电方式'] = '充电式'
    elif 'solar' in tl:
        specs['供电方式'] = '太阳能'
    elif 'plug' in tl or 'corded' in tl:
        specs['供电方式'] = '插电'
    elif 'battery' in tl:
        specs['供电方式'] = '电池供电'

    # === 新增参数 ===
    # 电池容量
    m = _re.search(r'(\d{3,5})\s*mah', tl)
    if m:
        specs['电池容量mAh'] = f"{m.group(1)}mAh"
    # 充电方式
    if 'usb-c' in tl or 'usb c' in tl or 'type-c' in tl or 'type c' in tl:
        specs['充电方式'] = 'USB-C'
    elif 'micro-usb' in tl or 'micro usb' in tl:
        specs['充电方式'] = 'Micro-USB'
    elif 'usb' in tl and 'charg' in tl:
        specs['充电方式'] = 'USB'
    # 光照模式数
    m = _re.search(r'(\d+)\s*(?:lighting\s+)?(?:mode|brightness\s+level|working\s+mode)', tl)
    if m:
        specs['光照模式数'] = f"{m.group(1)}种"
    # 磁力/固定方式
    fix_methods = []
    if 'magnet' in tl or 'magnetic' in tl:
        fix_methods.append('磁吸')
    if 'hook' in tl:
        fix_methods.append('挂钩')
    if 'clip' in tl or 'clamp' in tl:
        fix_methods.append('夹具')
    if 'tripod' in tl:
        fix_methods.append('三脚架')
    if 'carabiner' in tl:
        fix_methods.append('登山扣')
    if fix_methods:
        specs['磁力/固定方式'] = '+'.join(fix_methods)
    # 旋转角度
    m = _re.search(r'(\d{2,3})\s*(?:degree|°)\s*(?:rotation|rotate|swivel|adjustable|pivot)?', tl)
    if m:
        specs['旋转角度'] = f"{m.group(1)}°"
    # 材质
    materials = []
    if 'aluminum' in tl or 'aluminium' in tl:
        materials.append('铝合金')
    if 'abs' in tl:
        materials.append('ABS')
    if 'rubber' in tl or 'silicone' in tl:
        materials.append('橡胶/硅胶')
    if 'stainless steel' in tl:
        materials.append('不锈钢')
    if materials:
        specs['材质'] = '+'.join(materials)
    # 重量
    m = _re.search(r'(\d+\.?\d*)\s*(?:oz|ounces?)\b', tl)
    if m:
        specs['重量'] = f"{m.group(1)}oz"
    else:
        m = _re.search(r'(\d+\.?\d*)\s*(?:g|grams?)\b', tl)
        if m and float(m.group(1)) > 10:  # 排除误匹配
            specs['重量'] = f"{m.group(1)}g"
        else:
            m = _re.search(r'(\d+\.?\d*)\s*(?:lb|lbs|pounds?)\b', tl)
            if m:
                specs['重量'] = f"{m.group(1)}lb"
    return specs


def extract_all_specs(title, bullet_text):
    """合并标题和五点描述的规格提取结果，五点描述优先（更详细）"""
    title_specs = extract_specs_from_title(title)
    bullet_specs = extract_specs_from_bullets(bullet_text)
    merged = {}
    all_keys = set(list(title_specs.keys()) + list(bullet_specs.keys()))
    for k in all_keys:
        v_title = title_specs.get(k, '')
        v_bullet = bullet_specs.get(k, '')
        # 五点描述非空则优先，否则用标题提取
        merged[k] = v_bullet if v_bullet else v_title
    return merged


def aggregate_recommended_specs_from_spec_pack(df, packs, neg_counts=None, min_match_ratio: float = 0.10,
                                               product_type: str | None = None,
                                               min_same_type_samples: int = 10):
    """全品类通用版：按 SpecPack.spec_dimensions 的 extract_patterns 对 BSR 抓值，
    按命中率过滤低信号维度，返回 [(维度名, 建议规格, 数据依据, 优先级), ...]。

    数据源：
    - 维度清单 = LLM SpecAnalyzer 根据本次品类识别的 spec_dimensions（已注入 display_name 品类提示）
    - 每个 ASIN 的值 = 用 dim.extract_patterns 从 BSR 的 Product Title + Bullet Points/Product Overview 抓取
    - 命中阈值 = min_match_ratio（默认 10%），低于该阈值的维度不进报告，避免 LED 维度串到电池/充气机品类

    product_type:
    - 传入首推入场品类名时，先按 df['product_type'] == product_type 过滤再聚合，避免跨类型污染。
    - 同类型样本数 < min_same_type_samples 时自动退回全量聚合（样本太少算不出有效参数）。
    - 返回结果元组的"数据依据"列会注明 "基于 N 个同类型样本" 或 "样本不足回退全量"。

    不走 aggregate_recommended_specs() 那套 LED 专属 extract_specs_from_bullets 硬编码。
    """
    if packs is None or not packs.is_spec_real():
        return None  # 调用方用 None 判断是否 fallback 到 LED 老函数
    dims = list(packs.spec.spec_dimensions)
    if not dims:
        return None

    # 同类型筛选 + 样本数兜底：样本太少时退回全量，避免参数为空
    same_type_used = False
    if product_type and 'product_type' in df.columns:
        df_same = df[df['product_type'] == product_type]
        if len(df_same) >= min_same_type_samples:
            df = df_same
            same_type_used = True

    n = len(df)
    if n == 0:
        return []

    from llm.analyzers.spec_analyzer import extract_specs_by_dimensions
    import statistics as _stats
    from collections import Counter

    # 识别 title 列 + 所有长文本列（bullet/feature/overview/description/五点/产品描述）
    _LONG_KEYS = ('bullet', 'feature', 'overview', 'description', '五点', '产品描述', 'product detail')
    title_col = None
    long_cols: list[str] = []
    for c in df.columns:
        cl = str(c).lower()
        if title_col is None and ('product title' in cl or cl == 'title'):
            title_col = c
        if any(k in cl for k in _LONG_KEYS):
            long_cols.append(c)
    if title_col is None:
        title_col = 'Product Title' if 'Product Title' in df.columns else None
    if title_col is None:
        return []

    # 每个维度收集命中的值
    dim_values: dict[str, list[str]] = {d.name: [] for d in dims}
    for _, row in df.iterrows():
        title = str(row.get(title_col, '') or '')
        bullets = ' '.join(
            str(row[c]) for c in long_cols
            if pd.notna(row.get(c)) and str(row.get(c)).strip()
        )
        specs = extract_specs_by_dimensions(title, bullets, dims)
        for dname, dval in specs.items():
            if dval:
                dim_values[dname].append(str(dval))

    # 反向查找数字部分用于算中位/P75（有单位的去单位）
    def _numeric_part(s: str) -> float | None:
        if not s:
            return None
        m = _re.search(r'(\d+(?:\.\d+)?)', s)
        return float(m.group(1)) if m else None

    results: list[tuple[str, str, str, str]] = []
    if neg_counts is None:
        neg_counts = {}

    for d in dims:
        vals = dim_values.get(d.name, [])
        match_count = len(vals)
        # 回填到 SpecPack（供后续 UI/debug 用）
        try:
            d.match_count = match_count
            d.sample_values = vals[:10]
        except Exception:
            pass

        if match_count / n < min_match_ratio:
            # 低于阈值 —— 维度信号太弱，不进报告
            continue

        importance = (d.importance or '').strip()
        is_core = importance.startswith('核心') or importance == 'core'
        priority = 'P1-必备' if is_core else 'P2-重要'

        unit_suffix = d.unit or ''
        # 数值型 vs 分类型判别：以 vals 是否主体为"纯数字（可带小数点）"为准
        # 关键修复：避免 "lifepo4" / "3 aa" 这种字符串里含数字的值被误判为数值型 → median 抓出无意义数字
        # 纯数字 ≥50%：数值型；否则分类型（按 TOP 取值，保留原文）
        _pure_num_re = _re.compile(r'^\s*\d+(?:\.\d+)?\s*$')
        _pure_count = sum(1 for v in vals if _pure_num_re.match(str(v).strip()))
        _is_numeric_dim = (len(vals) > 0 and _pure_count / len(vals) >= 0.5)

        # 复合维度（unit 为空 + 数值型 vals 含字母尾部）：补全 unit 作为后缀
        # 例：vals=["3","3","4","3"]（纯数字，由完整短语 pattern 删 unit 后剩下）+ 原文常带 aa
        if _is_numeric_dim and not unit_suffix:
            _suf_counter: Counter = Counter()
            for v in vals:
                m_suf = _re.match(r'\s*\d+(?:\.\d+)?\s*(.*)$', str(v).strip(), flags=_re.IGNORECASE)
                if m_suf:
                    suf = m_suf.group(1).strip()
                    if suf:
                        _suf_counter[suf.lower()] += 1
            if _suf_counter:
                top_suf, top_cnt = _suf_counter.most_common(1)[0]
                if top_cnt / max(sum(_suf_counter.values()), 1) >= 0.5:
                    unit_suffix = ' ' + top_suf

        nums = [x for x in (_numeric_part(v) for v in vals) if x is not None and x > 0] if _is_numeric_dim else []
        if _is_numeric_dim and nums and len(nums) >= max(3, int(match_count * 0.5)):
            med = _stats.median(nums)
            p75 = sorted(nums)[int(len(nums) * 0.75)] if len(nums) >= 4 else max(nums)
            # 格式化数字，整数就不带小数
            def _fmt(x: float) -> str:
                return f'{int(x)}' if abs(x - int(x)) < 0.05 else f'{x:.1f}'
            rec_value = f'≥{_fmt(p75)}{unit_suffix}' if p75 > med else f'{_fmt(med)}{unit_suffix}'
            basis = f'{match_count}/{n}竞品标注，中位数 {_fmt(med)}{unit_suffix}，P75={_fmt(p75)}{unit_suffix}'
        else:
            # 分类型维度：取 TOP 1-3 值
            # 先按 + 拆组合（如 "磁吸+挂钩"）
            tokens: list[str] = []
            for v in vals:
                for t in str(v).split('+'):
                    t = t.strip()
                    if t:
                        tokens.append(t)
            if not tokens:
                continue
            ctr = Counter(tokens)
            top = ctr.most_common(3)
            top_names = [t for t, _ in top]
            rec_value = '+'.join(top_names[:2]) if len(top_names) >= 2 else top_names[0]
            basis = f'{match_count}/{n}竞品标注（TOP值：{top[0][0]}×{top[0][1]}）'

        # 在数据依据末尾注明样本范围，便于报告读者判断参数适用性
        if same_type_used and product_type:
            basis += f'｜样本范围：{product_type}（{n}个）'
        elif product_type:
            basis += f'｜样本范围：全类目（同类型样本不足{min_same_type_samples}回退全量，{n}个）'

        results.append((d.name, rec_value, basis, priority))

    # 按优先级排序，核心靠前
    results.sort(key=lambda x: 0 if x[3].startswith('P1') else (1 if x[3].startswith('P2') else 2))
    return results


def aggregate_recommended_specs(all_specs_list, neg_counts=None):
    """[Deprecated fallback]
    已被 `aggregate_recommended_specs_from_spec_pack()` 替代（走 LLM SpecPack 真值）。
    原函数里写死了 LED 专属维度（光通量/电池 mAh/防水 IP/光照模式/磁力/旋转角度），
    对非 LED 品类会产生跨品类污染，因此保留函数签名但返回空列表 —— 让 Sheet 10「推荐功能参数」
    段在 SpecPack 失败时整段跳过，比硬塞 LED 维度更安全。
    """
    import logging as _logging
    _logging.getLogger(__name__).warning(
        "[DeprecatedFallback] aggregate_recommended_specs() 返回空 —— "
        "SpecPack 未产出时 Sheet 10 推荐功能参数段会跳过渲染，不回退到 LED 硬编码"
    )
    return []


def _lc_dim_sales_yoy(market_data):
    """销量趋势同比。数据源：Market sell_trends（BSR 表无销量时序）。"""
    if not market_data or not isinstance(market_data, dict):
        return {'score': None, 'label': '数据缺失', 'note': '未上传 Market 文件', 'value': None}
    sell_df = market_data.get('sell_trends')
    if sell_df is None or len(sell_df) < 12:
        return {'score': None, 'label': '数据缺失',
                'note': 'Market 文件未含 sell_trends sheet 或月份数据 <12 月',
                'value': None}
    try:
        sell = sell_df.copy()
        sell.columns = [str(c).strip() for c in sell.columns]
        sales_c = sell.columns[1] if len(sell.columns) > 1 else None
        if sales_c is None:
            return {'score': None, 'label': '数据缺失', 'note': 'sell_trends 缺销量列', 'value': None}
        s = pd.to_numeric(sell[sales_c], errors='coerce').dropna().values
        if len(s) < 24:
            return {'score': None, 'label': '数据缺失',
                    'note': f'sell_trends 仅 {len(s)} 月（<24 月，无法算同比）',
                    'value': None}
        recent12 = float(s[-12:].sum())
        prev12 = float(s[-24:-12].sum())
        if prev12 <= 0:
            return {'score': None, 'label': '数据缺失', 'note': '上年同期销量为 0', 'value': None}
        yoy = (recent12 - prev12) / prev12
        if yoy > 0.10:
            sc, label = 1, f'+1 同比 +{yoy:.1%}'
        elif yoy < -0.10:
            sc, label = -1, f'-1 同比 {yoy:.1%}'
        else:
            sc, label = 0, f'0 同比 {yoy:+.1%}（平稳）'
        return {'score': sc, 'label': label, 'note': '来自 Market sell_trends', 'value': yoy}
    except Exception as _e:
        return {'score': None, 'label': '数据缺失', 'note': f'计算错误：{_e}', 'value': None}


def _lc_dim_search_yoy(market_data):
    """搜索量趋势同比。数据源：Market demand_trends 各核心词加总。"""
    if not market_data or not isinstance(market_data, dict):
        return {'score': None, 'label': '数据缺失', 'note': '未上传 Market 文件', 'value': None}
    demand_df = market_data.get('demand_trends')
    if demand_df is None or len(demand_df) < 12:
        return {'score': None, 'label': '数据缺失',
                'note': 'Market 文件未含 demand_trends sheet 或月份数据 <12 月',
                'value': None}
    try:
        d = demand_df.copy()
        d.columns = [str(c).strip() for c in d.columns]
        if len(d.columns) < 2:
            return {'score': None, 'label': '数据缺失', 'note': 'demand_trends 仅 1 列', 'value': None}
        month_c = d.columns[0]
        kw_cols = list(d.columns[1:])
        d[month_c] = d[month_c].astype(str)
        d = d[d[month_c].str.match(r'^\d{4}-\d{2}$', na=False)].sort_values(month_c)
        if len(d) < 24:
            return {'score': None, 'label': '数据缺失',
                    'note': f'demand_trends 仅 {len(d)} 月有效数据（<24 月）',
                    'value': None}
        total = d[kw_cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1).values
        recent12 = float(total[-12:].sum())
        prev12 = float(total[-24:-12].sum())
        if prev12 <= 0:
            return {'score': None, 'label': '数据缺失', 'note': '上年同期搜索量为 0', 'value': None}
        yoy = (recent12 - prev12) / prev12
        if yoy > 0.10:
            sc, label = 1, f'+1 同比 +{yoy:.1%}'
        elif yoy < -0.10:
            sc, label = -1, f'-1 同比 {yoy:.1%}'
        else:
            sc, label = 0, f'0 同比 {yoy:+.1%}（平稳）'
        return {'score': sc, 'label': label,
                'note': f'来自 Market demand_trends，{len(kw_cols)} 词加总',
                'value': yoy}
    except Exception as _e:
        return {'score': None, 'label': '数据缺失', 'note': f'计算错误：{_e}', 'value': None}


def _lc_dim_new_contribution(market_data):
    """新品贡献与成功率：近 2 年发布的产品销量占比 + 单品效率比。
    数据源：Market 文件 Publication Time Trends sheet（卖家精灵全市场聚合，非 BSR 反推）。
    - 销量贡献率：近 2 年（含当年）新品的销量占比之和
    - 效率比：新品销量占比 / 新品商品数占比 → >1 表示新品平均效率高于老品（成功），<1 反之
    """
    if not market_data or not isinstance(market_data, dict):
        return {'score': None, 'label': '数据缺失', 'note': '未上传 Market 文件', 'value': None}
    pub_df = market_data.get('publication_time_trends')
    if pub_df is None or len(pub_df) == 0:
        return {'score': None, 'label': '数据缺失',
                'note': 'Market 文件未含 Publication Time Trends sheet',
                'value': None}
    try:
        d = pub_df.copy()
        col_aliases = {
            'Launch Years': ['launch years', '发布年份', '发布年', '上架年份', '年份'],
            'Sales Proportion': ['sales proportion', '销量占比', '销售占比'],
            'Products': ['products', '商品数'],
        }
        for tgt, aliases in col_aliases.items():
            if tgt in d.columns:
                continue
            for c in d.columns:
                if str(c).strip().lower() in aliases:
                    d = d.rename(columns={c: tgt})
                    break
        if 'Launch Years' not in d.columns or 'Sales Proportion' not in d.columns:
            return {'score': None, 'label': '数据缺失',
                    'note': 'Publication Time Trends 缺 Launch Years / Sales Proportion 列',
                    'value': None}
        d['Launch Years'] = pd.to_numeric(d['Launch Years'], errors='coerce')
        d['Sales Proportion'] = pd.to_numeric(d['Sales Proportion'], errors='coerce')
        d = d.dropna(subset=['Launch Years', 'Sales Proportion'])
        if len(d) == 0:
            return {'score': None, 'label': '数据缺失',
                    'note': 'Launch Years 数据全为空', 'value': None}
        cur_year = datetime.now().year
        new_mask = d['Launch Years'] >= (cur_year - 1)
        new_sales_share = float(d.loc[new_mask, 'Sales Proportion'].sum())
        eff_note = ''
        eff_ratio = None
        new_products_share = None
        if 'Products' in d.columns:
            try:
                d['Products'] = pd.to_numeric(d['Products'], errors='coerce').fillna(0)
                tot_products = float(d['Products'].sum())
                new_products = float(d.loc[new_mask, 'Products'].sum())
                if tot_products > 0 and new_products > 0:
                    new_products_share = new_products / tot_products
                    eff_ratio = new_sales_share / new_products_share
                    eff_note = f'｜效率比 {eff_ratio:.2f}（销量占比 {new_sales_share:.0%} / 商品数占比 {new_products_share:.0%}）'
            except Exception:
                pass
        if new_sales_share > 0.30:
            sc = 1
            label = f'+1 近 2 年新品销量贡献 {new_sales_share:.1%}（市场对新品友好）{eff_note}'
        elif new_sales_share < 0.10:
            sc = -1
            label = f'-1 近 2 年新品销量贡献 {new_sales_share:.1%}（被老品垄断，新品难起）{eff_note}'
        else:
            sc = 0
            label = f'0 近 2 年新品销量贡献 {new_sales_share:.1%}（中等）{eff_note}'
        return {'score': sc, 'label': label,
                'note': '来自 Market 文件 Publication Time Trends',
                'value': new_sales_share,
                'efficiency_ratio': eff_ratio,
                'new_products_share': new_products_share}
    except Exception as _e:
        return {'score': None, 'label': '数据缺失', 'note': f'计算错误：{_e}', 'value': None}


def _lc_dim_brand_concentration(df, market_data=None):
    """品牌集中度 CR5：优先用 Market 文件 Sales Proportion（与全局 cr5 同源），
    缺失时退到 BSR Brand value_counts。高集中=负面（-1）。"""
    cr5 = None
    note = ''
    if isinstance(market_data, dict):
        bd = market_data.get('brand_concentration')
        if bd is not None and 'Sales Proportion' in getattr(bd, 'columns', []):
            try:
                sp = pd.to_numeric(bd['Sales Proportion'], errors='coerce').dropna()
                if len(sp) >= 5:
                    cr5 = float(sp.head(5).sum())
                    note = '来自 Market 文件 Sales Proportion'
            except Exception:
                pass
    if cr5 is None:
        if df is None or len(df) == 0:
            return {'score': None, 'label': '数据缺失', 'note': 'BSR 数据为空', 'value': None}
        if 'Brand' not in df.columns:
            return {'score': None, 'label': '数据缺失', 'note': 'BSR 表缺 Brand 列', 'value': None}
        try:
            brands = df['Brand'].dropna().astype(str)
            brands = brands[brands.str.strip() != '']
            if len(brands) == 0:
                return {'score': None, 'label': '数据缺失', 'note': 'Brand 列全为空', 'value': None}
            top5 = brands.value_counts().head(5).sum()
            cr5 = top5 / len(brands)
            note = '来自 BSR 表 Brand value_counts（兜底）'
        except Exception as _e:
            return {'score': None, 'label': '数据缺失', 'note': f'计算错误：{_e}', 'value': None}
    if cr5 > 0.60:
        sc, label = -1, f'-1 CR5={cr5:.0%}（高度集中，新品空间小）'
    elif cr5 < 0.40:
        sc, label = 1, f'+1 CR5={cr5:.0%}（分散，新品有空间）'
    else:
        sc, label = 0, f'0 CR5={cr5:.0%}（中等集中度）'
    return {'score': sc, 'label': label, 'note': note, 'value': cr5}


def _lc_dim_price_yoy(market_data):
    """价格趋势同比。数据源：Market sell_trends（销售额/销量=均价时序）。
    若 sell_trends 缺销售额列，标数据缺失。"""
    if not market_data or not isinstance(market_data, dict):
        return {'score': None, 'label': '数据缺失', 'note': '未上传 Market 文件', 'value': None}
    sell_df = market_data.get('sell_trends')
    if sell_df is None or len(sell_df) < 12:
        return {'score': None, 'label': '数据缺失',
                'note': 'Market 文件未含 sell_trends sheet 或月份数据 <12 月（无价格时序）',
                'value': None}
    try:
        sell = sell_df.copy()
        sell.columns = [str(c).strip() for c in sell.columns]
        if len(sell.columns) < 3:
            return {'score': None, 'label': '数据缺失',
                    'note': 'sell_trends 仅 2 列，无销售额无法算均价', 'value': None}
        sales_c = sell.columns[1]
        rev_c2 = sell.columns[2]
        s_sales = pd.to_numeric(sell[sales_c], errors='coerce').fillna(0).values
        s_rev = pd.to_numeric(sell[rev_c2], errors='coerce').fillna(0).values
        if len(s_sales) < 24:
            return {'score': None, 'label': '数据缺失',
                    'note': f'sell_trends 仅 {len(s_sales)} 月（<24 月，无法算同比）', 'value': None}
        recent_sales = s_sales[-12:].sum()
        recent_rev = s_rev[-12:].sum()
        prev_sales = s_sales[-24:-12].sum()
        prev_rev = s_rev[-24:-12].sum()
        if recent_sales <= 0 or prev_sales <= 0:
            return {'score': None, 'label': '数据缺失', 'note': '销量为 0 无法算均价', 'value': None}
        recent_avg = recent_rev / recent_sales
        prev_avg = prev_rev / prev_sales
        if prev_avg <= 0:
            return {'score': None, 'label': '数据缺失', 'note': '上年同期均价为 0', 'value': None}
        yoy = (recent_avg - prev_avg) / prev_avg
        if yoy > 0.05:
            sc, label = 1, f'+1 均价同比 +{yoy:.1%}（${recent_avg:.1f} vs ${prev_avg:.1f}）'
        elif yoy < -0.05:
            sc, label = -1, f'-1 均价同比 {yoy:.1%}（${recent_avg:.1f} vs ${prev_avg:.1f}，下行）'
        else:
            sc, label = 0, f'0 均价同比 {yoy:+.1%}（${recent_avg:.1f}，稳定）'
        return {'score': sc, 'label': label,
                'note': '来自 Market sell_trends 销售额/销量推算',
                'value': yoy}
    except Exception as _e:
        return {'score': None, 'label': '数据缺失', 'note': f'计算错误：{_e}', 'value': None}


def infer_lifecycle_stage(pub_trends_df, df=None, market_data=None):
    """
    多维度（5 维）综合判断类目生命周期。
    维度：销量趋势 / 搜索趋势 / 新品贡献率 / 品牌集中度 / 价格趋势
    每维评分 {+1, 0, -1, None=数据缺失}，「数据缺失」不计入分母。
    返回 (stage, reason_text, score_detail)
    - stage: 成长期 / 成熟期 / 成熟晚期 / 衰退期 / 数据不足
    - reason_text: 一行简短说明（向后兼容旧调用 stage, reason = ...）
    - score_detail: dict[dim_name -> {score, label, note, value}]，供 Sheet 8 渲染明细
    """
    dims = {
        '销量趋势': _lc_dim_sales_yoy(market_data),
        '搜索趋势': _lc_dim_search_yoy(market_data),
        '新品贡献率': _lc_dim_new_contribution(market_data),
        '品牌集中度': _lc_dim_brand_concentration(df, market_data=market_data),
        '价格趋势': _lc_dim_price_yoy(market_data),
    }
    valid_scores = [d['score'] for d in dims.values() if d['score'] is not None]
    valid_n = len(valid_scores)
    total_n = len(dims)
    if valid_n < 2:
        missing_dims = [k for k, v in dims.items() if v['score'] is None]
        reason = f'数据不足（仅 {valid_n}/{total_n} 维有效，缺失：{"、".join(missing_dims)}）'
        return ('数据不足', reason, dims)

    # 用决策表硬规则判定 stage（去掉 avg 阈值映射）
    stage, stage_reason, _matched_rule_id = _classify_stage_by_decision_table(dims)
    reason = stage_reason
    if df is not None and len(df) < 30:
        reason += f'｜⚠️ 样本量小（BSR n={len(df)}），结论参考性有限'
    return (stage, reason, dims)


_LIFECYCLE_DECISION_RULES = [
    # 顺序即优先级，第一条匹配的规则胜出
    {
        'id': 'R16',
        'sales': '*',
        'search': '*',
        'condition': '价格 -1 + 销量未涨',
        'stage': '衰退期',
        'reason': '价格 -1（均价同比下行）+ 销量未上涨 → 典型价格战衰退',
        'match': lambda s: s['price'] == -1 and s['sales_eff'] != 1,
    },
    {
        'id': 'R15',
        'sales': '-1',
        'search': '-1',
        'condition': '—',
        'stage': '衰退期',
        'reason': '销量 -1 + 搜索 -1 → 双向下行，趋势性衰退',
        'match': lambda s: s['sales_eff'] == -1 and s['search_eff'] == -1,
    },
    {
        'id': 'R12',
        'sales': '-1',
        'search': '+1',
        'condition': '—',
        'stage': '成熟晚期',
        'reason': '销量已跌但搜索回升，可能转折但未确认',
        'match': lambda s: s['sales_eff'] == -1 and s['search_eff'] == 1,
    },
    {
        'id': 'R13',
        'sales': '-1',
        'search': '0',
        'condition': '价格 +1',
        'stage': '成熟晚期',
        'reason': '销量在跌但价格上行，可能高端化转型',
        'match': lambda s: s['sales_eff'] == -1 and s['search_eff'] == 0 and s['price'] == 1,
    },
    {
        'id': 'R14',
        'sales': '-1',
        'search': '0',
        'condition': '其他',
        'stage': '衰退期',
        'reason': '销量 -1 + 搜索持平 + 无价格升级支撑 → 进入衰退',
        'match': lambda s: s['sales_eff'] == -1 and s['search_eff'] == 0,
    },
    {
        'id': 'R11',
        'sales': '0',
        'search': '-1',
        'condition': '—',
        'stage': '成熟晚期',
        'reason': '销量持平 + 搜索 -1 → 老用户复购但需求见顶',
        'match': lambda s: s['sales_eff'] == 0 and s['search_eff'] == -1,
    },
    {
        'id': 'R8',
        'sales': '0',
        'search': '0',
        'condition': '价格 +1',
        'stage': '成熟期',
        'reason': '销量与搜索持平 + 价格 +1 → 价值升级型成熟期',
        'match': lambda s: s['sales_eff'] == 0 and s['search_eff'] == 0 and s['price'] == 1,
    },
    {
        'id': 'R9',
        'sales': '0',
        'search': '0',
        'condition': '价格 -1',
        'stage': '成熟晚期',
        'reason': '销量与搜索持平 + 价格 -1 → 开始有价格压力',
        'match': lambda s: s['sales_eff'] == 0 and s['search_eff'] == 0 and s['price'] == -1,
    },
    {
        'id': 'R10',
        'sales': '0',
        'search': '0',
        'condition': '其他',
        'stage': '成熟晚期',
        'reason': '5 维全部中性或无信号，无增长动力',
        'match': lambda s: s['sales_eff'] == 0 and s['search_eff'] == 0,
    },
    {
        'id': 'R7',
        'sales': '0',
        'search': '+1',
        'condition': '—',
        'stage': '成长期',
        'reason': '销量未涨但搜索 +1 → 需求扩张窗口期（销量 3-6 月内可能跟上）',
        'match': lambda s: s['sales_eff'] == 0 and s['search_eff'] == 1,
    },
    {
        'id': 'R6',
        'sales': '+1',
        'search': '-1',
        'condition': '—',
        'stage': '成熟期',
        'reason': '销量 +1 但搜索 -1 → 老用户复购拉动需求见顶',
        'match': lambda s: s['sales_eff'] == 1 and s['search_eff'] == -1,
    },
    {
        'id': 'R4',
        'sales': '+1',
        'search': '0',
        'condition': '价格 +1',
        'stage': '成熟期',
        'reason': '销量 +1 + 搜索持平 + 价格 +1 → 消费升级型成熟期',
        'match': lambda s: s['sales_eff'] == 1 and s['search_eff'] == 0 and s['price'] == 1,
    },
    {
        'id': 'R5',
        'sales': '+1',
        'search': '0',
        'condition': '其他',
        'stage': '成长期',
        'reason': '销量 +1 + 搜索未跌 → 市场仍有动力',
        'match': lambda s: s['sales_eff'] == 1 and s['search_eff'] == 0,
    },
    {
        'id': 'R1',
        'sales': '+1',
        'search': '+1',
        'condition': 'CR5 > 60%（品牌已锁定）',
        'stage': '成熟期',
        'reason': '销量 +1 + 搜索 +1 但 CR5 高度集中 → 新品空间小，归成长后期/成熟期',
        'match': lambda s: s['sales_eff'] == 1 and s['search_eff'] == 1 and s['brand'] == -1,
    },
    {
        'id': 'R2',
        'sales': '+1',
        'search': '+1',
        'condition': '新品贡献高但效率比 < 1',
        'stage': '成熟期',
        'reason': '销量 +1 + 搜索 +1 + 新品贡献高但单品效率低 → 同质化卷归成熟期',
        'match': lambda s: (s['sales_eff'] == 1 and s['search_eff'] == 1
                            and s['new'] == 1 and s['eff_ratio'] is not None and s['eff_ratio'] < 1),
    },
    {
        'id': 'R3',
        'sales': '+1',
        'search': '+1',
        'condition': '其他（典型黄金窗口）',
        'stage': '成长期',
        'reason': '销量 +1 + 搜索 +1 → 双向正面，新品有空间且品牌未锁定',
        'match': lambda s: s['sales_eff'] == 1 and s['search_eff'] == 1,
    },
]


def _classify_stage_by_decision_table(score_detail):
    """硬规则决策表：5 维评分 → (stage, stage_reasoning, matched_rule_id)。

    遍历 _LIFECYCLE_DECISION_RULES 常量表，按优先级（顺序）匹配第一条命中规则。
    返回 (stage, reason_text, matched_rule_id)；未命中或数据不足时 matched_rule_id 为 None。
    """
    sales = (score_detail.get('销量趋势') or {}).get('score')
    search = (score_detail.get('搜索趋势') or {}).get('score')
    price = (score_detail.get('价格趋势') or {}).get('score')
    new = (score_detail.get('新品贡献率') or {}).get('score')
    brand = (score_detail.get('品牌集中度') or {}).get('score')
    eff_ratio = (score_detail.get('新品贡献率') or {}).get('efficiency_ratio')

    if sales is None and search is None:
        return ('数据不足', '销量与搜索两个核心维度均缺数据，无法判定阶段', None)

    state = {
        'sales_eff': sales if sales is not None else 0,
        'search_eff': search if search is not None else 0,
        'price': price,
        'new': new,
        'brand': brand,
        'eff_ratio': eff_ratio,
    }

    matched = None
    for rule in _LIFECYCLE_DECISION_RULES:
        try:
            if rule['match'](state):
                matched = rule
                break
        except Exception:
            continue

    if matched is None:
        return ('成熟晚期', '5 维组合未命中典型规则，谨慎归成熟晚期（兜底）', None)

    stage = matched['stage']
    reason = f"{matched['reason']}（{matched['id']}）"

    missing = [k for k in ['销量趋势', '搜索趋势', '价格趋势', '新品贡献率', '品牌集中度']
               if (score_detail.get(k) or {}).get('score') is None]
    if missing:
        reason += f'｜⚠️ 注：{", ".join(missing)} 数据缺失（按 0 中性参与判定）'

    return (stage, reason, matched['id'])


def _classify_market_pattern_by_rules(score_detail):
    """[规则兜底版] LLM 不可用时为每维生成一句基础分析 + 综合结论。

    返回 dict：
    {
        'dimension_analyses': [{'dimension': '<维度名>', 'analysis': '<1-2 句基础分析>'}, ...],
        'verdict': '<✅ 推荐 / ⚠️ 谨慎 / ❌ 不建议>',
        'overall_conclusion': '<3-5 句综合结论 + 建议>',
    }

    每维分析按 +1/0/-1 给一句模板话；综合结论按 5 维 avg 给基础判断。
    """
    def _s(name):
        d = score_detail.get(name) or {}
        return d.get('score')

    sales = _s('销量趋势')
    search = _s('搜索趋势')
    new = _s('新品贡献率')
    brand = _s('品牌集中度')
    price = _s('价格趋势')

    new_dim = score_detail.get('新品贡献率') or {}
    eff_ratio = new_dim.get('efficiency_ratio')

    # 每维基础分析模板（按 +1/0/-1）
    dim_templates = {
        '销量趋势': {
            1: '销量同比正向增长，市场需求处于扩张通道，对新进入者而言增量空间相对充足。',
            0: '销量同比基本平稳，市场已进入存量阶段，新入场者抢份额需要从老品牌手里夺。',
            -1: '销量同比下行，市场需求在收缩，存在见顶或衰退风险，新入场需要谨慎。',
        },
        '搜索趋势': {
            1: '搜索量同比上涨，意味着用户兴趣和潜在需求在扩大，是真实需求侧的正向信号。',
            0: '搜索量同比平稳，需求侧没有新增动力，但也未见显著疲态。',
            -1: '搜索量同比下滑，新流量在减少；若销量未同步下跌，多半是老用户复购在撑，市场可能见顶。',
        },
        '新品贡献率': {
            1: f'近 2 年新品销量贡献偏高，{("效率比 ≥ 1，新品单品平均效率优于市场平均，新品成功率高" if eff_ratio is not None and eff_ratio >= 1 else "若效率比 < 1 则说明新品多但单品难做爆")}',
            0: '新品贡献中等，市场对新品有一定接纳度但不显著，机会与挑战并存。',
            -1: '新品贡献偏低，市场被老品牌占据，新品成功率不高，新入场需要在差异化上花更大功夫。',
        },
        '品牌集中度': {
            1: 'CR5 偏低，品牌格局分散，头部尚未筑起明确护城河，新品有突围空间。',
            0: 'CR5 处于中等水平，头部品牌已有一定份额但未垄断，新品需要清晰差异化才能立足。',
            -1: 'CR5 偏高，市场已被头部品牌高度集中，新品突围难度大、广告竞争激烈。',
        },
        '价格趋势': {
            1: '均价同比上行，消费者愿意为这个品类付更多钱，存在价值升级 / 高端化机会。',
            0: '均价同比稳定，无价格战也无升级，跟随市场定价即可。',
            -1: '均价同比下行，市场存在价格战压力，利润空间被压缩，新品低价无优势。',
        },
    }

    dimension_analyses = []
    for dim_name, dim in (score_detail or {}).items():
        sc = dim.get('score')
        if sc is None:
            note = dim.get('note', '原因未知')
            dimension_analyses.append({
                'dimension': dim_name,
                'analysis': f'⚠️ 数据缺失（{note}）。本维度未参与综合判定。',
            })
        else:
            dimension_analyses.append({
                'dimension': dim_name,
                'analysis': dim_templates.get(dim_name, {}).get(sc, '（无模板）'),
            })

    # 综合结论 + 进入决策的 verdict（按 5 维 avg 给基础判断；
    # stage / stage_reasoning 由 classify_market_pattern() 主入口的决策表统一覆盖，
    # 这里仅算 verdict + 兜底 overall_conclusion）
    valid = [s for s in [sales, search, new, brand, price] if s is not None]
    avg = sum(valid) / len(valid) if valid else 0.0
    if avg >= 0.4:
        verdict = '✅ 推荐进入'
        conclusion = (f'5 维综合评分 +{avg:.2f}，整体偏正面，市场整体处于成长态势，新进入者有较好的红利窗口。'
                      '建议把握 3-6 个月内的入场时机，跟随市场均价、做差异化卖点，避免低价同质化竞争。'
                      '同时密切跟踪销量与搜索量增速，防止后期跟风者大量进入导致拥挤。')
    elif avg >= 0.0:
        verdict = '⚠️ 谨慎进入'
        conclusion = (f'5 维综合评分 +{avg:.2f}，整体偏稳定，市场已进入成长后期或成熟期。'
                      '存在增量但红利有限，新入场需要清晰的差异化定位，并在细分场景或人群上找突破口。'
                      '建议先小批量测试 SKU，跑通后再放量。')
    elif avg >= -0.4:
        verdict = '⚠️ 不建议作为重点'
        conclusion = (f'5 维综合评分 {avg:+.2f}，整体偏弱，市场处于成熟晚期或疲态阶段。'
                      '增长动力不足，新入场难度大，不建议作为主推方向。'
                      '已在该品类的卖家应评估收缩或转型策略。')
    else:
        verdict = '❌ 不建议进入'
        conclusion = (f'5 维综合评分 {avg:+.2f}，整体偏负面，市场处于衰退区间。'
                      '多个维度同时走弱，新入场承担接最后一棒风险，不建议进入。'
                      '已在该品类的卖家建议尽快收缩 SKU 或转移到衍生品类。')

    return {
        # stage / stage_reasoning 由调用方（classify_market_pattern 主入口）的决策表覆盖
        'stage': '',
        'stage_reasoning': '',
        'dimension_analyses': dimension_analyses,
        'verdict': verdict,
        'overall_conclusion': conclusion,
    }


def _build_dim_summary_row(dim_name, score_detail, market_data, df, packs):
    """对每维返回 (评分文本, 指标1, 指标2, 指标3, 数据来源) 5 个字段，供 Sheet 8「5 维数据总览」表渲染。

    指标按维度语义不同，统一为 3 列紧凑数字。无数据时该列填「-」。
    """
    dim = (score_detail or {}).get(dim_name) or {}
    score = dim.get('score')
    if score is None:
        score_txt = '⚠️ 缺数据'
    elif score == 1:
        score_txt = '+1 ↑'
    elif score == -1:
        score_txt = '-1 ↓'
    else:
        score_txt = '0 →'

    m1 = m2 = m3 = '-'
    src = dim.get('note', '-')

    try:
        if dim_name == '销量趋势':
            sell_df = market_data.get('sell_trends') if isinstance(market_data, dict) else None
            if sell_df is not None and len(sell_df) >= 12:
                sell = sell_df.copy()
                sell.columns = [str(c).strip() for c in sell.columns]
                if len(sell.columns) >= 2:
                    s = pd.to_numeric(sell[sell.columns[1]], errors='coerce').dropna().values
                    if len(s) >= 24:
                        recent = float(s[-12:].sum())
                        prev = float(s[-24:-12].sum())
                        m1 = f'近 12 月 {recent:,.0f} 件'
                        m2 = f'上 12 月 {prev:,.0f} 件'
                        m3 = f'同比 {((recent-prev)/prev):+.1%}' if prev > 0 else '上期 0'
        elif dim_name == '搜索趋势':
            demand_df = market_data.get('demand_trends') if isinstance(market_data, dict) else None
            if demand_df is not None and len(demand_df) >= 12:
                d = demand_df.copy()
                d.columns = [str(c).strip() for c in d.columns]
                month_c = d.columns[0]
                kw_cols = list(d.columns[1:])
                d[month_c] = d[month_c].astype(str)
                d = d[d[month_c].str.match(r'^\d{4}-\d{2}$', na=False)].sort_values(month_c)
                if len(d) >= 24:
                    total = d[kw_cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1).values
                    recent = float(total[-12:].sum())
                    prev = float(total[-24:-12].sum())
                    m1 = f'近 12 月 {recent:,.0f} 次'
                    m2 = f'上 12 月 {prev:,.0f} 次'
                    m3 = f'同比 {((recent-prev)/prev):+.1%}' if prev > 0 else '上期 0'
                    src = f'demand_trends（{len(kw_cols)} 词加总）'
        elif dim_name == '新品贡献率':
            v = dim.get('value')
            eff = dim.get('efficiency_ratio')
            nps = dim.get('new_products_share')
            if v is not None:
                m1 = f'近 2 年销量贡献 {v:.1%}'
            if nps is not None:
                m2 = f'商品数占比 {nps:.0%}'
            if eff is not None:
                m3 = f'效率比 {eff:.2f}' + ('（≥1 新品高效）' if eff >= 1 else '（<1 同质化）')
        elif dim_name == '品牌集中度':
            bd = market_data.get('brand_concentration') if isinstance(market_data, dict) else None
            if bd is not None and 'Sales Proportion' in getattr(bd, 'columns', []):
                sp = pd.to_numeric(bd['Sales Proportion'], errors='coerce').dropna()
                if len(sp) >= 3:
                    cr3 = float(sp.head(3).sum())
                    cr5 = float(sp.head(5).sum()) if len(sp) >= 5 else None
                    cr10 = float(sp.head(10).sum()) if len(sp) >= 10 else None
                    m1 = f'CR3 = {cr3:.1%}'
                    m2 = f'CR5 = {cr5:.1%}' if cr5 is not None else '-'
                    m3 = f'CR10 = {cr10:.1%}' if cr10 is not None else '-'
                    name_col = 'Brand' if 'Brand' in bd.columns else (bd.columns[1] if bd.shape[1] > 1 else None)
                    if name_col:
                        names = bd[name_col].iloc[:3].astype(str).tolist()
                        pairs = [f'{n} {sp.iloc[i]:.1%}' for i, n in enumerate(names)]
                        src = 'Top3：' + ' / '.join(pairs)
            elif df is not None and 'Brand' in df.columns:
                brands = df['Brand'].dropna().astype(str)
                brands = brands[brands.str.strip() != '']
                if len(brands) > 0:
                    vc = brands.value_counts().head(5)
                    total = len(brands)
                    cr5 = vc.sum() / total
                    m1 = f'CR5 = {cr5:.1%}（BSR 兜底）'
                    pairs = [f'{n} {(c/total):.1%}' for n, c in vc.head(3).items()]
                    src = 'Top3：' + ' / '.join(pairs)
        elif dim_name == '价格趋势':
            sell_df = market_data.get('sell_trends') if isinstance(market_data, dict) else None
            if sell_df is not None and len(sell_df) >= 12:
                sell = sell_df.copy()
                sell.columns = [str(c).strip() for c in sell.columns]
                if len(sell.columns) >= 3:
                    s_sales = pd.to_numeric(sell[sell.columns[1]], errors='coerce').fillna(0).values
                    s_rev = pd.to_numeric(sell[sell.columns[2]], errors='coerce').fillna(0).values
                    if len(s_sales) >= 24:
                        rs = s_sales[-12:].sum()
                        ps = s_sales[-24:-12].sum()
                        rr = s_rev[-12:].sum()
                        pr = s_rev[-24:-12].sum()
                        if rs > 0 and ps > 0:
                            ra = rr / rs
                            pa = pr / ps
                            m1 = f'近 12 月均价 ${ra:.2f}'
                            m2 = f'上 12 月均价 ${pa:.2f}'
                            m3 = f'同比 {((ra-pa)/pa):+.1%}' if pa > 0 else '上期 0'
    except Exception:
        pass

    return score_txt, m1, m2, m3, src


def _build_dim_data_detail(dim_name, score_detail, market_data, df, packs):
    """[已废弃 / 保留兼容] 改用 _build_dim_summary_row 紧凑表格替代。"""
    lines = []
    try:
        if dim_name == '销量趋势':
            sell_df = market_data.get('sell_trends') if isinstance(market_data, dict) else None
            if sell_df is not None and len(sell_df) >= 12:
                sell = sell_df.copy()
                sell.columns = [str(c).strip() for c in sell.columns]
                if len(sell.columns) >= 2:
                    sales_c = sell.columns[1]
                    s = pd.to_numeric(sell[sales_c], errors='coerce').dropna().values
                    if len(s) >= 24:
                        recent = float(s[-12:].sum())
                        prev = float(s[-24:-12].sum())
                        delta = recent - prev
                        lines.append(f'近 12 月销量合计：{recent:,.0f} 件')
                        lines.append(f'上 12 月销量合计：{prev:,.0f} 件')
                        if prev > 0:
                            lines.append(f'同比增量：{delta:+,.0f} 件（{(delta/prev):+.1%}）')
                    elif len(s) >= 6:
                        lines.append(f'近 {len(s)} 月销量合计：{float(s.sum()):,.0f} 件（数据不足 24 月，无法算同比）')
        elif dim_name == '搜索趋势':
            demand_df = market_data.get('demand_trends') if isinstance(market_data, dict) else None
            if demand_df is not None and len(demand_df) >= 12:
                d = demand_df.copy()
                d.columns = [str(c).strip() for c in d.columns]
                month_c = d.columns[0]
                kw_cols = list(d.columns[1:])
                d[month_c] = d[month_c].astype(str)
                d = d[d[month_c].str.match(r'^\d{4}-\d{2}$', na=False)].sort_values(month_c)
                if len(d) >= 24:
                    total = d[kw_cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1).values
                    recent = float(total[-12:].sum())
                    prev = float(total[-24:-12].sum())
                    delta = recent - prev
                    lines.append(f'近 12 月搜索量合计：{recent:,.0f} 次（{len(kw_cols)} 词加总）')
                    lines.append(f'上 12 月搜索量合计：{prev:,.0f} 次')
                    if prev > 0:
                        lines.append(f'同比增量：{delta:+,.0f} 次（{(delta/prev):+.1%}）')
                    if kw_cols:
                        lines.append(f'核心关键词（{len(kw_cols)} 词）：' + ', '.join(str(k) for k in kw_cols[:5]))
        elif dim_name == '新品贡献率':
            new_dim = (score_detail or {}).get('新品贡献率') or {}
            v = new_dim.get('value')
            eff = new_dim.get('efficiency_ratio')
            nps = new_dim.get('new_products_share')
            if v is not None:
                lines.append(f'近 2 年新品销量贡献：{v:.1%}')
            if nps is not None:
                lines.append(f'近 2 年新品商品数占比：{nps:.0%}')
            if eff is not None:
                lines.append(f'效率比（销量占比 ÷ 商品数占比）：{eff:.2f}'
                             + ('（≥1：新品平均效率优于市场，成功率高）' if eff >= 1
                                else '（<1：新品多但单品平均效率低于市场）'))
            lines.append('详细年份分布见上方「发布年份」表')
        elif dim_name == '品牌集中度':
            bd = market_data.get('brand_concentration') if isinstance(market_data, dict) else None
            if bd is not None and 'Sales Proportion' in getattr(bd, 'columns', []):
                sp = pd.to_numeric(bd['Sales Proportion'], errors='coerce').dropna()
                if len(sp) >= 3:
                    cr3 = float(sp.head(3).sum())
                    cr5 = float(sp.head(5).sum()) if len(sp) >= 5 else None
                    cr10 = float(sp.head(10).sum()) if len(sp) >= 10 else None
                    name_col = 'Brand' if 'Brand' in bd.columns else (bd.columns[1] if bd.shape[1] > 1 else None)
                    top_n = min(5, len(sp))
                    if name_col:
                        names = bd[name_col].iloc[:top_n].astype(str).tolist()
                        pairs = [f'{n} {sp.iloc[i]:.1%}' for i, n in enumerate(names)]
                        lines.append('Top 品牌：' + ' / '.join(pairs))
                    lines.append(f'CR3 = {cr3:.1%}'
                                 + (f' / CR5 = {cr5:.1%}' if cr5 is not None else '')
                                 + (f' / CR10 = {cr10:.1%}' if cr10 is not None else ''))
            # BSR 表 Brand 兜底
            if not lines and df is not None and 'Brand' in df.columns:
                brands = df['Brand'].dropna().astype(str)
                brands = brands[brands.str.strip() != '']
                if len(brands) > 0:
                    vc = brands.value_counts().head(5)
                    total = len(brands)
                    pairs = [f'{n} {(c/total):.1%}' for n, c in vc.items()]
                    lines.append('Top5 品牌（BSR 表 ASIN 数占比兜底）：' + ' / '.join(pairs))
                    cr5_bsr = vc.sum() / total
                    lines.append(f'CR5 = {cr5_bsr:.1%}（按 BSR TOP100 内 ASIN 数）')
        elif dim_name == '价格趋势':
            sell_df = market_data.get('sell_trends') if isinstance(market_data, dict) else None
            if sell_df is not None and len(sell_df) >= 12:
                sell = sell_df.copy()
                sell.columns = [str(c).strip() for c in sell.columns]
                if len(sell.columns) >= 3:
                    sales_c = sell.columns[1]
                    rev_c2 = sell.columns[2]
                    s_sales = pd.to_numeric(sell[sales_c], errors='coerce').fillna(0).values
                    s_rev = pd.to_numeric(sell[rev_c2], errors='coerce').fillna(0).values
                    if len(s_sales) >= 24:
                        rs = s_sales[-12:].sum()
                        ps = s_sales[-24:-12].sum()
                        rr = s_rev[-12:].sum()
                        pr = s_rev[-24:-12].sum()
                        if rs > 0 and ps > 0:
                            ra = rr / rs
                            pa = pr / ps
                            lines.append(f'近 12 月均价：${ra:.2f}（销售额 ${rr:,.0f} / 销量 {rs:,.0f} 件）')
                            lines.append(f'上 12 月均价：${pa:.2f}（销售额 ${pr:,.0f} / 销量 {ps:,.0f} 件）')
                            if pa > 0:
                                lines.append(f'均价同比：{(ra-pa)/pa:+.1%}（{"上行/价值升级" if ra>pa else "下行/价格压力" if ra<pa else "持平"}）')
    except Exception:
        pass
    return lines


def _build_lifecycle_llm_input(score_detail, packs, df):
    """从 packs / df / score_detail 抽取品类元数据，组装给 LifecycleAnalyzer 的入参。"""
    category_name = ''
    top_brands = []
    top_keywords = []
    price_distribution = None
    segments = []
    china_pct = None

    if packs is not None:
        category_name = (getattr(packs, 'display_name', '')
                         or getattr(packs, 'category_id', '')
                         or '')
        market = getattr(packs, 'market', None)
        if market is not None:
            segs = getattr(market, 'product_segments', None) or []
            segments = [getattr(s, 'name', '') for s in segs if getattr(s, 'name', '')][:8]
            ladder = getattr(market, 'price_ladder', None) or []
            if ladder:
                price_distribution = {
                    getattr(b, 'band', f'band_{i}'): getattr(b, 'description', '')
                    for i, b in enumerate(ladder[:6])
                }
        traffic = getattr(packs, 'traffic', None)
        if traffic is not None:
            kws = getattr(traffic, 'top_keywords', None) or getattr(traffic, 'keywords', None) or []
            for kw in kws[:10]:
                if isinstance(kw, str):
                    top_keywords.append(kw)
                elif isinstance(kw, dict) and kw.get('keyword'):
                    top_keywords.append(kw['keyword'])
                else:
                    name = getattr(kw, 'keyword', None) or getattr(kw, 'name', None)
                    if name:
                        top_keywords.append(str(name))

    if df is not None and len(df) > 0:
        if 'Brand' in df.columns:
            try:
                rev_col = None
                for c in ['Monthly Revenue($)', 'Revenue', '月销售额', 'Monthly Revenue']:
                    if c in df.columns:
                        rev_col = c
                        break
                if rev_col:
                    by_brand = (df.assign(_rev=pd.to_numeric(df[rev_col], errors='coerce').fillna(0))
                                  .groupby('Brand')['_rev'].sum()
                                  .sort_values(ascending=False))
                    top_brands = [str(b) for b in by_brand.head(8).index.tolist() if str(b).strip()]
                else:
                    top_brands = [str(b) for b in df['Brand'].dropna().value_counts().head(8).index.tolist()]
            except Exception:
                pass
        if 'BuyBox Location' in df.columns:
            try:
                loc = df['BuyBox Location'].dropna().astype(str).str.upper()
                china_pct = float((loc.str.contains('CN|CHINA', regex=True, na=False)).mean())
            except Exception:
                pass

    new_dim = (score_detail or {}).get('新品贡献率') or {}
    return {
        'category_name': category_name,
        'score_detail': score_detail,
        'top_brands': top_brands,
        'top_keywords': top_keywords,
        'price_distribution': price_distribution,
        'segments': segments,
        'china_pct': china_pct,
        'new_contribution': new_dim.get('value'),
        'eff_ratio': new_dim.get('efficiency_ratio'),
    }


def classify_market_pattern(score_detail, packs=None, df=None):
    """对外入口：LLM 优先 + 规则模板兜底。

    score_detail: 来自 infer_lifecycle_stage 第三个返回值
    packs: 可选，传入则尝试用 LLM 生成品类专属洞察文案
    df: 可选 BSR DataFrame，用于抽取 top_brands / china_pct

    返回 dict：含 dimension_analyses（每维一段详细分析） / verdict / overall_conclusion / source
    source 字段为 'LLM' 或 'fallback-template'，供渲染层选择性提示。
    """
    # 决策表硬规则定 stage（不分 LLM/兜底，stage 由规则统一覆盖，保证可追溯）
    rule_stage, rule_reasoning, matched_rule_id = _classify_stage_by_decision_table(score_detail)

    if packs is not None:
        try:
            from llm.analyzers.lifecycle_analyzer import LifecycleAnalyzer
            from llm.client import make_client
            client = make_client(api_key=getattr(packs, '_api_key', None))
            analyzer = LifecycleAnalyzer(client=client)
            input_data = _build_lifecycle_llm_input(score_detail, packs, df)
            # 把决策表 stage 作为 hint 传给 LLM，让 LLM 围绕该 stage 写分析
            input_data['rule_stage'] = rule_stage
            input_data['rule_stage_reasoning'] = rule_reasoning
            insight = analyzer.run(input_data)
            if (insight and not getattr(insight, 'is_fallback', False)
                    and insight.overall_conclusion and insight.dimension_analyses):
                return {
                    'stage': rule_stage,                    # 决策表硬覆盖
                    'stage_reasoning': rule_reasoning,      # 决策表给的判定逻辑
                    'matched_rule_id': matched_rule_id,     # 命中的规则编号（如 'R3'）
                    'dimension_analyses': [
                        {'dimension': da.dimension, 'analysis': da.analysis}
                        for da in insight.dimension_analyses
                    ],
                    'verdict': insight.verdict,
                    'overall_conclusion': insight.overall_conclusion,
                    'source': 'LLM+DecisionTable',
                }
        except Exception as e:
            print(f"[Lifecycle] LLM 调用失败，退回规则模板：{e}")

    rule_result = _classify_market_pattern_by_rules(score_detail)
    rule_result['stage'] = rule_stage           # 决策表硬覆盖
    rule_result['stage_reasoning'] = rule_reasoning
    rule_result['matched_rule_id'] = matched_rule_id
    rule_result['source'] = 'fallback-template'
    return rule_result


# ============================================================
# 核心报告生成函数
# ============================================================
def generate_report(bsr_path, review_paths, output_path, market_path=None, keyword_path=None, packs: Packs | None = None):
    """
    通用化版本：保留硬编码排版/数值算法，6 处品类相关内容由 LLM Insight Pack 驱动。
    packs 为 None 时自动跑 prepare_packs（LLM 不可用时各 analyzer 自动降级到硬编码统计）。
    """
    # --- 1. 读取BSR数据 ---
    df = pd.read_excel(bsr_path, sheet_name='US')
    # 卖家精灵 BSR 有中文列名版本（"商品标题"/"品牌"/"月销量"等），统一映射为英文标准名，
    # 避免下游 df['Product Title'] 等硬编码崩溃。已是英文的文件幂等无影响。
    from llm.analyzers.bsr_analyzer import normalize_bsr_columns
    df = normalize_bsr_columns(df)

    # --- 1a. 准备 Insight Packs（若调用方未传入）---
    if packs is None:
        try:
            packs = prepare_packs(
                bsr_path=bsr_path,
                review_paths=review_paths,
                keyword_path=keyword_path,
                market_path=market_path,
                bsr_df=df,
            )
            print(f"[Packs] category_id={packs.category_id} display={packs.display_name} "
                  f"market_real={packs.is_market_real()} voc_real={packs.is_voc_real()} "
                  f"traffic_real={packs.is_traffic_real()} trend_real={packs.is_trend_real()} "
                  f"synth_real={packs.is_synthesis_real()}")
        except Exception as e:
            print(f"[Packs] 全部 Pack 生成失败，所有内容走原硬编码降级: {e}")
            packs = None

    # --- 1b. 读取 Market 文件（可选） ---
    market_data = load_market_data(market_path)

    # --- 1c. 读取关键词文件（可选） ---
    keyword_data = load_keyword_data(keyword_path)

    # 动态识别列名
    price_col = None
    rev_col = None
    for c in df.columns:
        c_lower = c.lower()
        if price_col is None and 'price' in c_lower:
            price_col = c
        if rev_col is None and 'revenue' in c_lower and 'variation' not in c_lower:
            rev_col = c

    # 数值转换
    for c in [price_col, rev_col, 'Monthly Sales', 'Rating', 'Ratings',
              'Available days', 'Gross Margin', 'FBA($)']:
        if c and c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')

    # 产品分类：视觉 LLM 主导 + 本地兜底匹配（品类无关，不重调视觉 LLM）。
    # 视觉对每个 ASIN 看「图+标题」给出 segment_name + product_type_free（自由描述）。
    # 三路解析：
    # 1) cached segment_name 在当前 segments 列表里 → 直接采信
    # 2) cached segment_name 失效（旧 taxonomy 的名字、unknown、空）→ 用 cached product_type_free
    #    在本地做关键词重叠匹配回当前 segments（毫秒级，不调 LLM）
    # 3) 视觉无数据 / 本地匹配也未命中 → 落回 classify_with_packs 现有文本匹配链
    visual_label_map = {}
    if packs is not None and packs.is_market_real():
        for label in getattr(packs.market, 'visual_labels', []) or []:
            asin = str(getattr(label, 'asin', '') or '').strip()
            if asin:
                visual_label_map[asin] = label

    _current_segment_names = set()
    if packs is not None and packs.is_market_real():
        _current_segment_names = {s.name for s in (packs.market.product_segments or []) if s.name}

    # 确定性分桶（v7：删 name_score + 调权）：
    # 聚合 LLM 只产出桶 taxonomy（name + keywords + material_attribute + form_attribute），不分配 ASIN；
    # 这里用代码对每个 ASIN 的视觉描述+材质+形态做加权打分，确定归属桶。
    #
    # v6 → v7 关键改动：
    #   - 删 name_score：CJK 连续段会被 tokenizer 合成单一长 token（如"式电子互动存钱罐"），
    #     子串匹配几乎不可能命中。聚合 LLM 已把桶名核心特征写进 keywords，无需 name 维度。
    #   - 材质权重 ×2 → ×3：结构化标签稳定，应优先于自由描述匹配。
    #
    # 评分维度（最终版）：
    #   维度 1：keywords 子串命中数（×1，核心信号——形状/形态/功能词都在这里）
    # + 维度 2：材质完全相等 × 3（最强结构化信号）
    # + 维度 3：形态匹配 × 2（次强结构化信号）
    #     form 4 路命中规则（任一即算）：
    #       a) seg.form_attribute == vision.form_label
    #       b) 双向子串（seg form ⊂ prod form 或反之）
    #       c) vision.form_label 出现在 seg 任一 keyword 里（如 seg form 写"动物形"
    #          但 keyword 含"塑料猪形"，vision form="猪形" → "猪形" ⊂ "塑料猪形" → 命中）
    #       d) seg.form_attribute 出现在产品自由描述里
    # 置信门槛：(keyword 命中 ≥ 1 OR form 命中) 且 总分 ≥ 2

    def _match_via_visual_text(text: str, material: str = "", form: str = ""):
        if not _current_segment_names or packs is None or not packs.is_market_real():
            return None
        if not text and not material and not form:
            return None
        text_lower = str(text or "").lower()
        prod_mat = str(material or "").strip().lower()
        prod_form = str(form or "").strip().lower()

        best_seg = None
        best_score = 0
        best_strong_signal = False
        for seg in (packs.market.product_segments or []):
            if not seg.name or seg.name not in _current_segment_names:
                continue

            # 维度 1：keywords 子串命中（×1）
            kw_score = 0
            for kw in (seg.representative_keywords or []):
                kw_lower = str(kw or '').strip().lower()
                if not kw_lower or len(kw_lower) < 2:
                    continue
                if kw_lower in text_lower:
                    kw_score += 1

            # 维度 2：材质相等（×3 加权，否则 ×1 弱加权）
            # 单键 form 聚类后桶内可能混多种材质，seg.material_attribute 仅是多数派；
            # 非多数派材质 ASIN 不应直接 0 分（否则 form 命中也擦不过门槛），给 1 分弱信号
            seg_mat = str(seg.material_attribute or '').strip().lower()
            mat_score = 3 if (seg_mat and prod_mat and seg_mat == prod_mat) else 1

            # 维度 3：形态匹配（×2，4 路命中规则）
            seg_form = str(getattr(seg, 'form_attribute', '') or '').strip().lower()
            form_match = False
            if prod_form:
                if seg_form and seg_form == prod_form:
                    form_match = True
                elif seg_form and len(seg_form) >= 2 and (seg_form in prod_form or prod_form in seg_form):
                    form_match = True
                elif len(prod_form) >= 2 and any(
                    prod_form in str(kw or '').strip().lower()
                    for kw in (seg.representative_keywords or [])
                ):
                    form_match = True
                elif seg_form and len(seg_form) >= 2 and seg_form in text_lower:
                    form_match = True
            form_score = 2 if form_match else 0

            total = kw_score + mat_score + form_score
            strong_signal = (kw_score >= 1) or form_match
            if total > best_score:
                best_score = total
                best_strong_signal = strong_signal
                best_seg = seg.name
        if best_strong_signal and best_score >= 2:
            return best_seg
        return None

    # 预构建 ASIN → segment 反查（来自 BucketAssigner 的 LLM 语义判定结果）
    _asin_to_segment: dict[str, str] = {}
    if packs is not None and packs.is_market_real():
        for _seg in (packs.market.product_segments or []):
            for _a in (_seg.member_asins or []):
                _asin_str = str(_a or '').strip()
                if _asin_str:
                    _asin_to_segment[_asin_str] = _seg.name

    def _resolve_product_type(row):
        asin = str(row.get('ASIN', '') or '').strip()
        title = row.get('Product Title', '')
        # Path 0：BucketAssigner LLM 语义判桶结果（最高优先）
        # 同数据 → 同 LLM 输出（cache 稳定）→ 跨次复现
        if asin and asin in _asin_to_segment:
            return _asin_to_segment[asin]
        label = visual_label_map.get(asin)
        if label is not None:
            ptf = str(getattr(label, 'product_type_free', '') or '').strip()
            mat = str(getattr(label, 'material_label', '') or '').strip()
            form = str(getattr(label, 'form_label', '') or '').strip()
            # Path 1：cached vision segment_name 命中且仍在当前 segments 列表里 → 采信
            seg = str(getattr(label, 'segment_name', '') or '').strip()
            if seg and seg.lower() != 'unknown' and seg in _current_segment_names:
                return seg
            # Path 2：用 product_type_free + material + form 做加权打分匹配（v7 keyword 兜底）
            matched = _match_via_visual_text(ptf, material=mat, form=form)
            if matched:
                return matched
        # Path 3：classify_with_packs 兜底（基于标题 token 与桶 keywords 的重叠匹配）
        fallback = classify_with_packs(title, packs, classify, asin=asin)
        # Path 4：所有路径都没匹中 → 归到「其他」杂物桶（透明保底，业务可见漏的 ASIN）
        if not fallback or fallback in ('未分类', '其他/通用款') or fallback not in _current_segment_names:
            return '其他'
        return fallback

    df['product_type'] = df.apply(_resolve_product_type, axis=1)

    df['visual_material'] = df['ASIN'].astype(str).map(
        lambda a: str(getattr(visual_label_map.get(a), 'material_label', '') or '').strip() or '-'
    )
    # Sheet 3 N 列「产品类型」专用：视觉 LLM 自由描述产品类型（不受 BSR segment 列表限制）
    # 解决 BSR LLM 切的 4-8 个 segment 装不下细分类型的问题。空值 fallback 到 BSR segment.name。
    # 其他 sheet（4/5/6/10）按 df['product_type']（视觉主导分类）做分细分析。
    df['visual_product_type'] = df.apply(
        lambda r: (
            str(getattr(visual_label_map.get(str(r.get('ASIN', ''))), 'product_type_free', '') or '').strip()
            or str(r.get('product_type', '') or '').strip()
            or '-'
        ),
        axis=1,
    )

    # --- 2. 读取评论数据 ---
    # 支持中英文列名映射
    column_mapping = {
        'Rating': '星级',
        'Title': '标题',
        'Content': '内容',
        'Date': '评论时间',
        'source_asin': 'source_asin'
    }

    all_reviews = []
    for f in review_paths:
        try:
            df_r = pd.read_excel(f, sheet_name=0)
            # 动态识别Rating列（支持中英文）
            rating_col = None
            title_col = None
            content_col = None
            date_col = None
            
            for c in df_r.columns:
                c_lower = str(c).lower()
                if rating_col is None and ('rating' in c_lower or '星级' in c):
                    rating_col = c
                if title_col is None and ('title' in c_lower or '标题' in c):
                    title_col = c
                if content_col is None and ('content' in c_lower or '内容' in c):
                    content_col = c
                if date_col is None and ('date' in c_lower or '评论时间' in c or '时间' in c):
                    date_col = c
            
            if rating_col is None:
                continue
                
            asin = os.path.basename(f).split('-')[0]
            df_r['source_asin'] = asin
            
            # 重命名列以统一使用英文
            rename_dict = {}
            if rating_col and rating_col != 'Rating':
                rename_dict[rating_col] = 'Rating'
            if title_col and title_col != 'Title':
                rename_dict[title_col] = 'Title'
            if content_col and content_col != 'Content':
                rename_dict[content_col] = 'Content'
            if date_col and date_col != 'Date':
                rename_dict[date_col] = 'Date'
            
            if rename_dict:
                df_r = df_r.rename(columns=rename_dict)
            
            all_reviews.append(df_r)
        except Exception as e:
            print(f"读取评论文件出错: {f}, 错误: {e}")
            continue

    rev_df = pd.concat(all_reviews, ignore_index=True) if all_reviews else pd.DataFrame()
    if len(rev_df) > 0 and 'Rating' in rev_df.columns:
        rev_df['Rating'] = pd.to_numeric(rev_df['Rating'], errors='coerce')
    low_rev = rev_df[rev_df['Rating'] <= 2].copy() if len(rev_df) > 0 and 'Rating' in rev_df.columns else pd.DataFrame()
    high_rev = rev_df[rev_df['Rating'] >= 4].copy() if len(rev_df) > 0 and 'Rating' in rev_df.columns else pd.DataFrame()

    # --- 4. 统计汇总 ---
    total_rev = df[rev_col].sum() if rev_col else 0
    total_sales = df['Monthly Sales'].sum() if 'Monthly Sales' in df.columns else 0
    avg_price = df[price_col].mean() if price_col else 0
    median_price = df[price_col].median() if price_col else 0
    avg_rating = df['Rating'].mean() if 'Rating' in df.columns else 0
    brand_count = df['Brand'].nunique() if 'Brand' in df.columns else 0
    # 中国卖家占比 —— 统一口径：BuyBox Location ∈ {'CN', 'CN(HK)'}
    # 与 core/packs_runtime.py::_build_stats_for_synthesis 保持同一判定，
    # 各 Sheet 显示都用 cn_cnt / cn_pct_total / cn_rev_pct_total 这三个变量，
    # 避免老代码里 4 套 `== 'CN'` / `.contains('CN')` 各写一处导致数值不一致
    _loc_col = 'BuyBox Location' if 'BuyBox Location' in df.columns else None
    if _loc_col:
        _loc_up = df[_loc_col].fillna('').astype(str).str.strip().str.upper()
        cn_mask_df = _loc_up.isin(['CN', 'CN(HK)']) | _loc_up.str.contains('CHINA')
        us_mask_df = _loc_up.eq('US')
    else:
        cn_mask_df = pd.Series([False] * len(df), index=df.index)
        us_mask_df = pd.Series([False] * len(df), index=df.index)
    cn_rev = df[cn_mask_df][rev_col].sum() if rev_col else 0
    us_rev = df[us_mask_df][rev_col].sum() if rev_col else 0
    cn_cnt = int(cn_mask_df.sum())
    us_cnt = int(us_mask_df.sum())
    cn_pct_total = (cn_cnt / len(df)) if len(df) > 0 else 0.0
    cn_rev_pct_total = (float(cn_rev) / float(df[rev_col].sum())) if rev_col and df[rev_col].sum() > 0 else 0.0

    # 按 BuyBox Location 分别取头部 TOP3 品牌（Sheet 2 row 7-8 说明列）
    def _top3_in_mask(mask):
        if rev_col is None or not mask.any() or 'Brand' not in df.columns:
            return '—'
        sub = df[mask].groupby('Brand')[rev_col].sum().sort_values(ascending=False).head(3)
        names = [str(b) for b in sub.index.tolist() if str(b).strip()]
        return '、'.join(names) if names else '—'

    _cn_top3_brand_str = _top3_in_mask(cn_mask_df)
    _us_top3_brand_str = _top3_in_mask(us_mask_df)

    # 产品类型聚合
    type_agg = df.groupby('product_type').agg(
        count=('ASIN', 'count'),
        avg_price=(price_col, 'mean'),
        min_price=(price_col, 'min'),
        max_price=(price_col, 'max'),
        total_sales=('Monthly Sales', 'sum'),
        total_revenue=(rev_col, 'sum'),
        avg_rating=('Rating', 'mean'),
    ).sort_values('total_revenue', ascending=False).reset_index()
    # 过滤兜底 segment：「其他/通用款」是 _ensure_full_coverage 收容 BSR LLM 漏装 ASIN 的内部桶，
    # 异质混合（粉色塑料 / 陶瓷 / 木质等都被扫进来），作为"产品细分类型"出现在 Sheet 4 推荐入场价表、
    # Sheet 6 上新方向、Sheet 10 综合评分里没有业务意义。这里统一从聚合表里剔除，
    # 下游所有按 type_agg 聚合的展示自动受益（被装入此桶的 ASIN 仍在 Sheet 3 BSR TOP100 列表里展示个体）。
    type_agg = type_agg[type_agg['product_type'] != '其他/通用款'].reset_index(drop=True)

    # 品牌聚合
    brand_agg = df.groupby('Brand').agg(
        sku_count=('ASIN', 'count'),
        total_rev=(rev_col, 'sum'),
        avg_rating=('Rating', 'mean'),
    ).sort_values('total_rev', ascending=False).reset_index()

    # 全局 CR5/CR10 集中度（Market 文件 Sales Proportion 优先 → BSR brand_agg 兜底）
    # 提前到此处计算，让 Sheet1/Sheet2/Sheet8/Sheet10 共用同一份 cr5（避免不同 Sheet 显示不同 CR5 值）
    cr5 = cr10 = None
    _brand_df_global = market_data.get('brand_concentration') if isinstance(market_data, dict) else None
    if _brand_df_global is not None and len(_brand_df_global) > 5:
        try:
            _sp_global = pd.to_numeric(_brand_df_global['Sales Proportion'], errors='coerce').dropna()
            cr5 = float(_sp_global.head(5).sum())
            cr10 = float(_sp_global.head(10).sum())
        except Exception:
            pass
    _total_rev_for_cr = brand_agg['total_rev'].sum() if 'total_rev' in brand_agg.columns else 0
    if cr5 is None and _total_rev_for_cr:
        try:
            cr5 = float(brand_agg.head(5)['total_rev'].sum() / _total_rev_for_cr)
        except Exception:
            pass
    if cr10 is None and _total_rev_for_cr:
        try:
            cr10 = float(brand_agg.head(10)['total_rev'].sum() / _total_rev_for_cr)
        except Exception:
            pass

    # 价格分布
    bins = [0, 15, 25, 35, 50, 70, 100, 999]
    labels = ['<$15', '$15-25', '$25-35', '$35-50', '$50-70', '$70-100', '>$100']
    df['price_band'] = pd.cut(df[price_col], bins=bins, labels=labels)
    price_dist = df['price_band'].value_counts().sort_index()

    # 年龄分布
    new_df = df[df['Available days'] < 365] if 'Available days' in df.columns else pd.DataFrame()
    bins_age = [0, 90, 180, 365, 730, 1460, 9999]
    labels_age = ['<3个月', '3-6个月', '6-12个月', '1-2年', '2-4年', '>4年']
    df['age_band'] = pd.cut(df['Available days'], bins=bins_age, labels=labels_age)
    age_dist = df['age_band'].value_counts().sort_index()

    # --- 4. 关键词分析 ---
    # 差评/好评聚类全面依赖 VOC Pack（LLM 按本次品类的评论原文动态聚类）。
    # LLM 不可用时返回空 dict —— 下游的计数/排序段会检测到空 dict 并跳过相关统计，
    # 不再回落到 LED 专属词典（如"亮度不足/磁力不稳"），避免跨品类产生错误归类。
    neg_keywords = neg_keywords_dict(packs)
    pos_keywords = pos_keywords_dict(packs)

    import re as _re

    def _safe_pattern(kws):
        # LLM 聚类出的关键词可能含未转义的正则特殊字符（括号/方括号/?/* 等），
        # 直接 join 会导致 re.error: unbalanced parenthesis。
        parts = [_re.escape(k) for k in kws if isinstance(k, str) and k.strip()]
        return '|'.join(parts) if parts else None

    def _filter_voc_pain_by_asins(pain_clusters, asin_subset, reviews_df,
                                  rating_threshold: int = 2, top_n: int = 3,
                                  min_keywords: int = 2):
        """通用工具：在 ASIN 子集的低星评论文本里，按 cluster.keywords 命中数返回 Top N 痛点。

        用法（统一所有"声称针对某子品类"的痛点叙述）：
            sub = set(df.loc[df['product_type'] == _first_type, 'ASIN'].astype(str).str.strip())
            hits = _filter_voc_pain_by_asins(packs.voc.pain_clusters, sub, rev_df, top_n=4)
            names = [n for n, _ in hits] or [c.name for c in packs.voc.pain_clusters[:4]]  # 调用方决定 fallback

        子集为空 / 评论列缺失 / 关键词不足 / 全 0 命中 → 返回空 list，由调用方决定是否 fallback 到全局 Top N。
        """
        if not pain_clusters or not asin_subset:
            return []
        if reviews_df is None or len(reviews_df) == 0:
            return []
        if 'source_asin' not in reviews_df.columns \
                or 'Title' not in reviews_df.columns or 'Content' not in reviews_df.columns:
            return []
        sub = reviews_df[reviews_df['source_asin'].astype(str).str.strip().isin(asin_subset)].copy()
        if 'Rating' in sub.columns:
            sub = sub[sub['Rating'] <= rating_threshold]
        if len(sub) == 0:
            return []
        sub['_neg_text'] = (sub['Title'].fillna('').astype(str) + ' '
                            + sub['Content'].fillna('').astype(str)).str.lower()
        hits = []
        for _pc in pain_clusters:
            _name = (getattr(_pc, 'name', '') or '').strip()
            if not _name:
                continue
            _kws = [k.strip() for k in (getattr(_pc, 'keywords', None) or [])
                    if k and len(k.strip()) >= 2]
            if len(_kws) < min_keywords:
                continue
            _pat = _safe_pattern(_kws)
            if not _pat:
                continue
            _cnt = int(sub['_neg_text'].str.contains(_pat, na=False, regex=True).sum())
            if _cnt > 0:
                hits.append((_name, _cnt))
        hits.sort(key=lambda x: -x[1])
        return hits[:top_n]

    neg_counts = {}
    if len(low_rev) > 0 and 'Title' in low_rev.columns and 'Content' in low_rev.columns:
        neg_text = (low_rev['Title'].fillna('') + ' ' + low_rev['Content'].fillna('')).str.lower()
        for cat, kws in neg_keywords.items():
            pattern = _safe_pattern(kws)
            neg_counts[cat] = neg_text.str.contains(pattern, na=False, regex=True).sum() if pattern else 0
    else:
        neg_counts = {k: 0 for k in neg_keywords.keys()}

    pos_counts = {}
    if len(high_rev) > 0 and 'Title' in high_rev.columns and 'Content' in high_rev.columns:
        pos_text = (high_rev['Title'].fillna('') + ' ' + high_rev['Content'].fillna('')).str.lower()
        for cat, kws in pos_keywords.items():
            pattern = _safe_pattern(kws)
            pos_counts[cat] = pos_text.str.contains(pattern, na=False, regex=True).sum() if pattern else 0
    else:
        pos_counts = {k: 0 for k in pos_keywords.keys()}

    # --- 5.1 动态计算推荐入场价（基于实际BSR数据分析） ---
    # 先算 Python 综合评分（供 Sheet 6 下半决策矩阵 + _top_segment_name fallback 用）
    ranked_product_types, ranked_type_scores = rank_product_types(df, type_agg)

    # 首推品类 —— 4 个 Sheet（4/6/10 + Sheet 6 下半矩阵）唯一来源：全从 _top_segment_name 读
    # 避免 Sheet 4 用 ranked_types[0] / Sheet 6 用 _first_seg_name / Sheet 10 用 pricing_rec[0] 多路不一致
    # 优先取 LLM sheet6_priority_matrix 的 P1（排除"不推荐"条目）；LLM 不可用时回落 Python 综合评分 Top1
    # 杂物桶名集合：永远不能被选为 ★ 首推（既不是真实细分，也不能给业务推荐意义）
    _JUNK_SEGMENTS = {'未分类', '其他', '其他/通用款'}
    _top_segment_name = None
    if packs is not None and packs.is_synthesis_real():
        for _item in packs.synthesis.sheet6_priority_matrix or []:
            _ap = (_item.action_plan or '')
            if (_item.priority == 'P1' and '不推荐' not in _ap and _item.segment
                    and _item.segment not in _JUNK_SEGMENTS):
                _top_segment_name = _item.segment
                break
    if not _top_segment_name:
        if ranked_product_types:
            _top_segment_name = ranked_product_types[0]
        elif packs is not None and packs.is_market_real() and packs.market.product_segments:
            # 跳过杂物桶名（极端情况下 segments 列表第一个不该是 '其他'/'未分类'）
            for _seg in packs.market.product_segments:
                if _seg.name and _seg.name not in _JUNK_SEGMENTS:
                    _top_segment_name = _seg.name
                    break
            if not _top_segment_name:
                _top_segment_name = '未分类'
        else:
            _top_segment_name = '未分类'

    # 把统一决定的首推品类传给 pricing_recommendations，让 Sheet 4 的 ★ 标注和其他 Sheet 对齐
    # 同时传入 packs，让"理由"列拼上 LLM 按本品类给的每个细分的定性叙述
    pricing_recommendations = calculate_pricing_recommendations(
        df, price_col, type_agg, ranked_product_types,
        top_segment_name=_top_segment_name, packs=packs
    )

    # --- 5.1.5 提前并发触发 Sheet5ImprovementAnalyzer + LifecycleAnalyzer 的 LLM 调用 ---
    # 这两个 analyzer 原先在 Excel 渲染阶段串行触发（Sheet5 ~30s + Lifecycle ~30s），
    # 与 prepare_packs 已并发跑完的 Stage1/Stage2 结合后串行变成净瓶颈。
    # 此处所有依赖（pricing_recommendations / df['product_type'] / packs.voc / market_data）
    # 都已就绪，提前 fire-and-forget 启动；下游消费点改 future.result() 等待。
    # 节省：LLM 两次调用从串行 → 并行 + 与 Sheet1-4/6/7 渲染重叠，预计 30-60s。
    from concurrent.futures import ThreadPoolExecutor as _PreLLM_TPE
    _sheet5_future = None
    _lifecycle_future = None
    try:
        # Sheet5 早期输入：_first_type + _pain_list（与 L5081-5102 同口径）
        _first_rec_pre = pricing_recommendations[0] if pricing_recommendations else None
        _first_type_pre = _first_rec_pre.get('product_type', '') if _first_rec_pre else ''
        _pain_list_pre: list[tuple[str, int]] = []
        _total_low_pre = len(low_rev) if hasattr(low_rev, '__len__') else 0
        if (_first_type_pre and packs is not None and packs.is_voc_real() and packs.voc.pain_clusters
                and 'product_type' in df.columns):
            _type_asins_pre = set(df.loc[df['product_type'] == _first_type_pre, 'ASIN']
                                    .astype(str).str.strip())
            _pain_list_pre = list(_filter_voc_pain_by_asins(
                packs.voc.pain_clusters, _type_asins_pre, rev_df, top_n=4))
        if not _pain_list_pre and packs is not None and packs.is_voc_real() and packs.voc.pain_clusters:
            for _c_pre in packs.voc.pain_clusters[:4]:
                _pct_pre = float(_c_pre.frequency_pct or 0)
                if _total_low_pre > 0 and _pct_pre > 0:
                    _cnt_pre = int(round(_pct_pre * _total_low_pre / 100.0))
                elif neg_counts:
                    _cnt_pre = int(neg_counts.get(_c_pre.name, 0))
                else:
                    _cnt_pre = 0
                _pain_list_pre.append((_c_pre.name, _cnt_pre))
        if not _pain_list_pre and neg_counts:
            _pain_list_pre = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True)[:4]

        # Lifecycle 早期输入：pub_df → infer_lifecycle_stage 推 score_detail
        _pub_df_pre = None
        if 'Available days' in df.columns:
            try:
                _df_with_pub = df.copy()
                _df_with_pub['Launch Years'] = (datetime.now().year
                                                 - (_df_with_pub['Available days'] / 365).round().astype(int))
                if rev_col and rev_col in _df_with_pub.columns:
                    _agg_pre = _df_with_pub.groupby('Launch Years').agg(
                        Products=('ASIN', 'count'),
                        Sales=('Monthly Sales', 'sum'),
                        **{'Monthly Revenue($)': (rev_col, 'sum')},
                    ).reset_index()
                    _ts_pre = _agg_pre['Sales'].sum() or 1
                    _tr_pre = _agg_pre['Monthly Revenue($)'].sum() or 1
                    _agg_pre['Sales Proportion'] = _agg_pre['Sales'] / _ts_pre
                    _agg_pre['Revenue Proportion'] = _agg_pre['Monthly Revenue($)'] / _tr_pre
                    _agg_pre = _agg_pre[['Launch Years', 'Products', 'Sales', 'Sales Proportion',
                                         'Monthly Revenue($)', 'Revenue Proportion']]
                    _pub_df_pre = _agg_pre.sort_values('Launch Years').reset_index(drop=True)
            except Exception:
                _pub_df_pre = None

        _score_detail_pre = None
        try:
            _, _, _score_detail_pre = infer_lifecycle_stage(_pub_df_pre, df=df, market_data=market_data)
        except Exception:
            _score_detail_pre = None

        def _kick_sheet5():
            if not (_pain_list_pre and packs is not None and packs.is_voc_real() and packs.voc.pain_clusters):
                return None
            try:
                from llm.analyzers.sheet5_improvement_analyzer import Sheet5ImprovementAnalyzer
                from core.packs_runtime import _try_make_client
                _client_pre = _try_make_client(api_key=getattr(packs, '_api_key', None))
                if _client_pre is None:
                    return None
                _pain_by_name_pre = {pc.name: pc for pc in packs.voc.pain_clusters}
                _s5_pains = []
                for _pn_pre, _cnt2_pre in _pain_list_pre[:4]:
                    _pc_obj_pre = _pain_by_name_pre.get(_pn_pre)
                    _s5_pains.append({
                        "name": _pn_pre,
                        "frequency_pct": float(_pc_obj_pre.frequency_pct or 0) if _pc_obj_pre else 0.0,
                        "raw_quotes": list(_pc_obj_pre.raw_quotes or [])[:5] if _pc_obj_pre else [],
                    })
                _stats_pre = getattr(packs, 'synthesis_stats', None) or {}
                return Sheet5ImprovementAnalyzer(_client_pre).run({
                    "category_hint": (getattr(packs, 'display_name', '') or '').strip() or _first_type_pre or '未知品类',
                    "pain_clusters": _s5_pains,
                    "stats": {
                        "competitor_spec_p75": _stats_pre.get("competitor_spec_p75", {}),
                        "competitor_spec_medians": _stats_pre.get("competitor_spec_medians", {}),
                    },
                })
            except Exception as _e:
                print(f"[Sheet5 后台] LLM 调用失败：{_e}", file=__import__('sys').stderr, flush=True)
                return None

        def _kick_lifecycle():
            if _score_detail_pre is None:
                return None
            try:
                return classify_market_pattern(_score_detail_pre, packs=packs, df=df)
            except Exception as _e:
                print(f"[Lifecycle 后台] LLM 调用失败：{_e}", file=__import__('sys').stderr, flush=True)
                return None

        # 用 max_workers=2 起两个独立后台任务；不 shutdown，等渲染层 result() 后自然结束
        _pre_llm_pool = _PreLLM_TPE(max_workers=2)
        _sheet5_future = _pre_llm_pool.submit(_kick_sheet5)
        _lifecycle_future = _pre_llm_pool.submit(_kick_lifecycle)
    except Exception as _e_pre:
        print(f"[预启动 Sheet5/Lifecycle] 失败（不阻断，渲染层会回退到原行为）：{_e_pre}",
              file=__import__('sys').stderr, flush=True)
        _sheet5_future = None
        _lifecycle_future = None

    # --- 5.2 动态生成产品上新方向（基于实际差评分析和市场数据） ---
    new_directions = generate_product_directions(df, rev_df, neg_counts, type_agg, price_col, packs=packs)
    
    # 差评痛点排序（供多处使用）。
    # 优先用 VOC Pack 里的 pain_clusters（LLM 按本次品类聚类），否则退化到上面基于 neg_keywords 的计数。
    # 都无数据时留空字符串，下游调用处会做 if 保护。
    if packs is not None and packs.is_voc_real() and packs.voc.pain_clusters:
        top_pain_for_use = packs.voc.pain_clusters[0].name
        top_pain_cnt_for_use = int(round(packs.voc.pain_clusters[0].frequency_pct))
        neg_sorted_for_use = [(c.name, int(round(c.frequency_pct))) for c in packs.voc.pain_clusters]
    else:
        neg_sorted_for_use = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True) if neg_counts else []
        top_pain_for_use = neg_sorted_for_use[0][0] if neg_sorted_for_use else ''
        top_pain_cnt_for_use = neg_sorted_for_use[0][1] if neg_sorted_for_use else 0

    # --- 5. 构建Excel报告 ---
    wb = Workbook()

    # ===== Sheet 1: 市场分析 =====
    ws1 = wb.active
    ws1.title = '市场分析'
    ws1.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGH', [22, 18, 18, 18, 18, 18, 18, 18]):
        ws1.column_dimensions[col_letter].width = width

    # 标题
    ws1.merge_cells('A1:H1')
    c = ws1['A1']
    c.value = f"{display_name_for_title(packs, '选品')} 市场选品评估报告"
    c.font = Font(name='Arial', bold=True, size=16, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells('A2:H2')
    c = ws1['A2']
    c.value = f'数据来源：亚马逊 BSR TOP100  +  {len(rev_df)}条真实评论  |  生成时间：{datetime.now().strftime("%Y-%m-%d")}'
    c.font = Font(name='Arial', size=9, color='FF595959', italic=True)
    c.fill = PatternFill('solid', fgColor=C_GREY_LIGHT)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 18

    r = 4
    section_title(ws1, r, 1, '▌ 一、类目整体概况（BSR TOP100 样本）', span=8)
    r += 1

    hdr(ws1, r, 1, '指标', bg=C_BLUE_MID)
    hdr(ws1, r, 2, '数值', bg=C_BLUE_MID)
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    hdr(ws1, r, 3, '说明', bg=C_BLUE_MID)
    ws1.row_dimensions[r].height = 20
    r += 1

    # 毛利率计算
    gm_mean = df['Gross Margin'].mean() * 100 if 'Gross Margin' in df.columns else 0
    gm_median = df['Gross Margin'].median() * 100 if 'Gross Margin' in df.columns else 0

    overview_data = [
        ('样本商品数', f'{len(df)} 个', 'BSR TOP100 实时样本'),
        ('涉及品牌数', f'{df["Brand"].nunique()} 个', 'BSR TOP100 中独立品牌数（数值越小代表头部品牌可能越集中，最终判定见下方"集中度"行）'),
        ('月总销量', f'{int(total_sales):,} 件', 'BSR TOP100 样本合计月销'),
        ('月总销售额', f'${total_rev/10000:.1f}万', f'折合人民币约 ¥{total_rev*7.2/10000:.0f}万'),
        ('均价', f'${avg_price:.2f}', f'中位价 ${median_price:.2f}，P25-P75 = ${df[price_col].quantile(0.25):.0f}-${df[price_col].quantile(0.75):.0f}' if price_col else f'中位价 ${median_price:.2f}'),
        ('平均毛利率', f'{gm_mean:.1f}%', f'中位毛利率 {gm_median:.1f}%'),
        ('平均星级', f'{df["Rating"].mean():.2f}★', f'最低 {df["Rating"].min():.1f}★ · 最高 {df["Rating"].max():.1f}★' if 'Rating' in df.columns else ''),
        ('FBA占比', f'{(df["Fulfillment"]=="FBA").sum()}%' if 'Fulfillment' in df.columns else '-', 'FBA 配送占比（BSR 表 Fulfillment 列）'),
        ('中国卖家占比（BuyBox）', f'{cn_pct_total:.0%}',
            f'{cn_cnt} 个 SKU，月收入占比 {cn_rev_pct_total:.1%}（BuyBox Location ∈ CN / CN(HK)）'),
    ]
    for metric, value, note in overview_data:
        val(ws1, r, 1, metric, bold=True, bg=C_BLUE_LIGHT)
        c = ws1.cell(row=r, column=2, value=value)
        c.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor=C_WHITE)
        ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        val(ws1, r, 3, note, bg=C_WHITE, fg='FF595959')
        ws1.row_dimensions[r].height = 18
        r += 1

    apply_border(ws1, 5, r-1, 1, 8)
    r += 1

    # 价格分布
    section_title(ws1, r, 1, '▌ 二、价格分布', span=8)
    r += 1

    hdr(ws1, r, 1, '价格区间', bg=C_BLUE_MID)
    hdr(ws1, r, 2, '商品数量', bg=C_BLUE_MID)
    hdr(ws1, r, 3, '占比', bg=C_BLUE_MID)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    hdr(ws1, r, 4, '区间特征', bg=C_BLUE_MID)
    ws1.row_dimensions[r].height = 20
    r += 1

    # 区间特征列：纯 Python 从 BSR 源表统计每个 price band 内的事实性指标
    # （SKU 数、平均星级、新品占比、TOP 品牌）——全部字段 BSR xlsx 原生可读，无需 LLM。
    # 上轮交给 Market Pack.price_ladder 的叙述字段会漏填，导致大部分 band 描述列为空
    def _band_stats(band_name: str) -> str:
        sub = df[df['price_band'] == band_name]
        if len(sub) == 0:
            return ''
        parts = [f'{len(sub)}款']
        if 'Rating' in sub.columns:
            r_series = pd.to_numeric(sub['Rating'], errors='coerce').dropna()
            if len(r_series) > 0:
                parts.append(f'平均 {r_series.mean():.1f}★')
        if 'Available days' in sub.columns:
            ad = pd.to_numeric(sub['Available days'], errors='coerce').dropna()
            if len(ad) > 0:
                new_pct = (ad < 365).mean()
                parts.append(f'新品占比 {new_pct:.0%}')
        if 'Brand' in sub.columns:
            top_brands = sub['Brand'].fillna('').astype(str).value_counts().head(2).index.tolist()
            top_brands = [b for b in top_brands if b]
            if top_brands:
                parts.append('TOP 品牌 ' + '、'.join(top_brands))
        return ' · '.join(parts)

    main_count = sum([price_dist.get(b, 0) for b in ['<$15', '$15-25', '$25-35', '$35-50']])
    for band, cnt in price_dist.items():
        bg = C_YELLOW if '$25' in str(band) or '$35' in str(band) else C_WHITE
        val(ws1, r, 1, str(band), bold=True, bg=bg)
        ws1.cell(row=r, column=2, value=int(cnt)).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=r, column=2).fill = PatternFill('solid', fgColor=bg)
        c3 = ws1.cell(row=r, column=3, value=f'{cnt/len(df)*100:.0f}%')
        c3.alignment = Alignment(horizontal='center', vertical='center')
        c3.fill = PatternFill('solid', fgColor=bg)
        ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        val(ws1, r, 4, _band_stats(str(band)), fg='FF595959', bg=bg)
        ws1.row_dimensions[r].height = 18
        r += 1

    apply_border(ws1, r - len(price_dist) - 1, r-1, 1, 8)
    r += 1

    # 产品类型分布
    section_title(ws1, r, 1, '▌ 三、产品类型分布与收益分析', span=8)
    r += 1

    type_hdr = ['产品类型', 'SKU数', '平均价格', '价格区间', '月总销量', '月总销售额', '单SKU均收益', '平均星级']
    for i, h in enumerate(type_hdr):
        hdr(ws1, r, i+1, h, bg=C_BLUE_MID)
    ws1.row_dimensions[r].height = 20
    r += 1

    # 类型行背景色按销量 Top3 分配（不依赖具体品类名）：
    # Top1→黄、Top2→浅绿、Top3→橙，其余留白。全品类通用。
    # 杂物桶（"其他"等）排除在 Top3 配色外（不是真细分，不抢配色）。
    _JUNK_SEG_TABLE = {'未分类', '其他', '其他/通用款'}
    _sorted_types_for_color = [
        t for t in type_agg.sort_values('total_sales', ascending=False)['product_type'].tolist()
        if t not in _JUNK_SEG_TABLE
    ]
    _rank_colors = {}
    for _i, _pt in enumerate(_sorted_types_for_color[:3]):
        _rank_colors[_pt] = [C_YELLOW, C_GREEN_LIGHT, C_ORANGE][_i]
    for _, row_data in type_agg.iterrows():
        _ptype = row_data['product_type']
        _is_junk = _ptype in _JUNK_SEG_TABLE
        bg = C_GREY_LIGHT if _is_junk else _rank_colors.get(_ptype, C_WHITE)
        # 杂物桶名后追加"⚠ 异质混合 - 仅作分布参考"提示
        _name_display = f'{_ptype}  ⚠ 异质混合 - 仅作分布参考' if _is_junk else _ptype
        val(ws1, r, 1, _name_display, bold=True, bg=bg)
        rev_val = row_data['total_revenue'] if pd.notna(row_data['total_revenue']) else 0
        for ci, v in enumerate([
            int(row_data['count']),
            f"${row_data['avg_price']:.1f}",
            f"${row_data['min_price']:.0f}-${row_data['max_price']:.0f}",
            f"{int(row_data['total_sales']):,}",
            f"${rev_val/10000:.1f}万",
            f"${rev_val/row_data['count']:.0f}" if rev_val > 0 else 'N/A',
            f"{row_data['avg_rating']:.1f}★" if pd.notna(row_data['avg_rating']) else 'N/A',
        ], start=2):
            c = ws1.cell(row=r, column=ci, value=v)
            c.font = Font(name='Arial', size=10, italic=_is_junk, color='FF999999' if _is_junk else None)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[r].height = 18
        r += 1

    apply_border(ws1, r - len(type_agg) - 1, r-1, 1, 8)
    r += 1

    # 市场结论
    section_title(ws1, r, 1, '▌ 四、市场分析结论', span=8)
    r += 1

    single_sku = (brand_agg['sku_count'] == 1).sum()
    _sp = None
    if _brand_df_global is not None and 'Sales Proportion' in _brand_df_global.columns:
        _sp = pd.to_numeric(_brand_df_global['Sales Proportion'], errors='coerce').dropna()
    if _sp is not None and len(_sp) >= 10:
        top3_share = float(_sp.head(3).sum()) * 100
        top5_share = float(_sp.head(5).sum()) * 100
        top10_share = float(_sp.head(10).sum()) * 100
    else:
        top3_share = brand_agg.head(3)['total_rev'].sum() / total_rev * 100 if total_rev else 0
        top5_share = brand_agg.head(5)['total_rev'].sum() / total_rev * 100 if total_rev else 0
        top10_share = brand_agg.head(10)['total_rev'].sum() / total_rev * 100 if total_rev else 0

    # 主力品类行：按 type_agg 销量排序动态给出 Top1/Top2，不再写死 LED 品类名
    _top_by_sales = type_agg.sort_values('total_sales', ascending=False)
    _top1_line = ''
    if len(_top_by_sales) >= 1:
        _t1 = _top_by_sales.iloc[0]
        _top1_line = f'{_t1["product_type"]}是销量最大品类，月销{int(_t1["total_sales"]):,}件。'
    if len(_top_by_sales) >= 2:
        _t2 = _top_by_sales.iloc[1]
        _top1_line += f'{_t2["product_type"]}均价${_t2["avg_price"]:.0f}。'

    # A+B 混合模式：Python 阈值给兜底 tier + 数字，LLM narrative 覆盖时优先使用
    # 市场规模阈值（月销售额万元）
    _rev_wan = total_rev / 10000
    if _rev_wan < 100:
        _scale_tier = '小体量利基市场'
    elif _rev_wan < 500:
        _scale_tier = '中等体量稳定市场'
    elif _rev_wan < 2000:
        _scale_tier = '大体量成熟市场'
    else:
        _scale_tier = '超大体量红海市场'

    # 价格结构：找实际最密集的连续 2 个价格带区间
    try:
        _sorted_bands = sorted(price_dist.items(), key=lambda x: x[1], reverse=True)
        _top_bands = [b for b, c in _sorted_bands if c > 0][:2]
        _struct_range = '、'.join(_top_bands) if _top_bands else '$15-$50'
        _struct_pct = sum(c for b, c in price_dist.items() if b in _top_bands) / max(len(df), 1) * 100
    except Exception:
        _struct_range = '$15-$50'
        _struct_pct = 0

    # 集中度阈值：统一用 CR5（前 5 大品牌销量占比），与业界电商选品标准一致
    # ≥80% 极高度集中 / 50-80% 高度集中 / 30-50% 中度集中 / <30% 分散
    if top5_share >= 80:
        _conc_tier = '极高度集中（寡头垄断），新品几无入场空间'
    elif top5_share >= 50:
        _conc_tier = '头部高度集中，新品入场门槛高，需强差异化'
    elif top5_share >= 30:
        _conc_tier = '中度集中，存在差异化切入机会'
    else:
        _conc_tier = '市场分散，新品进入机会较大'

    _mc = sheet1_market_conclusions(packs)
    conclusions = [
        ('市场规模',
         _mc.get('scale')
         or f'月总销售额约${_rev_wan:.1f}万，月总销量{int(total_sales):,}件，为{_scale_tier}。'),
        ('价格结构',
         _mc.get('structure')
         or f'主力价格带集中在 {_struct_range}（约{_struct_pct:.0f}% SKU，共{main_count}款位于 $15-$50 内），具体分布见上表。'),
        ('主力品类', _top1_line or '—'),
        ('集中度',
         _mc.get('concentration')
         or f'CR5={top5_share:.1f}%（Top3占{top3_share:.1f}%、Top10占{top10_share:.1f}%），{df["Brand"].nunique()}个品牌参与竞争，{single_sku}个品牌仅有1个SKU，{_conc_tier}。'),
    ]
    for label, text in conclusions:
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        val(ws1, r, 1, f'【{label}】', bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        val(ws1, r, 3, text, wrap=True)
        ws1.row_dimensions[r].height = 36
        r += 1

    apply_border(ws1, r-len(conclusions), r-1, 1, 8)
    r += 1

    # ===== 市场分析 - 新增内容汇总 =====
    # 销量与价格关系分析（C 方案：替换鸡肋"分析结论"列为横向对比指标）
    section_title(ws1, r, 1, '▌ 五、销量与价格关系分析', span=8)
    ws1.row_dimensions[r].height = 24
    r += 1

    # 按月销量四分位切 3 段
    _ms = df['Monthly Sales']
    q75, q25 = _ms.quantile(0.75), _ms.quantile(0.25)
    high_sales = df[_ms > q75]
    low_sales = df[_ms < q25]
    mid_sales = df[(_ms >= q25) & (_ms <= q75)]

    # 字段动态识别（兼容缺列）
    _country_col = 'BuyBox Location' if 'BuyBox Location' in df.columns else None
    _ratings_col = 'Ratings' if 'Ratings' in df.columns else ('Reviews' if 'Reviews' in df.columns else None)
    _days_col = 'Available days' if 'Available days' in df.columns else None
    _brand_col = 'Brand' if 'Brand' in df.columns else None

    def _tier_metrics(sub):
        """计算单个销量分层的对比指标。"""
        n = len(sub)
        if n == 0:
            return dict(n=0, avg_price='-', avg_sales='-', cn_pct='-',
                        avg_rating='-', avg_reviews='-', avg_days='-', top_brands='-')
        cn_pct = '-'
        if _country_col:
            _cc = sub[_country_col].astype(str).str.upper()
            cn_pct = f"{(_cc.str.contains('CN') | _cc.str.contains('CHINA')).mean():.0%}"
        avg_rating = '-'
        if 'Rating' in sub.columns:
            _r = pd.to_numeric(sub['Rating'], errors='coerce').dropna()
            avg_rating = f"{_r.mean():.1f}★" if len(_r) > 0 else '-'
        avg_reviews = '-'
        if _ratings_col:
            _rv = pd.to_numeric(sub[_ratings_col], errors='coerce').dropna()
            avg_reviews = f"{int(_rv.mean()):,}" if len(_rv) > 0 else '-'
        avg_days = '-'
        if _days_col:
            _d = pd.to_numeric(sub[_days_col], errors='coerce').dropna()
            avg_days = f"{int(_d.mean())}天" if len(_d) > 0 else '-'
        top_brands = '-'
        if _brand_col:
            _bs = sub[_brand_col].dropna().astype(str)
            _bs = _bs[(_bs != '') & (_bs.str.lower() != 'nan')]
            if len(_bs) > 0:
                top_brands = ' / '.join(_bs.value_counts().head(3).index.tolist())
        return dict(
            n=n,
            avg_price=f"${sub[price_col].mean():.1f}",
            avg_sales=f"{int(sub['Monthly Sales'].mean()):,}件",
            cn_pct=cn_pct, avg_rating=avg_rating,
            avg_reviews=avg_reviews, avg_days=avg_days, top_brands=top_brands,
        )

    sales_price_data = [
        ('高销量产品(TOP25%)', _tier_metrics(high_sales), C_GREEN_LIGHT),
        ('中等销量产品',       _tier_metrics(mid_sales),  C_YELLOW),
        ('低销量产品(BOTTOM25%)', _tier_metrics(low_sales), C_RED_LIGHT),
    ]

    # 表头（8 列）
    headers = ['销量分层', 'SKU数', '均价', '平均月销', '中国卖家占比',
               '平均星级', '平均评论数', '平均上架天数']
    for i, h in enumerate(headers, 1):
        hdr(ws1, r, i, h, bg=C_BLUE_MID)
    ws1.row_dimensions[r].height = 20
    r += 1

    body_start = r
    for label, m, bg in sales_price_data:
        # 第 1 行：8 列指标
        val(ws1, r, 1, label, bold=True, bg=bg)
        for col_idx, key in enumerate(
            ['n', 'avg_price', 'avg_sales', 'cn_pct', 'avg_rating', 'avg_reviews', 'avg_days'],
            start=2,
        ):
            cell = ws1.cell(row=r, column=col_idx, value=m[key])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill('solid', fgColor=bg)
        ws1.row_dimensions[r].height = 22
        r += 1
        # 第 2 行：TOP3 品牌注释（合并 1-8 列，浅色小字）
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        val(ws1, r, 1, f'  └─ TOP3 品牌：{m["top_brands"]}', bg=C_WHITE, fg='FF595959')
        ws1.row_dimensions[r].height = 18
        r += 1
    apply_border(ws1, body_start-1, r-1, 1, 8)
    r += 1

    # ===== Sheet 2: 竞争分析 =====
    ws2 = wb.create_sheet('竞争分析')
    ws2.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGH', [22, 18, 15, 15, 15, 18, 15, 15]):
        ws2.column_dimensions[col_letter].width = width

    ws2.merge_cells('A1:H1')
    c = ws2['A1']
    c.value = f"{display_name_for_title(packs, '选品')} 竞争格局分析"
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 35

    r2 = 3
    section_title(ws2, r2, 1, '▌ 一、竞争指数概览', span=8)
    r2 += 1

    new_cnt = len(new_df)
    new_avg_rev = new_df[rev_col].mean() if len(new_df) > 0 and rev_col else 0

    # TOP3 品牌动态名（取代 R1601 硬编码 DeWalt/Klein；同时用于中高端/入门品牌判断）
    _top3_brand_names = brand_agg.head(3)['Brand'].astype(str).tolist() if 'Brand' in brand_agg.columns else []
    _top3_brand_str = '、'.join(_top3_brand_names[:3]) if _top3_brand_names else '（品牌数据缺失）'

    # 新品均收入 vs 整体均收入对比（取代"略低于均值，前期爬坡正常"）
    _overall_avg_rev = (df[rev_col].mean() if rev_col else 0) or 0
    if new_avg_rev > 0 and _overall_avg_rev > 0:
        _ratio = new_avg_rev / _overall_avg_rev
        if _ratio >= 1.1:
            _new_rev_note = f'高于品类均值 {(_ratio-1)*100:.0f}%，新品爬坡较快'
        elif _ratio >= 0.9:
            _new_rev_note = f'与品类均值持平（{_ratio:.0%}），新品进入期正常'
        else:
            _new_rev_note = f'低于品类均值 {(1-_ratio)*100:.0f}%，新品盈利需要时间积累'
    else:
        _new_rev_note = '新品样本量不足，参考性有限'

    # 评分门槛动态说明（取代"进入TOP100需保持4.4★以上"硬编码）
    _rating_p25 = df['Rating'].quantile(0.25) if 'Rating' in df.columns and len(df) > 0 else 0
    _rating_threshold = f'{_rating_p25:.1f}★' if _rating_p25 > 0 else '-'

    # 竞争指数 / 品牌集中度按 CR5 4 档动态判定（用全局 cr5，与 Sheet1/8/10 完全一致）
    _cr5_pct_s2 = float(cr5 * 100) if cr5 is not None else 0.0
    if _cr5_pct_s2 >= 80:
        _cr5_label_s2, _cr5_idx_str, _cr5_note_s2 = '极高 ★★★★★', '极高', '极高度集中（寡头垄断），新品几无入场空间'
    elif _cr5_pct_s2 >= 50:
        _cr5_label_s2, _cr5_idx_str, _cr5_note_s2 = '高 ★★★★☆', '高', '头部高度集中，新品入场门槛高，需强差异化突围'
    elif _cr5_pct_s2 >= 30:
        _cr5_label_s2, _cr5_idx_str, _cr5_note_s2 = '中 ★★★☆☆', '中', '中度集中，存在差异化切入机会'
    else:
        _cr5_label_s2, _cr5_idx_str, _cr5_note_s2 = '低 ★★☆☆☆', '低', '市场分散，新品进入壁垒较低'

    compete_kv = [
        ('竞争指数', _cr5_label_s2, f'CR5={_cr5_pct_s2:.1f}%；{_cr5_note_s2}'),
        ('品牌集中度', f'{_cr5_idx_str}（Top5 占 {_cr5_pct_s2:.0f}%）' if total_rev else '数据缺失', _cr5_note_s2),
        ('中国卖家BuyBox占比', f'{cn_cnt} 个 / {cn_pct_total:.0%}',
            f'月收入占比 {cn_rev_pct_total:.1%} | 头部：{_cn_top3_brand_str}'),
        ('美国本土卖家', f'{us_cnt} 个 / {us_rev/total_rev*100:.1f}%收入' if total_rev else f'{us_cnt} 个',
            f'头部：{_us_top3_brand_str}'),
        ('FBA比例', f'{(df["Fulfillment"]=="FBA").sum()}%' if 'Fulfillment' in df.columns else '-', '物流标准化，竞争主要在产品和广告'),
        ('新品存活率', f'{new_cnt}%（{new_cnt}/{len(df)}）', '新品（<1年）进入TOP100比例'),
        ('新品月均收益', f'${new_avg_rev:,.0f}', _new_rev_note),
        ('评分门槛', f'最低 {df["Rating"].min():.1f}★，均值 {df["Rating"].mean():.2f}★' if 'Rating' in df.columns else '-',
            f'P25 门槛 {_rating_threshold}（低于该水平会被流量边缘化）'),
    ]

    hdr(ws2, r2, 1, '竞争维度', bg=C_BLUE_MID)
    hdr(ws2, r2, 2, '数据', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=3, end_row=r2, end_column=8)
    hdr(ws2, r2, 3, '说明', bg=C_BLUE_MID)
    r2 += 1
    for i, (k, v, n) in enumerate(compete_kv):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        val(ws2, r2, 1, k, bold=True, bg=bg)
        c = ws2.cell(row=r2, column=2, value=v)
        c.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws2.merge_cells(start_row=r2, start_column=3, end_row=r2, end_column=8)
        val(ws2, r2, 3, n, bg=bg, fg='FF595959')
        ws2.row_dimensions[r2].height = 20
        r2 += 1
    apply_border(ws2, 4, r2-1, 1, 8)
    r2 += 1

    # TOP15品牌
    section_title(ws2, r2, 1, '▌ 二、TOP15 品牌收益排行', span=8)
    r2 += 1
    brand_hdr_cols = ['排名', '品牌', 'SKU数', '月总收益($)', '收入占比', '平均星级', '卖家所属地', '市场地位']
    for i, h in enumerate(brand_hdr_cols):
        hdr(ws2, r2, i+1, h, bg=C_BLUE_MID)
    r2 += 1

    # 每品牌主导 BuyBox Location（众数）：跨品类通用，无 BuyBox Location 列时回落 "—"
    def _brand_loc(brand_name) -> str:
        if 'BuyBox Location' not in df.columns:
            return '—'
        sub_locs = df[df['Brand'] == brand_name]['BuyBox Location'].fillna('').astype(str).str.strip().str.upper()
        sub_locs = sub_locs[sub_locs != '']
        if sub_locs.empty:
            return '—'
        counts = sub_locs.value_counts()
        top = str(counts.index[0])
        if len(counts) > 1 and counts.iloc[0] == counts.iloc[1]:
            return f'{top}/{counts.index[1]}'
        return top

    brand_colors = {0: C_ORANGE, 1: C_ORANGE, 2: C_ORANGE}
    top15 = brand_agg.head(15)
    for idx, (_, brow) in enumerate(top15.iterrows()):
        bg = brand_colors.get(idx, C_WHITE if idx % 2 == 0 else C_GREY_LIGHT)
        ws2.cell(row=r2, column=1, value=idx+1).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=r2, column=1).fill = PatternFill('solid', fgColor=bg)
        rev_val = brow['total_rev'] if pd.notna(brow['total_rev']) else 0
        for ci, v in enumerate([
            brow['Brand'],
            int(brow['sku_count']),
            f"${rev_val:,.0f}",
            f"{rev_val/total_rev*100:.1f}%" if total_rev else 'N/A',
            f"{brow['avg_rating']:.1f}★" if pd.notna(brow['avg_rating']) else 'N/A',
            _brand_loc(brow['Brand']),
        ], start=2):
            c = ws2.cell(row=r2, column=ci, value=v)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center' if ci >= 3 else 'left', vertical='center')
            c.font = Font(name='Arial', size=10)
        status_map = {0: 'TOP1 绝对领跑', 1: 'TOP2 品类强者', 2: 'TOP3 紧密跟随', 3: 'TOP4-5 中坚品牌'}
        status = status_map.get(idx, 'TOP10 活跃竞争')
        val(ws2, r2, 8, status, bg=bg, fg='FF595959')
        ws2.row_dimensions[r2].height = 18
        r2 += 1
    apply_border(ws2, r2-16, r2-1, 1, 8)
    r2 += 1

    # 卖家地区分布
    section_title(ws2, r2, 1, '▌ 三、卖家地区分布', span=8)
    r2 += 1
    loc_data = df['BuyBox Location'].value_counts()
    hdr(ws2, r2, 1, '卖家地区', bg=C_BLUE_MID)
    hdr(ws2, r2, 2, '数量', bg=C_BLUE_MID)
    hdr(ws2, r2, 3, '占比', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=6)
    hdr(ws2, r2, 4, '月收入', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
    hdr(ws2, r2, 7, '收入占比', bg=C_BLUE_MID)
    r2 += 1
    for loc, cnt in loc_data.items():
        loc_rev = df[df['BuyBox Location'] == loc][rev_col].sum()
        bg = C_YELLOW if loc == 'CN' else C_WHITE
        val(ws2, r2, 1, str(loc) if loc else '未知', bold=(loc == 'CN'), bg=bg)
        ws2.cell(row=r2, column=2, value=int(cnt)).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=r2, column=2).fill = PatternFill('solid', fgColor=bg)
        ws2.cell(row=r2, column=3, value=f'{cnt/len(df)*100:.0f}%').alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=r2, column=3).fill = PatternFill('solid', fgColor=bg)
        ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=6)
        ws2.cell(row=r2, column=4, value=f'${loc_rev:,.0f}')
        ws2.cell(row=r2, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=r2, column=4).fill = PatternFill('solid', fgColor=bg)
        ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
        ws2.cell(row=r2, column=7, value=f'{loc_rev/total_rev*100:.1f}%' if total_rev else 'N/A')
        ws2.cell(row=r2, column=7).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=r2, column=7).fill = PatternFill('solid', fgColor=bg)
        ws2.row_dimensions[r2].height = 18
        r2 += 1
    apply_border(ws2, r2-len(loc_data)-1, r2-1, 1, 8)
    r2 += 1

    # 新品存活分析
    section_title(ws2, r2, 1, '▌ 四、新品存活分析', span=8)
    r2 += 1
    hdr(ws2, r2, 1, '上架年龄段', bg=C_BLUE_MID)
    hdr(ws2, r2, 2, 'SKU数', bg=C_BLUE_MID)
    hdr(ws2, r2, 3, '占比', bg=C_BLUE_MID)
    hdr(ws2, r2, 4, '月均收益/SKU', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
    hdr(ws2, r2, 5, '解读', bg=C_BLUE_MID)
    r2 += 1
    interpret_map = {
        '<3个月': '极新品，流量正在爬坡',
        '3-6个月': '新品存活验证期',
        '6-12个月': '存活稳定，有竞争力',
        '1-2年': '成熟竞品，有流量积累',
        '2-4年': '稳固地位',
        '>4年': '行业老品，品牌沉淀强'
    }
    for age_band, cnt in age_dist.items():
        age_df_sub = df[df['age_band'] == age_band]
        avg_r = age_df_sub[rev_col].mean() if rev_col and len(age_df_sub) > 0 else 0
        bg = C_GREEN_LIGHT if '6-12' in str(age_band) or '3-6' in str(age_band) else C_WHITE
        val(ws2, r2, 1, str(age_band), bg=bg)
        for ci, v in enumerate([int(cnt), f'{cnt/len(df)*100:.0f}%', f'${avg_r:,.0f}'], start=2):
            c = ws2.cell(row=r2, column=ci, value=v)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = Font(name='Arial', size=10)
        ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
        val(ws2, r2, 5, interpret_map.get(str(age_band), ''), bg=bg, fg='FF595959')
        ws2.row_dimensions[r2].height = 18
        r2 += 1
    apply_border(ws2, r2-len(age_dist)-1, r2-1, 1, 8)
    r2 += 1
    
    # ===== 竞争分析 - 新增内容 =====
    # 竞争进入建议
    section_title(ws2, r2, 1, '▌ 五、竞争进入建议', span=8)
    ws2.row_dimensions[r2].height = 24
    r2 += 1
    
    # 计算关键竞争指标
    cn_sku = cn_cnt
    us_sku = us_cnt
    cn_rev_pct = cn_rev/total_rev*100 if total_rev else 0
    
    # 差异化方向：优先读 VOC.pain_clusters[0..2].name（LLM 从本次品类评论聚类，100% 品类相关），
    # 其次读 Synthesizer.differentiation_angles / upgrade_directions，
    # LLM 全挂时退化为"—"（不再写死 LED 专属的"电池+亮度+磁力"）
    _diff_judge = '—'
    if packs is not None and packs.is_voc_real() and packs.voc.pain_clusters:
        _diff_judge = '+'.join([c.name for c in packs.voc.pain_clusters[:3]])
    elif packs is not None and packs.is_synthesis_real():
        _angles = getattr(packs.synthesis, 'differentiation_angles', None) or []
        if _angles:
            _diff_judge = '+'.join(str(a)[:10] for a in _angles[:3])
        else:
            _upgrade_dirs = upgrade_directions(packs)
            if _upgrade_dirs:
                _diff_judge = '+'.join([d['dimension'] for d in _upgrade_dirs[:3]])
    # 价格策略：动态从首推品类的 pricing_recommendations 取，避免与 Sheet 4 矛盾
    # 兜底链：首推品类对应那条 → pricing_recommendations[0] → 全品类中位价区间
    _price_label = '参见 Sheet 4'
    _price_desc = '具体推荐入场价见 Sheet 4「推荐入场价」按品类细分给出'
    _top_rec = None
    if pricing_recommendations:
        _top_rec = next(
            (p for p in pricing_recommendations if p.get('product_type') == _top_segment_name),
            pricing_recommendations[0],
        )
    if _top_rec:
        _price_label = f'${_top_rec["rec_min"]:.0f}-${_top_rec["rec_max"]:.0f}'
        _price_desc = (
            f'首推「{_top_rec["product_type"]}」P25-P75 区间 '
            f'${_top_rec["p25"]:.1f}-${_top_rec["p75"]:.1f}（{_top_rec["sku_count"]} 个 SKU），'
            f'按中位 ${_top_rec["median"]:.1f} 上下浮动 10% 取入场价，详见 Sheet 4'
        )

    _cr5_pct_e = (cr5 * 100) if cr5 is not None else 0
    if _cr5_pct_e >= 80:
        _conc_phr = f'CR5={_cr5_pct_e:.0f}% 极高度集中（寡头垄断）'
    elif _cr5_pct_e >= 50:
        _conc_phr = f'CR5={_cr5_pct_e:.0f}% 头部高度集中'
    elif _cr5_pct_e >= 30:
        _conc_phr = f'CR5={_cr5_pct_e:.0f}% 头部中度集中'
    else:
        _conc_phr = f'CR5={_cr5_pct_e:.0f}% 市场分散，无绝对垄断'
    entry_suggestions = [
        ('市场机会', '中等偏高',
         f'中国卖家主导({cn_sku}个SKU占{cn_rev_pct:.0f}%收入)；{_conc_phr}'),
        ('入场门槛', '中等',
         f'平均评分{df["Rating"].mean():.1f}★，新品需保持4.4★以上才具备竞争力'),
        ('价格策略', _price_label, _price_desc),
        ('差异化方向', _diff_judge,
         f'TOP3差评痛点：{top_pain_for_use}等，解决这些问题是突破关键' if top_pain_for_use else '依据头部竞品差评痛点做功能升级'),
        ('风险提示', '注意以下几点',
         ('\n'.join(f'• {b}' for b in sheet2_risk_bullets(packs))
          if sheet2_risk_bullets(packs)
          else '• FBA费用上涨压缩利润\n• 广告CPC在新进入期较高\n• 避免纯价格战，选择品质路线')),
    ]
    
    hdr(ws2, r2, 1, '维度', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=3)
    hdr(ws2, r2, 2, '判断', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=8)
    hdr(ws2, r2, 4, '详细说明', bg=C_BLUE_MID)
    ws2.row_dimensions[r2].height = 20
    r2 += 1
    
    for i, (dim, judge, detail) in enumerate(entry_suggestions):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        val(ws2, r2, 1, dim, bold=True, bg=bg)
        ws2.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=3)
        c = ws2.cell(row=r2, column=2, value=judge)
        c.font = Font(name='Arial', bold=True, color='FF1F3864')
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=8)
        val(ws2, r2, 4, detail, bg=bg, fg='FF595959', wrap=True)
        ws2.row_dimensions[r2].height = 45
        r2 += 1
    apply_border(ws2, r2-len(entry_suggestions)-1, r2-1, 1, 8)
    r2 += 1
    
    # TOP品牌策略分析
    section_title(ws2, r2, 1, '▌ 六、TOP品牌竞争策略分析', span=8)
    ws2.row_dimensions[r2].height = 24
    r2 += 1
    
    hdr(ws2, r2, 1, '品牌层级', bg=C_BLUE_MID)
    hdr(ws2, r2, 2, '代表品牌', bg=C_BLUE_MID)
    hdr(ws2, r2, 3, '价格区间', bg=C_BLUE_MID)
    hdr(ws2, r2, 4, '客户群体', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
    hdr(ws2, r2, 5, '竞争策略建议', bg=C_BLUE_MID)
    ws2.row_dimensions[r2].height = 20
    r2 += 1
    
    # 品牌层级全品类动态生成 —— 数据源：BSR 的 Brand / Price / Monthly Revenue 列
    # 头部 = 按 total_rev 降序 TOP3 品牌（"代表品牌" 列填真实品牌名，不再硬编码 DeWalt/Milwaukee/Klein）
    # 中高端 = 价格中位数落在 P50-P75 且 SKU ≥ 2 的 TOP3 品牌
    # 入门级 = 价格 <P25 的 TOP3 品牌（或 SKU ≤ 1 的长尾）
    # 不再写死"Coquimbo/HOTLIGH"等 LED 品类的 OEM 品牌名
    brand_strategies: list[tuple[str, str, str, str, str]] = []

    try:
        _brand_rows: list[dict] = []
        for _, _br in brand_agg.iterrows():
            _bn = str(_br.get('Brand', '')).strip()
            if not _bn:
                continue
            _bdf = df[df['Brand'] == _br['Brand']]
            if len(_bdf) == 0:
                continue
            _bprices = pd.to_numeric(_bdf[price_col], errors='coerce').dropna() if price_col else pd.Series(dtype=float)
            _brand_rows.append({
                'name': _bn,
                'sku': len(_bdf),
                'total_rev': float(_br.get('total_rev', 0) or 0),
                'price_median': float(_bprices.median()) if len(_bprices) > 0 else 0,
                'price_min': float(_bprices.min()) if len(_bprices) > 0 else 0,
                'price_max': float(_bprices.max()) if len(_bprices) > 0 else 0,
            })
        # 全局 P25/P50/P75（价格）
        _all_prices = pd.to_numeric(df[price_col], errors='coerce').dropna() if price_col else pd.Series(dtype=float)
        _p25 = float(_all_prices.quantile(0.25)) if len(_all_prices) > 0 else 0
        _p50 = float(_all_prices.quantile(0.50)) if len(_all_prices) > 0 else 0
        _p75 = float(_all_prices.quantile(0.75)) if len(_all_prices) > 0 else 0

        # 客户群体 + 竞争策略优先用 LLM，缺失时兜底到跨品类套话
        _brand_llm = sheet2_brand_strategies(packs)
        _head_cs, _head_st = _brand_llm.get('head', ('品牌忠诚用户', '避开！定位高端 / 品牌溢价强，新品需长期积累'))
        _mid_cs, _mid_st = _brand_llm.get('mid', ('性价比敏感用户', '正面竞争需差异化（功能/品质），建议做细分场景'))
        _entry_cs, _entry_st = _brand_llm.get('entry', ('价格敏感用户', '利润薄，不建议价格战，可考虑功能升级切入'))
        # 头部：total_rev 降序 TOP3
        _head = sorted(_brand_rows, key=lambda b: b['total_rev'], reverse=True)[:3]
        if _head:
            _head_names = '/'.join(b['name'] for b in _head)
            _head_prices = [b['price_min'] for b in _head if b['price_min']] + [b['price_max'] for b in _head if b['price_max']]
            _head_range = f'${min(_head_prices):.0f}-${max(_head_prices):.0f}' if _head_prices else '-'
            brand_strategies.append(('头部品牌', _head_names, _head_range,
                _head_cs or '品牌忠诚用户',
                _head_st or '避开！定位高端 / 品牌溢价强，新品需长期积累'))
        # 中高端：价格中位在 [P50, P75] 且 SKU≥2 的 TOP3
        _mid = [b for b in _brand_rows if b['sku'] >= 2 and _p50 <= b['price_median'] <= _p75]
        _mid = sorted(_mid, key=lambda b: b['total_rev'], reverse=True)[:3]
        if _mid:
            _mid_names = '/'.join(b['name'] for b in _mid)
            brand_strategies.append(('中高端', _mid_names,
                f'${_p50:.0f}-${_p75:.0f}',
                _mid_cs or '性价比敏感用户',
                _mid_st or '正面竞争需差异化（功能/品质），建议做细分场景'))
        # 入门级：价格中位 < P25 的 TOP3
        _entry = [b for b in _brand_rows if b['price_median'] and b['price_median'] < _p25]
        _entry = sorted(_entry, key=lambda b: b['total_rev'], reverse=True)[:3]
        if _entry:
            _entry_names = '/'.join(b['name'] for b in _entry)
            brand_strategies.append(('入门级', _entry_names or '不知名品牌',
                f'<${_p25:.0f}',
                _entry_cs or '价格敏感用户',
                _entry_st or '利润薄，不建议价格战，可考虑功能升级切入'))
    except Exception as _e:
        # 完全失败则用 TOP3 通用兜底
        brand_strategies = [
            ('头部品牌', '（品牌数据解析失败）', '-', '-', '数据缺失'),
        ]
    
    for i, (tier, brand, price, customer, strategy) in enumerate(brand_strategies):
        bg = C_YELLOW if i == 1 else (C_ORANGE if i == 0 else C_GREEN_LIGHT)
        val(ws2, r2, 1, tier, bold=True, bg=bg)
        val(ws2, r2, 2, brand, bg=bg)
        val(ws2, r2, 3, price, bg=bg)
        val(ws2, r2, 4, customer, bg=bg)
        ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
        val(ws2, r2, 5, strategy, bg=bg, fg='FF595959')
        ws2.row_dimensions[r2].height = 35
        r2 += 1
    apply_border(ws2, r2-len(brand_strategies)-1, r2-1, 1, 8)
    r2 += 1

    # ---- 扩展段：关键词 + Market 文件 ----
    # ▌ 五、核心关键词竞争度
    section_title(ws2, r2, 1, '▌ 七、核心关键词竞争度', span=8)
    ws2.row_dimensions[r2].height = 24
    r2 += 1

    kw_df = keyword_data.get('keywords') if keyword_data.get('_available') else None
    kw_top3_for_conclusion = []
    kw_lowest_spr_for_conclusion = ''
    kw_avg_bid_for_conclusion = ''
    if kw_df is not None and len(kw_df) > 0:
        # --- 关键词文件真实数据版 ---
        kw = kw_df.copy()
        kw['M. Searches'] = pd.to_numeric(kw.get('M. Searches', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw['Title Density'] = pd.to_numeric(kw.get('Title Density', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw['SPR'] = pd.to_numeric(kw.get('SPR', pd.Series(dtype=float)), errors='coerce').fillna(0)
        # 点击总占比 = 前 3 ASIN 点击占比之和（卖家精灵 Click Share 列），头部垄断信号
        kw['Click Share'] = pd.to_numeric(kw.get('Click Share', pd.Series(dtype=float)), errors='coerce').fillna(0)
        # 转化总占比 = 前 3 ASIN 转化占比之和（卖家精灵 Conversion Share 列），头部垄断信号
        # 注意：Purchase Rate（=月购买量/月搜索量）是该词整体购买率，与"前 3 ASIN 转化占比之和"不是一回事，
        # 用作竞争度参考会误判，因此这里只接受 Conversion Share / Conversion Rate / Conversion 三个语义一致的列
        def _first_col(df, *names):
            lmap = {c.lower().strip(): c for c in df.columns}
            for n in names:
                if n.lower() in lmap:
                    return lmap[n.lower()]
            return None
        _conv_col = _first_col(kw, 'Conversion Share', 'Conversion Rate', 'Conversion')
        if _conv_col:
            kw['ConvShare'] = pd.to_numeric(kw[_conv_col], errors='coerce').fillna(0)
        else:
            kw['ConvShare'] = 0.0
        # 卖家精灵 ReverseASIN 表里此字段实际叫 'Sponsored ASINs'，旧代码硬编码 'Ads Competitor Count' 拿不到数据
        _ads_col = _first_col(kw, 'Ads Competitor Count', 'Sponsored ASINs', 'Sponsored Competitor')
        if _ads_col:
            kw['AdsCompetitorCount'] = pd.to_numeric(kw[_ads_col], errors='coerce').fillna(0)
        else:
            kw['AdsCompetitorCount'] = 0
        kw['Products'] = pd.to_numeric(kw.get('Products', pd.Series(dtype=float)), errors='coerce').fillna(0)
        # 选词主依据：流量占比（Traffic Share）+ 相关产品（Related Products），rank-sum 复合分。
        # 老数据源无这两列时整列回填 0，触发下方降级分支按月搜索量取 Top 20。
        kw['Traffic Share'] = pd.to_numeric(kw.get('Traffic Share', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw['Related Products'] = pd.to_numeric(kw.get('Related Products', pd.Series(dtype=float)), errors='coerce').fillna(0)
        has_ppc_bid = 'PPC Bid' in kw.columns
        has_suggested_bid = 'Suggested Bid' in kw.columns
        if kw['Traffic Share'].sum() == 0 and kw['Related Products'].sum() == 0:
            kw_top = kw.sort_values('M. Searches', ascending=False).head(20).reset_index(drop=True)
            kw_pick_method = 'fallback_msearches'
        else:
            ts_rank = kw['Traffic Share'].rank(method='min', ascending=False)
            rp_rank = kw['Related Products'].rank(method='min', ascending=False)
            kw['_pick_score'] = ts_rank + rp_rank
            kw_top = kw.sort_values(['_pick_score', 'M. Searches'],
                                    ascending=[True, False]).head(20).reset_index(drop=True)
            kw_pick_method = 'traffic_x_related'

        # 12 列：在「月搜索量」之后插入「流量占比 / 相关产品」两列，让用户直接看到选词依据
        headers = ['关键词', '月搜索量', '流量占比', '相关产品', '竞品ASIN数', 'SPR(8天上首页)',
                   '广告竞品数', '点击总占比', '转化总占比', 'PPC Bid', '建议Bid区间', '竞争度']
        for extra_col in ['C', 'D']:
            ws2.column_dimensions[extra_col].width = 12
        for extra_col in ['K', 'L']:
            ws2.column_dimensions[extra_col].width = 14
        for i, h in enumerate(headers, 1):
            hdr(ws2, r2, i, h, bg=C_BLUE_MID)
        ws2.row_dimensions[r2].height = 22
        r2 += 1
        row_start = r2
        bid_values = []

        def _comp_score(m_search, products, ads_comp, ppc_val, conv_share):
            """5 维加权打分：月搜索量 + 商品数 + 广告竞品数 + PPC 竞价 + 转化总占比，每维 0-3 分，总 0-15。"""
            s = 0
            s += 3 if m_search >= 100000 else (2 if m_search >= 10000 else (1 if m_search >= 1000 else 0))
            s += 3 if products >= 50000 else (2 if products >= 10000 else (1 if products >= 1000 else 0))
            s += 3 if ads_comp >= 50 else (2 if ads_comp >= 20 else (1 if ads_comp >= 5 else 0))
            s += 3 if ppc_val >= 3 else (2 if ppc_val >= 1 else (1 if ppc_val >= 0.5 else 0))
            s += 3 if conv_share >= 0.6 else (2 if conv_share >= 0.3 else (1 if conv_share >= 0.1 else 0))
            return s

        high_comp = 0
        for _, row in kw_top.iterrows():
            keyword = str(row.get('Keyword', '-'))
            m_search = float(row['M. Searches'])
            products = int(row['Products']) if row['Products'] > 0 else int(row['Title Density'])
            spr = float(row['SPR'])
            ads_comp = int(row['AdsCompetitorCount']) if row['AdsCompetitorCount'] > 0 else 0
            cs = float(row['Click Share'])
            conv_share = float(row['ConvShare'])
            ppc_bid = str(row.get('PPC Bid', '-')) if has_ppc_bid else '-'
            suggested_bid = str(row.get('Suggested Bid', '-')) if has_suggested_bid else '-'
            ppc_val = 0.0
            if ppc_bid != '-' and ppc_bid != 'nan':
                try:
                    ppc_val = float(ppc_bid.replace('$', '').strip())
                    bid_values.append(ppc_val)
                except (ValueError, AttributeError):
                    pass

            score = _comp_score(m_search, products, ads_comp, ppc_val, conv_share)
            if score >= 10:
                rating, rating_bg = '高', C_RED_LIGHT
                high_comp += 1
            elif score >= 6:
                rating, rating_bg = '中', C_YELLOW
            else:
                rating, rating_bg = '低', C_GREEN_LIGHT

            ts = float(row.get('Traffic Share', 0) or 0)
            rp = float(row.get('Related Products', 0) or 0)
            ts_disp = (f'{ts:.2%}' if ts > 1.5 else f'{ts*100:.2f}%') if ts > 0 else '-'
            rp_disp = f'{int(rp)}' if rp > 0 else '-'

            val(ws2, r2, 1, keyword, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
            val(ws2, r2, 2, f'{m_search:,.0f}')
            val(ws2, r2, 3, ts_disp, bg=C_RED_LIGHT if ts >= 0.05 else (C_YELLOW if ts >= 0.01 else C_WHITE))
            val(ws2, r2, 4, rp_disp, bg=C_RED_LIGHT if rp >= 5 else (C_YELLOW if rp >= 2 else C_WHITE))
            val(ws2, r2, 5, f'{products:,}', bg=C_RED_LIGHT if products > 10000 else (C_YELLOW if products > 1000 else C_WHITE))
            val(ws2, r2, 6, f'{spr:.0f}' if spr > 0 else '-')
            val(ws2, r2, 7, f'{ads_comp}' if ads_comp > 0 else '-',
                bg=C_RED_LIGHT if ads_comp >= 50 else (C_YELLOW if ads_comp >= 20 else C_WHITE))
            val(ws2, r2, 8, f'{cs:.2%}' if cs > 0 else '-',
                bg=C_RED_LIGHT if cs >= 0.6 else (C_YELLOW if cs >= 0.3 else C_WHITE))
            val(ws2, r2, 9, f'{conv_share:.2%}' if conv_share > 0 else '-',
                bg=C_RED_LIGHT if conv_share >= 0.6 else (C_YELLOW if conv_share >= 0.3 else C_WHITE))
            val(ws2, r2, 10, ppc_bid if ppc_bid != 'nan' else '-')
            val(ws2, r2, 11, suggested_bid if suggested_bid != 'nan' else '-')
            val(ws2, r2, 12, rating, bold=True, bg=rating_bg)
            ws2.row_dimensions[r2].height = 20
            r2 += 1
        if r2 > row_start:
            apply_border(ws2, row_start-1, r2-1, 1, 12)
        r2 += 1

        # 关键词小结（叙述性）
        top3 = kw_top.head(3)['Keyword'].tolist()
        kw_top3_for_conclusion = top3
        lowest_spr_row = kw_top[kw_top['SPR'] > 0].sort_values('SPR').head(1)
        lowest_spr_kw = lowest_spr_row.iloc[0]['Keyword'] if len(lowest_spr_row) > 0 else '—'
        lowest_spr_val = int(lowest_spr_row.iloc[0]['SPR']) if len(lowest_spr_row) > 0 else 0
        kw_lowest_spr_for_conclusion = f'{lowest_spr_kw}（SPR={lowest_spr_val}）'
        # high_comp 已在循环中按 5 维加权打分累计（评级=高的词数）
        avg_bid = f'${sum(bid_values)/len(bid_values):.2f}' if bid_values else 'N/A'
        kw_avg_bid_for_conclusion = avg_bid
        top1 = kw_top.iloc[0] if len(kw_top) > 0 else None
        if kw_pick_method == 'traffic_x_related':
            ts_sum_top = float(kw_top['Traffic Share'].sum())
            rp_avg_top = float(kw_top['Related Products'].mean())
            ts_unit = '%' if ts_sum_top > 1.5 else ''  # 兼容百分数(如 7.5)/小数(如 0.075)两种存储
            ts_display = f'{ts_sum_top:.1f}{ts_unit}' if ts_unit else f'{ts_sum_top*100:.1f}%'
            pick_line = f'• 选词依据：流量占比 + 相关产品 双维度 rank-sum 排序；Top 20 累计流量占比约 {ts_display}，平均相关产品 {rp_avg_top:.0f} 个'
        else:
            pick_line = '• 选词依据：数据源缺少「流量占比 / 相关产品」两列，已降级为按月搜索量取 Top 20'
        summary_items = [
            pick_line,
            f'• 核心大词「{top3[0]}」月搜索量 {int(kw_top.iloc[0]["M. Searches"]):,}，竞品 {int(kw_top.iloc[0]["Products"]):,} 个，竞争度高' if top1 is not None else '',
            f'• 最低门槛词「{lowest_spr_kw}」SPR 仅 {lowest_spr_val}，8 天需 {lowest_spr_val} 单即可上首页，适合新品切入',
            f'• Top 20 关键词平均 PPC Bid：{avg_bid}；{high_comp} 个词为高竞争度（5 维加权评分≥10 分）',
            f'• 长尾策略：优先投放 SPR < 30 且月搜索 > 5,000 的中等竞争词',
        ]
        summary_items = [s for s in summary_items if s]
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2+4, end_column=12)
        c = ws2.cell(row=r2, column=1)
        c.value = '\n'.join(summary_items)
        c.font = Font(name='Arial', size=10, color='FF1F3864')
        c.fill = PatternFill('solid', fgColor=C_BLUE_LIGHT)
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for rr in range(r2, r2+5):
            ws2.row_dimensions[rr].height = 22
        r2 += 5
    else:
        # --- 降级：Market 文件 5 词趋势 ---
        demand_df = market_data.get('demand_trends') if market_data.get('_available') else None
        if demand_df is not None and len(demand_df) > 2:
            d = demand_df.copy()
            month_col = d.columns[0]
            kw_cols = list(d.columns[1:])
            d[month_col] = d[month_col].astype(str)
            d = d.dropna(subset=[month_col])
            d = d[d[month_col].str.match(r'^\d{4}-\d{2}$', na=False)]
            d = d.sort_values(month_col)
            headers = ['核心关键词', '最新月搜索量', '12 月均值', '同比变化', '年内峰值月', '竞争度评级']
            for i, h in enumerate(headers, 1):
                hdr(ws2, r2, i, h, bg=C_BLUE_MID)
            ws2.row_dimensions[r2].height = 20
            r2 += 1
            row_start = r2
            for kw in kw_cols:
                try:
                    s = pd.to_numeric(d[kw], errors='coerce').dropna()
                    if len(s) == 0:
                        continue
                    latest = float(s.iloc[-1])
                    last12 = s.iloc[-12:] if len(s) >= 12 else s
                    avg12 = float(last12.mean())
                    if len(s) >= 24:
                        prev12 = s.iloc[-24:-12]
                        yoy = (last12.sum() - prev12.sum()) / prev12.sum() if prev12.sum() > 0 else 0
                        yoy_str = f'{yoy:+.1%}'
                    else:
                        yoy_str = 'N/A'
                    last12_months = d[month_col].iloc[-12:].tolist() if len(d) >= 12 else d[month_col].tolist()
                    peak_idx = last12.values.argmax()
                    peak_month = last12_months[peak_idx] if peak_idx < len(last12_months) else '--'
                    peak_month_m = peak_month.split('-')[-1] if '-' in peak_month else peak_month
                    if avg12 < 1000:
                        rating, rating_bg = '低', C_GREEN_LIGHT
                    elif avg12 < 10000:
                        rating, rating_bg = '中', C_YELLOW
                    else:
                        rating, rating_bg = '高', C_RED_LIGHT
                    val(ws2, r2, 1, str(kw).strip(), bold=True, bg=C_BLUE_LIGHT)
                    val(ws2, r2, 2, f'{latest:,.0f}')
                    val(ws2, r2, 3, f'{avg12:,.0f}')
                    val(ws2, r2, 4, yoy_str, fg='FF005A9E' if yoy_str.startswith('+') else 'FFC00000')
                    val(ws2, r2, 5, f'{peak_month_m} 月', bg=C_BLUE_LIGHT)
                    val(ws2, r2, 6, rating, bold=True, bg=rating_bg)
                    ws2.merge_cells(start_row=r2, start_column=6, end_row=r2, end_column=8)
                    ws2.row_dimensions[r2].height = 20
                    r2 += 1
                except Exception:
                    continue
            if r2 > row_start:
                apply_border(ws2, row_start-1, r2-1, 1, 8)
        else:
            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
            val(ws2, r2, 1, '⚠ 未上传关键词文件或 Market 分析文件，关键词竞争数据缺失', bold=True, bg=C_YELLOW, fg='FF7A4F01')
            ws2.row_dimensions[r2].height = 30
            r2 += 1
    r2 += 1

    # ▌ 六、新品推广难度评估
    section_title(ws2, r2, 1, '▌ 八、新品推广难度评估', span=8)
    ws2.row_dimensions[r2].height = 24
    r2 += 1

    # 数据来源：BSR + Market
    total_products = len(df)
    new_products_6m = len(df[df['Available days'] < 180]) if 'Available days' in df.columns else 0
    new_products_1y = len(df[df['Available days'] < 365]) if 'Available days' in df.columns else 0
    new_in_top100 = new_products_1y  # 已在 BSR TOP100 内
    new_ratio_top100 = new_in_top100 / total_products if total_products > 0 else 0

    # cr5 / cr10 已在 brand_agg 之后全局计算（Market 文件 Sales Proportion 优先 + BSR 兜底）
    # 这里仅补充 rating_50_ratio（与 cr5/cr10 不同源，仍单独算）
    rating_50_ratio = None
    brand_df = market_data.get('brand_concentration')

    # 中国卖家占比：直接用全局统一口径变量（见 generate_report 顶部）
    cn_seller_ratio = cn_pct_total

    diff_rows = [
        ('新品数量（<1 年）', f'{new_products_1y} 个', f'占 BSR TOP100 的 {new_ratio_top100:.1%}'),
        ('新品数量（<6 月）', f'{new_products_6m} 个', f'占 BSR TOP100 的 {new_products_6m/max(total_products,1):.1%}'),
        ('新品进 Top100 比例', f'{new_ratio_top100:.1%}', '越高=新品越有机会；<10% 推广难度高'),
        ('头部品牌 CR5', f'{cr5:.1%}' if cr5 is not None else '-', '≥50% 则头部垄断，需差异化切入（来源：Market 或 BSR brand_agg 兜底）'),
        ('头部品牌 CR10', f'{cr10:.1%}' if cr10 is not None else '-', '越高代表品牌集中度越强'),
        ('中国卖家占比', f'{cn_seller_ratio:.1%}', '>70% 则价格战激烈，利润被压缩（来源：BSR BuyBox Location）'),
        ('评分门槛参考', f'{int(avg_rating*10)/10:.1f}★', '低于均分会被流量边缘化'),
    ]

    hdr(ws2, r2, 1, '评估指标', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=2)
    hdr(ws2, r2, 3, '数值', bg=C_BLUE_MID)
    hdr(ws2, r2, 4, '含义说明', bg=C_BLUE_MID)
    ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=8)
    ws2.row_dimensions[r2].height = 20
    r2 += 1
    row_start = r2
    for label, value, note in diff_rows:
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=2)
        val(ws2, r2, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        val(ws2, r2, 3, value, bg=C_WHITE)
        ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=8)
        val(ws2, r2, 4, note, bg=C_WHITE, fg='FF595959')
        ws2.row_dimensions[r2].height = 20
        r2 += 1
    apply_border(ws2, row_start-1, r2-1, 1, 8)
    r2 += 1

    # 综合评级
    difficulty_level = '中'
    difficulty_bg = C_YELLOW
    reasons = []
    # CR5 阈值与全局四档对齐：≥80% 极高度集中 / 50-80% 高度集中 / 30-50% 中度 / <30% 分散
    # 50% 及以上拉高 difficulty 为"高"，30-50% 仅作中度提示不抬等级
    if cr5 is not None and cr5 >= 0.8:
        difficulty_level = '高'
        difficulty_bg = C_RED_LIGHT
        reasons.append(f'CR5={cr5:.0%} 极高度集中（寡头垄断）')
    elif cr5 is not None and cr5 >= 0.5:
        difficulty_level = '高'
        difficulty_bg = C_RED_LIGHT
        reasons.append(f'CR5={cr5:.0%} 头部高度集中')
    elif cr5 is not None and cr5 >= 0.3:
        reasons.append(f'CR5={cr5:.0%} 中度集中')
    if new_ratio_top100 < 0.1:
        difficulty_level = '高'
        difficulty_bg = C_RED_LIGHT
        reasons.append(f'新品进 Top100 仅 {new_ratio_top100:.0%}')
    if (cr5 is None or cr5 < 0.3) and new_ratio_top100 > 0.3:
        difficulty_level = '低'
        difficulty_bg = C_GREEN_LIGHT
        reasons.append(f'新品占比高 {new_ratio_top100:.0%}')
    if not reasons:
        reasons.append('各指标处于中等水平')
    cycle_map = {'低': '3-6 月可进入头部', '中': '6-9 月可进入头部', '高': '9-12 月需持续投入'}

    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=2)
    val(ws2, r2, 1, '综合推广难度', bold=True, bg=C_BLUE_DARK, fg='FFFFFFFF')
    val(ws2, r2, 3, difficulty_level, bold=True, bg=difficulty_bg)
    ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=8)
    val(ws2, r2, 4, f'依据：{"，".join(reasons)} | 预计周期：{cycle_map[difficulty_level]}', bg=C_WHITE)
    ws2.row_dimensions[r2].height = 25
    r2 += 1
    r2 += 1

    # ▌ 七、小结
    section_title(ws2, r2, 1, '▌ 九、竞争小结', span=8, bg=C_ORANGE)
    ws2.row_dimensions[r2].height = 24
    r2 += 1

    try:
        # brand_agg 在 groupby().reset_index() 后 index 是 RangeIndex，品牌名在 'Brand' 列，不在 index
        if 'Brand' in brand_agg.columns:
            top3_brands = brand_agg.head(3)['Brand'].astype(str).tolist()
        else:
            top3_brands = [str(b) for b in brand_agg.head(3).index.tolist()]
    except Exception:
        top3_brands = []
    cr5_txt = f'CR5={cr5:.0%}' if cr5 is not None else (
        f'CR5={brand_agg.head(5)["total_rev"].sum()/total_rev*100:.0f}%' if total_rev else 'CR5=数据缺失'
    )
    summary_text = (
        f'• 品牌集中度：前 3 品牌 {", ".join(top3_brands) if top3_brands else "见上表"}；{cr5_txt}\n'
        f'• 新品友好度：{new_ratio_top100:.1%}（<1 年）进入 BSR TOP100，{"新品机会大" if new_ratio_top100 > 0.2 else "新品突围难度大"}\n'
        f'• 综合推广难度：{difficulty_level}，建议 {cycle_map[difficulty_level]}'
    )
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2+2, end_column=8)
    c = ws2.cell(row=r2, column=1)
    c.value = summary_text
    c.font = Font(name='Arial', size=11, color='FF1F3864')
    c.fill = PatternFill('solid', fgColor=C_BLUE_LIGHT)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for rr in range(r2, r2+3):
        ws2.row_dimensions[rr].height = 22
    r2 += 3

    # ▌ 八、广告竞争格局（ReverseASIN）
    # 守卫：卖家精灵 ExpandKeywords 文件不含 Organic/Sponsored Share/DSR 列，
    # 若两列都缺就整段跳过，不再硬编码"自然流量为主"误导用户。
    # 源文件支持这些列时（ReverseASIN 的广告分析版或用户补充的数据）才渲染。
    _has_ad_share = (kw_df is not None and len(kw_df) > 0
                     and ('Organic Share' in kw_df.columns
                          or 'Sponsored Share' in kw_df.columns))
    if _has_ad_share:
        section_title(ws2, r2, 1, '▌ 十、广告竞争格局（自然流量 vs 广告流量）', span=10)
        ws2.row_dimensions[r2].height = 24
        r2 += 1

        kw_ad = kw_df.copy()
        kw_ad['Organic Share'] = pd.to_numeric(kw_ad.get('Organic Share', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_ad['Sponsored Share'] = pd.to_numeric(kw_ad.get('Sponsored Share', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_ad['Sponsored ASINs'] = pd.to_numeric(kw_ad.get('Sponsored ASINs', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_ad['DSR'] = pd.to_numeric(kw_ad.get('DSR', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_ad['M. Searches'] = pd.to_numeric(kw_ad.get('M. Searches', pd.Series(dtype=float)), errors='coerce').fillna(0)
        ad_top = kw_ad.sort_values('M. Searches', ascending=False).head(15).reset_index(drop=True)

        ad_headers = ['关键词', '月搜索量', '自然流量占比', '广告流量占比', '广告竞品数', 'DSR供需比', '广告策略']
        for i, h in enumerate(ad_headers, 1):
            hdr(ws2, r2, i, h, bg=C_BLUE_MID)
        ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
        ws2.row_dimensions[r2].height = 20
        r2 += 1
        ad_start = r2
        ad_dominant = 0
        organic_dominant = 0
        def _safe_num(v, default=0.0):
            try:
                return float(v) if pd.notna(v) else default
            except (TypeError, ValueError):
                return default
        for _, row in ad_top.iterrows():
            kw_name = str(row.get('Keyword', '-'))
            m_s = _safe_num(row.get('M. Searches'))
            org_s = _safe_num(row.get('Organic Share'))
            spon_s = _safe_num(row.get('Sponsored Share'))
            spon_asins = int(_safe_num(row.get('Sponsored ASINs')))
            dsr = _safe_num(row.get('DSR'))
            if spon_s > 0.6:
                strategy = '广告主导，必须投放'
                strat_bg = C_RED_LIGHT
                ad_dominant += 1
            elif spon_s > 0.3:
                strategy = '混合流量，建议投放'
                strat_bg = C_YELLOW
            else:
                strategy = '自然流量为主，SEO 优先'
                strat_bg = C_GREEN_LIGHT
                organic_dominant += 1
            val(ws2, r2, 1, kw_name, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
            val(ws2, r2, 2, f'{m_s:,.0f}')
            val(ws2, r2, 3, f'{org_s:.1%}' if org_s > 0 else '-', bg=C_GREEN_LIGHT if org_s > 0.5 else C_WHITE)
            val(ws2, r2, 4, f'{spon_s:.1%}' if spon_s > 0 else '-', bg=C_RED_LIGHT if spon_s > 0.6 else C_WHITE)
            val(ws2, r2, 5, f'{spon_asins}' if spon_asins > 0 else '-')
            val(ws2, r2, 6, f'{dsr:.1f}' if dsr > 0 else '-', bg=C_GREEN_LIGHT if dsr > 1 else (C_RED_LIGHT if dsr < 0.5 and dsr > 0 else C_WHITE))
            ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
            val(ws2, r2, 7, strategy, bold=True, bg=strat_bg)
            ws2.row_dimensions[r2].height = 20
            r2 += 1
        if r2 > ad_start:
            apply_border(ws2, ad_start-1, r2-1, 1, 8)
        r2 += 1
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
        val(ws2, r2, 1, f'小结：Top 15 词中 {ad_dominant} 个广告主导、{organic_dominant} 个自然流量为主。{"新品必须配合广告投放" if ad_dominant > organic_dominant else "部分词可通过 SEO 获取自然流量"}', bg=C_BLUE_LIGHT, fg='FF1F3864')
        ws2.row_dimensions[r2].height = 22
        r2 += 2

    # ▌ 九、关键词贡献度排名（ReverseASIN）
    if kw_df is not None and len(kw_df) > 0 and 'Units Sold' in kw_df.columns:
        section_title(ws2, r2, 1, '▌ 十一、关键词贡献度排名（按月购买量降序）', span=8)
        ws2.row_dimensions[r2].height = 24
        r2 += 1

        kw_sold = kw_df.copy()
        kw_sold['Units Sold'] = pd.to_numeric(kw_sold.get('Units Sold', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_sold['M. Searches'] = pd.to_numeric(kw_sold.get('M. Searches', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_sold['Purchase Rate'] = pd.to_numeric(kw_sold.get('Purchase Rate', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_sold['Traffic Share'] = pd.to_numeric(kw_sold.get('Traffic Share', pd.Series(dtype=float)), errors='coerce').fillna(0)
        kw_sold['Related Products'] = pd.to_numeric(kw_sold.get('Related Products', pd.Series(dtype=float)), errors='coerce').fillna(0)
        total_units = kw_sold['Units Sold'].sum()  # 全局分母不变，贡献占比仍代表占总销量比

        has_relevance = kw_sold['Traffic Share'].sum() > 0 or kw_sold['Related Products'].sum() > 0
        if has_relevance:
            # ① 硬地板：Related Products >= 2 排除明显跨品类噪声（如 "ego leaf blower" / "dewalt drill"）
            kw_rel = kw_sold[kw_sold['Related Products'] >= 2].copy()
            # ② rank-sum 候选池：Traffic Share + Related Products 双维度名次相加，取前 50%
            if len(kw_rel) >= 4:
                ts_rank = kw_rel['Traffic Share'].rank(method='min', ascending=False)
                rp_rank = kw_rel['Related Products'].rank(method='min', ascending=False)
                kw_rel['_pool_score'] = ts_rank + rp_rank
                kw_pool = kw_rel[kw_rel['_pool_score'] <= kw_rel['_pool_score'].median()]
            else:
                kw_pool = kw_rel
            pool_excluded = len(kw_sold) - len(kw_pool)
            contrib_method = 'relevance_gated'
        else:
            kw_pool = kw_sold
            pool_excluded = 0
            contrib_method = 'fallback'

        kw_sold_top = kw_pool[kw_pool['Units Sold'] > 0].sort_values('Units Sold', ascending=False).head(15).reset_index(drop=True)

        # 业务口径：月购买量 = Units Sold（卖家精灵预估带单量）；购买率 = 月购买量/月搜索量
        # 新增「流量占比 / 相关产品」列，让用户直接看到选词依据
        sold_headers = ['关键词', '月搜索量', '月购买量', '购买率', '流量占比', '相关产品', '贡献占比']
        for i, h in enumerate(sold_headers, 1):
            hdr(ws2, r2, i, h, bg=C_BLUE_MID)
        ws2.row_dimensions[r2].height = 20
        r2 += 1
        sold_start = r2
        for _, row in kw_sold_top.iterrows():
            units = float(row['Units Sold'])
            contrib = units / total_units if total_units > 0 else 0
            ts = float(row.get('Traffic Share', 0) or 0)
            rp = float(row.get('Related Products', 0) or 0)
            ts_disp = (f'{ts:.2%}' if ts > 1.5 else f'{ts*100:.2f}%') if ts > 0 else '-'
            rp_disp = f'{int(rp)}' if rp > 0 else '-'
            val(ws2, r2, 1, str(row.get('Keyword', '-')), bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
            val(ws2, r2, 2, f'{float(row["M. Searches"]):,.0f}')
            val(ws2, r2, 3, f'{int(units):,}')
            val(ws2, r2, 4, f'{float(row["Purchase Rate"]):.2%}' if float(row['Purchase Rate']) > 0 else '-')
            val(ws2, r2, 5, ts_disp, bg=C_RED_LIGHT if ts >= 0.05 else (C_YELLOW if ts >= 0.01 else C_WHITE))
            val(ws2, r2, 6, rp_disp, bg=C_RED_LIGHT if rp >= 5 else (C_YELLOW if rp >= 2 else C_WHITE))
            val(ws2, r2, 7, f'{contrib:.1%}', bold=True, bg=C_GREEN_LIGHT if contrib > 0.05 else C_WHITE)
            ws2.row_dimensions[r2].height = 20
            r2 += 1
        if r2 > sold_start:
            apply_border(ws2, sold_start-1, r2-1, 1, 7)
        top5_contrib = kw_sold_top.head(5)['Units Sold'].sum() / total_units if total_units > 0 else 0
        r2 += 1
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
        if contrib_method == 'relevance_gated':
            summary_text = (f'小结：Top 5 关键词贡献 {top5_contrib:.0%} 销量，总预估带单 {int(total_units):,} 件。'
                            f'（已排除 Related Products < 2 的 {pool_excluded} 个跨品类噪声词；Top 15 在「流量占比 + 相关产品」rank-sum 前 50% 候选池内）')
        else:
            summary_text = f'小结：Top 5 关键词贡献 {top5_contrib:.0%} 销量，总预估带单 {int(total_units):,} 件。'
        val(ws2, r2, 1, summary_text, bg=C_BLUE_LIGHT, fg='FF1F3864')
        ws2.row_dimensions[r2].height = 22
        r2 += 2

    # ▌ 十、Listing 标题优化建议（Unique Words）
    uw_df = keyword_data.get('unique_words') if keyword_data.get('_available') else None
    if uw_df is not None and len(uw_df) > 5:
        section_title(ws2, r2, 1, '▌ 十二、Listing 标题优化建议（高频词）', span=8)
        ws2.row_dimensions[r2].height = 24
        r2 += 1

        uw = uw_df.copy()
        uw['Frequency'] = pd.to_numeric(uw['Frequency'], errors='coerce').fillna(0)
        total_kw_count = len(kw_df) if kw_df is not None else 1000
        uw_top = uw.sort_values('Frequency', ascending=False).head(15).reset_index(drop=True)

        uw_headers = ['排名', '高频词', '出现次数', '覆盖率', '建议']
        for i, h in enumerate(uw_headers, 1):
            hdr(ws2, r2, i, h, bg=C_BLUE_MID)
        ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
        ws2.row_dimensions[r2].height = 20
        r2 += 1
        uw_start = r2
        must_words = []
        for idx, (_, row) in enumerate(uw_top.iterrows()):
            freq = int(row['Frequency'])
            coverage = freq / total_kw_count
            if coverage > 0.2:
                advice = '必填'
                advice_bg = C_RED_LIGHT
                must_words.append(str(row['Phrase']))
            elif coverage > 0.1:
                advice = '建议'
                advice_bg = C_YELLOW
            else:
                advice = '可选'
                advice_bg = C_WHITE
            val(ws2, r2, 1, f'{idx+1}', h_align='center')
            val(ws2, r2, 2, str(row['Phrase']), bold=True, bg=C_BLUE_LIGHT)
            val(ws2, r2, 3, f'{freq}')
            val(ws2, r2, 4, f'{coverage:.1%}')
            ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
            val(ws2, r2, 5, advice, bold=True, bg=advice_bg)
            ws2.row_dimensions[r2].height = 18
            r2 += 1
        if r2 > uw_start:
            apply_border(ws2, uw_start-1, r2-1, 1, 8)
        r2 += 1
        if must_words:
            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
            val(ws2, r2, 1, f'标题必含词：{", ".join(must_words)}', bold=True, bg=C_YELLOW, fg='FF7A4F01')
            ws2.row_dimensions[r2].height = 22
            r2 += 1
        r2 += 1

    # ▌ 十一、素材与配送门槛（US-Market）
    if market_data.get('_available'):
        has_data = False
        threshold_rows = []
        ff_df = market_data.get('fulfillment')
        if ff_df is not None and len(ff_df) > 0:
            try:
                for _, row in ff_df.iterrows():
                    if 'fba' in str(row.iloc[0]).lower():
                        fba_pct = float(pd.to_numeric(row.iloc[2], errors='coerce') or 0)
                        fba_sales_pct = float(pd.to_numeric(row.iloc[4], errors='coerce') or 0)
                        threshold_rows.append(('FBA 销量占比', f'{fba_sales_pct:.1%}', '必须走 FBA' if fba_sales_pct > 0.8 else ('FBA 为主' if fba_sales_pct > 0.5 else 'FBM 也可')))
                        has_data = True
                        break
            except Exception:
                pass
        ap_df = market_data.get('aplus_video')
        if ap_df is not None and len(ap_df) > 0:
            try:
                # 每档只保留第 1 条匹配行，避免同一 label 被 7 份数据行重复写入
                _aplus_seen = False
                _video_seen = False
                for _, row in ap_df.iterrows():
                    t_raw = str(row.iloc[0]).strip()
                    t = t_raw.lower()
                    pct = float(pd.to_numeric(row.iloc[2], errors='coerce') or 0)
                    if ('a+' in t or 'a plus' in t) and not _aplus_seen:
                        _label = f'A+ 页面覆盖率（{t_raw}）' if t_raw else 'A+ 页面覆盖率'
                        threshold_rows.append((_label, f'{pct:.1%}',
                            '建议制作 A+' if pct > 0.5 else 'A+ 非必需'))
                        has_data = True
                        _aplus_seen = True
                    elif 'video' in t and not _video_seen:
                        threshold_rows.append(('视频覆盖率', f'{pct:.1%}',
                            '建议制作视频' if pct > 0.3 else '视频非必需'))
                        has_data = True
                        _video_seen = True
            except Exception:
                pass
        # 中国卖家占比统一口径 —— 不再写 cn_seller_ratio 字面判断
        threshold_rows.append(('中国卖家占比', f'{cn_pct_total:.1%}',
            '价格竞争激烈' if cn_pct_total > 0.7 else '竞争多元'))
        has_data = True

        if has_data and threshold_rows:
            section_title(ws2, r2, 1, '▌ 十三、素材与配送门槛', span=8)
            ws2.row_dimensions[r2].height = 24
            r2 += 1
            hdr(ws2, r2, 1, '指标', bg=C_BLUE_MID)
            hdr(ws2, r2, 2, '数值', bg=C_BLUE_MID)
            ws2.merge_cells(start_row=r2, start_column=3, end_row=r2, end_column=8)
            hdr(ws2, r2, 3, '含义', bg=C_BLUE_MID)
            ws2.row_dimensions[r2].height = 20
            r2 += 1
            th_start = r2
            for label, value, meaning in threshold_rows:
                val(ws2, r2, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
                val(ws2, r2, 2, value, bold=True)
                ws2.merge_cells(start_row=r2, start_column=3, end_row=r2, end_column=8)
                val(ws2, r2, 3, meaning, bg=C_WHITE, fg='FF595959')
                ws2.row_dimensions[r2].height = 20
                r2 += 1
            apply_border(ws2, th_start-1, r2-1, 1, 8)
            r2 += 1

    # ▌ 十二、评论门槛（US-Market）
    ratings_df = market_data.get('ratings_dist') if market_data.get('_available') else None
    if ratings_df is not None and len(ratings_df) > 2:
        section_title(ws2, r2, 1, '▌ 十四、评论门槛分析', span=8)
        ws2.row_dimensions[r2].height = 24
        r2 += 1

        rt = ratings_df.copy()
        rt_headers = ['评论数区间', '商品数', '销量', '销量占比', '月销售额($)', '销售额占比']
        for i, h in enumerate(rt_headers, 1):
            hdr(ws2, r2, i, h, bg=C_BLUE_MID)
        ws2.merge_cells(start_row=r2, start_column=6, end_row=r2, end_column=8)
        ws2.row_dimensions[r2].height = 20
        r2 += 1
        rt_start = r2
        for _, row in rt.iterrows():
            try:
                band = str(row.iloc[0])
                if band == 'nan' or 'no rating' in band.lower():
                    continue
                products = int(pd.to_numeric(row.iloc[1], errors='coerce') or 0)
                sales = int(pd.to_numeric(row.iloc[2], errors='coerce') or 0)
                sp = float(pd.to_numeric(row.iloc[3], errors='coerce') or 0)
                rev = float(pd.to_numeric(row.iloc[4], errors='coerce') or 0)
                rp = float(pd.to_numeric(row.iloc[5], errors='coerce') or 0)
                val(ws2, r2, 1, band, bold=True, bg=C_BLUE_LIGHT)
                val(ws2, r2, 2, f'{products}')
                val(ws2, r2, 3, f'{sales:,}')
                val(ws2, r2, 4, f'{sp:.1%}')
                val(ws2, r2, 5, f'{rev:,.0f}')
                ws2.merge_cells(start_row=r2, start_column=6, end_row=r2, end_column=8)
                val(ws2, r2, 6, f'{rp:.1%}')
                ws2.row_dimensions[r2].height = 18
                r2 += 1
            except Exception:
                continue
        if r2 > rt_start:
            apply_border(ws2, rt_start-1, r2-1, 1, 8)
        r2 += 1
        # 计算门槛建议
        try:
            cumsum = 0
            threshold_band = '200+'
            for _, row in rt.iterrows():
                sp = float(pd.to_numeric(row.iloc[3], errors='coerce') or 0)
                cumsum += sp
                if cumsum >= 0.5:
                    threshold_band = str(row.iloc[0])
                    break
            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
            val(ws2, r2, 1, f'建议：进入 Top 50 销量阵营至少需要 {threshold_band} 条评论。新品前期重点积累评价。', bold=True, bg=C_YELLOW, fg='FF7A4F01')
            ws2.row_dimensions[r2].height = 22
            r2 += 1
        except Exception:
            pass
        r2 += 1

    # ===== Sheet 3: BSR TOP100 =====
    ws3 = wb.create_sheet('BSR TOP100')
    ws3.sheet_view.showGridLines = False

    # 第 3 列插入「主图」：兼容两种源表列名
    #   中文版导出：'商品主图' → 经 normalize 映射为 'Main Image'
    #   英文版导出：直接是 'Image URL'
    _img_col = 'Main Image' if 'Main Image' in df.columns else ('Image URL' if 'Image URL' in df.columns else None)
    _has_main_img = _img_col is not None
    # N 列「产品类型」= df['product_type']（聚合后的 segment 桶名，与 Sheet 1 三、产品类型分布 / Sheet 4/6/10 同源），
    # 业务先看粗类——这个 ASIN 归属哪个 segment 桶，可与其他 sheet 的统计对应。
    # O 列「所属细分」= visual_product_type（视觉 LLM 自由描述每 ASIN，不受 segment 列表限制），
    # 业务再看细描述——具体到产品级别的形态/材质/功能组合（per-ASIN 唯一）。
    display_cols = ['#', 'ASIN', _img_col or 'Main Image', 'Brand', 'Product Title', price_col, 'Monthly Sales', rev_col,
                    'Rating', 'Ratings', 'Available days', 'BuyBox Location', 'Fulfillment', 'product_type',
                    'visual_product_type', 'visual_material', 'Gross Margin']
    display_names = ['排名', 'ASIN', '主图', '品牌', '产品标题', '价格($)', '月销量', '月收入($)',
                     '星级', '评分数', '上架天数', '卖家地区', '配送', '产品类型',
                     '所属细分', '材质标签', '毛利率']
    widths_3 = [6, 14, 14, 14, 90, 10, 10, 13, 8, 10, 10, 10, 10, 22, 22, 12, 10]
    # 缺图列时降级移除该列，保持表头/列宽对齐
    if not _has_main_img:
        idx = display_cols.index('Main Image')
        display_cols.pop(idx); display_names.pop(idx); widths_3.pop(idx)
    # 图片列在主循环里通过列名识别，统一变量供后续 if/插入使用
    _img_col_name = _img_col
    last_col_letter = get_column_letter(len(display_cols))
    for i, w in enumerate(widths_3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.merge_cells(f'A1:{last_col_letter}1')
    c = ws3['A1']
    c.value = f'BSR TOP100 - {display_name_for_title(packs, "选品")} - {datetime.now().strftime("%Y-%m-%d")}'
    c.font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 30

    for i, name in enumerate(display_names):
        hdr(ws3, 2, i+1, name, bg=C_BLUE_MID)
    ws3.row_dimensions[2].height = 22

    available_cols = [c for c in display_cols if c in df.columns]
    # 预下载所有主图（并发，加速主循环）
    if _has_main_img and _img_col_name in df.columns:
        _prefetch_images([str(u) for u in df[_img_col_name].tolist() if pd.notna(u)])

    for row_idx, (_, drow) in enumerate(df[available_cols].iterrows(), start=3):
        bg = C_GREY_LIGHT if row_idx % 2 == 0 else C_WHITE
        if str(drow.get('BuyBox Location', '')) == 'CN':
            bg = C_BLUE_LIGHT if row_idx % 2 == 0 else C_WHITE
        for ci, col in enumerate(available_cols):
            if col == _img_col_name:
                # 嵌入图片：成功则单元格留空（图片浮在上面），失败则写 '-'
                url = drow[col] if pd.notna(drow[col]) else None
                ok = _insert_image_to_cell(ws3, row_idx, ci+1, str(url) if url else '') if url else False
                cell = ws3.cell(row=row_idx, column=ci+1, value='' if ok else '-')
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                continue
            v = drow[col]
            if col == 'Gross Margin' and pd.notna(v):
                v = f'{float(v)*100:.1f}%'
            elif col == rev_col and pd.notna(v):
                v = round(float(v), 1)
            c = ws3.cell(row=row_idx, column=ci+1, value=v)
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            # 标题/视觉依据列左对齐+折行；其他居中
            is_text_col = col == 'Product Title'
            h_align = 'left' if is_text_col else 'center'
            c.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=is_text_col)
        # 有主图时统一拉高行高让图片完整显示
        ws3.row_dimensions[row_idx].height = 65 if _has_main_img else 18

    apply_border(ws3, 2, 101, 1, len(available_cols))

    # 数据来源注释（行 103，放在主表 100 行数据 row 3-102 之后、数据统计汇总 row 104 之前）
    # 业务人员看 Sheet 3 时一眼明白 N/O/P 列是 LLM 判定的，不要当作权威结构化数据
    _note_row = 103
    _last_col_letter = get_column_letter(len(available_cols))
    ws3.merge_cells(f'A{_note_row}:{_last_col_letter}{_note_row}')
    _note_cell = ws3.cell(row=_note_row, column=1)
    _note_cell.value = (
        '注：N 列「产品类型」、O 列「所属细分」、P 列「材质标签」由视觉 LLM 综合产品图片+标题+卖点'
        '判定生成，并非源表直接字段（卖家精灵 BSR 导出里无现成材质/类型字段，详细参数列空缺率 >95%）；'
        '此类 LLM 判定字段会有 ~5% 误差（典型如透明材质难分辨亚克力 vs 玻璃），仅供参考。'
    )
    _note_cell.font = Font(name='Arial', size=9, italic=True, color='FF7A4F01')
    _note_cell.fill = PatternFill('solid', fgColor=C_YELLOW)
    _note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws3.row_dimensions[_note_row].height = 32

    # ===== BSR TOP100 - 新增内容汇总 =====
    r3 = 104
    
    # 统计汇总
    section_title(ws3, r3, 1, '▌ 数据统计汇总', span=14)
    ws3.row_dimensions[r3].height = 24
    r3 += 1
    
    # 计算各指标
    avg_sales = df['Monthly Sales'].mean()
    median_sales = df['Monthly Sales'].median()
    avg_rev_bsr = df[rev_col].mean() if rev_col else 0
    
    summary_stats = [
        ('月销量统计', f"均值: {int(avg_sales):,}件", f"中位数: {int(median_sales):,}件", f"最高: {int(df['Monthly Sales'].max()):,}件"),
        ('价格统计', f"均价: ${df[price_col].mean():.1f}", f"中位价: ${df[price_col].median():.1f}", f"最高: ${df[price_col].max():.0f}"),
        ('评分统计', f"均值: {df['Rating'].mean():.2f}★", f"中位数: {df['Rating'].median():.1f}★", f"最低: {df['Rating'].min():.1f}★"),
    ]
    
    for i, (stat_type, v1, v2, v3) in enumerate(summary_stats):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        val(ws3, r3, 1, stat_type, bold=True, bg=bg)
        for ci, v in enumerate([v1, v2, v3], start=2):
            c = ws3.cell(row=r3, column=ci, value=v)
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws3.row_dimensions[r3].height = 18
        r3 += 1
    apply_border(ws3, r3-len(summary_stats)-1, r3-1, 1, 4)
    r3 += 1
    
    # 高销量产品特征分析
    section_title(ws3, r3, 1, '▌ 高销量产品特征分析（TOP20）', span=14)
    ws3.row_dimensions[r3].height = 24
    r3 += 1
    
    hdr(ws3, r3, 1, '特征维度', bg=C_BLUE_MID)
    ws3.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=5)
    hdr(ws3, r3, 2, 'TOP20产品数据', bg=C_BLUE_MID)
    ws3.merge_cells(start_row=r3, start_column=6, end_row=r3, end_column=10)
    hdr(ws3, r3, 6, '整体市场数据', bg=C_BLUE_MID)
    ws3.merge_cells(start_row=r3, start_column=11, end_row=r3, end_column=14)
    hdr(ws3, r3, 11, '对比结论', bg=C_BLUE_MID)
    ws3.row_dimensions[r3].height = 20
    r3 += 1
    
    top20 = df.nlargest(20, 'Monthly Sales')
    top20_avg_price = top20[price_col].mean()
    top20_avg_rating = top20['Rating'].mean()
    top20_avg_sales = top20['Monthly Sales'].mean()
    
    feature_analysis = [
        ('平均价格', f"${top20_avg_price:.1f}", f"${df[price_col].mean():.1f}", 
         'TOP20价格略高' if top20_avg_price > df[price_col].mean() else 'TOP20价格更亲民'),
        ('平均评分', f"{top20_avg_rating:.2f}★", f"{df['Rating'].mean():.2f}★", 
         'TOP20评分更高' if top20_avg_rating > df['Rating'].mean() else '评分相近'),
        ('平均月销', f"{int(top20_avg_sales):,}件", f"{int(avg_sales):,}件", 
         '差异显著'),
    ]
    
    for i, (dim, top20_v, all_v, conclusion) in enumerate(feature_analysis):
        bg = C_YELLOW if i % 2 == 0 else C_WHITE
        val(ws3, r3, 1, dim, bold=True, bg=bg)
        ws3.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=5)
        c = ws3.cell(row=r3, column=2, value=top20_v)
        c.font = Font(name='Arial', bold=True, color='FF1F3864')
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws3.merge_cells(start_row=r3, start_column=6, end_row=r3, end_column=10)
        c = ws3.cell(row=r3, column=6, value=all_v)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws3.merge_cells(start_row=r3, start_column=11, end_row=r3, end_column=14)
        val(ws3, r3, 11, conclusion, bg=bg, fg='FF595959')
        ws3.row_dimensions[r3].height = 20
        r3 += 1
    apply_border(ws3, r3-len(feature_analysis)-1, r3-1, 1, 14)
    r3 += 1
    
    # TOP20 ASIN列表
    section_title(ws3, r3, 1, '▌ TOP20 ASIN速查（可直接竞品分析）', span=14)
    ws3.row_dimensions[r3].height = 24
    r3 += 1
    
    hdr(ws3, r3, 1, '排名', bg=C_BLUE_MID)
    hdr(ws3, r3, 2, 'ASIN', bg=C_BLUE_MID)
    hdr(ws3, r3, 3, '品牌', bg=C_BLUE_MID)
    hdr(ws3, r3, 4, '价格', bg=C_BLUE_MID)
    hdr(ws3, r3, 5, '月销量', bg=C_BLUE_MID)
    ws3.merge_cells(start_row=r3, start_column=6, end_row=r3, end_column=14)
    hdr(ws3, r3, 6, '产品标题', bg=C_BLUE_MID)
    ws3.row_dimensions[r3].height = 20
    r3 += 1
    
    for idx, (_, drow) in enumerate(top20.iterrows()):
        bg = C_ORANGE if idx < 3 else (C_YELLOW if idx < 10 else C_WHITE)
        ws3.cell(row=r3, column=1, value=idx+1).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=r3, column=1).fill = PatternFill('solid', fgColor=bg)
        ws3.cell(row=r3, column=2, value=drow['ASIN']).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=r3, column=2).fill = PatternFill('solid', fgColor=bg)
        ws3.cell(row=r3, column=3, value=drow['Brand']).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=r3, column=3).fill = PatternFill('solid', fgColor=bg)
        ws3.cell(row=r3, column=4, value=f"${drow[price_col]:.1f}").alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=r3, column=4).fill = PatternFill('solid', fgColor=bg)
        ws3.cell(row=r3, column=5, value=f"{int(drow['Monthly Sales']):,}").alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=r3, column=5).fill = PatternFill('solid', fgColor=bg)
        ws3.merge_cells(start_row=r3, start_column=6, end_row=r3, end_column=14)
        title_val = str(drow['Product Title'])[:80] if pd.notna(drow['Product Title']) else ''
        c = ws3.cell(row=r3, column=6, value=title_val)
        c.font = Font(name='Arial', size=9)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws3.row_dimensions[r3].height = 18
        r3 += 1
    apply_border(ws3, r3-21, r3-1, 1, 14)

    # ===== Sheet 4: 推荐入场价 =====
    ws4 = wb.create_sheet('推荐入场价')
    ws4.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGHI', [22, 14, 14, 14, 14, 16, 20, 25, 14]):
        ws4.column_dimensions[col_letter].width = width

    ws4.merge_cells('A1:I1')
    c = ws4['A1']
    c.value = f"{display_name_for_title(packs, '选品')} — 各产品类型推荐入场价区间"
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 35

    r4 = 3
    section_title(ws4, r4, 1, '▌ 推荐入场价逻辑：以市场P25-P75价位为参考，结合单品收益分析', span=9)
    ws4.row_dimensions[r4].height = 24
    r4 += 1

    price_hdr4 = ['产品类型', '市场MIN', '市场P25', '市场中位', '市场P75', '市场MAX', '推荐入场价', '理由', '备注']
    for i, h in enumerate(price_hdr4):
        hdr(ws4, r4, i+1, h, bg=C_BLUE_MID)
    ws4.row_dimensions[r4].height = 22
    r4 += 1

    # 使用动态计算的推荐入场价数据
    colors_row = [C_YELLOW, C_GREEN_LIGHT, C_WHITE, C_WHITE, C_BLUE_LIGHT, C_GREY_LIGHT]
    for i, rec in enumerate(pricing_recommendations):
        bg = colors_row[i] if i < len(colors_row) else C_WHITE
        row_p = (
            rec['product_type'],
            f"${rec['min_price']:.1f}",
            f"${rec['p25']:.1f}",
            f"${rec['median']:.1f}",
            f"${rec['p75']:.1f}",
            f"${rec['max_price']:.1f}",
            f"${rec['rec_min']:.2f}-${rec['rec_max']:.2f}",
            rec['reason'],
            rec['note']
        )
        for ci, v in enumerate(row_p):
            c = ws4.cell(row=r4, column=ci+1, value=v)
            c.fill = PatternFill('solid', fgColor=bg)
            c.font = Font(name='Arial', size=10, bold=(ci == 0 or ci == 6))
            c.alignment = Alignment(horizontal='center' if ci not in [0, 7, 8] else 'left',
                                    vertical='center', wrap_text=True)
        # 理由列内容含多段（P25/中位/P75 + SKU + 月收益 + 毛利中位 + 卖家精灵注释），
        # 列宽 25 下需要 4-5 行才能完整展示——按理由文本字数动态算高，避免被截断
        _reason_text = str(row_p[7] or '')
        # 列宽 25 字符对应每行约 20 个汉字，含中英文混排取保守值
        _est_lines = max(3, (len(_reason_text) // 18) + 1)
        ws4.row_dimensions[r4].height = max(45, _est_lines * 16)
        r4 += 1

    apply_border(ws4, 5, r4-1, 1, 9)
    r4 += 1

    # 注解行：解释 ★首推入场品类 的决策逻辑（与 Sheet 6 R10 + Sheet 10 R32 注解一致的统一口径）
    _uses_llm_matrix = (
        packs is not None and packs.is_synthesis_real()
        and any(
            it.priority == 'P1' and '不推荐' not in (it.action_plan or '') and it.segment
            for it in (packs.synthesis.sheet6_priority_matrix or [])
        )
    )
    if _uses_llm_matrix:
        _star_note = (
            '注：★首推入场品类由 LLM 综合策略矩阵 P1 决定（依据 Market/VOC/Traffic/Trend 4 份源数据深度分析）；'
            '若 LLM 不可用则回退到综合评分 = 需求(月销)×0.30 + 单品收益×0.25 + 新品占比×0.15 + 质量评分×0.15 + 竞争度×0.15'
        )
    else:
        _star_note = (
            '注：★首推入场品类由综合评分自动排序 = 需求(月销)×0.30 + 单品收益×0.25 + 新品占比×0.15 + 质量评分×0.15 + 竞争度×0.15'
            '（数据驱动，非固定）'
        )
    ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=9)
    c = ws4.cell(row=r4, column=1)
    c.value = _star_note
    c.font = Font(name='Arial', size=9, color='FF7A4F01', italic=True)
    c.fill = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws4.row_dimensions[r4].height = 32
    r4 += 2

    # 利润测算（基于首推品类的推荐价格）—— 必须使用统一 _top_segment_name 对应的那条 pricing_rec，
    # 不能用 pricing_recommendations[0]（那是 type_agg 迭代序，首条通常是 SKU 最多的品类而非 P1）
    first_rec = None
    if pricing_recommendations:
        for _pr in pricing_recommendations:
            if _pr.get('product_type') == _top_segment_name:
                first_rec = _pr
                break
        if first_rec is None:
            first_rec = pricing_recommendations[0]
    if first_rec:
        profit_price = (first_rec['rec_min'] + first_rec['rec_max']) / 2
        profit_title = f"▌ {first_rec['product_type']} 入场利润测算（${profit_price:.2f}档）"
    else:
        profit_price = 34.99
        # pricing_recommendations 为空时用品类中位价代替（动态品类名）
        profit_title = f"▌ {display_name_for_title(packs, '选品')} 入场利润测算（${profit_price:.2f}档）"
    
    section_title(ws4, r4, 1, profit_title, span=9)
    ws4.row_dimensions[r4].height = 24
    r4 += 1

    # 利润测算 —— 跨品类自适应：FBA 用 BSR 数据真实中位数，其他按行业普遍比例
    # 全部参数化，避免写死任何品类专属数字
    commission_rate = 0.15         # 行业普遍 15%（大多数品类落 8-17%，15% 是中位）
    ad_rate = 0.12                 # 新品期 CPC 常规 12%
    purchase_rate_low = 0.45       # 行业普遍采购占售价 45%（紧）
    purchase_rate_high = 0.55      # 55%（松）
    tax_rate = 0.075               # 跨境关税 + 杂费比例约 7.5%

    commission = profit_price * commission_rate
    ad_fee = profit_price * ad_rate
    purchase_cost_low = profit_price * purchase_rate_low
    purchase_cost_high = profit_price * purchase_rate_high
    other_fee = profit_price * tax_rate

    # FBA：从 BSR 数据 FBA($) 列取真实中位数（按品类自适应）；找不到列或全空 → 兜底 售价 × 25%
    fba_fee = profit_price * 0.25
    _fba_source_note = '兜底估算：售价 × 25%（BSR 数据无 FBA($) 列或全空）'
    if 'FBA($)' in df.columns:
        _fba_series = pd.to_numeric(df['FBA($)'], errors='coerce').dropna()
        _fba_series = _fba_series[_fba_series > 0]
        if len(_fba_series) > 0:
            fba_fee = float(_fba_series.median())
            _fba_source_note = f'从 BSR 数据 FBA($) 列取真实中位数（{len(_fba_series)} 个竞品实测，按品类自适应）'
    profit_low = profit_price - commission - fba_fee - ad_fee - purchase_cost_low - other_fee
    profit_high = profit_price - commission - fba_fee - ad_fee - purchase_cost_high - other_fee

    # 以 BSR Gross Margin 实际中位数为准
    _bsr_gm_pct = None
    if 'Gross Margin' in df.columns:
        try:
            _gm = pd.to_numeric(df['Gross Margin'], errors='coerce').dropna()
            if len(_gm) > 0:
                _gm_med = float(_gm.median())
                _bsr_gm_pct = _gm_med if _gm_med > 1.0 else _gm_med * 100
        except Exception:
            pass
    if _bsr_gm_pct is not None:
        profit_actual_low = profit_price * _bsr_gm_pct / 100
        profit_actual_high = profit_actual_low  # BSR 给的是中位值，不给区间
        gm_display = f'{_bsr_gm_pct:.0f}%（约 ${profit_actual_low:.2f}/件）'
        gm_basis = (
            f'市场上同价位竞品的 BSR 平台 Gross Margin 中位水平（仅扣 Amazon 平台费），'
            f'{len(df)} 个竞品实测。仅作市场参考，真实净利见下方「估算净利润」。'
        )
    else:
        gm_display = f'~{profit_low/profit_price*100:.0f}-{profit_high/profit_price*100:.0f}%（约 ${profit_low:.2f}~${profit_high:.2f}/件）'
        gm_basis = '估算区间，仅作市场参考，真实净利见下方「估算净利润」。'

    # 基于实际数据估算月销量
    if first_rec:
        est_sales_low = int(first_rec['avg_rev'] / first_rec['median'] * 0.6) if first_rec['median'] else 300
        est_sales_high = int(first_rec['avg_rev'] / first_rec['median'] * 1.2) if first_rec['median'] else 1200
    else:
        est_sales_low = 300
        est_sales_high = 500

    # 月毛利估算（用 BSR 实际毛利率）
    if _bsr_gm_pct is not None:
        month_profit_low = profit_actual_low * est_sales_low
        month_profit_high = profit_actual_high * est_sales_high
    else:
        month_profit_low = profit_low * est_sales_low
        month_profit_high = profit_high * est_sales_high

    # 下半扣项总和（用于注解里和上半的 BSR 实测毛利做对照，让读者看到两套视角的差异）
    _est_cost_low = commission + fba_fee + ad_fee + purchase_cost_low + other_fee
    _est_cost_high = commission + fba_fee + ad_fee + purchase_cost_high + other_fee
    _est_profit_low = profit_price - _est_cost_high
    _est_profit_high = profit_price - _est_cost_low
    _est_gm_low = _est_profit_low / profit_price * 100 if profit_price else 0
    _est_gm_high = _est_profit_high / profit_price * 100 if profit_price else 0

    _gap_note = (
        f'上半「毛利润 {gm_display}」= 售价 × BSR Gross Margin 中位 {_bsr_gm_pct:.0f}%，卖家精灵原始字段，'
        f'口径通常已扣除 Amazon 佣金 / FBA 等平台费用，是结果型口径。'
        f'下半按佣金+FBA+CPC+采购+税费独立估算单件扣项约 ${_est_cost_low:.2f}~${_est_cost_high:.2f}，'
        f'推出净利约 ${_est_profit_low:.2f}~${_est_profit_high:.2f}（毛利率 {_est_gm_low:.0f}-{_est_gm_high:.0f}%），'
        f'是结构型口径。'
        f'两者差异来自 BSR 毛利口径是否含广告费/采购成本 —— 下半作为结构拆解参考，最终以上半 BSR 实测为准。'
    ) if _bsr_gm_pct is not None else (
        f'上半毛利为按佣金/FBA/CPC 扣项估算（保守区间），下半单行罗列每项参考值便于核对。'
    )

    profit_items = [
        (f'亚马逊佣金({commission_rate:.0%})', f'-${commission:.2f}',
            f'按售价 × {commission_rate:.0%}（行业普遍标准——大多数品类 8-17%，15% 是中位；按官方 referral fee 表为准）'),
        ('FBA 费用', f'-${fba_fee:.2f}', _fba_source_note),
        (f'广告费(CPC 估算,{ad_rate:.0%})', f'-${ad_fee:.2f}',
            f'按售价 × {ad_rate:.0%}（新品期常规——成熟期可降到 8% 以下）'),
        ('采购成本（含运费）', f'-${purchase_cost_low:.2f}~${purchase_cost_high:.2f}',
            f'按售价 × {purchase_rate_low:.0%}~{purchase_rate_high:.0%}（行业普遍参考——以工厂报价为准）'),
        ('税费/其他', f'-${other_fee:.2f}',
            f'按售价 × {tax_rate:.1%}（跨境关税 + 杂费比例参考）'),
        ('—— 以上为需扣除的费用 ——', '',
            '除「亚马逊佣金」「FBA 费用」（平台规定/物流刚性，不可压）外，'
            '广告费/采购成本/税费**均有压价空间**——详见下方「成本结构优化空间」表 + 「盈亏平衡采购成本上限」多档表。'),
        ('—— 推荐入场售价 ——', f'${profit_price:.2f}',
            f'来源：首推品类中位价×0.9 ~ ×1.1 取中值。'
            f'**最终毛利润取决于你的压价能力**——按行业参考成本算{"亏损" if _est_profit_high < 0 else "微利"}，'
            f'但若把采购压到下方「盈亏平衡采购成本上限」对应档位以下，即可达到目标净利率（详见下方多档表）。'),
    ]

    hdr(ws4, r4, 1, '成本结构', bg=C_BLUE_MID)
    ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=5)
    hdr(ws4, r4, 2, '金额', bg=C_BLUE_MID)
    ws4.merge_cells(start_row=r4, start_column=6, end_row=r4, end_column=9)
    hdr(ws4, r4, 6, '说明', bg=C_BLUE_MID)
    r4 += 1
    for i, (item, amount, note) in enumerate(profit_items):
        if '推荐入场售价' in item or '估算净利润' in item:
            bg = C_BLUE_LIGHT
            bold = True
        else:
            bg = C_GREY_LIGHT if i % 2 == 0 else C_WHITE
            bold = False
        val(ws4, r4, 1, item, bold=bold, bg=bg)
        ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=5)
        c = ws4.cell(row=r4, column=2, value=amount)
        c.font = Font(name='Arial', bold=bold, size=11, color='FF1F3864' if bold else '000000')
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws4.merge_cells(start_row=r4, start_column=6, end_row=r4, end_column=9)
        val(ws4, r4, 6, note, bg=bg, fg='FF595959')
        # note 列宽 ~75 字符，按内容字数动态算高（1 行约 50 汉字）
        _note_lines = max(1, (len(str(note or '')) // 50) + 1)
        ws4.row_dimensions[r4].height = max(22, _note_lines * 16)
        r4 += 1
    apply_border(ws4, r4-len(profit_items)-1, r4-1, 1, 9)
    r4 += 2

    # ▌ 报告阅读说明：解释"推荐 vs 亏损"逻辑（业务读到亏损时直接看到为什么仍推荐）
    _reading_note_s4 = (
        '📌 本表数据为 LLM 基于行业普遍参考值（佣金、FBA、广告、采购、税）推算的估算结果，'
        '仅作参考。按此口径算下来若亏损，不代表品类做不了——本报告"推荐"基于市场机会判断'
        '（销量+需求+痛点+新品空间）；**真实盈利能力取决于你公司的实际费用结构以及供应链压价能力**。'
        '具体采购成本目标见下方「盈亏平衡采购成本上限（按目标净利率）」多档表。'
    )
    ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=9)
    c = ws4.cell(row=r4, column=1, value=_reading_note_s4)
    c.font = Font(name='Arial', size=10, bold=True, color='FF7B5800')
    c.fill = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws4.row_dimensions[r4].height = 50
    r4 += 2

    # ▌ 成本结构优化空间（跨品类通用：每项扣项的可压性 + 怎么压）
    section_title(ws4, r4, 1, '▌ 成本结构优化空间', span=9)
    ws4.row_dimensions[r4].height = 24
    r4 += 1

    opt_hdr = ['扣项', '当前值/估算', '可压性', '空间', '怎么压']
    for i, h in enumerate(opt_hdr):
        hdr(ws4, r4, i+1, h, bg=C_BLUE_MID)
    ws4.merge_cells(start_row=r4, start_column=5, end_row=r4, end_column=9)
    r4 += 1

    opt_rows = [
        ('亚马逊佣金',
            f'{commission_rate:.0%}（${commission:.2f}/件）',
            '不可压', '0',
            '平台规定的 referral fee，按品类固定费率'),
        ('FBA 费用',
            f'${fba_fee:.2f}/件（{"BSR 实测中位" if "FBA($)" in df.columns else "估算"}）',
            '弱可压', '$0-1',
            '优化包装尺寸/重量降一点（小件 < 1lb，标准 vs 超尺寸）'),
        ('广告费 CPC',
            f'{ad_rate:.0%}（${ad_fee:.2f}/件）',
            '中等可压', f'${ad_fee*0.3:.1f}-{ad_fee*0.5:.1f}',
            '优化 ACOS、长尾词、提高转化率走自然流量；新品期高 → 成熟期可降到 8% 以下'),
        ('采购成本',
            f'${purchase_cost_low:.2f}~${purchase_cost_high:.2f}/件',
            '强可压', f'${purchase_cost_low*0.3:.1f}+',
            '找工厂底价、1688 大量采购、自营供应链。**业务利润空间最大的杠杆**'),
        ('税费/其他',
            f'${other_fee:.2f}/件',
            '弱可压', '$0-0.5',
            '海关申报方式、保险、跨境财务规划'),
    ]
    for i, (item, val_, cmp_, sp, how) in enumerate(opt_rows):
        bg = C_GREY_LIGHT if i % 2 == 0 else C_WHITE
        val(ws4, r4, 1, item, bold=True, bg=bg)
        val(ws4, r4, 2, val_, bg=bg)
        val(ws4, r4, 3, cmp_, bold=(cmp_ in ('强可压', '不可压')), bg=bg)
        val(ws4, r4, 4, sp, bg=bg)
        ws4.merge_cells(start_row=r4, start_column=5, end_row=r4, end_column=9)
        val(ws4, r4, 5, how, bg=bg, fg='FF595959', wrap=True, size=9)
        # how 列宽 ~89 字符，按内容字数动态算高
        _how_lines = max(1, (len(str(how or '')) // 60) + 1)
        ws4.row_dimensions[r4].height = max(32, _how_lines * 14)
        r4 += 1
    apply_border(ws4, r4-len(opt_rows)-1, r4-1, 1, 9)
    r4 += 2

    # ▌ 盈亏平衡采购成本上限（按目标净利率）— 含单件净利 & 月毛利润
    section_title(ws4, r4, 1, '▌ 盈亏平衡采购成本上限（按目标净利率）', span=9)
    ws4.row_dimensions[r4].height = 24
    r4 += 1

    _breakeven_indep = profit_price - commission - fba_fee - ad_fee - other_fee
    _est_sales_mid = (est_sales_low + est_sales_high) / 2  # 月销中位用于算月利

    hdr(ws4, r4, 1, '目标净利率', bg=C_BLUE_MID)
    hdr(ws4, r4, 2, '采购上限', bg=C_BLUE_MID)
    hdr(ws4, r4, 3, '单件净利', bg=C_BLUE_MID)
    hdr(ws4, r4, 4, f'月毛利润（按月销 {int(_est_sales_mid)} 件）', bg=C_BLUE_MID)
    ws4.merge_cells(start_row=r4, start_column=5, end_row=r4, end_column=9)
    hdr(ws4, r4, 5, '说明', bg=C_BLUE_MID)
    r4 += 1

    _be_rows_indep = [
        (0.0,  '保本（0%）', '亏不亏的临界——采购压到这个数就是 0 利润'),
        (0.10, '保 10% 净利', '走量打法，薄利多销'),
        (0.15, '保 15% 净利', '行业常见目标'),
        (0.20, '保 20% 净利', '中等目标，需要供应链有一定竞争力'),
        (0.30, '保 30% 净利', '高目标，仅适合品牌溢价/独家货源'),
    ]
    for i, (rate, label, note) in enumerate(_be_rows_indep):
        upper = _breakeven_indep - profit_price * rate
        unit_profit = profit_price * rate
        monthly_profit = unit_profit * _est_sales_mid

        bg = C_YELLOW if i == 2 else (C_GREY_LIGHT if i % 2 == 0 else C_WHITE)
        bold = (i == 2)

        val(ws4, r4, 1, label, bold=bold, bg=bg)

        upper_str = f'≤${upper:.2f}' if upper >= 0 else '<$0（不可达）'
        c = ws4.cell(row=r4, column=2, value=upper_str)
        c.font = Font(name='Arial', bold=bold, size=11,
                      color='FF1F5C1F' if upper >= 0 else 'FF7B0000')
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')

        if upper < 0:
            val(ws4, r4, 3, '—', bg=bg)
            val(ws4, r4, 4, '—', bg=bg)
        else:
            c = ws4.cell(row=r4, column=3, value=f'${unit_profit:.2f}')
            c.font = Font(name='Arial', bold=bold, size=11, color='FF1F5C1F')
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')

            c = ws4.cell(row=r4, column=4, value=f'${monthly_profit:,.0f}')
            c.font = Font(name='Arial', bold=bold, size=11, color='FF1F5C1F')
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')

        ws4.merge_cells(start_row=r4, start_column=5, end_row=r4, end_column=9)
        val(ws4, r4, 5, note, bg=bg, fg='FF595959', size=9, wrap=True)
        # note 列宽 ~89 字符，按内容字数动态算高
        _note_lines = max(1, (len(str(note or '')) // 60) + 1)
        ws4.row_dimensions[r4].height = max(22, _note_lines * 14)
        r4 += 1
    apply_border(ws4, r4-len(_be_rows_indep)-1, r4-1, 1, 9)
    r4 += 2

    # ===== 推荐入场价 - 新增内容 =====
    # 首推品类决策依据
    section_title(ws4, r4, 1, '▌ 首推品类决策依据详细说明', span=9)
    ws4.row_dimensions[r4].height = 24
    r4 += 1
    
    if first_rec:
        rec_price_low = first_rec['rec_min']
        rec_price_high = first_rec['rec_max']
        rec_sku = first_rec['sku_count']
        rec_monthly_sales = int(first_rec.get('avg_rev', 0) / first_rec['median']) * rec_sku if rec_sku > 0 else 0
    
    # 首推品类销量：用 first_rec['product_type'] 动态查 type_agg，避免硬编码 LED 品类名
    _first_type = first_rec['product_type'] if first_rec else ''
    _first_sales = int(type_agg[type_agg['product_type'] == _first_type]['total_sales'].sum()) if _first_type else 0
    # 痛点 Top3：按"首推品类的 1-2★ 评论文本"对每个 cluster 做 keywords 命中计数
    # 通用工具 _filter_voc_pain_by_asins（适用于任意品类，不绑定具体子类型名）
    _pain_names = []
    if first_rec and _first_type and packs and packs.is_voc_real() and packs.voc.pain_clusters:
        _type_asins = set(df.loc[df['product_type'] == _first_type, 'ASIN']
                            .astype(str).str.strip())
        _pain_names = [n for n, _ in
                       _filter_voc_pain_by_asins(packs.voc.pain_clusters, _type_asins, rev_df, top_n=3)]
    # Fallback：子集匹配为空（评论缺失/cluster 无 keywords/全 0 命中）→ 全局 Top3
    if not _pain_names and packs and packs.is_voc_real() and packs.voc.pain_clusters:
        _pain_names = [c.name for c in packs.voc.pain_clusters[:3]]
    _pain_line = f'• 该品类差评痛点明确（{"/".join(_pain_names)}），产品改良空间大' if _pain_names else ''
    # 竞品 SKU 数改为用 first_rec 自己的 sku_count
    _rec_sku = first_rec['sku_count'] if first_rec else 0

    decision_basis = [
        ('为什么选择这个品类？',
         '\n'.join(filter(None, [
             f'• 月销量{_first_sales:,}件，{_rec_sku}个SKU有竞争空间' if _first_sales else '',
             f'• 均价${first_rec["median"]:.0f}，定价${rec_price_low:.0f}-${rec_price_high:.0f}位于市场主力区间' if first_rec else '',
             _pain_line,
         ]))),
        ('定价策略依据？',
         f'• P25=${first_rec["p25"]:.1f}，P75=${first_rec["p75"]:.1f}，中位价${first_rec["median"]:.0f}\n'
         f'• 推荐价=${rec_price_low:.2f}-${rec_price_high:.2f}位于P40-P60区间\n'
         f'• 价格层面有竞争力' if first_rec else '—'),
        ('风险评估与应对？',
         '\n'.join(filter(None, [
             f'• 风险1（市场层面）：竞品多（{_rec_sku}个SKU），需差异化突围' if _rec_sku else '',
             f'• 应对：聚焦差评痛点（{"/".join(_pain_names[:2])}），不做同质化竞争' if _pain_names else '',
             '• 风险2（运营层面）：广告 CPC 高，ACOS 初期可能超 30%',
             '• 应对：优化 listing 转化，配合秒杀活动',
             '• **风险3（成本层面）：按行业参考采购成本算下来亏损，必须把采购压到「盈亏平衡采购上限」以下**',
             '• 应对：多家工厂询价、考虑 1688 大厂直采、关注海运/包装成本优化',
         ]))),
    ]
    
    for i, (title, content) in enumerate(decision_basis):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=2)
        val(ws4, r4, 1, title, bold=True, bg=bg, fg='FF1F3864')
        ws4.merge_cells(start_row=r4, start_column=3, end_row=r4, end_column=9)
        val(ws4, r4, 3, content, bg=bg, fg='FF595959', wrap=True)
        ws4.row_dimensions[r4].height = 60
        r4 += 1
    apply_border(ws4, r4-len(decision_basis)-1, r4-1, 1, 9)
    r4 += 1
    
    # 各价格带竞争分析
    section_title(ws4, r4, 1, '▌ 各价格带竞争强度分析', span=9)
    ws4.row_dimensions[r4].height = 24
    r4 += 1
    
    hdr(ws4, r4, 1, '价格区间', bg=C_BLUE_MID)
    hdr(ws4, r4, 2, 'SKU数量', bg=C_BLUE_MID)
    hdr(ws4, r4, 3, '占比', bg=C_BLUE_MID)
    hdr(ws4, r4, 4, '平均月销', bg=C_BLUE_MID)
    hdr(ws4, r4, 5, '竞争强度', bg=C_BLUE_MID)
    ws4.merge_cells(start_row=r4, start_column=6, end_row=r4, end_column=9)
    hdr(ws4, r4, 6, '入场建议', bg=C_BLUE_MID)
    ws4.row_dimensions[r4].height = 20
    r4 += 1
    
    # 价格带与 Sheet 1 共用同一套 bins/labels（上游已定义），基于首推品类的中位价判断哪一档是"最优区间"
    # bins/labels 来自 Sheet 1 段：[0,15,25,35,50,70,100,999] / ['<$15','$15-25','$25-35','$35-50','$50-70','$70-100','>$100']
    _band_labels = list(labels)
    _rec_median = float(first_rec['median']) if first_rec else float(df[price_col].median())
    # 找出首推中位价落在哪一档
    _rec_band_idx = -1
    for _i, _b in enumerate(bins[:-1]):
        if bins[_i] <= _rec_median < bins[_i + 1]:
            _rec_band_idx = _i
            break
    _total_sku = len(df)
    # LLM 按本品类给的每档价格带的一句定性叙述
    try:
        from core.packs_runtime import price_band_narratives
        _band_narr_map = price_band_narratives(packs)
    except Exception:
        _band_narr_map = {}
    price_band_analysis: list[tuple[str, int, float, int, str, str]] = []
    for _i, _label in enumerate(_band_labels):
        _lo, _hi = bins[_i], bins[_i + 1]
        _sub = df[(df[price_col] >= _lo) & (df[price_col] < _hi)]
        _cnt = len(_sub)
        if _cnt == 0:
            continue
        if 'Monthly Sales' in _sub.columns:
            _mean_sales = pd.to_numeric(_sub['Monthly Sales'], errors='coerce').mean()
            _avg_sales = int(_mean_sales) if pd.notna(_mean_sales) else 0
        else:
            _avg_sales = 0
        # 竞争强度按 SKU 占比判断（纯数据驱动，不再硬编码）
        _share = _cnt / max(_total_sku, 1)
        if _share >= 0.35:
            _intensity = '高'
        elif _share >= 0.15:
            _intensity = '中'
        else:
            _intensity = '低'
        _advice_parts = []
        if _i == _rec_band_idx:
            _advice_parts.append('★ 首推中位价所在区间')
        if _intensity == '高':
            _advice_parts.append(f'{_cnt} 个 SKU 竞争激烈，需差异化突围')
        elif _intensity == '低':
            _advice_parts.append(f'{_cnt} 个 SKU，蓝海区间但需验证需求')
        else:
            _advice_parts.append(f'{_cnt} 个 SKU，竞争适中')
        # 拼 LLM 定性叙述（LLM 按本品类给的竞品画像 / 利润空间 / 新品入场可行性）
        # 尝试多种 band 格式兼容：原始 _label（如 '<$15'）或去掉 '<' '>' 前缀的版本
        _llm_narr = (_band_narr_map.get(_label)
                     or _band_narr_map.get(_label.lstrip('<>').strip())
                     or '')
        if _llm_narr:
            _advice_parts.append(_llm_narr)
        price_band_analysis.append(
            (_label, _cnt, _share, _avg_sales, _intensity, '；'.join(_advice_parts))
        )
    
    for i, (band, cnt, share, avg_s, intensity, advice) in enumerate(price_band_analysis):
        bg = C_GREEN_LIGHT if '★' in advice else (C_RED_LIGHT if intensity == '高' else C_WHITE)
        val(ws4, r4, 1, band, bold=True, bg=bg)
        ws4.cell(row=r4, column=2, value=cnt).alignment = Alignment(horizontal='center', vertical='center')
        ws4.cell(row=r4, column=2).fill = PatternFill('solid', fgColor=bg)
        # 占比列：直接显示已计算的 _share，让业务方理解竞争强度评级依据
        ws4.cell(row=r4, column=3, value=f'{share*100:.1f}%').alignment = Alignment(horizontal='center', vertical='center')
        ws4.cell(row=r4, column=3).fill = PatternFill('solid', fgColor=bg)
        ws4.cell(row=r4, column=4, value=f'{avg_s:,}件').alignment = Alignment(horizontal='center', vertical='center')
        ws4.cell(row=r4, column=4).fill = PatternFill('solid', fgColor=bg)
        intensity_color = 'FF1F3864' if intensity == '低' else ('FF7B0000' if intensity == '高' else 'FF595959')
        c = ws4.cell(row=r4, column=5, value=intensity)
        c.font = Font(name='Arial', bold=True, color=intensity_color)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws4.merge_cells(start_row=r4, start_column=6, end_row=r4, end_column=9)
        val(ws4, r4, 6, advice, bg=bg, fg='FF595959')
        ws4.row_dimensions[r4].height = 22
        r4 += 1
    apply_border(ws4, r4-len(price_band_analysis)-1, r4-1, 1, 9)
    # 说明注释：让业务方一眼看懂"占比"定义和"竞争强度"阈值
    ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=9)
    _note = ('注：占比 = 该价格档 SKU 数 ÷ BSR TOP100 总 SKU 数。'
             '竞争强度阈值：占比 ≥ 35% 为「高」（红海，需差异化突围）；'
             '15% ≤ 占比 < 35% 为「中」（适中竞争）；'
             '占比 < 15% 为「低」（蓝海机会，但需验证需求量）。'
             '本指标仅反映 SKU 数量分布，未考虑销量集中度／品牌集中度。')
    c = ws4.cell(row=r4, column=1)
    c.value = _note
    c.font = Font(name='Arial', size=9, color='FF7A4F01', italic=True)
    c.fill = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws4.row_dimensions[r4].height = 36
    r4 += 2

    # ===== Sheet 5: 竞品分析 =====
    ws5 = wb.create_sheet('竞品分析')
    ws5.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGHIJK', [22, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]):
        ws5.column_dimensions[col_letter].width = width

    ws5.merge_cells('A1:K1')
    c = ws5['A1']
    c.value = f'{display_name_for_title(packs, "选品")} — 竞品分析（基于{len(df)}个BSR样本 + {len(rev_df)}条真实评论）'
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 35

    r5 = 3

    # ▌ 一、Top 10 竞品参数横向对比
    section_title(ws5, r5, 1, '▌ 一、Top 10 竞品参数横向对比（按月销排序）', span=11)
    ws5.row_dimensions[r5].height = 24
    r5 += 1

    # 取 Top 10 竞品
    sort_key = None
    if 'Monthly Sales' in df.columns:
        sort_key = 'Monthly Sales'
    elif rev_col:
        sort_key = rev_col
    top10 = df.sort_values(sort_key, ascending=False).head(10).reset_index(drop=True) if sort_key else df.head(10).reset_index(drop=True)
    n_top = len(top10)

    def g(row, col_name, default=''):
        if col_name in top10.columns:
            v = row.get(col_name)
            if pd.isna(v):
                return default
            return v
        return default

    # 动态识别所有长文本列（BSR 源表里实际列名含大小写 + 多种命名：
    # 'Bullet Points' / 'Product Overview' / 'Features' / 'Product Description' / 'Five Point Description'）。
    # 旧代码只按 'bullet' 小写匹配会漏掉 'Bullet Points' 和 'Product Overview'，导致规格抽取命中率极低。
    _LONG_TEXT_KEYS = ('bullet', 'feature', 'overview', 'description', 'five point', '五点', '产品描述')

    def _collect_long_text(row) -> str:
        parts = []
        for c in row.index:
            cl = str(c).lower()
            if any(k in cl for k in _LONG_TEXT_KEYS):
                v = row[c]
                if pd.notna(v):
                    parts.append(str(v))
        return ' '.join(parts)

    # 提取 specs（标题 + 所有识别到的长文本列合并）
    specs_list = [extract_all_specs(
        g(top10.iloc[i], 'Product Title', ''),
        _collect_long_text(top10.iloc[i]),
    ) for i in range(n_top)]

    # 同时对全量 BSR 商品提取 specs（供 Sheet 10 聚合推荐使用）
    all_specs_for_agg = [extract_all_specs(
        str(df.iloc[i].get('Product Title', '')),
        _collect_long_text(df.iloc[i]),
    ) for i in range(len(df))]

    # 主图行：兼容 'Main Image'（中文版映射后名）/ 'Image URL'（英文版直出名）
    _top10_imgs = None
    _t10_img_col = 'Main Image' if 'Main Image' in top10.columns else ('Image URL' if 'Image URL' in top10.columns else None)
    if _t10_img_col:
        _top10_imgs = [str(g(top10.iloc[i], _t10_img_col, '')) for i in range(n_top)]
        _prefetch_images([u for u in _top10_imgs if u])

    # 各属性行
    attr_rows = [
        ('ASIN', [str(g(top10.iloc[i], 'ASIN', '-')) for i in range(n_top)]),
        ('品牌', [str(g(top10.iloc[i], 'Brand', '-')) for i in range(n_top)]),
        ('价格', [f'${g(top10.iloc[i], price_col, 0):.2f}' if price_col else '-' for i in range(n_top)]),
        ('月销量', [f'{int(g(top10.iloc[i], "Monthly Sales", 0)):,}' if 'Monthly Sales' in top10.columns else '-' for i in range(n_top)]),
        ('月销售额', [f'${g(top10.iloc[i], rev_col, 0):,.0f}' if rev_col else '-' for i in range(n_top)]),
        ('评分数', [f'{int(g(top10.iloc[i], "Ratings", 0)):,}' if 'Ratings' in top10.columns else '-' for i in range(n_top)]),
        ('星级', [f'{g(top10.iloc[i], "Rating", 0):.1f}★' if 'Rating' in top10.columns else '-' for i in range(n_top)]),
        ('在售天数', [f'{int(g(top10.iloc[i], "Available days", 0))}天' if 'Available days' in top10.columns else '-' for i in range(n_top)]),
        ('配送', [str(g(top10.iloc[i], 'Fulfillment', '-')) for i in range(n_top)]),
        ('卖家国家', [str(g(top10.iloc[i], 'BuyBox Location', '-')) for i in range(n_top)]),
        # 产品类型用 visual_product_type（视觉 LLM 自由描述每 ASIN）—— 比 BSR segment.name 精确，
        # 避免 BSR LLM 漏装兜底到「其他/通用款」时业务人员看不出产品类型的问题。
        # fallback 到 BSR segment 时也过滤「其他/通用款」（兜底桶展示无意义），改用产品标题前 30 字
        ('产品类型', [
            (lambda vpt, bsr_seg, title: (
                vpt
                or (bsr_seg if bsr_seg and bsr_seg != '其他/通用款' else (title[:30] if title else '-'))
            ))(
                str(g(top10.iloc[i], 'visual_product_type', '')).strip(),
                classify_with_packs(g(top10.iloc[i], 'Product Title', ''), packs, classify, asin=str(g(top10.iloc[i], 'ASIN', ''))),
                str(g(top10.iloc[i], 'Product Title', '')).strip(),
            )
            for i in range(n_top)
        ]),
    ]

    # 品类特异规格列：优先用 SpecPack 动态给的维度（LLM 按本次品类识别）；
    # LLM 不可用时回落到上游 specs_list 里已有的 LED 专属字段（如果有），否则只留基础列
    _spec_dims = spec_dimensions(packs)
    if _spec_dims:
        # 走 SpecPack 驱动：每个维度一列，Python 用 Pack 给的正则抓具体值
        # full_text 用 _collect_long_text 收集所有 bullet/overview/feature/description 列
        # 命中率过滤：辅助维度命中率 < 10% 隐藏（避免 LLM 给的过严正则导致整行 "—"）；核心维度即使 0 命中也保留
        _SPEC_MIN_RATIO = 0.10
        for _dim in _spec_dims:
            _dim_name = _dim['name']
            _importance = str(_dim.get('importance', '') or '')
            _vals = []
            for i in range(n_top):
                _title = str(g(top10.iloc[i], 'Product Title', ''))
                _full_text = _collect_long_text(top10.iloc[i])
                _spec_map = extract_specs_for_title(_title, _full_text, packs)
                _val = _spec_map.get(_dim_name, '')
                # 兜底 1：patterns 抓不到 → 用 LLM 的 representative_specs_by_asin（按 ASIN 匹配）
                # LLM 看 BSR 标题/bullet 用语义提取的样本数据，比正则字面匹配鲁棒
                if not _val and packs and getattr(packs, 'spec', None):
                    asin_top = str(g(top10.iloc[i], 'ASIN', '')).strip()
                    if asin_top:
                        for sample in (packs.spec.representative_specs_by_asin or []):
                            if str(getattr(sample, 'asin', '')).strip() == asin_top:
                                _v_llm = str((getattr(sample, 'specs', None) or {}).get(_dim_name, '') or '').strip()
                                if _v_llm:
                                    _val = _v_llm
                                break
                # 兜底 2：LLM 样本也没有 → 旧 LED legacy 正则兜底（键名近似匹配）
                if not _val and specs_list[i]:
                    for _legacy_k, _legacy_v in specs_list[i].items():
                        if _legacy_v and (_dim_name in _legacy_k or _legacy_k in _dim_name):
                            _val = _legacy_v
                            break
                _vals.append(_val)
            _hit_count = sum(1 for v in _vals if v)
            _ratio = _hit_count / max(n_top, 1)
            # 辅助维度低于阈值直接跳过，不进表格；核心维度保留（提示业务该品类此维度普遍未标注）
            if _importance != '核心' and _ratio < _SPEC_MIN_RATIO:
                continue
            attr_rows.append((_dim_name, _vals))
    elif specs_list:
        # Spec LLM 失败但 specs_list 里仍有 LED 风格字段（极端回退）——只显示几个基础字段
        for _fallback_key in ('光源', '功率W', '光通量lm', '电池容量mAh', '供电方式', '防水等级', '重量'):
            _vals = [specs_list[i].get(_fallback_key, '') for i in range(n_top)]
            if any(v for v in _vals):
                attr_rows.append((_fallback_key, _vals))

    # 表头行：#1-#10
    val(ws5, r5, 1, '属性', bold=True, bg=C_BLUE_MID, fg=C_WHITE)
    for j in range(n_top):
        val(ws5, r5, 2+j, f'#{j+1}', bold=True, bg=C_BLUE_MID, fg=C_WHITE)
    ws5.row_dimensions[r5].height = 22
    r5 += 1
    row_start = r5

    # 主图行（独立渲染，不进 attr_rows，因为图片要 add_image 而不是 cell.value）
    if _top10_imgs is not None:
        val(ws5, r5, 1, '主图', bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        for j, url in enumerate(_top10_imgs):
            ok = _insert_image_to_cell(ws5, r5, 2+j, url) if url else False
            cell = ws5.cell(row=r5, column=2+j, value='' if ok else '-')
            cell.fill = PatternFill('solid', fgColor=C_WHITE)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws5.row_dimensions[r5].height = 65
        r5 += 1

    for label, values in attr_rows:
        is_key = label in ('ASIN', '品牌', '价格', '月销量')
        val(ws5, r5, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        for j, v in enumerate(values):
            cell_bg = C_WHITE if not is_key else 'FFFFF8E7'
            val(ws5, r5, 2+j, str(v) if v != '' else '—', bg=cell_bg, fg='FF595959' if v == '' else 'FF000000')
        ws5.row_dimensions[r5].height = 20
        r5 += 1
    apply_border(ws5, row_start-1, r5-1, 1, 1+n_top)

    # 说明
    r5 += 1
    ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=11)
    c = ws5.cell(row=r5, column=1)
    c.value = '注：参数从产品标题 + Bullet Points（五点描述）正则提取，空值表示未标注（不代表实际不具备）'
    c.font = Font(name='Arial', size=10, color='FF7A4F01', italic=True)
    c.fill = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='left', vertical='center')
    r5 += 2

    # ▌ 二、正向卖点 + ▌ 三、差评（全市场 100 BSR 评论聚类，非细分子品类口径）
    section_title(ws5, r5, 1, '▌ 二、正向卖点提炼（全市场 4-5★评论）', span=4)
    section_title(ws5, r5, 5, '▌ 三、差评痛点分析（全市场 1-2★评论）', span=4, bg='FF7B2D00')
    ws5.row_dimensions[r5].height = 24
    r5 += 1

    hdr(ws5, r5, 1, '卖点维度', bg=C_BLUE_MID)
    hdr(ws5, r5, 2, '提及次数', bg=C_BLUE_MID)
    hdr(ws5, r5, 3, '占比', bg=C_BLUE_MID)
    hdr(ws5, r5, 4, '典型用语/消费者语言', bg=C_BLUE_MID)
    hdr(ws5, r5, 5, '差评痛点', bg='FFC00000', fg=C_WHITE)
    hdr(ws5, r5, 6, '提及次数', bg='FFC00000', fg=C_WHITE)
    hdr(ws5, r5, 7, '占比', bg='FFC00000', fg=C_WHITE)
    hdr(ws5, r5, 8, '典型反馈', bg='FFC00000', fg=C_WHITE)
    r5 += 1

    # 正向卖点 / 差评痛点两列：优先使用 VOC Pack 的 praise_clusters / pain_clusters
    # （LLM 按本次品类实际评论聚类），cluster 名和 raw_quotes 直接对齐
    # 全品类通用：命中统计只用 LLM 返回的 cluster.name + raw_quotes 做 str.contains，
    # 不依赖任何品类专属关键词表
    def _rows_from_voc_clusters(clusters, total_for_pct, text_series=None):
        """把 VOCPack 的 cluster 列表转成 [(name, count, pct, quotes_joined)] 行。
        - 若 LLM 给出 frequency_pct > 0（pain 几乎必有，praise 几乎没有）：按百分数换算条数
        - 若 LLM 未给 frequency_pct（praise 常态）：扫描 text_series 用 name/quote 做真实命中计数
          三级匹配：整 name 命中 → name 前 4 字 / 前 2 英文 token → raw_quotes[0] 前 8 字片段
        - 命中不足时显示真实命中数（含 0），不再用 '≥5' 占位
        """
        import re as _re_v
        rows_out = []
        # 把 text_series 拼成小写字符串列表，做子串计数
        texts_lower: list[str] = []
        if text_series is not None and len(text_series) > 0:
            try:
                texts_lower = [str(t).lower() for t in text_series.fillna('').astype(str).tolist()]
            except Exception:
                texts_lower = []

        def _count_hits(probe: str) -> int:
            probe = (probe or '').strip().lower()
            if not probe or not texts_lower:
                return 0
            return sum(1 for t in texts_lower if probe in t)

        for c in clusters or []:
            name = (c.name or '').strip()
            if not name:
                continue
            pct_val = getattr(c, 'frequency_pct', None)
            quotes = [q.strip() for q in (c.raw_quotes or [])[:3] if q and q.strip()]
            quotes_joined = ' | '.join(f'"{q[:120]}"' for q in quotes)

            # 优先：LLM 提供的 frequency_pct
            if pct_val is not None and pct_val > 0:
                pct_str = f'{pct_val:.1f}%'
                cnt_est = int(round((pct_val / 100.0) * max(total_for_pct, 1)))
                rows_out.append((name, cnt_est, pct_str, quotes_joined))
                continue

            # 兜底：扫描 text_series 做真实命中（全品类通用，不依赖关键词表）
            cnt = _count_hits(name)
            if cnt < 3:
                # 二级探针：name 的前 4 个中文字 / 前 2 个英文 token
                tokens = _re_v.findall(r'[a-z0-9]+', name.lower())
                if tokens:
                    probe = ' '.join(tokens[:2])
                    cnt = max(cnt, _count_hits(probe))
                else:
                    cnt = max(cnt, _count_hits(name[:4]))
            if cnt < 3 and quotes:
                # 三级探针：第一条 raw_quote 的前 8 字特征片段
                cnt = max(cnt, _count_hits(quotes[0][:8]))

            if total_for_pct > 0 and cnt > 0:
                pct_str = f'{cnt / total_for_pct * 100:.1f}%'
            else:
                pct_str = '—' if cnt == 0 else f'{cnt / max(total_for_pct, 1) * 100:.1f}%'
            rows_out.append((name, cnt, pct_str, quotes_joined))
        return rows_out

    voc_real = packs is not None and packs.is_voc_real()
    total_pos = len(high_rev)
    total_neg = len(low_rev)

    # 拼好评 / 差评文本 Series 供命中统计使用（Title + Content 小写）
    pos_text_series = None
    neg_text_series = None
    if total_pos > 0 and 'Title' in high_rev.columns and 'Content' in high_rev.columns:
        pos_text_series = (high_rev['Title'].fillna('') + ' ' + high_rev['Content'].fillna('')).str.lower()
    if total_neg > 0 and 'Title' in low_rev.columns and 'Content' in low_rev.columns:
        neg_text_series = (low_rev['Title'].fillna('') + ' ' + low_rev['Content'].fillna('')).str.lower()

    if voc_real:
        praise_rows = _rows_from_voc_clusters(packs.voc.praise_clusters, total_pos, pos_text_series)
        pain_rows = _rows_from_voc_clusters(packs.voc.pain_clusters, total_neg, neg_text_series)
        # LLM 合并阶段若因品类守门误判把 praise_clusters 全部清空，这里给一行黄色占位，
        # 与 P4 "本条 LLM 综合未覆盖"的体验对齐，避免无声空白
        if not praise_rows and total_pos > 20:
            praise_rows = [(
                '（LLM 合并时未保留正向卖点，建议刷新重跑）',
                0,
                '—',
                f'好评样本 {total_pos} 条充足，可能被 merge 阶段品类守门规则误删'
            )]
    else:
        # LLM 不可用：回退到旧 pos_counts / neg_counts 关键词计数（原文列留空不编）
        praise_rows = [(cat, cnt, f'{cnt/total_pos*100:.1f}%' if total_pos > 0 else '0%', '')
                       for cat, cnt in sorted(pos_counts.items(), key=lambda x: x[1], reverse=True) if cnt > 0]
        pain_rows = [(cat, cnt, f'{cnt/total_neg*100:.1f}%' if total_neg > 0 else '0%', '')
                     for cat, cnt in sorted(neg_counts.items(), key=lambda x: x[1], reverse=True) if cnt > 0]

    max_rows = max(len(praise_rows), len(pain_rows))
    for i in range(max_rows):
        bg_pos = C_GREEN_LIGHT if i % 2 == 0 else C_WHITE
        bg_neg = C_RED_LIGHT if i % 2 == 0 else 'FFFFF5F5'

        if i < len(praise_rows):
            name_p, cnt_p, pct_p, quotes_p = praise_rows[i]
            val(ws5, r5, 1, name_p, bold=True, bg=bg_pos)
            # cnt_p 现在始终是真实命中数（LLM frequency_pct 反推 或 Python 文本命中计数），不再用 "≥5" 占位
            _disp_cnt = cnt_p if cnt_p > 0 else '—'
            ws5.cell(row=r5, column=2, value=_disp_cnt).alignment = Alignment(horizontal='center', vertical='center')
            ws5.cell(row=r5, column=2).fill = PatternFill('solid', fgColor=bg_pos)
            ws5.cell(row=r5, column=3, value=pct_p or '—').alignment = Alignment(horizontal='center', vertical='center')
            ws5.cell(row=r5, column=3).fill = PatternFill('solid', fgColor=bg_pos)
            val(ws5, r5, 4, quotes_p or '—', fg='FF595959', bg=bg_pos, size=9)
        else:
            for ci in range(1, 5):
                ws5.cell(row=r5, column=ci).fill = PatternFill('solid', fgColor=C_WHITE)

        if i < len(pain_rows):
            name_n, cnt_n, pct_n, quotes_n = pain_rows[i]
            val(ws5, r5, 5, name_n, bold=True, bg=bg_neg, fg='FF7B0000')
            ws5.cell(row=r5, column=6, value=cnt_n if cnt_n > 0 else '—').alignment = Alignment(horizontal='center', vertical='center')
            ws5.cell(row=r5, column=6).fill = PatternFill('solid', fgColor=bg_neg)
            ws5.cell(row=r5, column=7, value=pct_n or '—').alignment = Alignment(horizontal='center', vertical='center')
            ws5.cell(row=r5, column=7).fill = PatternFill('solid', fgColor=bg_neg)
            val(ws5, r5, 8, quotes_n, fg='FF595959', bg=bg_neg, size=9)
        else:
            for ci in range(5, 9):
                ws5.cell(row=r5, column=ci).fill = PatternFill('solid', fgColor=C_WHITE)

        ws5.row_dimensions[r5].height = 28
        r5 += 1

    apply_border(ws5, 5, r5-1, 1, 8)
    r5 += 1

    # 各ASIN评论质量汇总
    section_title(ws5, r5, 1, '▌ 四、各ASIN评论质量汇总', span=8)
    r5 += 1
    asin_hdr = ['ASIN', '产品类型', '评论总数', '平均星级', '差评率', '5★占比', '主要差评类型']
    for i, h in enumerate(asin_hdr):
        hdr(ws5, r5, i+1, h, bg=C_BLUE_MID)
    ws5.merge_cells(start_row=r5, start_column=7, end_row=r5, end_column=8)
    r5 += 1

    # 产品类型用 visual_product_type（视觉 LLM 自由描述）作为展示主路径，
    # fallback 到 BSR segment.name 但跳过兜底桶「其他/通用款」（异质混合，业务展示无意义）
    def _resolve_display_type(vpt: str, pt: str) -> str:
        v = str(vpt).strip()
        if v:
            return v
        p = str(pt).strip()
        return p if p and p != '其他/通用款' else '-'
    asin_type_map = {
        str(asin): _resolve_display_type(vpt, pt)
        for asin, vpt, pt in zip(df['ASIN'], df.get('visual_product_type', df['product_type']), df['product_type'])
    }

    # 处理评论数据为空的情况
    if len(rev_df) > 0 and 'source_asin' in rev_df.columns:
        rev_asin_count = len(rev_df['source_asin'].unique())
        for source_asin in rev_df['source_asin'].unique():
            adf = rev_df[rev_df['source_asin'] == source_asin]
            if len(adf) == 0:
                continue
            # 平均星级：优先取 BSR 源表中该 ASIN 的前台展示星级（与 Sheet3、亚马逊详情页同源）
            avg_r = 0
            _bsr_row = df[df['ASIN'].astype(str).str.strip() == str(source_asin).strip()] if 'ASIN' in df.columns else None
            if _bsr_row is not None and len(_bsr_row) > 0 and 'Rating' in _bsr_row.columns:
                _v = _bsr_row['Rating'].iloc[0]
                if pd.notna(_v):
                    try:
                        avg_r = float(_v)
                    except (TypeError, ValueError):
                        avg_r = 0
            # Fallback：BSR 中该 ASIN 缺失或无 Rating 时才用评论样本均值
            if avg_r == 0 and 'Rating' in adf.columns and len(adf) > 0:
                _m = adf['Rating'].mean()
                if pd.notna(_m):
                    avg_r = float(_m)
            low_pct = (adf['Rating'] <= 2).sum() / len(adf) * 100 if 'Rating' in adf.columns else 0
            five_pct = (adf['Rating'] == 5).sum() / len(adf) * 100 if 'Rating' in adf.columns else 0

            # 主要差评类型：用 LLM 产出的 cluster keywords 在该 ASIN 评论文本中真实匹配
            # 取该 ASIN 的差评文本（仅 1-2 星，按业务确认）
            complaints: list[str] = []
            adf_low = adf[adf['Rating'] <= 2].copy() if 'Rating' in adf.columns else pd.DataFrame()
            if packs is not None and packs.is_voc_real() and packs.voc.pain_clusters and len(adf_low) > 0 \
                    and 'Title' in adf_low.columns and 'Content' in adf_low.columns:
                adf_low['neg_text'] = (adf_low['Title'].fillna('').astype(str) + ' '
                                       + adf_low['Content'].fillna('').astype(str)).str.lower()
                cluster_hits = []
                for _pc in packs.voc.pain_clusters:
                    _pc_name = (_pc.name or '').strip()
                    if not _pc_name:
                        continue
                    # 过滤噪声：长度 < 2 字符的关键词剔除
                    kws = [k.strip() for k in (_pc.keywords or []) if k and len(k.strip()) >= 2]
                    # 有效 keywords < 2 个时跳过该 cluster（避免单关键词误命中）
                    if len(kws) < 2:
                        continue
                    pattern = _safe_pattern(kws)
                    if not pattern:
                        continue
                    cnt = int(adf_low['neg_text'].str.contains(pattern, na=False, regex=True).sum())
                    if cnt > 0:
                        cluster_hits.append((_pc_name, cnt))
                cluster_hits.sort(key=lambda x: -x[1])
                complaints = [f'{n}({c})' for n, c in cluster_hits[:3]]
            # Fallback：LLM 不可用 / cluster 无 keywords / 该 ASIN 0 命中 → 走旧的 neg_keywords 关键词扫描
            if not complaints:
                if len(adf_low) > 0 and 'Title' in adf_low.columns and 'Content' in adf_low.columns:
                    if 'neg_text' not in adf_low.columns:
                        adf_low['neg_text'] = adf_low['Title'].fillna('') + ' ' + adf_low['Content'].fillna('')
                    for cat, kws in neg_keywords.items():
                        pattern = _safe_pattern(kws)
                        if not pattern:
                            continue
                        cnt = adf_low['neg_text'].str.lower().str.contains(pattern, na=False, regex=True).sum()
                        if cnt > 0:
                            complaints.append(f'{cat}({cnt})')
            complaint_str = '、'.join(complaints[:3]) if complaints else '无明显差评集中'
            ptype = asin_type_map.get(source_asin, '未知')
            bg = C_RED_LIGHT if avg_r < 3.5 else (C_YELLOW if avg_r < 4.0 else C_WHITE)
            vals5 = [source_asin, ptype, len(adf), f'{avg_r:.1f}★', f'{low_pct:.1f}%', f'{five_pct:.1f}%']
            for ci, v in enumerate(vals5):
                c = ws5.cell(row=r5, column=ci+1, value=v)
                c.fill = PatternFill('solid', fgColor=bg)
                c.font = Font(name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws5.merge_cells(start_row=r5, start_column=7, end_row=r5, end_column=8)
            val(ws5, r5, 7, complaint_str, bg=bg, fg='FF595959', size=9)
            ws5.row_dimensions[r5].height = 18
            r5 += 1
        apply_border(ws5, r5 - rev_asin_count - 1, r5-1, 1, 8)
    else:
        val(ws5, r5, 1, '暂无评论数据', bg=C_YELLOW)
        ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=8)
        r5 += 1
    r5 += 1

    # ▌ 五、重点ASIN参数深度分析（有评论数据的ASIN）
    review_asins = rev_df['source_asin'].unique().tolist() if len(rev_df) > 0 and 'source_asin' in rev_df.columns else []
    # 筛选同时在BSR数据中的ASIN
    review_asins_in_bsr = [a for a in review_asins if a in df['ASIN'].values]

    if review_asins_in_bsr:
        # 按月销量排序，最多取8个
        asin_sales = []
        for a in review_asins_in_bsr:
            row_match = df[df['ASIN'] == a]
            if len(row_match) > 0:
                ms = row_match.iloc[0].get('Monthly Sales', 0)
                ms = ms if pd.notna(ms) else 0
                asin_sales.append((a, ms))
        asin_sales.sort(key=lambda x: x[1], reverse=True)
        focus_asins = [a for a, _ in asin_sales[:8]]
        n_focus = len(focus_asins)

        section_title(ws5, r5, 1, f'▌ 五、重点ASIN参数深度分析（{n_focus}个有评论数据的ASIN）', span=max(n_focus+1, 8))
        ws5.row_dimensions[r5].height = 24
        r5 += 1

        # 提取每个ASIN的参数
        focus_specs = []
        focus_rows_data = []
        for a in focus_asins:
            row_match = df[df['ASIN'] == a].iloc[0]
            title_val = str(row_match.get('Product Title', ''))
            # 与 Top10 段保持一致：收集所有长文本列（bullet/overview/feature/description）
            full_text_val = _collect_long_text(row_match)
            sp = extract_all_specs(title_val, full_text_val)
            focus_specs.append(sp)
            focus_rows_data.append(row_match)

        # 预构建 VOC 聚类的"关键词袋"：每个 cluster 用 name + raw_quotes 关键短语作匹配词
        # 这样每个 ASIN 的 TOP3 差评就是 "自己的 1-2★ 评论文本命中哪些 cluster 最多"，按本 ASIN 真实命中数排序
        # 避免直接用 affected_asins 导致所有重点 ASIN 显示相同的前 3 个 pain cluster
        _praise_clusters_terms: list[tuple[str, list[str]]] = []  # [(name, [关键短语...]), ...]
        _pain_clusters_terms: list[tuple[str, list[str]]] = []
        if packs is not None and packs.is_voc_real():
            import re as _re_cluster
            def _extract_terms(cluster):
                """把 cluster.name + raw_quotes 拆成短语（2-6 字的核心词），用于子串匹配"""
                terms: set[str] = set()
                name = (cluster.name or '').strip()
                if name:
                    terms.add(name.lower())
                    # name 里按 "/、，" 分出几个关键词
                    for seg in _re_cluster.split(r'[/、，,\s]+', name):
                        s = seg.strip().lower()
                        if 2 <= len(s) <= 8:
                            terms.add(s)
                for q in (cluster.raw_quotes or []):
                    q = (q or '').strip()
                    if not q:
                        continue
                    # 短 quote（≤10 字）整条作为关键词；长 quote 只取其中的 2-6 字的名词词
                    ql = q.lower()
                    if len(ql) <= 12:
                        terms.add(ql)
                return [t for t in terms if t]
            for _pc in packs.voc.praise_clusters or []:
                _praise_clusters_terms.append(((_pc.name or '').strip(), _extract_terms(_pc)))
            for _pc in packs.voc.pain_clusters or []:
                _pain_clusters_terms.append(((_pc.name or '').strip(), _extract_terms(_pc)))

        def _top3_by_cluster_hits(asin_text_series, clusters_terms):
            """在该 ASIN 的评论文本上扫描每个 cluster 的关键短语，按命中总次数排 TOP3"""
            if asin_text_series is None or len(asin_text_series) == 0 or not clusters_terms:
                return []
            text_all = ' '.join(asin_text_series.fillna('').astype(str).str.lower().tolist())
            if not text_all.strip():
                return []
            hits = []
            for name, terms in clusters_terms:
                if not name or not terms:
                    continue
                cnt = sum(text_all.count(t) for t in terms if t)
                if cnt > 0:
                    hits.append((name, cnt))
            hits.sort(key=lambda x: x[1], reverse=True)
            return hits[:3]

        # 每个ASIN的评论统计
        focus_review_stats = []
        for a in focus_asins:
            adf = rev_df[rev_df['source_asin'] == a]
            total_rev_count = len(adf)
            # 平均星级：优先取 BSR 源表中该 ASIN 的前台展示星级（与 R47 上半部、Sheet3、亚马逊详情页同源）
            avg_r = 0
            _bsr_row = df[df['ASIN'].astype(str).str.strip() == str(a).strip()] if 'ASIN' in df.columns else None
            if _bsr_row is not None and len(_bsr_row) > 0 and 'Rating' in _bsr_row.columns:
                _v = _bsr_row['Rating'].iloc[0]
                if pd.notna(_v):
                    try:
                        avg_r = float(_v)
                    except (TypeError, ValueError):
                        avg_r = 0
            # Fallback：BSR 中该 ASIN 缺失或无 Rating 时才用评论样本均值
            if avg_r == 0 and 'Rating' in adf.columns and len(adf) > 0:
                _m = adf['Rating'].mean()
                if pd.notna(_m):
                    avg_r = float(_m)
            low_pct = (adf['Rating'] <= 2).sum() / len(adf) * 100 if 'Rating' in adf.columns and len(adf) > 0 else 0

            # Top3 正向：扫本 ASIN 的 4-5★ 评论文本对 praise_clusters 的命中
            high_adf = adf[adf['Rating'] >= 4] if 'Rating' in adf.columns else pd.DataFrame()
            high_text_series = pd.Series(dtype=str)
            if len(high_adf) > 0 and 'Title' in high_adf.columns and 'Content' in high_adf.columns:
                high_text_series = (high_adf['Title'].fillna('') + ' ' + high_adf['Content'].fillna(''))
            pos_top3 = _top3_by_cluster_hits(high_text_series, _praise_clusters_terms)

            # Top3 负向：同上，扫本 ASIN 的 1-2★ 评论对 pain_clusters
            low_adf = adf[adf['Rating'] <= 2] if 'Rating' in adf.columns else pd.DataFrame()
            low_text_series = pd.Series(dtype=str)
            if len(low_adf) > 0 and 'Title' in low_adf.columns and 'Content' in low_adf.columns:
                low_text_series = (low_adf['Title'].fillna('') + ' ' + low_adf['Content'].fillna(''))
            neg_top3_asin = _top3_by_cluster_hits(low_text_series, _pain_clusters_terms)

            # 回退：VOC cluster 全部没命中 → 走 Python 关键词扫描（旧路径）
            if not pos_top3 and len(high_adf) > 0 and 'Title' in high_adf.columns and 'Content' in high_adf.columns:
                high_adf_text = high_adf['Title'].fillna('') + ' ' + high_adf['Content'].fillna('')
                _tmp = []
                for cat, kws in pos_keywords.items():
                    pattern = _safe_pattern(kws)
                    if not pattern:
                        continue
                    cnt = high_adf_text.str.lower().str.contains(pattern, na=False, regex=True).sum()
                    if cnt > 0:
                        _tmp.append((cat, cnt))
                _tmp.sort(key=lambda x: x[1], reverse=True)
                pos_top3 = _tmp[:3]
            if not neg_top3_asin and len(low_adf) > 0 and 'Title' in low_adf.columns and 'Content' in low_adf.columns:
                low_adf_text = low_adf['Title'].fillna('') + ' ' + low_adf['Content'].fillna('')
                _tmp = []
                for cat, kws in neg_keywords.items():
                    pattern = _safe_pattern(kws)
                    if not pattern:
                        continue
                    cnt = low_adf_text.str.lower().str.contains(pattern, na=False, regex=True).sum()
                    if cnt > 0:
                        _tmp.append((cat, cnt))
                _tmp.sort(key=lambda x: x[1], reverse=True)
                neg_top3_asin = _tmp[:3]

            focus_review_stats.append({
                'total': total_rev_count,
                'avg_rating': avg_r,
                'neg_rate': low_pct,
                'pos_top3': pos_top3,
                'neg_top3': neg_top3_asin,
            })

        # 构建属性行
        focus_attr_rows = [
            ('ASIN', [a for a in focus_asins]),
            ('品牌', [str(focus_rows_data[i].get('Brand', '-')) for i in range(n_focus)]),
            ('价格', [f'${focus_rows_data[i].get(price_col, 0):.2f}' if price_col and pd.notna(focus_rows_data[i].get(price_col)) else '-' for i in range(n_focus)]),
            ('月销量', [f'{int(focus_rows_data[i].get("Monthly Sales", 0)):,}' if pd.notna(focus_rows_data[i].get('Monthly Sales')) else '-' for i in range(n_focus)]),
            ('评分/评论数', [f'{focus_review_stats[i]["avg_rating"]:.1f}★({focus_review_stats[i]["total"]}条)' for i in range(n_focus)]),
        ]
        # Sheet 5 重点 ASIN 参数深度分析 —— 全品类通用：
        # 优先用 SpecPack（LLM 按本品类识别的维度）+ extract_specs_by_dimensions 按 ASIN 抓值
        # SpecPack 不可用时降级到 extract_all_specs 的 LED 字段（仅 LED 场景有意义）
        spec_attr_rows: list[tuple[str, list[str]]] = []
        if packs is not None and packs.is_spec_real() and packs.spec.spec_dimensions:
            from llm.analyzers.spec_analyzer import extract_specs_by_dimensions as _esbd
            _dims = list(packs.spec.spec_dimensions)
            # 对每个重点 ASIN 用 LLM 维度抽取一次值（走 title + bullets）
            _focus_dim_specs: list[dict] = []
            for i in range(n_focus):
                _row = focus_rows_data[i]
                _title = str(_row.get('Product Title', ''))
                _bullets = _collect_long_text(_row)
                _focus_dim_specs.append(_esbd(_title, _bullets, _dims))
            # LLM 样本兜底（patterns 抓不到 → 从 representative_specs_by_asin 按 ASIN 匹配）
            _llm_samples_by_asin = {
                str(getattr(s, 'asin', '')).strip(): (getattr(s, 'specs', None) or {})
                for s in (packs.spec.representative_specs_by_asin or [])
                if str(getattr(s, 'asin', '')).strip()
            }
            # 按 SpecPack 维度顺序显示
            for _d in _dims:
                _vals_focus = []
                for i in range(n_focus):
                    _v = _focus_dim_specs[i].get(_d.name, '')
                    if not _v:
                        # patterns 没抓到 → LLM 样本兜底
                        _asin_f = str(focus_asins[i]).strip()
                        _v_llm = str(_llm_samples_by_asin.get(_asin_f, {}).get(_d.name, '') or '').strip()
                        if _v_llm:
                            _v = _v_llm
                    _vals_focus.append(_v)
                spec_attr_rows.append((_d.name, _vals_focus))
        else:
            # LLM 降级兜底：LED 专属字段（跨品类时多数会显示空，但不写死为报告默认）
            spec_attr_rows = [
                ('光通量', [focus_specs[i].get('光通量lm', '') for i in range(n_focus)]),
                ('电池容量', [focus_specs[i].get('电池容量mAh', '') for i in range(n_focus)]),
                ('充电方式', [focus_specs[i].get('充电方式', '') for i in range(n_focus)]),
                ('防水等级', [focus_specs[i].get('防水等级', '') for i in range(n_focus)]),
                ('光照模式', [focus_specs[i].get('光照模式数', '') for i in range(n_focus)]),
                ('磁力/固定', [focus_specs[i].get('磁力/固定方式', '') for i in range(n_focus)]),
                ('材质', [focus_specs[i].get('材质', '') for i in range(n_focus)]),
                ('续航', [focus_specs[i].get('续航', '') for i in range(n_focus)]),
                ('旋转角度', [focus_specs[i].get('旋转角度', '') for i in range(n_focus)]),
                ('重量', [focus_specs[i].get('重量', '') for i in range(n_focus)]),
            ]
        review_attr_rows = [
            ('好评TOP3', [', '.join([f'{cat}({cnt})' for cat, cnt in focus_review_stats[i]['pos_top3']]) or '—' for i in range(n_focus)]),
            ('差评TOP3', [', '.join([f'{cat}({cnt})' for cat, cnt in focus_review_stats[i]['neg_top3']]) or '—' for i in range(n_focus)]),
            ('差评率', [f'{focus_review_stats[i]["neg_rate"]:.1f}%' for i in range(n_focus)]),
        ]

        # 写表头
        val(ws5, r5, 1, '属性', bold=True, bg=C_BLUE_MID, fg=C_WHITE)
        for j in range(n_focus):
            val(ws5, r5, 2+j, f'ASIN-{j+1}', bold=True, bg=C_BLUE_MID, fg=C_WHITE)
        ws5.row_dimensions[r5].height = 22
        r5 += 1
        focus_start = r5

        # 基础信息行
        for label, values in focus_attr_rows:
            val(ws5, r5, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
            for j, v in enumerate(values):
                val(ws5, r5, 2+j, str(v) if v else '—', bg='FFFFF8E7', fg='FF000000')
            ws5.row_dimensions[r5].height = 20
            r5 += 1

        # 参数分隔行
        val(ws5, r5, 1, '—— 产品参数（Bullet Points提取）——', bold=True, bg='FFD9E2F3', fg='FF1F3864')
        for j in range(n_focus):
            ws5.cell(row=r5, column=2+j).fill = PatternFill('solid', fgColor='FFD9E2F3')
        ws5.row_dimensions[r5].height = 20
        r5 += 1

        # 参数行：即使某维度多数 ASIN 没标注，该属性维度仍保留（SpecPack 识别出的维度都体现本品类关注点）
        # 单个 ASIN 没匹配到就显示 "—"，代表"未在标题/Bullet Points 中标注"而非"不具备"
        for label, values in spec_attr_rows:
            val(ws5, r5, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
            for j, v in enumerate(values):
                val(ws5, r5, 2+j, str(v) if v else '—', bg=C_WHITE, fg='FF595959' if not v else 'FF000000')
            ws5.row_dimensions[r5].height = 20
            r5 += 1

        # 评论分隔行
        val(ws5, r5, 1, '—— 评论洞察 ——', bold=True, bg='FFFCE4D6', fg='FF7B2D00')
        for j in range(n_focus):
            ws5.cell(row=r5, column=2+j).fill = PatternFill('solid', fgColor='FFFCE4D6')
        ws5.row_dimensions[r5].height = 20
        r5 += 1

        # 评论行
        for label, values in review_attr_rows:
            is_neg = '差评' in label
            row_bg = C_RED_LIGHT if is_neg else C_GREEN_LIGHT
            val(ws5, r5, 1, label, bold=True, bg=row_bg, fg='FF7B0000' if is_neg else 'FF1F5C1F')
            for j, v in enumerate(values):
                val(ws5, r5, 2+j, str(v), bg=row_bg, fg='FF7B0000' if is_neg else 'FF595959', size=9)
            ws5.row_dimensions[r5].height = 22
            r5 += 1

        apply_border(ws5, focus_start-1, r5-1, 1, 1+n_focus)
    else:
        section_title(ws5, r5, 1, '▌ 五、重点ASIN参数深度分析', span=8)
        ws5.row_dimensions[r5].height = 24
        r5 += 1
        val(ws5, r5, 1, '暂无评论数据（上传评论文件后可显示重点ASIN的参数+评论深度分析）', bg=C_YELLOW)
        ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=8)
        r5 += 1
    r5 += 1

    # ===== 竞品分析 - 改进方向 =====
    # 核心改进方向总结 —— LLM 驱动的完整改进计划（根因诊断 + 落地建议 + 目标指标）
    # 口径：按"首推子品类的 ASIN 子集"过滤痛点，避免与第二/三节的全市场表混淆
    _s5_six_title = (f'▌ 六、核心改进方向总结（聚焦首推子品类：{_first_type}）'
                     if _first_type else '▌ 六、核心改进方向总结')
    section_title(ws5, r5, 1, _s5_six_title, span=8)
    ws5.row_dimensions[r5].height = 24
    r5 += 1

    # 表头：优先级 | 痛点（差评数） | 根因诊断 | 改进建议 + 目标指标
    hdr(ws5, r5, 1, '优先级', bg=C_BLUE_MID)
    hdr(ws5, r5, 2, '痛点 · 差评数', bg=C_BLUE_MID)
    hdr(ws5, r5, 3, '根因诊断', bg=C_BLUE_MID)
    ws5.merge_cells(start_row=r5, start_column=4, end_row=r5, end_column=8)
    hdr(ws5, r5, 4, '改进建议 + 目标指标', bg=C_BLUE_MID)
    ws5.row_dimensions[r5].height = 22
    r5 += 1

    # 改进计划 **只从 LLM 拿**（Synthesizer.sheet5_improvement_plan）——
    # root_cause / action_items / target_metric 全部由 LLM 基于当前上传评论生成，
    # Python 不再从 sheet6_priority_matrix 反推、不再拼兜底文案。
    # LLM 未覆盖的 pain 行显示明显的"本条 LLM 综合未覆盖"黄色提示，让用户看得出是 LLM 缺漏。
    improvement_priority: list[tuple[str, str, str, str, bool]] = []  # (priority, pain_title, root_cause, solution, is_missing)

    # 痛点索引：按"首推品类 ASIN 子集"的 1-2★ 评论文本命中数排 Top4，供 P1-P4 标题 + 差评数引用
    # 修复 sheet5 第六节"核心改进方向"原本用全局 pain_clusters[:4] 的跨品类污染（用户红字 J72 同根因）
    _pain_list: list[tuple[str, int]] = []
    _total_low = len(low_rev) if 'low_rev' in dir() and hasattr(low_rev, '__len__') else 0
    if (_first_type and packs is not None and packs.is_voc_real() and packs.voc.pain_clusters
            and 'product_type' in df.columns):
        _type_asins_s5 = set(df.loc[df['product_type'] == _first_type, 'ASIN']
                                .astype(str).str.strip())
        _pain_list = list(_filter_voc_pain_by_asins(
            packs.voc.pain_clusters, _type_asins_s5, rev_df, top_n=4))
    # Fallback 1：子集匹配为空 → 全局 Top4，差评条数按 frequency_pct × total_low 估算
    if not _pain_list and packs is not None and packs.is_voc_real() and packs.voc.pain_clusters:
        for _c in packs.voc.pain_clusters[:4]:
            _pct = float(_c.frequency_pct or 0)
            if _total_low > 0 and _pct > 0:
                _cnt = int(round(_pct * _total_low / 100.0))
            elif neg_counts:
                _cnt = int(neg_counts.get(_c.name, 0))
            else:
                _cnt = 0
            _pain_list.append((_c.name, _cnt))
    # Fallback 2：LLM 不可用 → 老 neg_keywords 关键词扫描
    if not _pain_list and neg_counts:
        _pain_list = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True)[:4]

    _labels = ['P1 首要改进', 'P2 重要改进', 'P3 品质改进', 'P4 体验改进']

    # LLM 主路径：读 sheet5_improvement_plan，按 pain_name 匹配到 _pain_list
    # 规范化匹配（去掉空白/标点/全半角差异），防止 P1/P2 因细微字符差异显示"—"
    import re as _re5
    def _norm5(s: str) -> str:
        return _re5.sub(r'[\s/（）()·•、，。·・．\.:：;；\-\–\—_]+', '', str(s or '')).lower()

    # 优先取 5.1.5 阶段提前并发跑出来的 Sheet5 future 结果（已与 prepare_packs 阶段重叠完成）；
    # future 不可用 / 超时 / 拿到 None 时回退到现场 LLM 调用，保证语义不变。
    _new_path_plans: list[dict] = []
    _s5_result = None
    if _sheet5_future is not None:
        try:
            _s5_result = _sheet5_future.result(timeout=180)
        except Exception as _e_s5_await:
            print(f"[Sheet5 渲染层] 后台 future 取结果失败：{_e_s5_await}", file=__import__('sys').stderr, flush=True)
            _s5_result = None
    if _s5_result is None and _pain_list and packs is not None and packs.is_voc_real() and packs.voc.pain_clusters:
        # 兜底：5.1.5 阶段 _pain_list 早期口径与此处可能略有差异（如 _first_type 在 5.1.5 后被改写场景），
        # future 没产物时按渲染层完整 _pain_list 现场再跑一次。
        try:
            from llm.analyzers.sheet5_improvement_analyzer import Sheet5ImprovementAnalyzer
            from core.packs_runtime import _try_make_client
            _s5_client = _try_make_client(api_key=getattr(packs, '_api_key', None))
            if _s5_client is not None:
                _pain_by_name = {pc.name: pc for pc in packs.voc.pain_clusters}
                _s5_input_pains = []
                for _pn, _cnt in _pain_list[:4]:
                    _pc_obj = _pain_by_name.get(_pn)
                    _s5_input_pains.append({
                        "name": _pn,
                        "frequency_pct": float(_pc_obj.frequency_pct or 0) if _pc_obj else 0.0,
                        "raw_quotes": list(_pc_obj.raw_quotes or [])[:5] if _pc_obj else [],
                    })
                _stats = getattr(packs, 'synthesis_stats', None) or {}
                _s5_input = {
                    "category_hint": (getattr(packs, 'display_name', '') or '').strip() or _first_type or '未知品类',
                    "pain_clusters": _s5_input_pains,
                    "stats": {
                        "competitor_spec_p75": _stats.get("competitor_spec_p75", {}),
                        "competitor_spec_medians": _stats.get("competitor_spec_medians", {}),
                    },
                }
                _s5_result = Sheet5ImprovementAnalyzer(_s5_client).run(_s5_input)
        except Exception as _e_s5:
            import sys as _sys
            import traceback as _tb_s5
            print(f"[Sheet5 渲染层] 兜底 LLM 调用失败：{_e_s5}", file=_sys.stderr, flush=True)
            _tb_s5.print_exc(file=_sys.stderr)
            try:
                log.exception("[Sheet5 渲染层] 兜底 LLM 调用失败")
            except Exception:
                pass

    if _s5_result and not _s5_result.is_fallback and _s5_result.plans:
        _new_path_plans = [
            {
                "priority": p.priority,
                "pain_name": p.pain_name,
                "root_cause": p.root_cause,
                "action_items": list(p.action_items),
                "target_metric": p.target_metric,
                "supporting_fields": list(p.supporting_fields),
            }
            for p in _s5_result.plans
        ]

    _llm_plans_by_pain: dict[str, dict] = {}
    _llm_plans_normalized: dict[str, dict] = {}
    # 优先用新路径产出的 plans（pain 列表与渲染层完全一致），失败/空才回落到旧 helper（synthesis 兜底）
    _plans_to_use = _new_path_plans
    if not _plans_to_use:
        try:
            from core.packs_runtime import sheet5_improvement_plan as _s5ip
            _plans_to_use = list(_s5ip(packs))
        except Exception:
            _plans_to_use = []
    for _plan in _plans_to_use:
        pn = (_plan.get("pain_name") or "").strip()
        if pn:
            _llm_plans_by_pain[pn] = _plan
            _llm_plans_normalized[_norm5(pn)] = _plan

    for i in range(min(4, len(_pain_list))):
        pain_name, pain_cnt = _pain_list[i]
        pain_title = f'{pain_name}\n（差评 {pain_cnt} 条）' if pain_cnt else f'{pain_name}\n（差评 — 条）'

        # 精确 → 规范化 → 子串 三级匹配
        _llm_plan = _llm_plans_by_pain.get(pain_name)
        if _llm_plan is None:
            _n = _norm5(pain_name)
            _llm_plan = _llm_plans_normalized.get(_n)
            if _llm_plan is None:
                for _k_norm, _v in _llm_plans_normalized.items():
                    if _n and _k_norm and (_n in _k_norm or _k_norm in _n):
                        _llm_plan = _v
                        break

        if _llm_plan:
            root_cause = (_llm_plan.get("root_cause") or "").strip()
            items = _llm_plan.get("action_items") or []
            target = (_llm_plan.get("target_metric") or "").strip()
            if not root_cause and not items and not target:
                # LLM 返回空壳 plan，当作缺失（用户决策：明确占位、不用 Python 假装 LLM 输出）
                improvement_priority.append((_labels[i], pain_title,
                    '（本条 LLM 综合未覆盖——建议刷新重跑）', '（本条 LLM 综合未覆盖——建议刷新重跑）', True))
                continue
            solution_lines = [f'• {x}' for x in items[:5]]
            if target:
                solution_lines.append(f'🎯 目标：{target}')
            solution = '\n'.join(solution_lines) if solution_lines else '—'
            improvement_priority.append((_labels[i], pain_title, root_cause or '—', solution, False))
        else:
            # LLM 未为该 pain 生成计划——明确占位让用户知道 LLM 缺失（黄色背景由下游标记）
            # 用户决策：不写 Python 兜底假装 LLM 已输出，而是明示"未覆盖"，配合堆栈打印（synthesizer.py
            # 与 app.py:5062 已加 traceback）让我们能从根因修 LLM 调用本身
            miss_msg = '（本条 LLM 综合未覆盖——可能因跨品类 pain 被 LLM 跳过，建议刷新重跑）'
            improvement_priority.append((_labels[i], pain_title, miss_msg, miss_msg, True))

    if improvement_priority:
        _section_start = r5
        for i, (priority, pain_title, root_cause, solution, is_missing) in enumerate(improvement_priority):
            # LLM 未覆盖的行统一用黄色，让缺漏显眼
            if is_missing:
                bg = C_YELLOW
                fg_rc = 'FF7B0000'
            else:
                bg = C_RED_LIGHT if i == 0 else (C_YELLOW if i == 1 else (C_BLUE_LIGHT if i == 2 else C_WHITE))
                fg_rc = 'FF595959'
            val(ws5, r5, 1, priority, bold=True, bg=bg, fg='FF7B0000' if i == 0 else 'FF1F3864')
            val(ws5, r5, 2, pain_title, bold=True, bg=bg, wrap=True)
            val(ws5, r5, 3, root_cause, bg=bg, fg=fg_rc, wrap=True, size=9)
            ws5.merge_cells(start_row=r5, start_column=4, end_row=r5, end_column=8)
            val(ws5, r5, 4, solution, bg=bg, fg=fg_rc, wrap=True)
            ws5.row_dimensions[r5].height = 75
            r5 += 1
        apply_border(ws5, _section_start-1, r5-1, 1, 8)
    else:
        val(ws5, r5, 1, '暂无评论数据或 LLM 综合降级；上传评论文件后可生成改进优先级', bg=C_YELLOW)
        ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=8)
        r5 += 1
    r5 += 1

    # ===== Sheet 6: 产品上新方向 =====
    ws6 = wb.create_sheet('产品上新方向')
    ws6.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGHI', [5, 26, 30, 28, 16, 14, 14, 18, 20]):
        ws6.column_dimensions[col_letter].width = width

    ws6.merge_cells('A1:I1')
    c = ws6['A1']
    c.value = f"{display_name_for_title(packs, '选品')} — 产品上新方向建议（基于数据缺口推导）"
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws6.row_dimensions[1].height = 35

    r6 = 3
    section_title(ws6, r6, 1, '▌ 数据支撑逻辑：差评痛点 × 市场缺口 × 价格空白 → 新品机会', span=9)
    ws6.row_dimensions[r6].height = 24
    r6 += 1

    new_prod_hdr = ['优先级', '产品方向', '核心改进点', '数据支撑依据', '目标售价', '预计月销',
                    '竞争难度', '启动方式', '预期亮点']
    for i, h in enumerate(new_prod_hdr):
        hdr(ws6, r6, i+1, h, bg=C_BLUE_MID)
    ws6.row_dimensions[r6].height = 22
    r6 += 1

    # top_pain / top_pain_cnt 已在上游从 VOC Pack 或 neg_counts 计算（变量名 top_pain_for_use）
    top_pain = top_pain_for_use or '—'
    top_pain_cnt = top_pain_cnt_for_use or 0

    # 动态生成产品上新方向数据失败时的占位行 —— 统一用 _top_segment_name（和 Sheet 4/10 对齐）
    default_directions = [
        ('P1\n首推', f'{_top_segment_name}（基于实际数据推荐）',
         f'• 针对TOP痛点优化：{top_pain}({top_pain_cnt}条差评)\n• 根据实际竞品分析生成改进建议',
         f'基于{len(df)}个BSR样本和{len(rev_df)}条评论分析',
         '$29.99-$39.99', '300-500件/月', '中等', 'FBA直发，首批500pcs',
         '根据实际数据推荐'),
    ] if len(df) > 0 else []

    # 合并动态数据和默认数据
    display_directions = new_directions if new_directions else default_directions

    priority_colors = {'P1\n首推': C_YELLOW, 'P2\n推荐': C_GREEN_LIGHT, 'P3\n参考': C_BLUE_LIGHT, 'P4\n备选': C_GREY_LIGHT}
    for row_dir in display_directions:
        prio = row_dir[0]
        bg = priority_colors.get(prio, C_WHITE)
        for ci, v in enumerate(row_dir):
            c = ws6.cell(row=r6, column=ci+1, value=v)
            c.fill = PatternFill('solid', fgColor=bg)
            c.font = Font(name='Arial', size=9, bold=(ci == 0 or ci == 1))
            c.alignment = Alignment(horizontal='center' if ci in [0, 4, 5, 6] else 'left',
                                    vertical='center', wrap_text=True)
        ws6.row_dimensions[r6].height = 65
        r6 += 1

    apply_border(ws6, 5, r6-1, 1, 9)

    # 评分逻辑注解
    ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=9)
    c = ws6.cell(row=r6, column=1)
    c.value = '注：优先级由综合评分自动排序 = 需求(月销)×0.30 + 单品收益×0.25 + 新品占比×0.15 + 质量评分×0.15 + 竞争度×0.15（数据驱动，非固定）'
    c.font = Font(name='Arial', size=9, color='FF7A4F01', italic=True)
    c.fill = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws6.row_dimensions[r6].height = 16
    r6 += 2

    # 决策矩阵
    section_title(ws6, r6, 1, '▌ 上新决策矩阵 — 优先级说明', span=9)
    ws6.row_dimensions[r6].height = 24
    r6 += 1

    # 品类特征关键字：
    # 优先取 packs.market.product_segments[*].description（LLM 按本次品类生成的细分特征描述）；
    # LLM 不可用时 type_traits 为空，_type_strategy 里对应字段留空 —— 不再硬编码 LED 专属"磁力固定/三脚架/T8灯管"
    type_traits: dict[str, str] = {}
    if packs is not None and packs.is_market_real() and packs.market.product_segments:
        for _seg in packs.market.product_segments:
            if _seg.description:
                type_traits[_seg.name] = _seg.description

    neg_sorted_for_desc = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True)
    top_pains_cat_for_desc = [p[0] for p in neg_sorted_for_desc[:2] if p[1] > 0]
    pain_str_for_desc = '+'.join([p.replace('/', '') for p in top_pains_cat_for_desc]) if top_pains_cat_for_desc else ''
    top_pain_cnt_for_desc = int(neg_sorted_for_desc[0][1]) if neg_sorted_for_desc and neg_sorted_for_desc[0][1] > 0 else 0

    # 基于实际排名动态生成决策建议（结合数据 + 品类特征 + 自然叙述）
    def _type_strategy(ptype, rank_idx, all_scores):
        sc = all_scores.get(ptype, {})
        sku_n = sc.get('sku_count', 0)
        new_r = sc.get('new_ratio', 0)
        rating_v = sc.get('avg_rating', 0)
        avg_rev_v = sc.get('avg_rev', 0)
        total_s = int(sc.get('total_sales', 0))

        all_skus = [s.get('sku_count', 0) for s in all_scores.values()]
        max_sku = max(all_skus) if all_skus else 1
        all_revs = [s.get('avg_rev', 0) for s in all_scores.values()]
        max_rev = max(all_revs) if all_revs else 0
        all_sales = [s.get('total_sales', 0) for s in all_scores.values()]
        max_sales = max(all_sales) if all_sales else 0

        sku_ratio = sku_n / max_sku if max_sku else 0
        rev_ratio = avg_rev_v / max_rev if max_rev else 0
        sales_ratio = total_s / max_sales if max_sales else 0

        trait = type_traits.get(ptype, '核心差异化点')

        if rank_idx == 0:
            # P1 首推
            demand_phrase = '销量最大品类' if sales_ratio >= 0.9 else ('销量靠前' if sales_ratio >= 0.3 else '市场需求明确')
            stage_phrase = '市场需求成熟' if new_r < 0.2 else ('处于成长红利期' if new_r > 0.3 else '需求稳定')
            price_sense = '价格敏感度中等' if rating_v >= 4.3 else '用户对品质敏感'
            pain_phrase = f'差评痛点明确（{pain_str_for_desc}，{top_pain_cnt_for_desc}条）' if top_pain_cnt_for_desc else '差评数据有限'
            return (f'{ptype}是{demand_phrase}，{stage_phrase}，{pain_phrase}，'
                    f'围绕{trait}改良空间大，{price_sense}，适合首批快速验证。目标：6 个月内进入类目 TOP50。')

        elif rank_idx == 1:
            # P2 推荐
            if rev_ratio >= 0.7:
                rev_phrase = '客单价高、单品收益好'
            elif sales_ratio >= 0.5:
                rev_phrase = '销量稳定、走量型赛道'
            else:
                rev_phrase = '综合表现第二梯队'
            quality_req = '对产品品质要求更严格' if rating_v >= 4.5 else '对质量容忍度中等'
            return (f'{ptype}{rev_phrase}，但{quality_req}（{trait}），'
                    f'需要工厂定制/差异化升级支持，适合首推品类验证后扩展。')

        elif rank_idx <= 3:
            # P3 机会
            if sku_ratio < 0.25:
                comp_phrase = f'竞争极少（仅 {sku_n} 个 SKU），是短期内可快速占位的小类目'
            elif new_r > 0.3:
                comp_phrase = '处于成长期，新品红利仍在'
            else:
                comp_phrase = '需求稳定，可作为 SKU 丰富产品线'
            return f'{ptype}{comp_phrase}，围绕{trait}做差异化，降低产品线集中风险。'

        else:
            # P4 备选
            if sales_ratio < 0.1:
                vol_phrase = f'市场体量偏小（月销 {total_s:,} 件）'
            else:
                vol_phrase = '综合评分偏弱'
            return (f'{ptype}{vol_phrase}，日常 SKU 变动少，'
                    f'主要靠包装和节日/场景运营推动转化，适合节前短期投入，控投入控风险。')

    # 排序来源统一：优先走 LLM sheet6_priority_matrix 的 P1-P4 顺序（与 Sheet 6 上半表 + Sheet 10 矩阵对齐）
    # LLM 不可用时回落到 Python ranked_product_types 综合评分
    _ordered_types: list[str] = []
    if packs is not None and packs.is_synthesis_real() and packs.synthesis.sheet6_priority_matrix:
        _prio_rank = {'P1': 0, 'P2': 1, 'P3': 2, 'P4': 3}
        _llm_items = sorted(
            [it for it in packs.synthesis.sheet6_priority_matrix if it.segment],
            key=lambda it: (_prio_rank.get(it.priority, 9), 0)
        )
        _seen = set()
        for it in _llm_items:
            if it.segment not in _seen and it.segment in ranked_type_scores:
                _ordered_types.append(it.segment)
                _seen.add(it.segment)
        # LLM 没覆盖到的剩余细分按 Python 综合评分补齐到末尾
        for pt in ranked_product_types:
            if pt not in _seen:
                _ordered_types.append(pt)
                _seen.add(pt)
    if not _ordered_types:
        _ordered_types = list(ranked_product_types)

    decision_notes = []
    labels = ['P1 首推（立即启动）', 'P2 推荐（第二梯队）', 'P3 参考（机会品类）', 'P3 参考（机会品类）', 'P4 备选（节日运营）']
    # 若 LLM 矩阵可用，直接按 LLM 给的优先级标签，避免和上半表的 P1/P2/P3/P4 不一致
    _llm_prio_by_seg: dict[str, str] = {}
    if packs is not None and packs.is_synthesis_real():
        for it in packs.synthesis.sheet6_priority_matrix or []:
            if it.segment and it.priority:
                _llm_prio_by_seg[it.segment] = it.priority
    _prio_label_map = {
        'P1': 'P1 首推（立即启动）', 'P2': 'P2 推荐（第二梯队）',
        'P3': 'P3 参考（机会品类）', 'P4': 'P4 备选（节日运营）',
    }
    for i, pt in enumerate(_ordered_types[:5]):
        if pt in _llm_prio_by_seg:
            lbl = _prio_label_map.get(_llm_prio_by_seg[pt], labels[i] if i < len(labels) else 'P4 备选')
        else:
            lbl = labels[i] if i < len(labels) else 'P4 备选'
        decision_notes.append((lbl, _type_strategy(pt, i, ranked_type_scores)))
    if not decision_notes:
        decision_notes = [('—', '暂无品类评分数据（BSR 数据不足）')]
    for label, note in decision_notes:
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=2)
        val(ws6, r6, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        ws6.merge_cells(start_row=r6, start_column=3, end_row=r6, end_column=9)
        val(ws6, r6, 3, note, wrap=True)
        ws6.row_dimensions[r6].height = 40
        r6 += 1
    apply_border(ws6, r6-len(decision_notes), r6-1, 1, 9)
    r6 += 1
    
    # ===== 产品上新方向 - 新增内容 =====
    # 行动计划总结
    section_title(ws6, r6, 1, '▌ 三、行动计划总结（立即执行）', span=9)
    ws6.row_dimensions[r6].height = 24
    r6 += 1
    
    hdr(ws6, r6, 1, '阶段', bg=C_BLUE_MID)
    hdr(ws6, r6, 2, '时间节点', bg=C_BLUE_MID)
    hdr(ws6, r6, 3, '关键任务', bg=C_BLUE_MID)
    ws6.merge_cells(start_row=r6, start_column=4, end_row=r6, end_column=6)
    hdr(ws6, r6, 4, '执行要点', bg=C_BLUE_MID)
    ws6.merge_cells(start_row=r6, start_column=7, end_row=r6, end_column=9)
    hdr(ws6, r6, 7, '预期目标', bg=C_BLUE_MID)
    ws6.row_dimensions[r6].height = 20
    r6 += 1
    
    action_plan = [
        ('选品开发', '第1-2周', '竞品调研+产品定义',
         '• 深入分析 TOP10 竞品优缺点\n• 确定差异化功能点\n• 与工厂沟通打样',
         '完成产品定义文档'),
        ('产品优化', '第3-4周', '针对差评痛点优化',
         '• 根据差评数据确定改进规格\n• 打样并功能测试\n• 确认成本与利润',
         '样品通过功能测试'),
        ('Listing准备', '第5-6周', 'Listing + 主图设计',
         '• 标题埋词优化\n• 五点描述突出差异化\n• A+ 页面设计',
         'Listing 优化完成'),
        ('测试期运营', '第7-12周', '新品推广',
         '• 初期测评+QA\n• 广告冷启动\n• 报秒杀/优惠券活动',
         f'评分≥{avg_rating:.1f}★'),
        ('稳定期', '第3-6月', '稳定排名',
         '• 广告精细化运营\n• 持续优化转化率\n• 拓展变体',
         '进入类目 TOP50'),
    ]
    
    priority_plan_colors = [C_YELLOW, C_YELLOW, C_BLUE_LIGHT, C_BLUE_LIGHT, C_GREEN_LIGHT]
    for i, (phase, time, task, points, goal) in enumerate(action_plan):
        bg = priority_plan_colors[i]
        val(ws6, r6, 1, phase, bold=True, bg=bg)
        val(ws6, r6, 2, time, bg=bg)
        val(ws6, r6, 3, task, bg=bg)
        ws6.merge_cells(start_row=r6, start_column=4, end_row=r6, end_column=6)
        val(ws6, r6, 4, points, bg=bg, fg='FF595959', size=9, wrap=True)
        ws6.merge_cells(start_row=r6, start_column=7, end_row=r6, end_column=9)
        val(ws6, r6, 7, goal, bg=bg, fg='FF1F3864')
        ws6.row_dimensions[r6].height = 55
        r6 += 1
    apply_border(ws6, r6-len(action_plan)-1, r6-1, 1, 9)
    r6 += 1
    
    # 执行要点与风险提示
    section_title(ws6, r6, 1, '▌ 四、执行要点与风险提示', span=9)
    ws6.row_dimensions[r6].height = 24
    r6 += 1
    
    risk_notes = [
        ('供应链准备', 'C_RED_LIGHT',
         '• 首批备货建议 300-500pcs，避免库存压力\n• 选择有出口经验的工厂，确保交期\n• 要求工厂提供样品进行功能测试'),
        ('品质把控', 'C_YELLOW',
         '• 重点测试差评高频维度（见上表改进优先级）\n• 跌落/防水/耐久等品控标准按品类确定\n• 要求工厂提供第三方检测报告'),
        ('Listing 优化', 'C_BLUE_LIGHT',
         '• 标题格式：[核心词]+[功能]+[规格]\n• 五点描述按痛点→解决方案→产品优势排列\n• 主图突出差异化功能'),
        ('风险预警', 'C_RED_LIGHT',
         '• 风险1：广告 ACOS 过高\n  → 应对：优化 Listing 转化，配合优惠券\n• 风险2：差评导致评分下滑\n  → 应对：主动联系差评用户，及时处理\n• 风险3：竞品价格战\n  → 应对：不参与价格战，用品质和评价区隔'),
    ]
    
    for title, color, content in risk_notes:
        bg = globals().get(color, C_WHITE)
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=2)
        val(ws6, r6, 1, title, bold=True, bg=bg, fg='FF1F3864')
        ws6.merge_cells(start_row=r6, start_column=3, end_row=r6, end_column=9)
        val(ws6, r6, 3, content, bg=bg, fg='FF595959', wrap=True)
        ws6.row_dimensions[r6].height = 65
        r6 += 1
    apply_border(ws6, r6-len(risk_notes)-1, r6-1, 1, 9)

    # ===== Sheet 7: 评论数据汇总 =====
    ws7 = wb.create_sheet('评论数据汇总')
    ws7.sheet_view.showGridLines = False
    ws7.column_dimensions['A'].width = 14
    ws7.column_dimensions['B'].width = 14
    ws7.column_dimensions['C'].width = 8
    ws7.column_dimensions['D'].width = 35
    ws7.column_dimensions['E'].width = 55
    ws7.column_dimensions['F'].width = 10

    ws7.merge_cells('A1:F1')
    c = ws7['A1']
    c.value = f'评论原始数据（{len(rev_df)}条评论）'
    c.font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws7.row_dimensions[1].height = 28

    for i, h in enumerate(['ASIN', '来源文件', '星级', '标题', '评论内容（摘要）', '日期']):
        hdr(ws7, 2, i+1, h, bg=C_BLUE_MID)
    ws7.row_dimensions[2].height = 20

    if len(rev_df) > 0 and 'source_asin' in rev_df.columns:
        # 只选择存在的列
        display_cols = ['source_asin', 'Rating', 'Title', 'Content', 'Date']
        display_cols = [c for c in display_cols if c in rev_df.columns]
        rev_display = rev_df[display_cols].head(800)
        for row_idx, (_, revrow) in enumerate(rev_display.iterrows(), start=3):
            rating = revrow.get('Rating', 3)
            bg = C_RED_LIGHT if (pd.notna(rating) and float(rating) <= 2) else (C_WHITE if (pd.notna(rating) and float(rating) <= 3) else C_GREEN_LIGHT if row_idx % 2 == 0 else C_WHITE)
            content_preview = str(revrow.get('Content', ''))[:200] if pd.notna(revrow.get('Content', '')) else ''
            title_val = revrow.get('Title', '') if 'Title' in rev_df.columns else ''
            date_val = revrow.get('Date', '') if 'Date' in rev_df.columns else ''
            row_vals = [revrow.get('source_asin', ''), '', rating, title_val, content_preview, date_val]
            for ci, v in enumerate(row_vals):
                c = ws7.cell(row=row_idx, column=ci+1, value=str(v) if v else '')
                c.fill = PatternFill('solid', fgColor=bg)
                c.font = Font(name='Arial', size=8)
                c.alignment = Alignment(horizontal='center' if ci in [0, 2, 5] else 'left', vertical='center', wrap_text=(ci == 4))
            ws7.row_dimensions[row_idx].height = 14
        apply_border(ws7, 2, min(800+2, len(rev_df)+2), 1, 6)
    else:
        val(ws7, 3, 1, '暂无评论数据', bg=C_YELLOW)
        ws7.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    
    # ===== 评论数据汇总 - 新增内容 =====
    r7 = 803 if len(rev_df) > 0 else 5
    
    # 评分分布统计
    section_title(ws7, r7, 1, '▌ 评论评分分布统计', span=6)
    ws7.row_dimensions[r7].height = 24
    r7 += 1
    
    if len(rev_df) > 0 and 'Rating' in rev_df.columns:
        rating_dist = rev_df['Rating'].value_counts().sort_index()
        total_reviews = len(rev_df)
    else:
        rating_dist = {}
        total_reviews = 0
    
    hdr(ws7, r7, 1, '星级', bg=C_BLUE_MID)
    hdr(ws7, r7, 2, '评论数', bg=C_BLUE_MID)
    hdr(ws7, r7, 3, '占比', bg=C_BLUE_MID)
    hdr(ws7, r7, 4, '累计占比', bg=C_BLUE_MID)
    ws7.merge_cells(start_row=r7, start_column=5, end_row=r7, end_column=6)
    hdr(ws7, r7, 5, '质量评估', bg=C_BLUE_MID)
    ws7.row_dimensions[r7].height = 20
    r7 += 1
    
    cumsum = 0
    rating_colors_rev = {5: C_GREEN_LIGHT, 4: C_YELLOW, 3: 'FFFFF2CC', 2: C_ORANGE, 1: C_RED_LIGHT}
    quality_notes = {5: '★★★★★ 优秀', 4: '★★★★ 良好', 3: '★★★ 一般', 2: '★★ 较差', 1: '★ 差评集中'}
    
    for rating in [5, 4, 3, 2, 1]:
        cnt = rating_dist.get(rating, 0)
        pct = cnt / total_reviews * 100 if total_reviews > 0 else 0
        cumsum += pct
        bg = rating_colors_rev.get(rating, C_WHITE)
        val(ws7, r7, 1, f'{rating}★', bold=True, bg=bg)
        ws7.cell(row=r7, column=2, value=int(cnt)).alignment = Alignment(horizontal='center', vertical='center')
        ws7.cell(row=r7, column=2).fill = PatternFill('solid', fgColor=bg)
        ws7.cell(row=r7, column=3, value=f'{pct:.1f}%').alignment = Alignment(horizontal='center', vertical='center')
        ws7.cell(row=r7, column=3).fill = PatternFill('solid', fgColor=bg)
        ws7.cell(row=r7, column=4, value=f'{cumsum:.1f}%').alignment = Alignment(horizontal='center', vertical='center')
        ws7.cell(row=r7, column=4).fill = PatternFill('solid', fgColor=bg)
        ws7.merge_cells(start_row=r7, start_column=5, end_row=r7, end_column=6)
        val(ws7, r7, 5, quality_notes.get(rating, ''), bg=bg)
        ws7.row_dimensions[r7].height = 18
        r7 += 1
    apply_border(ws7, r7-5, r7-1, 1, 6)
    r7 += 1
    
    # 高频痛点词Top10
    section_title(ws7, r7, 1, '▌ 高频痛点词Top10（差评分析）', span=6)
    ws7.row_dimensions[r7].height = 24
    r7 += 1
    
    # 提取差评中的高频词
    if len(low_rev) > 0:
        all_low_text = (low_rev['Title'].fillna('') + ' ' + low_rev['Content'].fillna('')).str.lower()
        
        pain_words = ['battery', 'charge', 'died', 'broke', 'broken', 'dead', 'dim', 
                      'not bright', 'weak', 'stopped', 'cheap', 'return', 'refund', 
                      ' defective', 'failed', 'waste', 'garbage', 'terrible', 'worst']
        
        word_counts = {}
        for word in pain_words:
            word_counts[word] = all_low_text.str.contains(word, na=False).sum()
        
        top_pain_words = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    else:
        top_pain_words = []
    
    hdr(ws7, r7, 1, '排名', bg=C_BLUE_MID)
    hdr(ws7, r7, 2, '痛点词', bg=C_BLUE_MID)
    hdr(ws7, r7, 3, '出现次数', bg=C_BLUE_MID)
    ws7.merge_cells(start_row=r7, start_column=4, end_row=r7, end_column=6)
    hdr(ws7, r7, 4, '翻译/说明', bg=C_BLUE_MID)
    ws7.row_dimensions[r7].height = 20
    r7 += 1
    
    word_translations = {
        'battery': '电池问题', 'charge': '充电问题', 'died': '故障/坏了',
        'broke': '损坏', 'broken': '损坏', 'dead': '无法使用',
        'dim': '亮度不足', 'not bright': '不够亮', 'weak': '太弱',
        'stopped': '停止工作', 'cheap': '质量差', 'return': '退货',
        'refund': '退款', 'defective': '有缺陷', 'failed': '失效',
        'waste': '浪费钱', 'garbage': '垃圾', 'terrible': '很差', 'worst': '最差'
    }
    
    for idx, (word, cnt) in enumerate(top_pain_words):
        bg = C_RED_LIGHT if idx < 3 else C_WHITE
        ws7.cell(row=r7, column=1, value=idx+1).alignment = Alignment(horizontal='center', vertical='center')
        ws7.cell(row=r7, column=1).fill = PatternFill('solid', fgColor=bg)
        val(ws7, r7, 2, word, bold=(idx < 3), bg=bg)
        ws7.cell(row=r7, column=3, value=cnt).alignment = Alignment(horizontal='center', vertical='center')
        ws7.cell(row=r7, column=3).fill = PatternFill('solid', fgColor=bg)
        ws7.merge_cells(start_row=r7, start_column=4, end_row=r7, end_column=6)
        val(ws7, r7, 4, word_translations.get(word, ''), bg=bg, fg='FF595959')
        ws7.row_dimensions[r7].height = 18
        r7 += 1
    apply_border(ws7, r7-len(top_pain_words)-1, r7-1, 1, 6)
    r7 += 1
    
    # 高频好评词Top10
    section_title(ws7, r7, 1, '▌ 高频好评词Top10（好评分析）', span=6)
    ws7.row_dimensions[r7].height = 24
    r7 += 1
    
    if len(high_rev) > 0:
        all_high_text = (high_rev['Title'].fillna('') + ' ' + high_rev['Content'].fillna('')).str.lower()
        
        good_words = ['bright', 'easy', 'great', 'perfect', 'love', 'good', 'excellent',
                      'recommend', 'useful', 'strong', 'compact', 'portable', 'durable', 'quality']
        
        good_word_counts = {}
        for word in good_words:
            good_word_counts[word] = all_high_text.str.contains(word, na=False).sum()
        
        top_good_words = sorted(good_word_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    else:
        top_good_words = []
    
    hdr(ws7, r7, 1, '排名', bg=C_BLUE_MID)
    hdr(ws7, r7, 2, '好评词', bg=C_BLUE_MID)
    hdr(ws7, r7, 3, '出现次数', bg=C_BLUE_MID)
    ws7.merge_cells(start_row=r7, start_column=4, end_row=r7, end_column=6)
    hdr(ws7, r7, 4, '翻译/说明', bg=C_BLUE_MID)
    ws7.row_dimensions[r7].height = 20
    r7 += 1
    
    good_translations = {
        'bright': '亮度高', 'easy': '易用', 'great': '很棒', 'perfect': '完美',
        'love': '喜欢', 'good': '好', 'excellent': '优秀', 'recommend': '推荐',
        'useful': '实用', 'strong': '强劲', 'compact': '小巧', 'portable': '便携',
        'durable': '耐用', 'quality': '品质好'
    }
    
    for idx, (word, cnt) in enumerate(top_good_words):
        bg = C_GREEN_LIGHT if idx < 3 else C_WHITE
        ws7.cell(row=r7, column=1, value=idx+1).alignment = Alignment(horizontal='center', vertical='center')
        ws7.cell(row=r7, column=1).fill = PatternFill('solid', fgColor=bg)
        val(ws7, r7, 2, word, bold=(idx < 3), bg=bg)
        ws7.cell(row=r7, column=3, value=cnt).alignment = Alignment(horizontal='center', vertical='center')
        ws7.cell(row=r7, column=3).fill = PatternFill('solid', fgColor=bg)
        ws7.merge_cells(start_row=r7, start_column=4, end_row=r7, end_column=6)
        val(ws7, r7, 4, good_translations.get(word, ''), bg=bg, fg='FF595959')
        ws7.row_dimensions[r7].height = 18
        r7 += 1
    apply_border(ws7, r7-len(top_good_words)-1, r7-1, 1, 6)

    # 预初始化 Sheet 8 变量（供 Sheet 10 引用）
    sell_summary = '—'
    lifecycle_stage = '未知'
    lifecycle_reason = ''
    seasonality_peak_months = []

    # ===== Sheet 8: 类目趋势 =====
    ws8 = wb.create_sheet('类目趋势')
    ws8.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGH', [20, 14, 14, 14, 14, 14, 14, 40]):
        ws8.column_dimensions[col_letter].width = width
    ws8.merge_cells('A1:H1')
    c = ws8['A1']
    c.value = f"{display_name_for_title(packs, '选品')} — 类目趋势 · 季节性 · 生命周期分析"
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws8.row_dimensions[1].height = 35
    r8 = 3

    if not market_data.get('_available'):
        ws8.merge_cells(start_row=r8, start_column=1, end_row=r8+2, end_column=8)
        c = ws8.cell(row=r8, column=1)
        c.value = '⚠ 未上传 Market 分析文件，无法生成类目趋势 / 季节性 / 生命周期分析。\n请在首页第 3 步上传卖家精灵 US-Market 文件（可选）后重新生成报告。'
        c.font = Font(name='Arial', bold=True, size=14, color='FF7A4F01')
        c.fill = PatternFill('solid', fgColor=C_YELLOW)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for rr in range(r8, r8+3):
            ws8.row_dimensions[rr].height = 40
    else:
        # ▌ 一、市场销售趋势
        section_title(ws8, r8, 1, '▌ 一、近 2 年市场销售趋势', span=8)
        ws8.row_dimensions[r8].height = 24
        r8 += 1

        sell_df = market_data.get('sell_trends')
        sell_summary = '—'
        if sell_df is not None and len(sell_df) > 2:
            sell = sell_df.copy()
            sell.columns = [str(c).strip() for c in sell.columns]
            month_c = sell.columns[0]
            sales_c = sell.columns[1] if len(sell.columns) > 1 else None
            rev_c2 = sell.columns[2] if len(sell.columns) > 2 else None
            hdr(ws8, r8, 1, '月份', bg=C_BLUE_MID)
            hdr(ws8, r8, 2, '月销量', bg=C_BLUE_MID)
            hdr(ws8, r8, 3, '月销售额($)', bg=C_BLUE_MID)
            if len(sell.columns) > 3:
                hdr(ws8, r8, 4, '平均BSR', bg=C_BLUE_MID)
            r8 += 1
            data_start = r8
            for _, row in sell.iterrows():
                try:
                    m = str(row[month_c])
                    if m.replace('.', '').replace('0', '').isdigit() and len(m) >= 6:
                        m = m.split('.')[0]
                        m = f'{m[:4]}-{m[4:6]}'
                    val(ws8, r8, 1, m, bg=C_WHITE)
                    if sales_c:
                        c = ws8.cell(row=r8, column=2, value=float(row[sales_c]))
                        c.number_format = '#,##0'
                        c.fill = PatternFill('solid', fgColor=C_WHITE)
                        c.font = Font(name='Arial', size=10)
                        c.alignment = Alignment(horizontal='left', vertical='center')
                    if rev_c2:
                        c = ws8.cell(row=r8, column=3, value=float(row[rev_c2]))
                        c.number_format = '#,##0'
                        c.fill = PatternFill('solid', fgColor=C_WHITE)
                        c.font = Font(name='Arial', size=10)
                        c.alignment = Alignment(horizontal='left', vertical='center')
                    if len(sell.columns) > 3:
                        c = ws8.cell(row=r8, column=4, value=float(row[sell.columns[3]]))
                        c.number_format = '#,##0'
                        c.fill = PatternFill('solid', fgColor=C_WHITE)
                        c.font = Font(name='Arial', size=10)
                        c.alignment = Alignment(horizontal='left', vertical='center')
                    ws8.row_dimensions[r8].height = 18
                    r8 += 1
                except Exception:
                    continue
            data_end = r8 - 1
            apply_border(ws8, data_start-1, data_end, 1, 4)
            # 计算同比
            try:
                sales_series = pd.to_numeric(sell[sales_c], errors='coerce').dropna().values
                if len(sales_series) >= 24:
                    recent12 = sales_series[-12:].sum()
                    prev12 = sales_series[-24:-12].sum()
                    yoy = (recent12 - prev12) / prev12 if prev12 > 0 else 0
                    sell_summary = f'近 12 月销量 {recent12:,.0f} 件，同比{"增长" if yoy > 0 else "下降"} {abs(yoy):.1%}'
                elif len(sales_series) >= 6:
                    sell_summary = f'近 {len(sales_series)} 月销量合计 {sales_series.sum():,.0f} 件（数据不足 2 年，无法计算同比）'
            except Exception:
                pass
            # 折线图
            try:
                chart = LineChart()
                chart.title = '月度销量 / 销售额趋势'
                chart.height = 8
                chart.width = 18
                data_ref = Reference(ws8, min_col=2, max_col=3, min_row=data_start-1, max_row=data_end)
                cats_ref = Reference(ws8, min_col=1, max_col=1, min_row=data_start, max_row=data_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws8.add_chart(chart, f'F{data_start-1}')
            except Exception:
                pass
            r8 += 1
        else:
            val(ws8, r8, 1, '（Industry Sell Trends 无数据）', bg=C_WHITE, fg='FF999999')
            r8 += 1
        r8 += 1

        # ▌ 二、核心关键词搜索趋势
        section_title(ws8, r8, 1, '▌ 二、核心关键词搜索量趋势（近 12 月 vs 上年同期）', span=8)
        ws8.row_dimensions[r8].height = 24
        r8 += 1

        demand_df = market_data.get('demand_trends')
        keyword_peaks = []
        if demand_df is not None and len(demand_df) > 12:
            d = demand_df.copy()
            d.columns = [str(c).strip() for c in d.columns]
            month_c = d.columns[0]
            kw_cols = list(d.columns[1:])
            d[month_c] = d[month_c].astype(str)
            d = d[d[month_c].str.match(r'^\d{4}-\d{2}$', na=False)].sort_values(month_c)

            hdr(ws8, r8, 1, '关键词', bg=C_BLUE_MID)
            hdr(ws8, r8, 2, '近 12 月总搜索', bg=C_BLUE_MID)
            hdr(ws8, r8, 3, '上年同期', bg=C_BLUE_MID)
            hdr(ws8, r8, 4, '同比', bg=C_BLUE_MID)
            hdr(ws8, r8, 5, '峰值月', bg=C_BLUE_MID)
            hdr(ws8, r8, 6, '峰值搜索', bg=C_BLUE_MID)
            ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
            hdr(ws8, r8, 7, '趋势判断', bg=C_BLUE_MID)
            r8 += 1
            kw_start = r8
            for kw in kw_cols:
                try:
                    s = pd.to_numeric(d[kw], errors='coerce').fillna(0)
                    if len(s) < 12:
                        continue
                    last12 = s.iloc[-12:]
                    prev12 = s.iloc[-24:-12] if len(s) >= 24 else None
                    tot_last = last12.sum()
                    tot_prev = prev12.sum() if prev12 is not None else 0
                    yoy = ((tot_last - tot_prev) / tot_prev) if tot_prev > 0 else None
                    peak_idx = int(last12.values.argmax())
                    months = d[month_c].iloc[-12:].tolist()
                    peak_month = months[peak_idx] if peak_idx < len(months) else '-'
                    peak_val = float(last12.iloc[peak_idx])
                    if yoy is None:
                        trend = '数据不足'
                        trend_bg = C_WHITE
                    elif yoy > 0.15:
                        trend = '明显增长'
                        trend_bg = C_GREEN_LIGHT
                    elif yoy > 0:
                        trend = '小幅增长'
                        trend_bg = C_GREEN_LIGHT
                    elif yoy > -0.15:
                        trend = '小幅下滑'
                        trend_bg = C_YELLOW
                    else:
                        trend = '明显下滑'
                        trend_bg = C_RED_LIGHT
                    val(ws8, r8, 1, kw, bold=True, bg=C_BLUE_LIGHT)
                    val(ws8, r8, 2, f'{tot_last:,.0f}')
                    val(ws8, r8, 3, f'{tot_prev:,.0f}' if tot_prev > 0 else 'N/A')
                    val(ws8, r8, 4, f'{yoy:+.1%}' if yoy is not None else 'N/A', fg='FF005A9E' if (yoy or 0) > 0 else 'FFC00000')
                    val(ws8, r8, 5, peak_month)
                    val(ws8, r8, 6, f'{peak_val:,.0f}')
                    ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
                    val(ws8, r8, 7, trend, bold=True, bg=trend_bg)
                    ws8.row_dimensions[r8].height = 20
                    r8 += 1
                    # 收集峰值月（用于季节性）
                    if peak_month != '-':
                        mm = peak_month.split('-')[-1]
                        keyword_peaks.append(mm)
                except Exception:
                    continue
            apply_border(ws8, kw_start-1, r8-1, 1, 8)
        else:
            val(ws8, r8, 1, '（Industry Demand and Trends 无数据）', bg=C_WHITE, fg='FF999999')
            r8 += 1
        r8 += 1

        # ▌ 三、季节性波动
        section_title(ws8, r8, 1, '▌ 三、季节性波动（历年同月均值）', span=8)
        ws8.row_dimensions[r8].height = 24
        r8 += 1

        seasonality_peak_months = []
        if demand_df is not None and len(demand_df) > 24:
            d = demand_df.copy()
            d.columns = [str(c).strip() for c in d.columns]
            month_c = d.columns[0]
            kw_cols = list(d.columns[1:])
            d[month_c] = d[month_c].astype(str)
            d = d[d[month_c].str.match(r'^\d{4}-\d{2}$', na=False)].copy()
            d['_m'] = d[month_c].str.slice(5, 7)
            agg = d.groupby('_m')[kw_cols].apply(lambda x: pd.to_numeric(x.stack(), errors='coerce').mean())
            hdr(ws8, r8, 1, '月份', bg=C_BLUE_MID)
            for j in range(2, 9):
                ws8.cell(row=r8, column=j).fill = PatternFill('solid', fgColor=C_BLUE_MID)
            hdr(ws8, r8, 2, '历年同月均搜索量（5 词合计）', bg=C_BLUE_MID)
            ws8.merge_cells(start_row=r8, start_column=2, end_row=r8, end_column=4)
            hdr(ws8, r8, 5, '相对均值', bg=C_BLUE_MID)
            ws8.merge_cells(start_row=r8, start_column=5, end_row=r8, end_column=6)
            hdr(ws8, r8, 7, '旺/淡季', bg=C_BLUE_MID)
            ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
            r8 += 1
            season_start = r8
            mean_val = agg.mean() if len(agg) else 0
            agg_sorted = agg.sort_index()
            for mm, v in agg_sorted.items():
                try:
                    rel = (v / mean_val - 1) if mean_val else 0
                    if rel > 0.15:
                        tag = '🔥 旺季'
                        bg = C_RED_LIGHT
                        seasonality_peak_months.append(mm)
                    elif rel < -0.15:
                        tag = '❄ 淡季'
                        bg = C_BLUE_LIGHT
                    else:
                        tag = '平稳'
                        bg = C_WHITE
                    val(ws8, r8, 1, f'{mm} 月', bold=True, bg=C_BLUE_LIGHT)
                    ws8.merge_cells(start_row=r8, start_column=2, end_row=r8, end_column=4)
                    val(ws8, r8, 2, f'{v:,.0f}', bg=C_WHITE)
                    ws8.merge_cells(start_row=r8, start_column=5, end_row=r8, end_column=6)
                    val(ws8, r8, 5, f'{rel:+.1%}', bg=bg, fg='FFC00000' if rel > 0 else 'FF005A9E')
                    ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
                    val(ws8, r8, 7, tag, bold=True, bg=bg)
                    ws8.row_dimensions[r8].height = 20
                    r8 += 1
                except Exception:
                    continue
            apply_border(ws8, season_start-1, r8-1, 1, 8)
        else:
            val(ws8, r8, 1, '（Industry Demand and Trends 不足 2 年，无法计算季节性）', bg=C_WHITE, fg='FF999999')
            r8 += 1
        r8 += 1

        # ▌ 四、生命周期判断
        section_title(ws8, r8, 1, '▌ 四、生命周期判断（销量趋势 + 搜索趋势 + 新品贡献率 + 品牌集中度 + 价格趋势 5 维综合评分）', span=8)
        ws8.row_dimensions[r8].height = 24
        r8 += 1

        pub_df = market_data.get('publication_time_trends')
        lifecycle_stage = '未知'
        lifecycle_reason = ''

        # Market 文件的 Publication Time Trends 缺失时，用 BSR 的 Available days 反算发布年份做兜底
        # 卖家精灵 BSR 表固定含 Available days 列（在售天数），反算 launch year = now.year - ceil(days/365)
        if (pub_df is None or len(pub_df) <= 3) and 'Available days' in df.columns:
            try:
                _ad = pd.to_numeric(df['Available days'], errors='coerce').dropna()
                _rev = pd.to_numeric(df[rev_col], errors='coerce').fillna(0) if rev_col else pd.Series([0] * len(df))
                _sales = pd.to_numeric(df.get('Monthly Sales', 0), errors='coerce').fillna(0)
                _launch_years = (datetime.now().year - (_ad / 365).apply(lambda x: int(x) if pd.notna(x) else 0)).astype(int)
                _tmp = pd.DataFrame({
                    'Launch Years': _launch_years.values,
                    'Products': 1,
                    'Sales': _sales.reindex(_launch_years.index, fill_value=0).values,
                    'Monthly Revenue($)': _rev.reindex(_launch_years.index, fill_value=0).values,
                })
                _agg = _tmp.groupby('Launch Years').agg(
                    Products=('Products', 'sum'),
                    Sales=('Sales', 'sum'),
                    **{'Monthly Revenue($)': ('Monthly Revenue($)', 'sum')},
                ).reset_index()
                _total_sales = _agg['Sales'].sum() or 1
                _total_rev = _agg['Monthly Revenue($)'].sum() or 1
                _agg['Sales Proportion'] = _agg['Sales'] / _total_sales
                _agg['Revenue Proportion'] = _agg['Monthly Revenue($)'] / _total_rev
                _agg = _agg[['Launch Years', 'Products', 'Sales', 'Sales Proportion',
                             'Monthly Revenue($)', 'Revenue Proportion']]
                _agg = _agg.sort_values('Launch Years').reset_index(drop=True)
                pub_df = _agg
            except Exception as _e:
                print(f"[Lifecycle] BSR Available days 反算失败: {_e}")

        stage, reason, score_detail = infer_lifecycle_stage(pub_df, df=df, market_data=market_data)
        lifecycle_stage = stage
        lifecycle_reason = reason

        if pub_df is not None and len(pub_df) > 3:
            hdr(ws8, r8, 1, '发布年份', bg=C_BLUE_MID)
            hdr(ws8, r8, 2, '商品数', bg=C_BLUE_MID)
            hdr(ws8, r8, 3, '销量', bg=C_BLUE_MID)
            hdr(ws8, r8, 4, '销量占比', bg=C_BLUE_MID)
            hdr(ws8, r8, 5, '销售额($)', bg=C_BLUE_MID)
            hdr(ws8, r8, 6, '销售额占比', bg=C_BLUE_MID)
            ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
            hdr(ws8, r8, 7, '备注', bg=C_BLUE_MID)
            r8 += 1
            p_start = r8
            for _, row in pub_df.iterrows():
                try:
                    y = row.iloc[0]
                    if pd.isna(y):
                        continue
                    y_int = int(float(y))
                    products = int(pd.to_numeric(row.iloc[1], errors='coerce') or 0)
                    sales = int(pd.to_numeric(row.iloc[2], errors='coerce') or 0)
                    sp = float(pd.to_numeric(row.iloc[3], errors='coerce') or 0)
                    rev_v = float(pd.to_numeric(row.iloc[4], errors='coerce') or 0)
                    rp = float(pd.to_numeric(row.iloc[5], errors='coerce') or 0)
                    age = datetime.now().year - y_int
                    if age < 3:
                        note = '新品代'
                        bg = C_GREEN_LIGHT
                    elif age < 6:
                        note = '中生代'
                        bg = C_YELLOW
                    else:
                        note = '老品'
                        bg = C_BLUE_LIGHT
                    val(ws8, r8, 1, f'{y_int}年', bold=True, bg=bg)
                    val(ws8, r8, 2, f'{products}', bg=C_WHITE)
                    val(ws8, r8, 3, f'{sales:,}', bg=C_WHITE)
                    val(ws8, r8, 4, f'{sp:.1%}', bg=C_WHITE)
                    val(ws8, r8, 5, f'{rev_v:,.0f}', bg=C_WHITE)
                    val(ws8, r8, 6, f'{rp:.1%}', bg=C_WHITE)
                    ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
                    val(ws8, r8, 7, note, bold=True, bg=bg)
                    ws8.row_dimensions[r8].height = 18
                    r8 += 1
                except Exception:
                    continue
            apply_border(ws8, p_start-1, r8-1, 1, 8)
        else:
            val(ws8, r8, 1, '（Publication Time Trends 无数据，仅作辅助展示，不影响下方 5 维生命周期评分）', bg=C_WHITE, fg='FF999999')
            r8 += 1

        # 先调 LLM 拿到生命周期阶段 + 每维详细分析（提前到顶部主结论行之前）。
        # 优先取 5.1.5 阶段提前并发跑出来的 future 结果；不可用时现场调用兜底。
        pattern_info = None
        if _lifecycle_future is not None:
            try:
                pattern_info = _lifecycle_future.result(timeout=180)
            except Exception as _e_lc_await:
                print(f"[Lifecycle 渲染层] 后台 future 取结果失败：{_e_lc_await}",
                      file=__import__('sys').stderr, flush=True)
                pattern_info = None
        if pattern_info is None:
            try:
                pattern_info = classify_market_pattern(score_detail, packs=packs, df=df)
            except Exception as _e:
                pattern_info = None
        dim_analysis_map = {}
        if pattern_info and pattern_info.get('dimension_analyses'):
            for da in pattern_info['dimension_analyses']:
                dim_analysis_map[da.get('dimension', '')] = da.get('analysis', '')

        # 用 LLM 给的 stage 覆盖 avg 写死映射（兜底仍用 lifecycle_stage）
        llm_stage = (pattern_info or {}).get('stage') or ''
        if llm_stage:
            lifecycle_stage = llm_stage  # 后续 L5564 / L5979 等处也用这个

        # 顶部主结论行（生命周期阶段标签 + reason）
        r8 += 1
        ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
        c = ws8.cell(row=r8, column=1)
        c.value = f'🎯 生命周期判断：{lifecycle_stage}   |   {lifecycle_reason}'
        c.font = Font(name='Arial', bold=True, size=12, color='FF1F3864')
        c.fill = PatternFill('solid', fgColor=C_GREEN_LIGHT)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws8.row_dimensions[r8].height = 28
        r8 += 1

        # 主结论下方小字：判定逻辑（说明为什么是这个阶段）
        stage_reasoning = (pattern_info or {}).get('stage_reasoning', '').strip()
        if stage_reasoning:
            ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
            sr = ws8.cell(row=r8, column=1)
            sr.value = f'   📍 判定逻辑：{stage_reasoning}'
            sr.font = Font(name='Arial', size=9, italic=True, color='FF595959')
            sr.fill = PatternFill('solid', fgColor='FFF6F8FB')
            sr.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            line_estimate = max(2, len(stage_reasoning) // 70 + 1)
            ws8.row_dimensions[r8].height = max(28, line_estimate * 16)
            r8 += 1

        # ── 子段：生命周期判定标准（决策表 16 条规则）──
        matched_rule_id = (pattern_info or {}).get('matched_rule_id')
        r8 += 1
        ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
        dt_title = ws8.cell(row=r8, column=1)
        dt_title.value = ('📚 生命周期判定标准（决策表）'
                          + (f'   ◀ 当前命中：{matched_rule_id}' if matched_rule_id else ''))
        dt_title.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
        dt_title.fill = PatternFill('solid', fgColor=C_BLUE_LIGHT)
        dt_title.alignment = Alignment(horizontal='left', vertical='center')
        ws8.row_dimensions[r8].height = 22
        r8 += 1

        # 表头：编号 | 销量 | 搜索 | 修正条件 | 判定阶段 | 业务说明
        hdr(ws8, r8, 1, '编号', bg=C_BLUE_MID)
        hdr(ws8, r8, 2, '销量', bg=C_BLUE_MID)
        hdr(ws8, r8, 3, '搜索', bg=C_BLUE_MID)
        ws8.merge_cells(start_row=r8, start_column=4, end_row=r8, end_column=5)
        hdr(ws8, r8, 4, '修正条件', bg=C_BLUE_MID)
        hdr(ws8, r8, 6, '判定阶段', bg=C_BLUE_MID)
        ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
        hdr(ws8, r8, 7, '业务说明', bg=C_BLUE_MID)
        ws8.row_dimensions[r8].height = 22
        r8 += 1

        dt_start = r8
        for rule in _LIFECYCLE_DECISION_RULES:
            is_hit = (rule['id'] == matched_rule_id)
            row_bg = 'FFB7E1A1' if is_hit else C_WHITE  # 命中行深绿，其他白
            row_fg = 'FF0F5132' if is_hit else 'FF333333'
            row_bold = is_hit

            val(ws8, r8, 1, rule['id'], bold=True if is_hit else False, bg=row_bg, fg=row_fg)
            val(ws8, r8, 2, rule['sales'], bold=row_bold, bg=row_bg, fg=row_fg)
            val(ws8, r8, 3, rule['search'], bold=row_bold, bg=row_bg, fg=row_fg)
            ws8.merge_cells(start_row=r8, start_column=4, end_row=r8, end_column=5)
            val(ws8, r8, 4, rule['condition'], bold=row_bold, bg=row_bg, fg=row_fg)
            val(ws8, r8, 6, rule['stage'], bold=True if is_hit else False, bg=row_bg, fg=row_fg)
            ws8.merge_cells(start_row=r8, start_column=7, end_row=r8, end_column=8)
            val(ws8, r8, 7, rule['reason'], bold=row_bold, bg=row_bg, fg=row_fg)
            ws8.row_dimensions[r8].height = 20
            r8 += 1
        apply_border(ws8, dt_start - 1, r8 - 1, 1, 8)

        # 决策表说明小字
        ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
        dt_note = ws8.cell(row=r8, column=1)
        dt_note.value = ('   ℹ️ 决策表按优先级（自上而下）匹配第一条命中规则；'
                         '深绿色高亮行 = 当前品类的判定结果。规则覆盖典型 5 维组合，'
                         '未命中则归为「成熟晚期」兜底。')
        dt_note.font = Font(name='Arial', size=9, italic=True, color='FF595959')
        dt_note.fill = PatternFill('solid', fgColor=C_WHITE)
        dt_note.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws8.row_dimensions[r8].height = 28
        r8 += 1

        lc_start = r8
        # ── 子段一：5 维数据总览表（紧凑横表）──
        ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
        title_c = ws8.cell(row=r8, column=1)
        title_c.value = '📊 5 维数据总览'
        title_c.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
        title_c.fill = PatternFill('solid', fgColor=C_BLUE_LIGHT)
        title_c.alignment = Alignment(horizontal='left', vertical='center')
        ws8.row_dimensions[r8].height = 22
        r8 += 1

        # 表头：维度 | 评分 | 关键指标 1 | 关键指标 2 | 关键指标 3 | 数据来源（跨 3 列）
        hdr(ws8, r8, 1, '维度', bg=C_BLUE_MID)
        hdr(ws8, r8, 2, '评分', bg=C_BLUE_MID)
        hdr(ws8, r8, 3, '关键指标 1', bg=C_BLUE_MID)
        hdr(ws8, r8, 4, '关键指标 2', bg=C_BLUE_MID)
        hdr(ws8, r8, 5, '关键指标 3', bg=C_BLUE_MID)
        ws8.merge_cells(start_row=r8, start_column=6, end_row=r8, end_column=8)
        hdr(ws8, r8, 6, '数据来源 / 备注', bg=C_BLUE_MID)
        ws8.row_dimensions[r8].height = 22
        r8 += 1

        ov_start = r8
        for dim_name in score_detail.keys():
            try:
                score_txt, m1, m2, m3, src = _build_dim_summary_row(
                    dim_name, score_detail, market_data, df, packs
                )
            except Exception:
                score_txt, m1, m2, m3, src = '⚠️', '-', '-', '-', '-'

            score = (score_detail.get(dim_name) or {}).get('score')
            row_bg = C_WHITE
            row_fg = 'FF1F3864'
            if score is None:
                row_bg = 'FFFCFCFC'
                row_fg = 'FF999999'
            elif score == 1:
                row_bg = 'FFE8F5E8'  # 浅绿
            elif score == -1:
                row_bg = 'FFFFE7E7'  # 浅红
            else:
                row_bg = 'FFFFFCE5'  # 浅黄

            val(ws8, r8, 1, dim_name, bold=True, bg=row_bg, fg=row_fg)
            val(ws8, r8, 2, score_txt, bold=True, bg=row_bg, fg=row_fg)
            val(ws8, r8, 3, m1, bg=C_WHITE)
            val(ws8, r8, 4, m2, bg=C_WHITE)
            val(ws8, r8, 5, m3, bg=C_WHITE)
            ws8.merge_cells(start_row=r8, start_column=6, end_row=r8, end_column=8)
            val(ws8, r8, 6, src, bg=C_WHITE, fg='FF595959')
            ws8.row_dimensions[r8].height = 22
            r8 += 1
        apply_border(ws8, ov_start - 1, r8 - 1, 1, 8)

        # ── 子段二：每维 LLM 解读（精简版，每维一段，不再有 📌 数据明细块）──
        if dim_analysis_map:
            r8 += 1
            ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
            llm_title = ws8.cell(row=r8, column=1)
            llm_title.value = '💡 每维 LLM 解读'
            llm_title.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
            llm_title.fill = PatternFill('solid', fgColor=C_BLUE_LIGHT)
            llm_title.alignment = Alignment(horizontal='left', vertical='center')
            ws8.row_dimensions[r8].height = 22
            r8 += 1

            llm_start = r8
            for dim_name in score_detail.keys():
                analysis_text = dim_analysis_map.get(dim_name, '').strip()
                if not analysis_text:
                    continue
                ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
                ac = ws8.cell(row=r8, column=1)
                ac.value = f'  • {dim_name}：{analysis_text}'
                ac.font = Font(name='Arial', size=10, color='FF333333')
                ac.fill = PatternFill('solid', fgColor=C_WHITE)
                ac.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                line_estimate = max(2, len(analysis_text) // 60 + 1)
                ws8.row_dimensions[r8].height = max(36, line_estimate * 18)
                r8 += 1
            apply_border(ws8, llm_start - 1, r8 - 1, 1, 8)

        # 底部：综合结论 + 进入决策（一段彩色横条）
        if pattern_info and pattern_info.get('overall_conclusion'):
            r8 += 1
            verdict_text = pattern_info.get('verdict', '')
            ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
            vc = ws8.cell(row=r8, column=1)
            vc.value = f'🎯 综合结论与进入决策：{verdict_text}'
            vc.font = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
            vc.fill = PatternFill('solid', fgColor='FF1F3864')
            vc.alignment = Alignment(horizontal='left', vertical='center')
            ws8.row_dimensions[r8].height = 28
            r8 += 1

            ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
            ec = ws8.cell(row=r8, column=1)
            ec.value = pattern_info['overall_conclusion']
            ec.font = Font(name='Arial', size=11, color='FF1F3864')
            ec.fill = PatternFill('solid', fgColor='FFE8F5E8')
            ec.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            conc_text = pattern_info['overall_conclusion']
            conc_lines = max(3, len(conc_text) // 60 + 1)
            ws8.row_dimensions[r8].height = max(50, conc_lines * 20)
            r8 += 1

            # avg 计算公式注解（让用户知道 +0.40 这种数字怎么来的）
            try:
                _avg_parts = []
                _avg_sum = 0
                _avg_n = 0
                for _dn, _di in (score_detail or {}).items():
                    _sc = _di.get('score') if isinstance(_di, dict) else None
                    if _sc is None:
                        _avg_parts.append(f'({_dn} 数据缺失)')
                    else:
                        sign = '+' if _sc > 0 else ('' if _sc == 0 else '')
                        _avg_parts.append(f'({_dn} {sign}{_sc})')
                        _avg_sum += _sc
                        _avg_n += 1
                if _avg_n > 0:
                    _avg_val = _avg_sum / _avg_n
                    _formula = (
                        f'ℹ️ 综合评分算法：5 维评分（每维 +1/0/-1）相加 ÷ 有效维度数 = 加权平均分\n'
                        f'   本品类：{" + ".join(_avg_parts)} = {_avg_sum:+d}，÷ {_avg_n} 维有效 = {_avg_val:+.2f}\n'
                        f'   说明：此 +{_avg_val:.2f} 仅用于 verdict 决策（推/谨慎/不推）；'
                        f'生命周期阶段（如「成长期」）由上方决策表决定，不依赖 avg'
                    )
                    ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
                    fc = ws8.cell(row=r8, column=1)
                    fc.value = _formula
                    fc.font = Font(name='Arial', size=9, italic=True, color='FF595959')
                    fc.fill = PatternFill('solid', fgColor='FFF6F8FB')
                    fc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    ws8.row_dimensions[r8].height = 56
                    r8 += 1
            except Exception:
                pass

            # 来源小字（让用户知道是 LLM 还是兜底模板）
            # classify_market_pattern 成功路径返回 'LLM+DecisionTable'、兜底返回 'fallback-template'，
            # 任何以 'LLM' 打头的来源都视为 LLM 真实生成（含 LLM+DecisionTable 这种组合标识）。
            src = pattern_info.get('source', '') or ''
            src_label = ('（业务洞察由 LLM 生成）' if src.startswith('LLM')
                         else '（LLM 不可用，使用规则模板兜底）')
            ws8.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=8)
            sc_cell = ws8.cell(row=r8, column=1)
            sc_cell.value = src_label
            sc_cell.font = Font(name='Arial', size=9, italic=True, color='FF999999')
            sc_cell.fill = PatternFill('solid', fgColor=C_WHITE)
            sc_cell.alignment = Alignment(horizontal='right', vertical='center')
            ws8.row_dimensions[r8].height = 16
            r8 += 1

        apply_border(ws8, lc_start-1, r8-1, 1, 8)
        r8 += 1

        # ▌ 五、类目集中度
        section_title(ws8, r8, 1, '▌ 五、类目集中度', span=8)
        ws8.row_dimensions[r8].height = 24
        r8 += 1

        brand_df = market_data.get('brand_concentration')
        listing_df = market_data.get('listing_concentration')
        conc_rows = []
        _got_brand = False
        if brand_df is not None and len(brand_df) >= 5:
            try:
                sp = pd.to_numeric(brand_df['Sales Proportion'], errors='coerce').dropna()
                cr3_b = sp.head(3).sum()
                cr5_b = sp.head(5).sum()
                cr10_b = sp.head(10).sum()
                top3_b = brand_df.iloc[:3, 1].astype(str).tolist() if brand_df.shape[1] > 1 else []
                conc_rows.append(('品牌 CR3', f'{cr3_b:.1%}', f'Top3 品牌：{", ".join(top3_b)}'))
                conc_rows.append(('品牌 CR5', f'{cr5_b:.1%}', '前 5 品牌销量合计占比（数据源：Market 品牌集中度）'))
                conc_rows.append(('品牌 CR10', f'{cr10_b:.1%}', '前 10 品牌销量合计占比（数据源：Market 品牌集中度）'))
                _got_brand = True
            except Exception:
                pass
        # Market 文件缺失或解析失败时，用 BSR brand_agg 兜底
        if not _got_brand:
            try:
                if 'brand_agg' in dir() or 'brand_agg' in locals():
                    _ba = locals().get('brand_agg')
                    _tr = locals().get('total_rev')
                    if _ba is not None and _tr and 'total_rev' in _ba.columns and len(_ba) >= 3:
                        _cr3 = float(_ba.head(3)['total_rev'].sum() / _tr)
                        _cr5 = float(_ba.head(5)['total_rev'].sum() / _tr) if len(_ba) >= 5 else None
                        _cr10 = float(_ba.head(10)['total_rev'].sum() / _tr) if len(_ba) >= 10 else None
                        _top3_names = _ba.head(3)['Brand'].astype(str).tolist() if 'Brand' in _ba.columns else []
                        conc_rows.append(('品牌 CR3', f'{_cr3:.1%}', f'Top3 品牌：{", ".join(_top3_names)}（来源：BSR brand_agg 兜底）'))
                        if _cr5 is not None:
                            conc_rows.append(('品牌 CR5', f'{_cr5:.1%}', '前 5 品牌销售收入合计占比（来源：BSR brand_agg 兜底）'))
                        if _cr10 is not None:
                            conc_rows.append(('品牌 CR10', f'{_cr10:.1%}', '前 10 品牌销售收入合计占比（来源：BSR brand_agg 兜底）'))
            except Exception:
                pass
        _got_listing = False
        if listing_df is not None and len(listing_df) >= 10:
            try:
                sp = pd.to_numeric(listing_df['Sales Proportion'], errors='coerce').dropna()
                cr10_l = sp.head(10).sum()
                cr20_l = sp.head(20).sum()
                conc_rows.append(('Listing CR10', f'{cr10_l:.1%}', '前 10 ASIN 销量合计占比（数据源：Market 商品集中度）'))
                conc_rows.append(('Listing CR20', f'{cr20_l:.1%}', '前 20 ASIN 销量合计占比（数据源：Market 商品集中度）'))
                _got_listing = True
            except Exception:
                pass
        # Listing 集中度同样用 BSR 单 ASIN 收入兜底
        if not _got_listing:
            try:
                _rev_col = locals().get('rev_col') or 'Parent Revenue'
                _tr = locals().get('total_rev')
                if _rev_col in df.columns and _tr and len(df) >= 10:
                    _sorted_rev = pd.to_numeric(df[_rev_col], errors='coerce').fillna(0).sort_values(ascending=False)
                    _cr10_l = float(_sorted_rev.head(10).sum() / _tr)
                    _cr20_l = float(_sorted_rev.head(20).sum() / _tr) if len(_sorted_rev) >= 20 else None
                    conc_rows.append(('Listing CR10', f'{_cr10_l:.1%}', '前 10 ASIN 收入合计占比（来源：BSR 兜底）'))
                    if _cr20_l is not None:
                        conc_rows.append(('Listing CR20', f'{_cr20_l:.1%}', '前 20 ASIN 收入合计占比（来源：BSR 兜底）'))
            except Exception:
                pass
        if conc_rows:
            hdr(ws8, r8, 1, '指标', bg=C_BLUE_MID)
            hdr(ws8, r8, 2, '数值', bg=C_BLUE_MID)
            ws8.merge_cells(start_row=r8, start_column=3, end_row=r8, end_column=8)
            hdr(ws8, r8, 3, '说明', bg=C_BLUE_MID)
            r8 += 1
            c_start = r8
            for label, v, note in conc_rows:
                val(ws8, r8, 1, label, bold=True, bg=C_BLUE_LIGHT)
                val(ws8, r8, 2, v, bold=True, bg=C_WHITE)
                ws8.merge_cells(start_row=r8, start_column=3, end_row=r8, end_column=8)
                val(ws8, r8, 3, note, bg=C_WHITE, fg='FF595959')
                ws8.row_dimensions[r8].height = 20
                r8 += 1
            apply_border(ws8, c_start-1, r8-1, 1, 8)
        else:
            val(ws8, r8, 1, '（Brand/Listing Concentration 无数据）', bg=C_WHITE, fg='FF999999')
            r8 += 1
        r8 += 1

        # ▌ 六、小结
        section_title(ws8, r8, 1, '▌ 六、趋势与生命周期小结', span=8, bg=C_ORANGE)
        ws8.row_dimensions[r8].height = 24
        r8 += 1
        peak_summary = '；'.join([f'{m} 月' for m in sorted(set(seasonality_peak_months))]) if seasonality_peak_months else '无明显旺季'
        summary_lines = [
            f'• 销售趋势：{sell_summary}',
            f'• 季节性：旺季集中在 {peak_summary}',
            f'• 生命周期：{lifecycle_stage} —— {lifecycle_reason}',
        ]
        ws8.merge_cells(start_row=r8, start_column=1, end_row=r8+2, end_column=8)
        c = ws8.cell(row=r8, column=1)
        c.value = '\n'.join(summary_lines)
        c.font = Font(name='Arial', size=11, color='FF1F3864')
        c.fill = PatternFill('solid', fgColor=C_BLUE_LIGHT)
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for rr in range(r8, r8+3):
            ws8.row_dimensions[rr].height = 24
        r8 += 3

    # ===== Sheet 9: 风险预估 =====
    ws9 = wb.create_sheet('风险预估')
    ws9.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGH', [22, 24, 18, 18, 50, 10, 10, 10]):
        ws9.column_dimensions[col_letter].width = width
    ws9.merge_cells('A1:H1')
    c = ws9['A1']
    c.value = f"{display_name_for_title(packs, '选品')} — 上架前风险预估清单（通用模板 + 品类自适应）"
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor='FFB43E21')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws9.row_dimensions[1].height = 35
    r9 = 3

    # 基于 BSR Title 做关键词命中统计
    titles_all = ' '.join(str(t).lower() for t in df['Product Title'].astype(str).tolist()) if 'Product Title' in df.columns else ''
    def count_hits(words):
        return sum(1 for t in df['Product Title'].astype(str) if any(w in t.lower() for w in words)) if 'Product Title' in df.columns else 0

    # 全品类通用的风险项（知识产权/平台政策/供应链）保留硬编码 —— 这些跨品类都适用
    # 「供应链与海运」按品类特征动态组合文案（电池 / 液体 / 磁铁 / 大件 / 易碎 / 儿童 / 无线电子 / 食品接触）
    _cat_name_risk = (getattr(packs, 'display_name', '') or getattr(packs, 'category_id', '') or '').lower()
    _cert_names_all = ''
    if packs is not None and packs.is_compliance_real():
        _cert_names_all = ' '.join([c.name for c in packs.compliance.required_certifications]).lower()

    def _has_trait(kws, cert_kws=None):
        """标题命中 ≥5 条或 ≥30%，或品类名含关键词，或认证列表命中对应认证 → 认为本批 ASIN 属于该品类特征。"""
        thresh = max(5, int(len(df) * 0.3)) if len(df) else 5
        if any(k.lower() in _cat_name_risk for k in kws):
            return True
        if count_hits([k.lower() for k in kws]) >= thresh:
            return True
        if cert_kws and any(k in _cert_names_all for k in cert_kws):
            return True
        return False

    _traits = {
        'battery':   _has_trait(['battery', 'charger', 'lithium', 'rechargeable', 'cordless', '电池', '充电', '锂电'],
                                cert_kws=['un38.3', 'un 38.3', '62368', '62133', 'battery']),
        'liquid':    _has_trait(['spray', 'aerosol', 'liquid', 'oil', 'cleaner', 'paint', 'glue', 'adhesive', 'ink',
                                 '喷雾', '液体', '清洁剂', '油', '胶水'],
                                cert_kws=['msds', 'hazmat']),
        'magnetic':  _has_trait(['magnet', 'magnetic', '磁', '磁铁']),
        'bulky':     _has_trait(['mower', 'generator', 'ladder', 'treadmill', 'scooter', 'grill', 'inflator',
                                 '电机', '发电机', '跑步机', '割草机', '充气机', '洗车']),
        'fragile':   _has_trait(['glass', 'mirror', 'ceramic', 'porcelain', 'crystal',
                                 '玻璃', '陶瓷', '镜', '水晶']),
        'kids':      _has_trait(['kid', 'child', 'baby', 'toddler', 'infant', 'toy',
                                 '儿童', '宝宝', '玩具', '婴儿'],
                                cert_kws=['cpsia', 'astm f963']),
        'wireless':  _has_trait(['bluetooth', 'wifi', 'wi-fi', 'remote', 'wireless', 'smart', 'zigbee',
                                 '蓝牙', '无线', '遥控', '智能'],
                                cert_kws=['fcc part 15', 'fcc id']),
        'food':      _has_trait(['food', 'kitchen', 'cookware', 'mug', 'bottle', 'cup', 'bowl', 'utensil',
                                 '食品', '厨房', '餐具', '水杯', '水壶'],
                                cert_kws=['fda', 'lfgb']),
    }

    _advice_parts = ['头程建议预留 6-8 周']
    if _traits['battery']:
        _advice_parts.append('含电池类仅限海卡/美森等合规通道，需 UN38.3 + MSDS + 非危包证')
    if _traits['liquid']:
        _advice_parts.append('含液体/气雾剂需提供 MSDS + 非危鉴定报告，禁飞或走海卡')
    if _traits['magnetic']:
        _advice_parts.append('含磁性部件需磁检报告，选择接受磁货的承运')
    if _traits['bulky']:
        _advice_parts.append('大件走海运整柜/拼箱降本，注意体积重与超尺寸附加费')
    if _traits['fragile']:
        _advice_parts.append('易碎品加强防震包装（EPE/气泡膜/角撑），首批走空派降低破损率')
    if _traits['kids']:
        _advice_parts.append('儿童用品包装需含 Tracking Label + 小零件警告')
    if _traits['wireless']:
        _advice_parts.append('含无线/电子模块需 FCC ID 标签且 labeling 与 listing 参数一致')
    if _traits['food']:
        _advice_parts.append('食品接触品需提供食品级材质证明（FDA/LFGB），包装避免混装化工品')
    if len(_advice_parts) == 1:
        _advice_parts.append('首批建议空派快速铺货，稳定后切海运降本；出货前核实工厂 MSDS / 材质报告 / 非危鉴定')
    _supply_advice = '；'.join(_advice_parts)
    _is_battery_cat = _traits['battery']  # 给下方 checklist 的 shipping_line 沿用

    risk_items = [
        ('知识产权 · 专利 · 商标', '默认必查', '全品类', '高',
         '查询 Amazon Brand Registry / Google Patents / USPTO / Trademark Electronic Search System；避开明显仿款',
         1),
        ('Amazon 平台政策', '默认', '全品类', '中',
         '避开 Restricted Category；使用品牌备案减少跟卖；禁用 overclaim；准备好发票应对品牌投诉',
         1),
        ('供应链与海运', '默认', '全品类', '低',
         _supply_advice,
         1),
    ]

    # 品类特异的认证风险由 CompliancePack 动态生成（LLM 按本次品类识别）；
    # LLM 不可用时回落到基于关键词命中的通用认证列表（保留跨品类普适的 UN38.3/FCC/Prop 65/CPSIA/FDA）
    _llm_certs = compliance_certs(packs)
    import re as __re_risk
    if _llm_certs:
        for _c in _llm_certs:
            _name = _c['name']
            _applies = _c['applies_to'] or '本品类'
            _level = '高' if _c['mandatory'] else '中'
            _advice = _c['risk_if_missing'] or '参见亚马逊合规页面'
            # 触发关键词显示规则：
            # - 优先取 applies_to 英文原文前 50 字（如 "Contains AC power adapter"）
            # - 若 applies_to 为中文或为空 → 显示"按品类识别"
            # 不再把认证名（"FCC Part 15"）切成 token 塞进来（会拆成"FCC, Part"碎片）
            _en_tokens = __re_risk.findall(r'[A-Za-z][A-Za-z0-9]{2,}', _name)
            _hit_kws = [t.lower() for t in _en_tokens if len(t) > 2][:5]
            _hits = count_hits(_hit_kws) if _hit_kws else 0

            _applies_str = str(_applies).strip()
            _has_english = bool(__re_risk.search(r'[A-Za-z]{3,}', _applies_str))
            if _has_english and _applies_str not in ('本品类', ''):
                _trigger_disp = _applies_str[:50]
            else:
                _trigger_disp = '按品类识别'
            risk_items.append((_name, _trigger_disp, _applies, _level, _advice, _hits))
    else:
        # 通用认证清单（非 LED 专属）—— 删除了原 UL8750/UL1598/UL60950/退货率"工作灯 5-8%" 硬句
        risk_items.extend([
            ('电池安全 · UN38.3', ['battery', 'rechargeable', 'cordless', 'lithium'], '含电池类', '高',
             '锂电产品需要 UN38.3 测试报告 + MSDS + 非危险品鉴定；海运需走危包证',
             count_hits(['battery', 'rechargeable', 'cordless', 'lithium'])),
            ('FCC 认证', ['wireless', 'bluetooth', 'remote', 'wifi', 'rf'], '无线设备', '中',
             '含无线模块产品必须提供 FCC ID；亚马逊抽查较多',
             count_hits(['wireless', 'bluetooth', 'remote', 'wifi', 'rf'])),
            ('加州 Prop 65', ['plastic', 'pvc', 'paint', 'cable'], '化学材料', '中',
             '含塑料 / 电缆 / 涂层产品需在 listing 添加 Prop 65 警告标签',
             count_hits(['plastic', 'pvc', 'paint', 'cable'])),
            ('CPSC / CPSIA', ['kid', 'child', 'baby', 'toddler'], '儿童用品', '高',
             '儿童用品需 CPSIA 重金属测试 + 小零件警告 + Tracking Label',
             count_hits(['kid', 'child', 'baby', 'toddler'])),
            ('FDA 注册', ['food', 'kitchen', 'contact'], '食品接触类', '中',
             '接触食品类需 FDA 设施注册 + 预警通知 PN',
             count_hits(['food', 'kitchen', 'contact'])),
        ])

    # 退货率：优先用 CompliancePack 给的品类典型退货率 + 退货原因（LLM 按品类识别）
    _ret_rate, _ret_reasons = compliance_return_info(packs)
    if _ret_rate > 0 or _ret_reasons:
        _ret_advice_parts = []
        if _ret_rate > 0:
            _ret_advice_parts.append(f'本品类典型退货率约 {_ret_rate:.1f}%')
        if _ret_reasons:
            _ret_advice_parts.append('常见退货原因：' + '、'.join(_ret_reasons[:5]))
        _ret_advice_parts.append('listing 需标明真实规格并提供产品真实照片')
        risk_items.append(('退货率预估', '—', '本品类', '中', '；'.join(_ret_advice_parts), 0))

    hdr(ws9, r9, 1, '风险类别', bg='FFB43E21', fg=C_WHITE)
    hdr(ws9, r9, 2, '触发关键词', bg='FFB43E21', fg=C_WHITE)
    hdr(ws9, r9, 3, '适用范围', bg='FFB43E21', fg=C_WHITE)
    hdr(ws9, r9, 4, '风险等级', bg='FFB43E21', fg=C_WHITE)
    ws9.merge_cells(start_row=r9, start_column=5, end_row=r9, end_column=7)
    hdr(ws9, r9, 5, '规避建议', bg='FFB43E21', fg=C_WHITE)
    hdr(ws9, r9, 8, '命中数', bg='FFB43E21', fg=C_WHITE)
    ws9.row_dimensions[r9].height = 22
    r9 += 1

    total_hits = 0
    risk_rows_start = r9
    for cat, kws, scope, level, advice, hits in risk_items:
        level_bg = {'高': C_RED_LIGHT, '中': C_YELLOW, '低': C_GREEN_LIGHT}.get(level, C_WHITE)
        if isinstance(kws, list):
            kw_disp = ', '.join(kws)
        else:
            kw_disp = str(kws)
        val(ws9, r9, 1, cat, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        val(ws9, r9, 2, kw_disp, bg=C_WHITE, fg='FF595959')
        val(ws9, r9, 3, scope, bg=C_WHITE)
        val(ws9, r9, 4, level, bold=True, bg=level_bg)
        ws9.merge_cells(start_row=r9, start_column=5, end_row=r9, end_column=7)
        val(ws9, r9, 5, advice, bg=C_WHITE)
        val(ws9, r9, 8, f'{hits}' if hits > 0 else '—', bold=True,
            bg=C_RED_LIGHT if hits > 10 else (C_YELLOW if hits > 0 else C_WHITE))
        ws9.row_dimensions[r9].height = 40
        if hits > 0 and level in ('高', '中'):
            total_hits += 1
        r9 += 1
    apply_border(ws9, risk_rows_start-1, r9-1, 1, 8)
    r9 += 1

    # 通用合规清单
    section_title(ws9, r9, 1, '▌ 通用合规清单（上新前自查）', span=8, bg=C_ORANGE)
    ws9.row_dimensions[r9].height = 24
    r9 += 1
    # 合规端 / 海运端：按本次品类动态生成，避免非本品类出现不相关提示
    _cert_tail = ''
    if packs is not None and packs.is_compliance_real():
        _mand = [c.name for c in packs.compliance.required_certifications if c.mandatory][:3]
        if _mand:
            _cert_tail = '（本次品类重点：' + '、'.join(_mand) + '）'
    _ship_tips = []
    if _traits['battery']:
        _ship_tips.append('电池类 UN38.3 + MSDS + 非危包证')
    if _traits['liquid']:
        _ship_tips.append('液体/气雾剂 MSDS + 非危鉴定')
    if _traits['magnetic']:
        _ship_tips.append('磁性部件磁检报告')
    if _traits['bulky']:
        _ship_tips.append('大件走海运并核算体积重 / 超尺寸')
    if _traits['fragile']:
        _ship_tips.append('易碎品加强防震包装')
    if _ship_tips:
        _shipping_line = '✓ 海运端：' + '、'.join(_ship_tips) + '；选择合规承运'
    else:
        _shipping_line = '✓ 海运端：核实非危鉴定 / MSDS / 木质包装 IPPC 等基础出货资料，首批走空派提速'
    checklist = [
        '✓ 采购端：索要 1688/工厂的 材质报告、MSDS、发票',
        '✓ 品牌端：Amazon Brand Registry 备案（TM+R 商标）+ Transparency 计划防跟卖',
        f'✓ 合规端：按品类准备认证报告{_cert_tail or "（UL/FCC/CE/CPSIA 等按品类选取）"}',
        _shipping_line,
        '✓ Listing 端：避免夸大卖点 / 对标词、补充真实规格图 / 使用说明',
        '✓ 法务端：查 Google Patents + USPTO + 商标全球搜索，避开 3 年内专利',
    ]
    for item in checklist:
        ws9.merge_cells(start_row=r9, start_column=1, end_row=r9, end_column=8)
        val(ws9, r9, 1, item, bg=C_WHITE, fg='FF1F3864')
        ws9.row_dimensions[r9].height = 20
        r9 += 1

    # ===== Sheet 10: 选品综合结论 =====
    ws10 = wb.create_sheet('选品综合结论')
    ws10.sheet_view.showGridLines = False
    for col_letter, width in zip('ABCDEFGH', [20, 14, 14, 14, 50, 14, 14, 14]):
        ws10.column_dimensions[col_letter].width = width
    ws10.merge_cells('A1:H1')
    c = ws10['A1']
    c.value = f"{display_name_for_title(packs, '选品')} — 选品综合评估与最终建议"
    c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws10.row_dimensions[1].height = 35
    r10 = 3

    # 报告阅读说明 — 把"推荐 = 市场机会，盈利 = 压价能力"逻辑贴在结论页顶部
    _reading_note = (
        '📌 报告阅读说明：本报告"推荐"基于市场机会判断（销量+需求+痛点+新品空间）；'
        '报告中各项费用为 LLM 基于行业普遍参考值的推算估算（仅作参考），'
        '**真实盈利能力取决于你公司的实际费用结构以及供应链压价能力**。'
        '具体采购成本目标见 Sheet 4「盈亏平衡采购成本上限（按目标净利率）」多档表。'
    )
    ws10.merge_cells(start_row=r10, start_column=1, end_row=r10, end_column=8)
    c = ws10.cell(row=r10, column=1, value=_reading_note)
    c.font = Font(name='Arial', size=10, bold=True, color='FF7B5800')
    c.fill = PatternFill('solid', fgColor=C_YELLOW)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws10.row_dimensions[r10].height = 50
    r10 += 2

    # 1) 计算 8 维度评分
    def score_market_volume():
        if total_sales >= 200000: return 5, f'月总销量 {total_sales:,} 件，体量大'
        if total_sales >= 100000: return 4, f'月总销量 {total_sales:,} 件，体量良好'
        if total_sales >= 50000: return 3, f'月总销量 {total_sales:,} 件，体量中等'
        if total_sales >= 20000: return 2, f'月总销量 {total_sales:,} 件，体量偏小'
        return 1, f'月总销量 {total_sales:,} 件，体量有限'

    def score_demand_trend():
        if not market_data.get('_available'):
            return 3, '缺少 Market 文件，按中性评估'
        sell_df2 = market_data.get('sell_trends')
        if sell_df2 is None or len(sell_df2) < 13:
            return 3, '销售趋势数据不足'
        try:
            s = pd.to_numeric(sell_df2.iloc[:, 1], errors='coerce').dropna()
            if len(s) < 13:
                return 3, '销售趋势数据不足'
            recent = s.iloc[-12:].sum()
            prev = s.iloc[-24:-12].sum() if len(s) >= 24 else s.iloc[:-12].sum() * (12 / max(len(s)-12, 1))
            yoy = (recent - prev) / prev if prev > 0 else 0
            if yoy > 0.2: return 5, f'近 12 月销量同比 +{yoy:.0%}，高速增长'
            if yoy > 0.05: return 4, f'同比 +{yoy:.0%}，稳步增长'
            if yoy > -0.05: return 3, f'同比 {yoy:+.0%}，基本持平'
            if yoy > -0.15: return 2, f'同比 {yoy:+.0%}，小幅下滑'
            return 1, f'同比 {yoy:+.0%}，明显下滑'
        except Exception:
            return 3, '趋势计算失败'

    def score_competition():
        if cr5 is None:
            # 用 BSR 品牌数兜底
            if brand_count < 30: return 2, f'{brand_count} 品牌在争，竞争集中'
            if brand_count < 50: return 3, f'{brand_count} 品牌，中等竞争'
            return 4, f'{brand_count} 品牌分散，新品有机会'
        # 四档阈值与全局统一：≥80% 极高度集中（1 分极难）/ 50-80% 高度集中（2 分难）
        # / 30-50% 中度集中（3 分中等）/ <30% 分散（4 分易）
        if cr5 >= 0.8: return 1, f'CR5={cr5:.0%}，极高度集中（寡头垄断）'
        if cr5 >= 0.5: return 2, f'CR5={cr5:.0%}，头部高度集中'
        if cr5 >= 0.3: return 3, f'CR5={cr5:.0%}，中度集中'
        return 4, f'CR5={cr5:.0%}，竞争分散'

    def score_margin():
        # 结构净利率口径（跟 Sheet 4 入场利润测算一致）：
        # 真实净利率 = BSR Gross Margin（已扣佣金+FBA）- 广告 - 采购 - 税
        # 跨品类参数（行业普遍参考，跟 Sheet 4 commission/ad/purchase/tax 比例一致）
        _AD = 0.12
        _PURCHASE = 0.50  # 采购中位（45%~55% 取中）
        _TAX = 0.075
        if 'Gross Margin' in df.columns:
            try:
                gm_series = pd.to_numeric(df['Gross Margin'], errors='coerce').dropna()
                gm_bsr = gm_series.mean()
                net = gm_bsr - _AD - _PURCHASE - _TAX  # 结构净利率
                gross_label = f'BSR 毛利 {gm_bsr:.0%}'
                net_label = f'结构净利率 {net:.0%}（扣广告/采购/税）'
                if net > 0.20: return 5, f'{gross_label}，{net_label}，空间大'
                if net > 0.10: return 4, f'{gross_label}，{net_label}，健康'
                if net > 0:    return 3, f'{gross_label}，{net_label}，保本以上'
                if net > -0.10: return 2, f'{gross_label}，{net_label}，需小幅压采购'
                return 1, f'{gross_label}，{net_label}，必须压采购才能盈利（详见 Sheet 4）'
            except Exception:
                pass
        return 3, '毛利数据缺失，按中性评估'

    # 供应链维度：分数走启发式规则（只识别合规密集的食品/医疗/药品），
    # 文案优先取 Synthesizer LLM 的"供应链"维度理由（Sheet 10「一、推荐理由」那段也会用到）；
    # LLM 不可用时回退到基于 compliance 认证数量的事实性叙述（所有数据都来自源文件/Pack，不写空字符串）。
    def _supply_chain_llm_reason() -> str:
        if packs is None or not packs.is_synthesis_real():
            return ''
        for dr in packs.synthesis.sheet10_final_verdict.dimension_reasons or []:
            if dr.dimension == '供应链' and (dr.reason_with_evidence or '').strip():
                return dr.reason_with_evidence
        return ''

    def score_supply_chain():
        t = titles_all
        if 'food' in t or 'medical' in t or 'pharma' in t:
            base_sc, base_reason = 2, '合规密集品类，供应链门槛高'
        else:
            base_sc, base_reason = 3, ''
        llm_reason = _supply_chain_llm_reason()
        if llm_reason:
            return base_sc, llm_reason
        if base_reason:
            return base_sc, base_reason
        # 基于 compliance 认证数量的事实兜底
        _certs = compliance_certs(packs)
        _mandatory = [c['name'] for c in _certs if c.get('mandatory')]
        if _mandatory:
            return base_sc, f'涉及 {len(_mandatory)} 项强制认证（{"、".join(_mandatory[:3])}），需提前排查工厂资质'
        return base_sc, '中性评估（无品类特异信号）'

    def score_promotion():
        if difficulty_level == '低': return 4, '新品推广难度低，进入期 3-6 月'
        if difficulty_level == '中': return 3, '推广难度中等，进入期 6-9 月'
        return 2, '推广难度高，需 9-12 月周期'

    def score_risk():
        if total_hits <= 1: return 5, f'仅 {total_hits} 项中高风险命中，风险可控'
        if total_hits <= 2: return 4, f'{total_hits} 项中高风险，需逐项准备'
        if total_hits <= 4: return 3, f'{total_hits} 项中高风险，合规成本上升'
        return 2, f'{total_hits} 项中高风险，需法务/认证前置'

    def score_diff():
        try:
            if price_col:
                prices = pd.to_numeric(df[price_col], errors='coerce').dropna()
                if len(prices) > 5:
                    cv = prices.std() / prices.mean() if prices.mean() else 0
                    if cv > 0.5: return 4, f'价格变异系数 {cv:.2f}，产品差异化空间大'
                    if cv > 0.3: return 3, f'价格变异系数 {cv:.2f}，差异化可做'
                    return 2, f'价格变异系数 {cv:.2f}，同质化严重'
        except Exception:
            pass
        return 3, '差异化数据缺失，按中性评估'

    dims = [
        ('市场体量', score_market_volume(), 0.15),
        ('需求趋势', score_demand_trend(), 0.15),
        ('竞争难度', score_competition(), 0.15),
        ('利润率', score_margin(), 0.15),
        ('供应链', score_supply_chain(), 0.10),
        ('推广压力', score_promotion(), 0.10),
        ('风险可控性', score_risk(), 0.10),
        ('差异化机会', score_diff(), 0.10),
    ]
    total_score = sum(s[0] * w for _, s, w in dims)
    if total_score >= 4.2:
        overall_label = '强烈推荐'
        overall_bg = C_GREEN_LIGHT
    elif total_score >= 3.5:
        overall_label = '推荐'
        overall_bg = C_GREEN_LIGHT
    elif total_score >= 2.8:
        overall_label = '可尝试'
        overall_bg = C_YELLOW
    elif total_score >= 2.0:
        overall_label = '谨慎进入'
        overall_bg = C_YELLOW
    else:
        overall_label = '不推荐'
        overall_bg = C_RED_LIGHT
    stars_full = int(round(total_score))
    stars_str = '★' * stars_full + '☆' * (5 - stars_full)
    def star_label(sc):
        return '★' * sc + '☆' * (5 - sc)

    # ---- 各维度数据汇总 ----
    # 中国卖家占比直接用全局统一值，不再读 Market 文件返回的 cn_seller_ratio
    cn_seller_txt = f'中国卖家占比 {cn_pct_total:.0%}'
    cr5_txt2 = f'CR5={cr5:.0%}' if cr5 is not None else f'Top5 占 {brand_agg.head(5)["total_rev"].sum()/total_rev*100:.0f}%' if total_rev else '-'
    neg_top3 = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True)[:3]
    neg_top3_txt = '、'.join([f'{cat}({int(cnt)}条)' for cat, cnt in neg_top3 if cnt > 0])
    peak_months_txt = '、'.join([f'{m}月' for m in sorted(set(seasonality_peak_months))]) if seasonality_peak_months else '无明显旺季'
    kw_summary_txt = f'{", ".join(kw_top3_for_conclusion)}' if kw_top3_for_conclusion else '（需关键词文件）'
    gm_val = gm_mean / 100 if gm_mean else 0
    promo_cycle = '3-6' if difficulty_level == '低' else ('6-9' if difficulty_level == '中' else '9-12')

    # ▌ 一、推荐理由总结
    section_title(ws10, r10, 1, '▌ 一、推荐理由总结', span=8)
    ws10.row_dimensions[r10].height = 24
    r10 += 1
    hdr(ws10, r10, 1, '分析维度', bg=C_BLUE_MID)
    ws10.merge_cells(start_row=r10, start_column=2, end_row=r10, end_column=8)
    hdr(ws10, r10, 2, '数据支撑与分析结论', bg=C_BLUE_MID)
    ws10.row_dimensions[r10].height = 20
    r10 += 1
    reason_start = r10
    # 八维推荐理由（与下方"二、综合评级"的八维评分一一对齐）：
    # - 前 4 维（市场体量/需求趋势/竞争难度/利润率）由 Python 从源数据直接算句子，无 LLM 参与
    # - 后 4 维（供应链/推广压力/风险可控性/差异化机会）优先取 Synthesizer LLM 输出；
    #   LLM 不可用时回落到本地的启发式兜底句，不再写 "LLM 不可用，无叙述" 这类废话
    _diff_pain = neg_top3_txt or '—'
    _default_reason_rows = [
        ('市场体量', f'BSR TOP100 月总销量 {total_sales:,} 件，月销售额 ${total_rev:,.0f}。{sell_summary}'),
        ('需求趋势', f'生命周期：{lifecycle_stage} —— {lifecycle_reason}；旺季 {peak_months_txt}'),
        ('竞争难度', f'{brand_count} 个品牌，{cr5_txt2}。新品(<1年)占比 {new_ratio_top100:.1%}' + (f'。{cn_seller_txt}' if cn_seller_txt else '')),
        ('利润率',
            (lambda _g, _net: (
                f'均价 ${avg_price:.0f}，BSR 毛利均值 {_g:.1%}（仅扣佣金+FBA），'
                f'结构净利率 {_net:.1%}（扣广告/采购/税）。'
                + ("可支撑推广投入" if _net > 0.10 else
                   ("利润健康" if _net > 0.05 else
                    ("利润偏薄，需小幅压采购" if _net > 0 else
                     "按行业参考成本算下来亏损，必须压采购（详见 Sheet 4「盈亏平衡采购成本上限」）")))
            ))(gm_val, gm_val - 0.12 - 0.50 - 0.075)),
        ('供应链', _build_supply_chain_fallback_text(packs, difficulty_level)),
        ('推广压力', _build_promotion_pressure_fallback_text(packs, kw_summary_txt, kw_lowest_spr_for_conclusion)),
        ('风险可控性', f'{total_hits} 项中高风险命中。需完成认证排查 + 专利检索'),
        ('差异化机会', f'差评 TOP3 痛点：{_diff_pain}' + ('。从痛点切入做差异化改进' if _diff_pain != '—' else '')),
    ]
    # 混合模式：前 4 维 Python / 后 4 维 LLM
    reason_rows = sheet10_dimension_reasons(packs, _default_reason_rows)
    for label, text in reason_rows:
        if not (text or '').strip():
            continue  # 空内容直接跳过，不写占位行
        val(ws10, r10, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        ws10.merge_cells(start_row=r10, start_column=2, end_row=r10, end_column=8)
        val(ws10, r10, 2, text, bg=C_WHITE, fg='FF595959')
        ws10.row_dimensions[r10].height = 28
        r10 += 1
    apply_border(ws10, reason_start-1, r10-1, 1, 8)
    r10 += 1

    # ▌ 二、综合评级
    section_title(ws10, r10, 1, '▌ 二、综合评级', span=8)
    ws10.row_dimensions[r10].height = 24
    r10 += 1
    ws10.merge_cells(start_row=r10, start_column=1, end_row=r10, end_column=2)
    val(ws10, r10, 1, stars_str, bold=True, bg=overall_bg, size=14, h_align='center')
    val(ws10, r10, 3, f'{total_score:.2f} / 5.0', bold=True, bg=overall_bg, size=14, h_align='center')
    ws10.merge_cells(start_row=r10, start_column=4, end_row=r10, end_column=8)
    val(ws10, r10, 4, f'{overall_label}', bold=True, bg=overall_bg, size=14, h_align='center')
    ws10.row_dimensions[r10].height = 32
    r10 += 1

    hdr(ws10, r10, 1, '维度', bg=C_BLUE_MID)
    hdr(ws10, r10, 2, '评分', bg=C_BLUE_MID)
    hdr(ws10, r10, 3, '星级', bg=C_BLUE_MID)
    hdr(ws10, r10, 4, '权重', bg=C_BLUE_MID)
    ws10.merge_cells(start_row=r10, start_column=5, end_row=r10, end_column=8)
    hdr(ws10, r10, 5, '评分依据', bg=C_BLUE_MID)
    r10 += 1
    dim_start = r10
    for name, (sc, reason), weight in dims:
        bg = C_GREEN_LIGHT if sc >= 4 else (C_YELLOW if sc >= 3 else C_RED_LIGHT)
        val(ws10, r10, 1, name, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        val(ws10, r10, 2, f'{sc} / 5', bold=True, bg=bg, h_align='center')
        val(ws10, r10, 3, star_label(sc), bg=bg, h_align='center')
        val(ws10, r10, 4, f'{weight:.0%}', bg=C_WHITE, h_align='center')
        ws10.merge_cells(start_row=r10, start_column=5, end_row=r10, end_column=8)
        val(ws10, r10, 5, reason, bg=C_WHITE, fg='FF595959')
        ws10.row_dimensions[r10].height = 22
        r10 += 1
    apply_border(ws10, dim_start-1, r10-1, 1, 8)
    r10 += 1

    # 预计算推荐开发参数（供 ▌三 和 ▌五 复用）
    # 全品类通用优先：从 LLM SpecAnalyzer 识别的本品类维度 + Python 正则抓值 + 命中率 ≥10% 过滤；
    # LLM 不可用或 SpecPack 降级时，回退到老 LED 专属 aggregate_recommended_specs。
    # product_type=_top_segment_name：仅用首推品类对应的同类型 ASIN 聚合参数，避免跨类型污染；
    # 同类型样本数 < 10 时函数内部自动退回全量聚合，basis 文案会注明样本范围。
    rec_specs = aggregate_recommended_specs_from_spec_pack(
        df, packs, neg_counts, product_type=_top_segment_name
    )
    if rec_specs is None:
        rec_specs = aggregate_recommended_specs(all_specs_for_agg, neg_counts)

    # ▌ 三、首推入场品类（独立模块）
    section_title(ws10, r10, 1, '▌ 三、首推入场品类', span=8, bg=C_BLUE_DARK)
    ws10.row_dimensions[r10].height = 24
    r10 += 1
    # 首批备货量 & 月销预估：默认值在 pricing_recommendations 缺失时也能给行动建议用
    est_low, est_high = 300, 1200
    _stock_txt = '300-500 pcs'
    if pricing_recommendations:
        # 必须用 _top_segment_name 对应的那条（统一 P1），而不是 pricing_recommendations[0]（type_agg 迭代序）
        top_rec = next(
            (p for p in pricing_recommendations if p.get('product_type') == _top_segment_name),
            pricing_recommendations[0]
        )
        est_low = int(top_rec.get('avg_sales', 0) * 0.6) if top_rec.get('avg_sales') else 300
        est_high = int(top_rec.get('avg_sales', 0) * 1.2) if top_rec.get('avg_sales') else 1200
        # 首批备货量：覆盖 15-20 天销量，按 est_low*0.5 - est_high*0.3 推导
        _stock_low = max(100, int(est_low * 0.5))
        _stock_high = max(_stock_low + 50, int(est_high * 0.3))
        _stock_txt = f'{_stock_low}-{_stock_high} pcs'
        entry_fields = [
            ('首推品类', f'{_top_segment_name}', C_RED_LIGHT),
            ('建议价格区间', f'${top_rec["rec_min"]:.0f}-${top_rec["rec_max"]:.0f}', C_WHITE),
            ('目标月销量', f'{est_low}-{est_high}', C_WHITE),
            ('首批备货量', _stock_txt, C_WHITE),
            ('预计回款周期', f'{promo_cycle} 个月', C_WHITE),
        ]
        entry_start = r10
        for label, value, bg in entry_fields:
            val(ws10, r10, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
            ws10.merge_cells(start_row=r10, start_column=2, end_row=r10, end_column=8)
            val(ws10, r10, 2, value, bold=True, bg=bg, fg='FF7B0000' if bg == C_RED_LIGHT else 'FF1F3864')
            ws10.row_dimensions[r10].height = 24
            r10 += 1
        apply_border(ws10, entry_start-1, r10-1, 1, 8)

        # 评分逻辑注解
        ws10.merge_cells(start_row=r10, start_column=1, end_row=r10, end_column=8)
        c = ws10.cell(row=r10, column=1)
        c.value = '注：首推品类由综合评分自动排序 = 需求(月销)×0.30 + 单品收益×0.25 + 新品占比×0.15 + 质量评分×0.15 + 竞争度×0.15（数据驱动，非固定）'
        c.font = Font(name='Arial', size=9, color='FF7A4F01', italic=True)
        c.fill = PatternFill('solid', fgColor=C_YELLOW)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws10.row_dimensions[r10].height = 16
        r10 += 2

        # 子模块A：推荐功能参数
        if rec_specs:
            ws10.merge_cells(start_row=r10, start_column=1, end_row=r10, end_column=8)
            c = ws10.cell(row=r10, column=1)
            c.value = '—— 推荐功能参数 ——'
            c.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
            c.fill = PatternFill('solid', fgColor='FFD9E2F3')
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws10.row_dimensions[r10].height = 22
            r10 += 1

            hdr(ws10, r10, 1, '参数名称', bg=C_BLUE_MID)
            ws10.merge_cells(start_row=r10, start_column=2, end_row=r10, end_column=3)
            hdr(ws10, r10, 2, '建议规格', bg=C_BLUE_MID)
            ws10.merge_cells(start_row=r10, start_column=4, end_row=r10, end_column=6)
            hdr(ws10, r10, 4, '数据依据', bg=C_BLUE_MID)
            ws10.merge_cells(start_row=r10, start_column=7, end_row=r10, end_column=8)
            hdr(ws10, r10, 7, '优先级', bg=C_BLUE_MID)
            ws10.row_dimensions[r10].height = 20
            r10 += 1
            param_start = r10

            for param_name, rec_value, basis, priority in rec_specs:
                if 'P1' in priority:
                    p_bg = C_RED_LIGHT
                    p_fg = 'FF7B0000'
                elif 'P2' in priority:
                    p_bg = C_YELLOW
                    p_fg = 'FF7A4F01'
                else:
                    p_bg = C_BLUE_LIGHT
                    p_fg = 'FF1F3864'
                val(ws10, r10, 1, param_name, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
                ws10.merge_cells(start_row=r10, start_column=2, end_row=r10, end_column=3)
                val(ws10, r10, 2, rec_value, bold=True, bg=C_WHITE, fg='FF000000')
                ws10.merge_cells(start_row=r10, start_column=4, end_row=r10, end_column=6)
                val(ws10, r10, 4, basis, bg=C_WHITE, fg='FF595959', size=9)
                ws10.merge_cells(start_row=r10, start_column=7, end_row=r10, end_column=8)
                val(ws10, r10, 7, priority, bold=True, bg=p_bg, fg=p_fg, h_align='center')
                ws10.row_dimensions[r10].height = 24
                r10 += 1

            apply_border(ws10, param_start-1, r10-1, 1, 8)
            r10 += 1

        # 子模块B：差异化升级方向
        # 优先来源：VOC.pain_clusters（LLM 按本品类评论聚类，带 frequency_pct，比 neg_counts 更准）
        # neg_counts 的 key 在之前版本是 LED 硬编码（"电池/充电问题"等），跨品类时对不上 → 所有条数显示 0
        # 现在 neg_counts 的 key 已改为来自 VOC.praise_clusters/pain_clusters 的真实聚类名，但仍以 VOC 优先
        diff_lines = []
        _dif_pairs: list[tuple[str, int]] = []
        if packs is not None and packs.is_voc_real() and packs.voc.pain_clusters:
            _total_low_rev = max(len(low_rev), 1) if 'low_rev' in dir() else 100
            for _pc in packs.voc.pain_clusters[:3]:
                _cnt = int(round(_pc.frequency_pct * _total_low_rev / 100)) if _pc.frequency_pct <= 100 else int(_pc.frequency_pct)
                if _cnt <= 0:
                    _cnt = int(neg_counts.get(_pc.name, 0)) or 0
                _dif_pairs.append((_pc.name, _cnt))
        if not _dif_pairs and neg_counts:
            _dif_pairs = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True)[:3]

        if len(_dif_pairs) >= 1:
            t1 = _dif_pairs[0]
            diff_lines.append(f'① 首要差异化：针对"{t1[0]}"（{int(t1[1])}条差评），重点突破竞品短板')
        if len(_dif_pairs) >= 2:
            t2 = _dif_pairs[1]
            diff_lines.append(f'② 次要差异化：改善"{t2[0]}"（{int(t2[1])}条差评），提升用户体验')
        if len(_dif_pairs) >= 3:
            t3 = _dif_pairs[2]
            diff_lines.append(f'③ 加分项：优化"{t3[0]}"（{int(t3[1])}条差评），建立口碑优势')
        if diff_lines:
            ws10.merge_cells(start_row=r10, start_column=1, end_row=r10, end_column=8)
            c = ws10.cell(row=r10, column=1)
            c.value = '差异化升级方向：\n' + '\n'.join(diff_lines)
            c.font = Font(name='Arial', size=10, color='FF1F3864')
            c.fill = PatternFill('solid', fgColor='FFD9E2F3')
            c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws10.row_dimensions[r10].height = 20 + len(diff_lines) * 18
            r10 += 1

    r10 += 1

    # ▌ 四、行动建议（精简，不重复数据）
    section_title(ws10, r10, 1, '▌ 四、行动建议', span=8, bg=C_BLUE_DARK)
    ws10.row_dimensions[r10].height = 24
    r10 += 1
    # 认证清单：动态从 CompliancePack 拼具体认证名（避免写死 "UL/FCC/UN38.3" 在非电池品类串味）
    _mand_cert_names = []
    if packs is not None and packs.is_compliance_real():
        _mand_cert_names = [c.name for c in packs.compliance.required_certifications
                            if c.mandatory][:3]
    _cert_str = '、'.join(_mand_cert_names) if _mand_cert_names else '按品类选取 UL/FCC/CE/CPSIA 等适用认证'
    advice_rows = [
        ('【切入策略】', f'从差评痛点切入做差异化，初期低价冲排名，积累评价后提价至 ${avg_price*0.85:.0f}-${avg_price*1.05:.0f}'),
        ('【推广节奏】', f'前 {promo_cycle.split("-")[0]} 月广告跑量 + 优惠券，重点投放 SPR 低的长尾词；评分需达 {avg_rating:.1f}★ 门槛'),
        ('【风险前置】', f'上新前完成认证排查（{_cert_str}）+ 专利检索，避免 Listing 下架'),
        ('【备货节奏】', f'旺季 {peak_months_txt}，提前 1-2 个月到仓；首批 {_stock_txt} 控风险'),
    ]
    adv_start = r10
    for label, text in advice_rows:
        ws10.merge_cells(start_row=r10, start_column=1, end_row=r10, end_column=2)
        val(ws10, r10, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
        ws10.merge_cells(start_row=r10, start_column=3, end_row=r10, end_column=8)
        val(ws10, r10, 3, text, bg=C_WHITE)
        ws10.row_dimensions[r10].height = 28
        r10 += 1
    apply_border(ws10, adv_start-1, r10-1, 1, 8)
    r10 += 1

    # ▌ 五、LLM 综合升级建议（仅当 Synthesizer 实际成功时呈现）
    _upgrade = upgrade_directions(packs)
    _matrix = sheet6_priority_matrix(packs)
    # 渲染前最后一道 scrub：清掉 LLM 偶尔泄露在 justification / action_plan 里的 schema 字段路径
    from core.packs_runtime import _scrub_schema_paths as _scrub_paths
    if _upgrade:
        for _u in _upgrade:
            _u['justification'] = _scrub_paths(_u.get('justification', ''))
            _u['target_spec'] = _scrub_paths(_u.get('target_spec', ''))
    if _matrix:
        for _m in _matrix:
            _m['action_plan'] = _scrub_paths(_m.get('action_plan', ''))
            _m['improvements'] = [_scrub_paths(s) for s in (_m.get('improvements') or [])]
    if _upgrade or _matrix:
        section_title(ws10, r10, 1, '▌ 五、LLM 综合升级建议（基于 4 份源数据深度分析）', span=8, bg=C_ORANGE)
        ws10.row_dimensions[r10].height = 24
        r10 += 1
        if _upgrade:
            hdr(ws10, r10, 1, '升级维度', bg=C_BLUE_MID)
            hdr(ws10, r10, 2, '目标规格', bg=C_BLUE_MID)
            ws10.merge_cells(start_row=r10, start_column=3, end_row=r10, end_column=8)
            hdr(ws10, r10, 3, '升级依据（含数据引用）', bg=C_BLUE_MID)
            r10 += 1
            up_start = r10
            for u in _upgrade[:6]:
                val(ws10, r10, 1, u['dimension'], bold=True, bg=C_BLUE_LIGHT)
                val(ws10, r10, 2, u['target_spec'], bg=C_WHITE)
                ws10.merge_cells(start_row=r10, start_column=3, end_row=r10, end_column=8)
                val(ws10, r10, 3, u['justification'], bg=C_WHITE)
                ws10.row_dimensions[r10].height = 32
                r10 += 1
            apply_border(ws10, up_start-1, r10-1, 1, 8)
            r10 += 1
        if _matrix:
            section_title(ws10, r10, 1, '▌ 上新优先级矩阵（基于 LLM 综合分析）', span=8)
            ws10.row_dimensions[r10].height = 22
            r10 += 1
            hdr(ws10, r10, 1, '细分', bg=C_BLUE_MID)
            hdr(ws10, r10, 2, '优先级', bg=C_BLUE_MID)
            ws10.merge_cells(start_row=r10, start_column=3, end_row=r10, end_column=5)
            hdr(ws10, r10, 3, '行动计划', bg=C_BLUE_MID)
            ws10.merge_cells(start_row=r10, start_column=6, end_row=r10, end_column=8)
            hdr(ws10, r10, 6, '改进点', bg=C_BLUE_MID)
            r10 += 1
            mtx_start = r10
            for item in _matrix[:6]:
                pri = item['priority']
                pri_bg = C_GREEN_LIGHT if pri == 'P1' else (C_YELLOW if pri == 'P2' else C_WHITE)
                val(ws10, r10, 1, item['segment'], bold=True, bg=C_BLUE_LIGHT)
                val(ws10, r10, 2, pri, bold=True, bg=pri_bg, h_align='center')
                ws10.merge_cells(start_row=r10, start_column=3, end_row=r10, end_column=5)
                val(ws10, r10, 3, item['action_plan'], bg=C_WHITE)
                ws10.merge_cells(start_row=r10, start_column=6, end_row=r10, end_column=8)
                val(ws10, r10, 6, '\n'.join(item['improvements'][:4]), bg=C_WHITE)
                ws10.row_dimensions[r10].height = 24 + len(item['improvements']) * 14
                r10 += 1
            apply_border(ws10, mtx_start-1, r10-1, 1, 8)
            r10 += 1
    elif packs is not None and not packs.is_synthesis_real():
        ws10.merge_cells(start_row=r10, start_column=1, end_row=r10, end_column=8)
        val(ws10, r10, 1, '【提示】综合策略由 LLM 生成失败，本节使用降级模板。',
            bg=C_GREY_LIGHT if 'C_GREY_LIGHT' in dir() else 'FFEEEEEE', fg='FF999999')
        r10 += 1

    # ===== 保存 =====
    wb.save(output_path)
    return output_path


# ============================================================
# Flask 路由
# ============================================================
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    if 'bsr_file' not in request.files:
        flash('请上传BSR数据文件', 'error')
        return redirect(url_for('index'))

    bsr_file = request.files['bsr_file']
    if bsr_file.filename == '':
        flash('请上传BSR数据文件', 'error')
        return redirect(url_for('index'))

    if not allowed_file(bsr_file.filename):
        flash('BSR文件必须是Excel格式(.xlsx或.xls)', 'error')
        return redirect(url_for('index'))

    review_files = request.files.getlist('review_files')
    review_files = [f for f in review_files if f.filename and allowed_file(f.filename)]

    market_file = request.files.get('market_file')
    market_path = None

    # 前端可在 form 字段传 session_id，让 /status 轮询与本次请求一一对应（双标签并发时不串扰）；
    # 缺省则后端生成。仅接受 [a-zA-Z0-9]{4,32} 防止路径注入。
    client_sid = (request.form.get('session_id') or '').strip()
    if client_sid and len(client_sid) <= 32 and all(c.isalnum() for c in client_sid):
        session_id = client_sid
    else:
        session_id = str(uuid.uuid4())[:8]
    temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    os.makedirs(temp_dir, exist_ok=True)

    try:
        bsr_filename = secure_filename(bsr_file.filename)
        bsr_path = os.path.join(temp_dir, bsr_filename)
        bsr_file.save(bsr_path)

        review_paths = []
        for f in review_files:
            if f.filename:
                filename = secure_filename(f.filename)
                path = os.path.join(temp_dir, filename)
                f.save(path)
                review_paths.append(path)

        if market_file and market_file.filename and allowed_file(market_file.filename):
            market_filename = secure_filename(market_file.filename)
            market_path = os.path.join(temp_dir, market_filename)
            market_file.save(market_path)

        keyword_file = request.files.get('keyword_file')
        keyword_path = None
        if keyword_file and keyword_file.filename and allowed_file(keyword_file.filename):
            keyword_filename = secure_filename(keyword_file.filename)
            keyword_path = os.path.join(temp_dir, keyword_filename)
            keyword_file.save(keyword_path)

        report_filename = f'选品评估报告_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{session_id}.xlsx'
        report_path = os.path.join(app.config['REPORT_FOLDER'], report_filename)

        # 写入生成状态（前端轮询检测）。文件名按 session_id 隔离，支持多浏览器标签并发。
        status_file = os.path.join(app.config['REPORT_FOLDER'], f'_status_{session_id}.txt')
        def _write_status(msg: str) -> None:
            with open(status_file, 'w', encoding='utf-8') as sf:
                sf.write(msg)

        _write_status(f'phase:读取文件 ({len(review_paths)} 份评论)')

        # 从 key pool 为本次请求分配一个 api_key，让 N 个浏览器标签并发跑时各自独占一把 key
        # （pool 空 → None → 各 analyzer fallback 到环境变量 DASHSCOPE_API_KEY，行为同从前）
        from llm.key_pool import get_pool
        _request_api_key = get_pool().acquire()
        if _request_api_key:
            log.warning("[upload] session=%s 分配 key 末四位 …%s", session_id, _request_api_key[-4:])

        # 先生成 packs（用于回传 LLM 状态给前端），再生成报告复用同一份 packs
        try:
            _write_status('phase:调用 LLM 分析 BSR / 评论 / 关键词 / Market（并发）')
            # 用户重传 = 重跑：清掉文本类 LLM 缓存，保留 vision_classify_*（同图分类不变且冷启动 60-180s）。
            # LLM_CACHE_REUSE=1 dev 模式下复用所有缓存，省 LLM 钱方便迭代 Excel 渲染。
            if os.environ.get('LLM_CACHE_REUSE', '').lower() in ('1', 'true', 'yes'):
                print('[upload] LLM_CACHE_REUSE=1，跳过清缓存（dev 模式：复用所有缓存）')
            else:
                try:
                    from llm.cache import LLMCache as _LLMCache
                    _purged = _LLMCache(os.getenv('LLM_CACHE_DIR', 'llm_cache')).purge_except('vision_classify_')
                    if _purged > 0:
                        print(f'[upload] 清掉 {_purged} 份非 vision 缓存（确保本次重跑生成全新分析）')
                except Exception as _ce:
                    print(f'[upload] 清缓存失败（不阻断）: {_ce}')
            packs = prepare_packs(
                bsr_path=bsr_path,
                review_paths=review_paths,
                keyword_path=keyword_path,
                market_path=market_path,
                api_key=_request_api_key,
            )
            packs_status = {
                'category_id': packs.category_id,
                'display_name': packs.display_name,
                'market_real': packs.is_market_real(),
                'voc_real': packs.is_voc_real(),
                'traffic_real': packs.is_traffic_real(),
                'trend_real': packs.is_trend_real(),
                'synthesis_real': packs.is_synthesis_real(),
            }
        except Exception as pe:
            print(f'[upload] prepare_packs 整体失败: {pe}')
            packs = None
            packs_status = {'error': str(pe)}

        _write_status('phase:生成 10-Sheet Excel 报告')
        generate_report(bsr_path, review_paths, report_path, market_path, keyword_path, packs=packs)

        shutil.rmtree(temp_dir, ignore_errors=True)

        # 写入完成状态 + 报告路径（供前端轮询获取）
        _write_status(f'done:{report_filename}')

        # 返回 JSON，前端收到后显示下载按钮
        return jsonify({
            'success': True,
            'filename': report_filename,
            'message': '报告生成完成，点击下方按钮下载',
            'packs_status': packs_status,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        # 出错时清除状态（写入对应 session 的状态文件）
        try:
            status_file = os.path.join(app.config['REPORT_FOLDER'], f'_status_{session_id}.txt')
            with open(status_file, 'w', encoding='utf-8') as sf:
                sf.write(f'error:{str(e)}')
        except:
            pass
        flash(f'生成报告时出错: {str(e)}', 'error')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/detect_category', methods=['POST'])
def detect_category():
    """前端上传 BSR 文件后，立即返回识别到的 category_id 和 display_name 供预览。
    不调 LLM，仅做文件名解析（毫秒级响应）。"""
    bsr_file = request.files.get('bsr_file')
    if not bsr_file or not bsr_file.filename:
        return jsonify({'error': '请上传 BSR 文件'}), 400
    from utils.category_id import extract_from_bsr_filename
    category_id = extract_from_bsr_filename(bsr_file.filename)
    display_name = category_id.replace('_', ' ').title()
    return jsonify({
        'category_id': category_id,
        'display_name': display_name,
        'note': '系统将根据 BSR + 评论 + 关键词 + Market 数据由 LLM 自动生成定制化分析'
    })


@app.route('/status')
def status():
    """前端轮询状态接口。需要 ?sid=<session_id>，按 session 隔离，避免多标签并发时进度互串。"""
    sid = (request.args.get('sid') or '').strip()
    if not sid or len(sid) > 32 or not all(c.isalnum() for c in sid):
        return {'status': 'idle'}
    status_file = os.path.join(app.config['REPORT_FOLDER'], f'_status_{sid}.txt')
    try:
        with open(status_file, 'r', encoding='utf-8') as f:
            content = f.read().strip()
        return {'status': content}
    except:
        return {'status': 'idle'}


@app.route('/generate_asin_list', methods=['POST'])
def generate_asin_list():
    """两步法第 1 步：仅基于 BSR + 可选 Market，规则筛 12-18 个高优先级 ASIN，
    输出"重点 ASIN 评论采集清单 .xlsx"，让运营定向去卖家精灵下载评论。
    不调 LLM，秒级返回。
    """
    if 'bsr_file' not in request.files:
        return jsonify({'success': False, 'error': '请上传 BSR 数据文件'}), 400
    bsr_file = request.files['bsr_file']
    if not bsr_file or bsr_file.filename == '':
        return jsonify({'success': False, 'error': '请上传 BSR 数据文件'}), 400
    if not allowed_file(bsr_file.filename):
        return jsonify({'success': False, 'error': 'BSR 文件必须是 Excel 格式(.xlsx 或 .xls)'}), 400

    market_file = request.files.get('market_file')

    session_id = str(uuid.uuid4())[:8]
    temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    os.makedirs(temp_dir, exist_ok=True)

    try:
        bsr_filename = secure_filename(bsr_file.filename)
        bsr_path = os.path.join(temp_dir, bsr_filename)
        bsr_file.save(bsr_path)

        market_path = None
        if market_file and market_file.filename and allowed_file(market_file.filename):
            market_filename = secure_filename(market_file.filename)
            market_path = os.path.join(temp_dir, market_filename)
            market_file.save(market_path)

        from llm.analyzers.bsr_analyzer import normalize_bsr_columns
        from core.packs_runtime import _load_market
        from core.asin_collection_planner import build_asin_collection_list
        from core.asin_collection_xlsx import write_asin_collection_xlsx
        from utils.category_id import extract_from_bsr_filename

        try:
            xl = pd.ExcelFile(bsr_path)
            target_sheet = next((s for s in xl.sheet_names if s.upper().startswith('US')), xl.sheet_names[0])
            bsr_df = pd.read_excel(bsr_path, sheet_name=target_sheet)
            bsr_df = normalize_bsr_columns(bsr_df)
        except Exception as e:
            return jsonify({'success': False, 'error': f'读取 BSR 文件失败：{e}'}), 400

        market_data = _load_market(market_path) if market_path else {}

        category_id = extract_from_bsr_filename(bsr_file.filename) or '未知品类'
        category_name = category_id.replace('_', ' ').title()

        df = build_asin_collection_list(bsr_df, market_data=market_data, top_n=15)

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_filename = f'重点ASIN评论采集清单_{category_name}_{ts}.xlsx'
        out_path = os.path.join(app.config['REPORT_FOLDER'], out_filename)
        write_asin_collection_xlsx(df, out_path, category_name)

        shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({
            'success': True,
            'filename': out_filename,
            'count': int(len(df)),
            'category_name': category_name,
            'message': f'已生成 {len(df)} 条 ASIN 采集清单',
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/download/<filename>')
def download_report(filename):
    """手动下载报告接口"""
    report_path = os.path.join(app.config['REPORT_FOLDER'], filename)
    if not os.path.exists(report_path):
        flash('报告文件不存在或已过期，请重新生成', 'error')
        return redirect(url_for('index'))
    return send_file(report_path, as_attachment=True, download_name=filename)


@app.route('/health')
def health():
    return {'status': 'ok', 'message': '选品报告生成服务运行中'}


if __name__ == '__main__':
    import os as _os
    _port = int(_os.environ.get('FLASK_PORT', '8000'))
    print(f"""
╔═══════════════════════════════════════════════════════════════╗
║  通用品类选品报告生成系统 v3.0 (LLM 驱动 · 任意品类自适应)    ║
║  访问地址: http://localhost:{_port:<5}                              ║
╚═══════════════════════════════════════════════════════════════╝
    """)
    # use_reloader=False：避免编辑 .py 时 Flask 重启杀掉正在跑的 upload 请求。
    # 仍保留 debug=True 以打印错误栈；改代码后需要手动重启服务生效。
    _debug = _os.environ.get('FLASK_DEBUG', '1') == '1'
    _reload = _os.environ.get('FLASK_RELOAD', '0') == '1'
    app.run(host='0.0.0.0', port=_port, debug=_debug, use_reloader=_reload, threaded=True)
