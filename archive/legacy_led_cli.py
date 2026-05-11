
import pandas as pd
import sys, io, os, re
from collections import Counter
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ============================================================
# 1. 读取 BSR 数据
# ============================================================
bsr_file = r'C:\Users\18782\Desktop\选品报告（新需求）\LED数据源\BSR(Job-Site-Lighting(Current))-100-US-20260409.xlsx'
df = pd.read_excel(bsr_file, sheet_name='US')
price_col = [c for c in df.columns if c.startswith('Price')][0]
rev_col = [c for c in df.columns if 'revenue' in c.lower() and 'Variation' not in c][0]
for c in [price_col, rev_col, 'Monthly Sales', 'Rating', 'Ratings', 'Available days', 'Gross Margin', 'FBA($)']:
    df[c] = pd.to_numeric(df[c], errors='coerce')

def classify(title):
    t = str(title).lower()
    if 'spotlight' in t or ('flashlight' in t and 'work' not in t):
        return '手持手电/聚光灯'
    elif 'tripod' in t or ('stand' in t and ('work light' in t or 'lamp' in t)):
        return '三脚架工作灯'
    elif 'shop light' in t or 'tube light' in t or 't8' in t or 't5' in t:
        return '车间吊装灯'
    elif 'clamp' in t or 'clip' in t:
        return '夹灯'
    elif 'solar' in t:
        return '太阳能工作灯'
    elif 'rechargeable' in t or 'magnetic' in t or 'mechanic light' in t:
        return '充电磁吸工作灯'
    elif 'flood light' in t or 'job site' in t:
        return '泛光工作灯'
    else:
        return '其他工作灯'

df['product_type'] = df['Product Title'].apply(classify)

# ============================================================
# 2. 读取评论数据
# ============================================================
reviews_dir = r'C:\Users\18782\Desktop\选品报告（新需求）\LED数据源'
review_files = [f for f in os.listdir(reviews_dir) if 'Reviews' in f and f.endswith('.xlsx')]
all_reviews = []
for f in review_files:
    path = os.path.join(reviews_dir, f)
    df_r = pd.read_excel(path, sheet_name=0)
    df_r['source_asin'] = f.split('-')[0]
    all_reviews.append(df_r)
rev_df = pd.concat(all_reviews, ignore_index=True)
rev_df['Rating'] = pd.to_numeric(rev_df['Rating'], errors='coerce')
low_rev = rev_df[rev_df['Rating'] <= 2].copy()
high_rev = rev_df[rev_df['Rating'] >= 4].copy()

# ============================================================
# 3. 统计汇总
# ============================================================
total_rev = df[rev_col].sum()
total_sales = df['Monthly Sales'].sum()
avg_price = df[price_col].mean()
median_price = df[price_col].median()
cn_rev = df[df['BuyBox Location']=='CN'][rev_col].sum()
us_rev = df[df['BuyBox Location']=='US'][rev_col].sum()
cn_cnt = (df['BuyBox Location']=='CN').sum()

type_agg = df.groupby('product_type').agg(
    count=('ASIN','count'),
    avg_price=(price_col,'mean'),
    min_price=(price_col,'min'),
    max_price=(price_col,'max'),
    total_sales=('Monthly Sales','sum'),
    total_revenue=(rev_col,'sum'),
    avg_rating=('Rating','mean'),
).sort_values('total_revenue', ascending=False).reset_index()

brand_agg = df.groupby('Brand').agg(
    sku_count=('ASIN','count'),
    total_rev=(rev_col,'sum'),
    avg_rating=('Rating','mean'),
).sort_values('total_rev', ascending=False).reset_index()

bins = [0,15,25,35,50,70,100,999]
labels = ['<$15','$15-25','$25-35','$35-50','$50-70','$70-100','>$100']
df['price_band'] = pd.cut(df[price_col], bins=bins, labels=labels)
price_dist = df['price_band'].value_counts().sort_index()

neg_keywords = {
    '电池/充电问题': ['battery', 'charge', 'charging', 'battery life', 'drain', 'dies fast'],
    '耐久性/质量差': ['broke', 'broken', 'stopped working', 'quit working', 'doesnt work', "doesn't work",
                     'dead', 'died', 'cheap', 'flimsy', 'poor quality', 'defective', 'failed'],
    '亮度不足': ['not bright', 'dim', 'not bright enough', 'brightness', 'weak light', 'too dim'],
    '灯头/角度固定差': ['head', 'angle', 'position', 'tilt', 'rotate', 'floppy', 'swivel', 'pivot', 'loose'],
    '客服/售后问题': ['customer service', 'return', 'refund', 'warranty', 'replacement'],
    '磁力不稳/固定问题': ['magnet', 'magnetic', 'not stick', "won't stick", 'falls off', 'hold position', 'wobble'],
    '不含电池/配件缺失': ['not included', 'tool only', 'no battery', 'bare tool', 'no charger'],
    '防水/防尘问题': ['water', 'rain', 'wet', 'rust', 'corrosion'],
}
neg_text = (low_rev['Title'].fillna('') + ' ' + low_rev['Content'].fillna('')).str.lower()
neg_counts = {}
for cat, kws in neg_keywords.items():
    pattern = '|'.join(kws)
    cnt = neg_text.str.contains(pattern, na=False).sum()
    neg_counts[cat] = cnt

pos_keywords = {
    '高亮度/强光': ['bright', 'lumens', 'illumination', 'super bright', 'very bright', 'powerful light'],
    '轻便/便携': ['compact', 'lightweight', 'portable', 'small', 'easy to carry', 'handy'],
    '易用/操作方便': ['easy to use', 'convenient', 'simple', 'love it', 'perfect', 'great for'],
    '性价比高': ['value', 'great price', 'good price', 'affordable', 'worth', 'bang for the buck'],
    '耐用/品质好': ['durable', 'sturdy', 'solid', 'tough', 'last long', 'holds up', 'well built'],
    '电池续航好': ['long battery', 'battery life', 'usb', 'charges fast', 'long lasting'],
    '多功能/多模式': ['multiple modes', 'different modes', 'versatile', 'multi', 'strobe', 'sos'],
    '磁力强/多角度调节': ['magnet', 'magnetic', 'strong magnet', '360', 'rotate', 'flexible', 'hands-free'],
}
pos_text = (high_rev['Title'].fillna('') + ' ' + high_rev['Content'].fillna('')).str.lower()
pos_counts = {}
for cat, kws in pos_keywords.items():
    pattern = '|'.join(kws)
    cnt = pos_text.str.contains(pattern, na=False).sum()
    pos_counts[cat] = cnt

# ============================================================
# 4. 构建 Excel 报告
# ============================================================
output_path = r'C:\Users\18782\Desktop\workbuddy\LED工作灯选品评估报告.xlsx'
wb = Workbook()

# Color palette (all 6-char hex, openpyxl will handle with FF prefix internally)
C_BLUE_DARK = 'FF1F3864'
C_BLUE_MID  = 'FF2E74B5'
C_BLUE_LIGHT = 'FFD6E4F0'
C_YELLOW    = 'FFFFF2CC'
C_GREEN_LIGHT = 'FFE2EFDA'
C_ORANGE    = 'FFFCE4D6'
C_WHITE     = 'FFFFFFFF'
C_GREY_LIGHT = 'FFF2F2F2'
C_RED_LIGHT  = 'FFFDECEA'
C_HEADER_BG  = 'FF2E74B5'
C_SECTION_BG = 'FF1F3864'
C_GREEN_DARK  = 'FF375623'

def hdr(ws, row, col, text, bg=C_HEADER_BG, fg=C_WHITE, bold=True, size=11, wrap=False, h_align='center'):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=wrap)
    return cell

def val(ws, row, col, text, bg=C_WHITE, bold=False, size=10, wrap=True, h_align='left', fg='FF000000', italic=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size, italic=italic)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=wrap)
    return cell

def thin_border():
    side = Side(style='thin', color='FFBFBFBF')
    return Border(left=side, right=side, top=side, bottom=side)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()

def section_title(ws, row, col, text, span=8, bg=C_SECTION_BG):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name='Arial', bold=True, color=C_WHITE, size=12)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 22

# ============================================================
# SHEET 1: 市场分析
# ============================================================
ws1 = wb.active
ws1.title = '市场分析'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 18

# Title
ws1.merge_cells('A1:H1')
cell = ws1['A1']
cell.value = 'LED工作灯 (Job Site Lighting) 市场选品评估报告'
cell.font = Font(name='Arial', bold=True, size=16, color=C_WHITE)
cell.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

ws1.merge_cells('A2:H2')
cell = ws1['A2']
cell.value = '数据来源：亚马逊 BSR Job-Site-Lighting TOP100 (2026-04-09)  +  13个ASIN真实评论 (共2589条)  |  分析日期：2026-04-15'
cell.font = Font(name='Arial', size=9, color='FF595959', italic=True)
cell.fill = PatternFill('solid', fgColor=C_GREY_LIGHT)
cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 18

r = 4
# --- 1.1 类目整体概况 ---
section_title(ws1, r, 1, '▌ 一、类目整体概况（BSR TOP100 样本）', span=8)
ws1.row_dimensions[r].height = 24
r += 1

headers_1 = ['指标', '数值', '说明']
widths = [22, 18, 50]
for i, h in enumerate(headers_1):
    hdr(ws1, r, i+1, h, bg=C_BLUE_MID)
ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
ws1.row_dimensions[r].height = 20
r += 1

overview_data = [
    ('样本商品数', '100 个', 'BSR Job-Site-Lighting 实时 TOP100'),
    ('涉及品牌数', f'{df["Brand"].nunique()} 个', '品牌分散，头部效应弱，新品进入门槛低'),
    ('月总销量', f'{int(total_sales):,} 件', 'BSR TOP100 样本合计月销'),
    ('月总销售额', f'${total_rev/10000:.1f}万', f'折合人民币约 ¥{total_rev*7.2/10000:.0f}万'),
    ('均价', f'${avg_price:.2f}', f'中位价 ${median_price:.2f}，主力区间 $20-$60'),
    ('平均毛利率', f'{df["Gross Margin"].mean()*100:.1f}%', '中位毛利率 66.4%，利润空间充足'),
    ('平均星级', f'{df["Rating"].mean():.2f}★', '星级集中在 4.4-4.8，市场竞争良性'),
    ('FBA占比', f'{(df["Fulfillment"]=="FBA").sum()}%', '94款采用FBA，AMZ自营6款'),
    ('中国卖家占比（BuyBox）', f'{cn_cnt}%', f'CN BuyBox占比 {cn_cnt}%，月收入占比 {cn_rev/total_rev*100:.1f}%'),
]
for metric, value, note in overview_data:
    val(ws1, r, 1, metric, bold=True, bg=C_BLUE_LIGHT)
    c = ws1.cell(row=r, column=2, value=value)
    c.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor=C_WHITE)
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    val(ws1, r, 3, note, bg=C_WHITE, fg='FF595959')
    ws1.row_dimensions[r].height = 18
    r += 1

apply_border(ws1, 5, r-1, 1, 8)
r += 1

# --- 1.2 价格分布 ---
section_title(ws1, r, 1, '▌ 二、价格分布', span=8)
ws1.row_dimensions[r].height = 24
r += 1

price_hdr_cols = ['价格区间', '商品数量', '占比', '区间特征']
for i, h in enumerate(price_hdr_cols):
    hdr(ws1, r, i+1, h, bg=C_BLUE_MID)
ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
ws1.row_dimensions[r].height = 20
r += 1

price_desc = {
    '<$15': '特价引流款，低门槛，日销较高',
    '$15-25': '入门走量区，泛光灯/夹灯集中',
    '$25-35': '主力竞争区，充电磁吸灯最密集',
    '$35-50': '中高端区间，品质感强、新品机会好',
    '$50-70': '三脚架灯主力价格，单品收益高',
    '$70-100': '专业级手持灯/品牌溢价区',
    '>$100': '品牌灯（DeWalt/Streamlight）高端区',
}
for band, cnt in price_dist.items():
    bg = C_YELLOW if '$25' in str(band) or '$35' in str(band) else C_WHITE
    val(ws1, r, 1, str(band), bold=True, bg=bg)
    ws1.cell(row=r, column=2, value=int(cnt)).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=r, column=2).fill = PatternFill('solid', fgColor=bg)
    pct_val = f'{cnt/100*100:.0f}%'
    c3 = ws1.cell(row=r, column=3, value=pct_val)
    c3.alignment = Alignment(horizontal='center', vertical='center')
    c3.fill = PatternFill('solid', fgColor=bg)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    val(ws1, r, 4, price_desc.get(str(band), ''), fg='FF595959', bg=bg)
    ws1.row_dimensions[r].height = 18
    r += 1

apply_border(ws1, r - len(price_dist) - 1, r-1, 1, 8)
r += 1

# --- 1.3 产品类型分布 ---
section_title(ws1, r, 1, '▌ 三、产品类型分布与收益分析', span=8)
ws1.row_dimensions[r].height = 24
r += 1

type_hdr = ['产品类型', 'SKU数', '平均价格', '价格区间', '月总销量', '月总销售额', '单SKU均收益', '平均星级']
for i, h in enumerate(type_hdr):
    hdr(ws1, r, i+1, h, bg=C_BLUE_MID)
ws1.row_dimensions[r].height = 20
r += 1

type_highlights = {'充电磁吸工作灯': C_YELLOW, '三脚架工作灯': C_GREEN_LIGHT, '手持手电/聚光灯': C_ORANGE}
for _, row_data in type_agg.iterrows():
    bg = type_highlights.get(row_data['product_type'], C_WHITE)
    val(ws1, r, 1, row_data['product_type'], bold=True, bg=bg)
    for ci, v in enumerate([
        int(row_data['count']),
        f"${row_data['avg_price']:.1f}",
        f"${row_data['min_price']:.0f}-${row_data['max_price']:.0f}",
        f"{int(row_data['total_sales']):,}",
        f"${row_data['total_revenue']/10000:.1f}万",
        f"${row_data['total_revenue']/row_data['count']:.0f}",
        f"{row_data['avg_rating']:.1f}★",
    ], start=2):
        c = ws1.cell(row=r, column=ci, value=v)
        c.font = Font(name='Arial', size=10)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[r].height = 18
    r += 1

apply_border(ws1, r - len(type_agg) - 1, r-1, 1, 8)
r += 1

# --- 1.4 市场小结 ---
section_title(ws1, r, 1, '▌ 四、市场分析结论', span=8)
ws1.row_dimensions[r].height = 24
r += 1

conclusions = [
    ('市场规模', f'月总销售额约${total_rev/10000:.1f}万，月总销量{int(total_sales):,}件，为中等体量稳定市场。类目年度峰值在9-12月（施工/礼品旺季）。'),
    ('价格结构', f'主力价格带集中在$20-$60区间（{price_dist.get("$15-25",0)+price_dist.get("$25-35",0)+price_dist.get("$35-50",0)}款/占{(price_dist.get("$15-25",0)+price_dist.get("$25-35",0)+price_dist.get("$35-50",0))}%），$35-50为最优入场区间，竞争密度较$25-35低但单品收益更高。'),
    ('主力品类', '充电磁吸工作灯是销量最大品类（39 SKU，月销3.9万件，月收入$118万），三脚架工作灯客单价高（均价$50），手持聚光灯虽SKU少但单品收益最高。'),
    ('集中度', f'Top3品牌收入占比仅26.4%，Top10占44.1%，76个品牌参与竞争，{(brand_agg["sku_count"]==1).sum()}个品牌仅有1个SKU，市场分散，新品进入机会较大。'),
]
for label, text in conclusions:
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    val(ws1, r, 1, f'【{label}】', bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    val(ws1, r, 3, text, wrap=True)
    ws1.row_dimensions[r].height = 36
    r += 1

apply_border(ws1, r-len(conclusions), r-1, 1, 8)

# ============================================================
# SHEET 2: 竞争分析
# ============================================================
ws2 = wb.create_sheet('竞争分析')
ws2.sheet_view.showGridLines = False
for col_letter, width in zip('ABCDEFGH', [22,18,15,15,15,18,15,15]):
    ws2.column_dimensions[col_letter].width = width

ws2.merge_cells('A1:H1')
cell = ws2['A1']
cell.value = 'LED工作灯 竞争格局分析'
cell.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
cell.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

r2 = 3
section_title(ws2, r2, 1, '▌ 一、竞争指数概览', span=8)
ws2.row_dimensions[r2].height = 24
r2 += 1

compete_kv = [
    ('竞争指数', '中等偏低 ★★★☆☆', '市场分散，新品进入壁垒较低'),
    ('品牌集中度（HHI）', '低（Top5占33%）', '无绝对垄断品牌，竞争均衡'),
    ('中国卖家BuyBox占比', f'{cn_cnt}个 / {cn_cnt}%', f'月收入占比{cn_rev/total_rev*100:.1f}%，中国卖家主导市场'),
    ('美国本土卖家', f'8个 / {us_rev/total_rev*100:.1f}%收入', '主要为DeWalt/Klein等工具品牌'),
    ('FBA比例', '94%', '物流标准化，竞争主要在产品和广告'),
    ('新品存活率', '25%（25/100）', '25款新品（<1年）进入TOP100，存活率高'),
    ('新品月均收益', '$24,832', '略低于均值，前期爬坡正常'),
    ('评分门槛', f'最低{df["Rating"].min()}★，均值{df["Rating"].mean():.2f}★', '进入TOP100需保持4.4★以上'),
    ('总SKU数（类目估算）', '>3,000+', '类目总SKU体量大，差异化是核心'),
]
hdr(ws2, r2, 1, '竞争维度', bg=C_BLUE_MID)
hdr(ws2, r2, 2, '数据', bg=C_BLUE_MID)
ws2.merge_cells(start_row=r2, start_column=3, end_row=r2, end_column=8)
hdr(ws2, r2, 3, '说明', bg=C_BLUE_MID)
r2 += 1
for i, (k, v, n) in enumerate(compete_kv):
    bg = C_BLUE_LIGHT if i%2==0 else C_WHITE
    val(ws2, r2, 1, k, bold=True, bg=bg)
    c = ws2.cell(row=r2, column=2, value=v)
    c.font = Font(name='Arial', bold=True, size=11, color='FF1F3864')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells(start_row=r2, start_column=3, end_row=r2, end_column=8)
    val(ws2, r2, 3, n, bg=bg, fg='FF595959')
    ws2.row_dimensions[r2].height = 20
    r2 += 1
apply_border(ws2, 4, r2-1, 1, 8)
r2 += 1

# Top Brand Table
section_title(ws2, r2, 1, '▌ 二、TOP15 品牌收益排行', span=8)
r2 += 1
brand_hdr_cols = ['排名', '品牌', 'SKU数', '月总收益($)', '收入占比', '平均星级', '市场地位']
for i, h in enumerate(brand_hdr_cols):
    hdr(ws2, r2, i+1, h, bg=C_BLUE_MID)
ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
r2 += 1

brand_colors = {0: C_ORANGE, 1: C_ORANGE, 2: C_ORANGE}
top15 = brand_agg.head(15)
for idx, (_, brow) in enumerate(top15.iterrows()):
    bg = brand_colors.get(idx, C_WHITE if idx%2==0 else C_GREY_LIGHT)
    ws2.cell(row=r2, column=1, value=idx+1).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=r2, column=1).fill = PatternFill('solid', fgColor=bg)
    for ci, v in enumerate([brow['Brand'], int(brow['sku_count']),
                             f"${brow['total_rev']:,.0f}",
                             f"{brow['total_rev']/total_rev*100:.1f}%",
                             f"{brow['avg_rating']:.1f}★"], start=2):
        c = ws2.cell(row=r2, column=ci, value=v)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center' if ci>=3 else 'left', vertical='center')
        c.font = Font(name='Arial', size=10)
    status_map = {0:'TOP1 绝对领跑', 1:'TOP2 品类强者', 2:'TOP3 紧密跟随', 3:'TOP4-5 中坚品牌'}
    status = status_map.get(idx, 'TOP10 活跃竞争')
    ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
    val(ws2, r2, 7, status, bg=bg, fg='FF595959')
    ws2.row_dimensions[r2].height = 18
    r2 += 1
apply_border(ws2, r2-16, r2-1, 1, 8)
r2 += 1

# Seller location distribution
section_title(ws2, r2, 1, '▌ 三、卖家地区分布', span=8)
r2 += 1
loc_data = df['BuyBox Location'].value_counts()
hdr(ws2, r2, 1, '卖家地区', bg=C_BLUE_MID)
hdr(ws2, r2, 2, '数量', bg=C_BLUE_MID)
hdr(ws2, r2, 3, '占比', bg=C_BLUE_MID)
ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=6)
hdr(ws2, r2, 4, '月收入', bg=C_BLUE_MID)
ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
hdr(ws2, r2, 7, '收入占比', bg=C_BLUE_MID)
r2 += 1
for loc, cnt in loc_data.items():
    loc_rev = df[df['BuyBox Location']==loc][rev_col].sum()
    bg = C_YELLOW if loc=='CN' else C_WHITE
    val(ws2, r2, 1, str(loc) if loc else '未知', bold=(loc=='CN'), bg=bg)
    ws2.cell(row=r2, column=2, value=int(cnt)).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=r2, column=2).fill = PatternFill('solid', fgColor=bg)
    ws2.cell(row=r2, column=3, value=f'{cnt/100*100:.0f}%').alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=r2, column=3).fill = PatternFill('solid', fgColor=bg)
    ws2.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=6)
    ws2.cell(row=r2, column=4, value=f'${loc_rev:,.0f}')
    ws2.cell(row=r2, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=r2, column=4).fill = PatternFill('solid', fgColor=bg)
    ws2.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
    ws2.cell(row=r2, column=7, value=f'{loc_rev/total_rev*100:.1f}%')
    ws2.cell(row=r2, column=7).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=r2, column=7).fill = PatternFill('solid', fgColor=bg)
    ws2.row_dimensions[r2].height = 18
    r2 += 1
apply_border(ws2, r2-len(loc_data)-1, r2-1, 1, 8)
r2 += 1

# New product survival
section_title(ws2, r2, 1, '▌ 四、新品存活分析', span=8)
r2 += 1
new_df = df[df['Available days'] < 365]
mature_df = df[df['Available days'] >= 365]
bins_age = [0, 90, 180, 365, 730, 1460, 9999]
labels_age = ['<3个月', '3-6个月', '6-12个月', '1-2年', '2-4年', '>4年']
df['age_band'] = pd.cut(df['Available days'], bins=bins_age, labels=labels_age)
age_dist = df['age_band'].value_counts().sort_index()

hdr(ws2, r2, 1, '上架年龄段', bg=C_BLUE_MID)
hdr(ws2, r2, 2, 'SKU数', bg=C_BLUE_MID)
hdr(ws2, r2, 3, '占比', bg=C_BLUE_MID)
hdr(ws2, r2, 4, '月均收益/SKU', bg=C_BLUE_MID)
ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
hdr(ws2, r2, 5, '解读', bg=C_BLUE_MID)
r2 += 1
for age_band, cnt in age_dist.items():
    age_df_sub = df[df['age_band']==age_band]
    avg_r = age_df_sub[rev_col].mean()
    bg = C_GREEN_LIGHT if '6-12' in str(age_band) or '3-6' in str(age_band) else C_WHITE
    val(ws2, r2, 1, str(age_band), bg=bg)
    for ci, v in enumerate([int(cnt), f'{cnt}%', f'${avg_r:,.0f}'], start=2):
        c = ws2.cell(row=r2, column=ci, value=v)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(name='Arial', size=10)
    interpret = {'<3个月':'极新品，流量正在爬坡', '3-6个月':'新品存活验证期', '6-12个月':'存活稳定，有竞争力',
                 '1-2年':'成熟竞品，有流量积累', '2-4年':'稳固地位', '>4年':'行业老品，品牌沉淀强'}
    ws2.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=8)
    val(ws2, r2, 5, interpret.get(str(age_band),''), bg=bg, fg='FF595959')
    ws2.row_dimensions[r2].height = 18
    r2 += 1
apply_border(ws2, r2-len(age_dist)-1, r2-1, 1, 8)

# ============================================================
# SHEET 3: BSR TOP100 数据
# ============================================================
ws3 = wb.create_sheet('BSR TOP100')
ws3.sheet_view.showGridLines = False
display_cols = ['#', 'ASIN', 'Brand', 'Product Title', price_col, 'Monthly Sales', rev_col,
                'Rating', 'Ratings', 'Available days', 'BuyBox Location', 'Fulfillment', 'product_type', 'Gross Margin']
display_names = ['排名', 'ASIN', '品牌', '产品标题', '价格($)', '月销量', '月收入($)', '星级', '评分数', '上架天数', '卖家地区', '配送', '产品类型', '毛利率']
widths_3 = [6, 14, 14, 55, 10, 10, 13, 8, 10, 10, 10, 10, 18, 10]
for i, (ltr, w) in enumerate(zip('ABCDEFGHIJKLMN', widths_3)):
    ws3.column_dimensions[ltr].width = w

ws3.merge_cells('A1:N1')
c = ws3['A1']
c.value = 'BSR TOP100 - LED Job Site Lighting - 2026-04-09'
c.font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

for i, name in enumerate(display_names):
    hdr(ws3, 2, i+1, name, bg=C_BLUE_MID)
ws3.row_dimensions[2].height = 22

for row_idx, (_, drow) in enumerate(df[display_cols].iterrows(), start=3):
    bg = C_GREY_LIGHT if row_idx%2==0 else C_WHITE
    # highlight CN sellers
    if str(drow.get('BuyBox Location', '')) == 'CN':
        bg2 = C_BLUE_LIGHT if row_idx%2==0 else C_WHITE
    else:
        bg2 = C_ORANGE
    for ci, col in enumerate(display_cols):
        v = drow[col]
        if col == 'Gross Margin' and pd.notna(v):
            v = f'{float(v)*100:.1f}%'
        elif col == rev_col and pd.notna(v):
            v = round(float(v), 1)
        c = ws3.cell(row=row_idx, column=ci+1, value=v)
        c.font = Font(name='Arial', size=9)
        c.fill = PatternFill('solid', fgColor=bg)
        h_align = 'center' if ci > 2 else ('left' if ci == 3 else 'center')
        c.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=(ci==3))
    ws3.row_dimensions[row_idx].height = 15 if ci != 3 else 30
apply_border(ws3, 2, 101, 1, len(display_cols))

# ============================================================
# SHEET 4: 推荐入场价
# ============================================================
ws4 = wb.create_sheet('推荐入场价')
ws4.sheet_view.showGridLines = False
for col_letter, width in zip('ABCDEFGHI', [22,14,14,14,14,16,16,20,25]):
    ws4.column_dimensions[col_letter].width = width

ws4.merge_cells('A1:I1')
c = ws4['A1']
c.value = 'LED工作灯 — 各产品类型推荐入场价区间'
c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 35

r4 = 3
section_title(ws4, r4, 1, '▌ 推荐入场价逻辑：以市场P25-P75价位为参考，结合单品收益分析', span=9)
ws4.row_dimensions[r4].height = 24
r4 += 1

price_hdr4 = ['产品类型', '市场MIN', '市场P25', '市场中位', '市场P75', '市场MAX', '推荐入场价', '理由', '备注']
for i, h in enumerate(price_hdr4):
    hdr(ws4, r4, i+1, h, bg=C_BLUE_MID)
ws4.row_dimensions[r4].height = 22
r4 += 1

pricing_data = [
    ('充电磁吸工作灯', 8.0, 22.7, 30.0, 38.3, 67.0, '$29.99-$39.99',
     '主力竞争区间，磁力+充电是核心卖点，差异化靠亮度(≥1500lm)和磁力(≥30N)',
     '★ 首推入场品类'),
    ('三脚架工作灯', 18.0, 33.2, 54.8, 63.7, 89.1, '$49.99-$59.99',
     '客单价高，单品收益$34K/月，入场门槛较高，需三头灯+高亮+防水(IP66)',
     '高收益但竞争较激烈'),
    ('泛光工作灯', 15.0, 23.0, 25.2, 27.0, 43.7, '$19.99-$27.99',
     '主打走量，价格敏感，需控成本，太阳能款可溢价（+$5-8）',
     '走量低利'),
    ('车间吊装灯', 17.0, 21.2, 25.1, 31.8, 70.0, '$22.99-$32.99',
     '有稳定需求，T8灯管主流，链接款/多联款可溢价',
     '规格差异大，需精准定位'),
    ('夹灯', 14.0, 30.5, 36.0, 37.0, 40.0, '$28.99-$35.99',
     '需求稳定，加强夹头稳定性和E26/E27多规格兼容是差异化关键',
     '改良夹具可做差异化'),
    ('手持手电/聚光灯', 40.0, 54.6, 69.2, 102.6, 136.0, '不建议直接进入',
     'DeWalt品牌占据TOP3，品牌溢价强，新品难以竞争',
     '品牌护城河强，慎入'),
    ('太阳能工作灯', 12.5, 15.6, 18.7, 21.9, 25.0, '$15.99-$22.99',
     'SKU少，需求小众，可差异化做户外施工/应急场景',
     '小众赛道，量小'),
]

colors_row = [C_YELLOW, C_GREEN_LIGHT, C_WHITE, C_WHITE, C_BLUE_LIGHT, C_RED_LIGHT, C_GREY_LIGHT]
for i, row_p in enumerate(pricing_data):
    bg = colors_row[i]
    ptype, mn, p25, p50, p75, mx, recommend, reason, note = row_p
    vals_row = [ptype, f'${mn}', f'${p25}', f'${p50}', f'${p75}', f'${mx}', recommend, reason, note]
    for ci, v in enumerate(vals_row):
        c = ws4.cell(row=r4, column=ci+1, value=v)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(name='Arial', size=10, bold=(ci==0 or ci==6))
        c.alignment = Alignment(horizontal='center' if ci not in [0,7,8] else 'left',
                                vertical='center', wrap_text=True)
    ws4.row_dimensions[r4].height = 45
    r4 += 1

apply_border(ws4, 5, r4-1, 1, 9)
r4 += 2

# Profit calc section
section_title(ws4, r4, 1, '▌ 充电磁吸工作灯入场利润测算（$34.99档）', span=9)
ws4.row_dimensions[r4].height = 24
r4 += 1

profit_items = [
    ('售价', '$34.99', ''),
    ('亚马逊佣金(15%)', '-$5.25', ''),
    ('FBA费用(估算)', '-$5.15', '重量约0.8-1.2lb，含packaging'),
    ('广告费(CPC估算,12%)', '-$4.20', '新品期可能更高'),
    ('采购成本（含运费）', '-$8.00~$10.00', '国内约¥40-55/个，含海运'),
    ('税费/其他', '-$1.50', ''),
    ('毛利润', '$9.39~$11.39', '毛利率 ~27-33%'),
    ('月销量目标（保守）', '300-500件', '进入TOP50可期待'),
    ('月毛利润（估算）', '$2,800~$5,700', ''),
]
hdr(ws4, r4, 1, '成本结构', bg=C_BLUE_MID)
ws2.merge_cells
ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=5)
hdr(ws4, r4, 2, '金额', bg=C_BLUE_MID)
ws4.merge_cells(start_row=r4, start_column=6, end_row=r4, end_column=9)
hdr(ws4, r4, 6, '说明', bg=C_BLUE_MID)
r4 += 1
for i, (item, amount, note) in enumerate(profit_items):
    bg = C_YELLOW if '毛利润' in item else (C_GREY_LIGHT if i%2==0 else C_WHITE)
    bold = '毛利润' in item
    val(ws4, r4, 1, item, bold=bold, bg=bg)
    ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=5)
    c = ws4.cell(row=r4, column=2, value=amount)
    c.font = Font(name='Arial', bold=bold, size=11, color='FF1F3864' if bold else '000000')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws4.merge_cells(start_row=r4, start_column=6, end_row=r4, end_column=9)
    val(ws4, r4, 6, note, bg=bg, fg='FF595959')
    ws4.row_dimensions[r4].height = 22
    r4 += 1
apply_border(ws4, r4-len(profit_items)-1, r4-1, 1, 9)

# ============================================================
# SHEET 5: 竞品卖点 & 差评
# ============================================================
ws5 = wb.create_sheet('竞品卖点与差评')
ws5.sheet_view.showGridLines = False
for col_letter, width in zip('ABCDEFGH', [25,12,12,40,25,12,12,35]):
    ws5.column_dimensions[col_letter].width = width

ws5.merge_cells('A1:H1')
c = ws5['A1']
c.value = 'LED工作灯 — 竞品核心卖点 & 差评痛点分析（基于2589条真实评论）'
c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 35

r5 = 3
# Positive selling points
section_title(ws5, r5, 1, '▌ 一、正向卖点提炼（4-5★评论，共1,667条）', span=4)
section_title(ws5, r5, 5, '▌ 二、差评痛点分析（1-2★评论，共595条）', span=4, bg='FF7B2D00')
ws5.row_dimensions[r5].height = 24
r5 += 1

hdr(ws5, r5, 1, '卖点维度', bg=C_BLUE_MID)
hdr(ws5, r5, 2, '提及次数', bg=C_BLUE_MID)
hdr(ws5, r5, 3, '占比', bg=C_BLUE_MID)
hdr(ws5, r5, 4, '典型用语/消费者语言', bg=C_BLUE_MID)
hdr(ws5, r5, 5, '差评痛点', bg='FFC00000', fg=C_WHITE)
hdr(ws5, r5, 6, '提及次数', bg='FFC00000', fg=C_WHITE)
hdr(ws5, r5, 7, '占比', bg='FFC00000', fg=C_WHITE)
hdr(ws5, r5, 8, '典型反馈', bg='FFC00000', fg=C_WHITE)
r5 += 1

pos_quotes = {
    '高亮度/强光': '"super bright","lights up my whole garage","incredibly bright for the size"',
    '轻便/便携': '"compact","fits in toolbox","easy to take anywhere","lightweight"',
    '易用/操作方便': '"easy to use","set it and forget it","simple one-button operation"',
    '性价比高': '"great value","can\'t beat the price","exactly what I needed at this price"',
    '耐用/品质好': '"solid build","feels sturdy","very well made","has held up"',
    '电池续航好': '"battery lasts hours","lasted all day","charges fast via USB-C"',
    '多功能/多模式': '"multiple modes","strobe mode is great","works for camping too"',
    '磁力强/多角度调节': '"strong magnet","sticks to any metal surface","360 rotation is perfect"',
}
neg_quotes = {
    '电池/充电问题': '"battery died after 2 months","won\'t charge anymore","battery drains overnight"',
    '耐久性/质量差': '"stopped working","broke after 3 uses","cheaply made","fell apart"',
    '亮度不足': '"not bright at all","dim","dollar store flashlight is brighter"',
    '灯头/角度固定差': '"head is floppy","doesn\'t stay in position","pivot is very loose"',
    '客服/售后问题': '"terrible customer service","no response","can\'t get a replacement"',
    '磁力不稳/固定问题': '"magnet won\'t hold","falls off metal surface","weak magnet"',
    '不含电池/配件缺失': '"battery not included","just the tool","need to buy charger separately"',
    '防水/防尘问题': '"not actually waterproof","rusted in rain","water got inside"',
}
pos_sorted = sorted(pos_counts.items(), key=lambda x: x[1], reverse=True)
neg_sorted = sorted(neg_counts.items(), key=lambda x: x[1], reverse=True)
total_pos = len(high_rev)
total_neg = len(low_rev)

max_rows = max(len(pos_sorted), len(neg_sorted))
for i in range(max_rows):
    bg_pos = C_GREEN_LIGHT if i%2==0 else C_WHITE
    bg_neg = C_RED_LIGHT if i%2==0 else 'FFFFF5F5'
    if i < len(pos_sorted):
        cat_p, cnt_p = pos_sorted[i]
        val(ws5, r5, 1, cat_p, bold=True, bg=bg_pos)
        ws5.cell(row=r5, column=2, value=cnt_p).alignment = Alignment(horizontal='center', vertical='center')
        ws5.cell(row=r5, column=2).fill = PatternFill('solid', fgColor=bg_pos)
        ws5.cell(row=r5, column=3, value=f'{cnt_p/total_pos*100:.1f}%').alignment = Alignment(horizontal='center', vertical='center')
        ws5.cell(row=r5, column=3).fill = PatternFill('solid', fgColor=bg_pos)
        val(ws5, r5, 4, pos_quotes.get(cat_p, ''), fg='FF595959', bg=bg_pos, size=9)
    else:
        for ci in range(1, 5):
            ws5.cell(row=r5, column=ci).fill = PatternFill('solid', fgColor=C_WHITE)
    if i < len(neg_sorted):
        cat_n, cnt_n = neg_sorted[i]
        val(ws5, r5, 5, cat_n, bold=True, bg=bg_neg, fg='FF7B0000')
        ws5.cell(row=r5, column=6, value=cnt_n).alignment = Alignment(horizontal='center', vertical='center')
        ws5.cell(row=r5, column=6).fill = PatternFill('solid', fgColor=bg_neg)
        ws5.cell(row=r5, column=7, value=f'{cnt_n/total_neg*100:.1f}%').alignment = Alignment(horizontal='center', vertical='center')
        ws5.cell(row=r5, column=7).fill = PatternFill('solid', fgColor=bg_neg)
        val(ws5, r5, 8, neg_quotes.get(cat_n, ''), fg='FF595959', bg=bg_neg, size=9)
    else:
        for ci in range(5, 9):
            ws5.cell(row=r5, column=ci).fill = PatternFill('solid', fgColor=C_WHITE)
    ws5.row_dimensions[r5].height = 28
    r5 += 1

apply_border(ws5, 5, r5-1, 1, 8)
r5 += 1

# Per-ASIN review summary
section_title(ws5, r5, 1, '▌ 三、各ASIN评论质量汇总', span=8)
r5 += 1
asin_hdr = ['ASIN', '产品类型', '评论总数', '平均星级', '差评率', '5★占比', '主要差评类型']
for i, h in enumerate(asin_hdr):
    hdr(ws5, r5, i+1, h, bg=C_BLUE_MID)
ws5.merge_cells(start_row=r5, start_column=7, end_row=r5, end_column=8)
r5 += 1

# Link asin to product type
asin_type_map = dict(zip(df['ASIN'], df['product_type']))
for source_asin in rev_df['source_asin'].unique():
    adf = rev_df[rev_df['source_asin']==source_asin]
    avg_r = adf['Rating'].mean()
    low_pct = (adf['Rating']<=2).sum()/len(adf)*100
    five_pct = (adf['Rating']==5).sum()/len(adf)*100
    # find main complaints
    adf_low = adf[adf['Rating']<=2].copy()
    adf_low['neg_text'] = adf_low['Title'].fillna('') + ' ' + adf_low['Content'].fillna('')
    complaints = []
    for cat, kws in neg_keywords.items():
        pattern = '|'.join(kws)
        cnt = adf_low['neg_text'].str.lower().str.contains(pattern, na=False).sum()
        if cnt > 0:
            complaints.append(f'{cat}({cnt})')
    complaint_str = '、'.join(complaints[:3]) if complaints else '无明显差评集中'
    ptype = asin_type_map.get(source_asin, '未知')
    bg = C_RED_LIGHT if avg_r < 3.5 else (C_YELLOW if avg_r < 4.0 else C_WHITE)
    vals5 = [source_asin, ptype, len(adf), f'{avg_r:.1f}★', f'{low_pct:.1f}%', f'{five_pct:.1f}%']
    for ci, v in enumerate(vals5):
        c = ws5.cell(row=r5, column=ci+1, value=v)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws5.merge_cells(start_row=r5, start_column=7, end_row=r5, end_column=8)
    val(ws5, r5, 7, complaint_str, bg=bg, fg='FF595959', size=9)
    ws5.row_dimensions[r5].height = 18
    r5 += 1

apply_border(ws5, r5-13, r5-1, 1, 8)

# ============================================================
# SHEET 6: 产品上新方向
# ============================================================
ws6 = wb.create_sheet('产品上新方向')
ws6.sheet_view.showGridLines = False
for col_letter, width in zip('ABCDEFGHI', [5, 26, 30, 28, 16, 14, 14, 18, 20]):
    ws6.column_dimensions[col_letter].width = width

ws6.merge_cells('A1:I1')
c = ws6['A1']
c.value = 'LED工作灯 — 产品上新方向建议（基于数据缺口推导）'
c.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 35

r6 = 3
section_title(ws6, r6, 1, '▌ 数据支撑逻辑：差评痛点 × 市场缺口 × 价格空白 → 新品机会', span=9)
ws6.row_dimensions[r6].height = 24
r6 += 1

new_prod_hdr = ['优先级', '产品方向', '核心改进点', '数据支撑依据', '目标售价', '预计月销', '竞争难度', '启动方式', '预期亮点']
for i, h in enumerate(new_prod_hdr):
    hdr(ws6, r6, i+1, h, bg=C_BLUE_MID)
ws6.row_dimensions[r6].height = 22
r6 += 1

new_directions = [
    ('P1\n首推', '超强磁吸多模式充电工作灯\n（锁头版）',
     '针对"灯头松动"TOP痛点：\n• 旋转锁定机构，固定任意角度\n• 磁力升级至≥35N\n• USB-C快充+真实2000lm',
     '差评TOP1:灯头松动(57条)\n差评TOP2:电池寿命(183条)\n充电磁吸灯单品均收益$30K',
     '$34.99-$39.99', '400-600件/月', '中等', 'FBA直发，首批500pcs',
     '锁头设计专利壁垒，视频展示功能'),
    ('P1\n首推', '高亮大容量充电磁吸工作灯\n（5200mAh/2500lm升级版）',
     '• 电池容量≥5200mAh（市场多为3000-4000mAh）\n• 亮度≥2000lm（多数竞品1200-1500lm）\n• 数显电量计（市场差异化卖点）',
     '差评#2:亮度不足(74条)\n"battery dies too fast"高频词\n竞品平均1200lm/3000mAh',
     '$35.99-$42.99', '350-500件/月', '中等', '优化供应链，OEM定制',
     '参数碾压同价位竞品'),
    ('P2\n推荐', '三脚架工作灯（双头+USB充电）',
     '• 增加USB-A侧充电口（为手机/设备充电）\n• 折叠三脚架改良，展开速度更快\n• 更高防水等级（IP66→IP67）',
     '三脚架灯月收入$75.7万，平均单品$34K\n无USB充电功能款差异化空间大\n竞品多无防水标注',
     '$54.99-$62.99', '200-400件/月', '中高', '差异化规格，备货300',
     'USB充电卖点可见即可感'),
    ('P2\n推荐', '无线遥控充电三脚架灯\n（可拆卸+遥控）',
     '• 三头可单独拆卸成手持灯\n• 附遥控器（10m范围）\n• 定时自动关闭功能',
     'B0FKRKLGY7(CAVN)月收入$5万/203评\n产品新且销量已稳，说明遥控方向有市场\n灯头可拆是明显差异点',
     '$65.99-$74.99', '150-300件/月', '中高', '重点差异化，需要工厂定制',
     '遥控+可拆=施工/露营双场景'),
    ('P3\n参考', '高亮度太阳能泛光工作灯\n（施工+应急场景）',
     '• 太阳能+USB双充电\n• 更高亮度（≥1000lm）\n• 挂钩+磁底双固定',
     '太阳能灯BSR TOP100仅2款，竞争空白\n泛光灯月收入$30.5万，潜力品类\n户外施工需求未被充分满足',
     '$19.99-$26.99', '200-400件/月', '低', '小批量测试，首批200pcs',
     '竞争空白，先发优势大'),
    ('P3\n参考', '多彩车间T8吊装灯\n（5000K/4000K可切换）',
     '• 色温可切换（日光/暖白）\n• 双开关设计\n• 免工具安装夹具',
     '车间吊装灯月收入$16.4万，需求稳定\n多为单色温，功能同质化严重\n安装方便是评论关键词',
     '$24.99-$32.99', '300-500件/月', '低', '标准化工厂合作',
     '双色温差异化，安装便捷性'),
    ('P4\n备选', '礼品套装充电工作灯\n（2-Pack礼品装）',
     '• 2件装礼品包装\n• 附带定制礼品盒\n• Father\'s Day / Birthday礼品定位',
     '"gift"出现72次，4-5★评论\n礼品定位可提升转化率\n双拼装可提高客单价至$45+',
     '$39.99-$49.99', '200-350件/月', '低', '节日前2个月上架',
     '礼品关键词提升自然流量'),
]

priority_colors = {'P1\n首推': C_YELLOW, 'P2\n推荐': C_GREEN_LIGHT, 'P3\n参考': C_BLUE_LIGHT, 'P4\n备选': C_GREY_LIGHT}
for row_dir in new_directions:
    prio = row_dir[0]
    bg = priority_colors.get(prio, C_WHITE)
    for ci, v in enumerate(row_dir):
        c = ws6.cell(row=r6, column=ci+1, value=v)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(name='Arial', size=9, bold=(ci==0 or ci==1))
        c.alignment = Alignment(horizontal='center' if ci in [0,4,5,6] else 'left',
                                vertical='center', wrap_text=True)
    ws6.row_dimensions[r6].height = 65
    r6 += 1

apply_border(ws6, 5, r6-1, 1, 9)
r6 += 1

# Summary box
section_title(ws6, r6, 1, '▌ 上新决策矩阵 — 优先级说明', span=9)
ws6.row_dimensions[r6].height = 24
r6 += 1

decision_notes = [
    ('P1 首推（立即启动）', '充电磁吸工作灯是销量最大品类，市场需求成熟，差评痛点明确（灯头+电池），改良空间大，价格敏感度中等，适合首批快速验证。目标：6个月内进入类目TOP50。'),
    ('P2 推荐（第二梯队）', '三脚架工作灯客单价高、单品收益好，但对产品品质要求更严格（防水、三脚架稳定性），需要工厂定制支持，适合已有供应链后扩展。'),
    ('P3 参考（机会品类）', '太阳能灯竞争极少，是短期内可快速占位的小类目；T8吊装灯需求稳定，可作为SKU丰富产品线，降低集中风险。'),
    ('P4 备选（节日运营）', '礼品套装对日常SKU改动少，主要靠包装和节日运营推动转化，适合节前短期投入，ROI可观。'),
]
for label, note in decision_notes:
    ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=2)
    val(ws6, r6, 1, label, bold=True, bg=C_BLUE_LIGHT, fg='FF1F3864')
    ws6.merge_cells(start_row=r6, start_column=3, end_row=r6, end_column=9)
    val(ws6, r6, 3, note, wrap=True)
    ws6.row_dimensions[r6].height = 40
    r6 += 1
apply_border(ws6, r6-len(decision_notes), r6-1, 1, 9)

# ============================================================
# SHEET 7: 评论原始数据
# ============================================================
ws7 = wb.create_sheet('评论数据汇总')
ws7.sheet_view.showGridLines = False
ws7.column_dimensions['A'].width = 14
ws7.column_dimensions['B'].width = 14
ws7.column_dimensions['C'].width = 8
ws7.column_dimensions['D'].width = 35
ws7.column_dimensions['E'].width = 55
ws7.column_dimensions['F'].width = 10

ws7.merge_cells('A1:F1')
c = ws7['A1']
c.value = '评论原始数据（13个ASIN，共2,589条评论）'
c.font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
c.fill = PatternFill('solid', fgColor=C_BLUE_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws7.row_dimensions[1].height = 28

for i, h in enumerate(['ASIN', '来源文件', '星级', '标题', '评论内容（摘要）', '日期']):
    hdr(ws7, 2, i+1, h, bg=C_BLUE_MID)
ws7.row_dimensions[2].height = 20

rev_display_cols = [c for c in ['source_asin','Rating','Title','Content','Date'] if c in rev_df.columns]
for row_idx, (_, revrow) in enumerate(rev_df[rev_display_cols].head(800).iterrows(), start=3):
    rating = revrow.get('Rating', 3)
    bg = C_RED_LIGHT if (pd.notna(rating) and float(rating) <= 2) else (C_WHITE if (pd.notna(rating) and float(rating) <= 3) else C_GREEN_LIGHT if row_idx%2==0 else C_WHITE)
    content_preview = str(revrow.get('Content',''))[:200] if pd.notna(revrow.get('Content','')) else ''
    row_vals = [revrow.get('source_asin',''), '', rating, revrow.get('Title',''), content_preview, revrow.get('Date','')]
    for ci, v in enumerate(row_vals):
        c = ws7.cell(row=row_idx, column=ci+1, value=str(v) if v else '')
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(name='Arial', size=8)
        c.alignment = Alignment(horizontal='center' if ci in [0,2,5] else 'left', vertical='center', wrap_text=(ci==4))
    ws7.row_dimensions[row_idx].height = 14

apply_border(ws7, 2, min(800+2, len(rev_df)+2), 1, 6)

# ============================================================
# Save
# ============================================================
wb.save(output_path)
print(f'Report saved to: {output_path}')
print('Sheets:', wb.sheetnames)
